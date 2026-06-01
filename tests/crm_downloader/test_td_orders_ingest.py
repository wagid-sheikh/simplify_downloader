from __future__ import annotations

import io
from datetime import datetime, timedelta
from decimal import Decimal
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
import pytest
import sqlalchemy as sa

from app.common.db import session_scope
from app.crm_downloader.td_orders_sync.ingest import (
    TdOrdersIngestResult,
    _expected_headers,
    _coerce_input_row,
    _orders_table,
    _stg_td_orders_table,
    ingest_td_orders_rows,
    ingest_td_orders_workbook,
)
from app.dashboard_downloader.json_logger import JsonLogger, get_logger


def test_coerce_input_row_maps_customer_phone_alias() -> None:
    raw = {
        "orderNo": "ORD-API-ALIAS",
        "orderDate": "2025-05-10 09:30",
        "customerPhone": "+91 88889-99762",
    }

    coerced = _coerce_input_row(raw)

    assert coerced["Order No."] == "ORD-API-ALIAS"
    assert coerced["Order Date / Time"] == "2025-05-10 09:30"
    assert coerced["Phone"] == "+91 88889-99762"



def _build_sample_workbook(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = list(_expected_headers())
    ws.append(headers)
    ws.append(
        [
            "2025-05-10 09:30",
            "ORD-001",
            "C001",
            "Alice",
            "123 Street",
            "+91 99999-88888",
            "Pickup",
            "",
            "2025-05-12 10:00",
            2,
            1.5,
            "1,200.00",
            "100",
            "50",
            "1050",
            None,
            500,
            "25.50",
            550,
            None,
            None,
            "Bob",
            "Note A",
            "Order note",
            "Yes",
            "North",
            "Charlie",
            "GST123",
            "App",
            "POS1",
            "Yes",
            "TypeA",
            "Pack1",
            "Positive",
            "tag1",
            "Comment",
            "Dry Clean",
            "Extra",
            "Pending",
            "2025-05-11 08:00",
            "Info",
            "CPN1",
        ]
    )
    ws.append(
        [
            "2025-05-11",
            "ORD-002",
            "C002",
            "Bob",
            "456 Lane",
            "12345",
            "Delivery",
            "2025-05-15",
            None,
            3,
            2.0,
            900,
            "0",
            "0",
            "900",
            0,
            0,
            "-10.25",
            900,
            0,
            0,
            "Dana",
            None,
            None,
            "No",
            None,
            None,
            None,
            None,
            "No",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ]
    )
    wb.save(path)
    return path


def _build_sample_workbook_with_invalid_data(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = list(_expected_headers())
    ws.append(headers)
    ws.append(
        [
            "2025-05-10 09:30",
            "ORD-001",
            "C001",
            "Alice",
            "123 Street",
            "A668--7051",
            "Pickup",
            "",
            "not-a-date",
            2,
            1.5,
            "1,200.00",
            "100",
            "50",
            "1050",
            None,
            500,
            0,
            550,
            None,
            None,
            "Bob",
            "Note A",
            "Order note",
            "Yes",
            "North",
            "Charlie",
            "GST123",
            "App",
            "POS1",
            "Yes",
            "TypeA",
            "Pack1",
            "Positive",
            "tag1",
            "Comment",
            "Dry Clean",
            "Extra",
            "Pending",
            "2025-05-11 08:00",
            "Info",
            "CPN1",
        ]
    )
    ws.append(
        [
            "2025-05-11",
            "ORD-002",
            "C002",
            "Bob",
            "456 Lane",
            "+91 99999-88888",
            "Delivery",
            "2025-05-15",
            None,
            3,
            2.0,
            900,
            "0",
            "0",
            "900",
            0,
            0,
            0,
            900,
            0,
            0,
            "Dana",
            None,
            None,
            "No",
            None,
            None,
            None,
            None,
            "No",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ]
    )
    wb.save(path)
    return path


def _build_sample_workbook_with_footer(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = list(_expected_headers())
    ws.append(headers)
    ws.append(
        [
            "2025-05-10 09:30",
            "ORD-001",
            "C001",
            "Alice",
            "123 Street",
            "+91 99999-88888",
            "Pickup",
            "",
            "2025-05-12 10:00",
            2,
            1.5,
            "1,200.00",
            "100",
            "50",
            "1050",
            None,
            500,
            0,
            550,
            None,
            None,
            "Bob",
            "Note A",
            "Order note",
            "Yes",
            "North",
            "Charlie",
            "GST123",
            "App",
            "POS1",
            "Yes",
            "TypeA",
            "Pack1",
            "Positive",
            "tag1",
            "Comment",
            "Dry Clean",
            "Extra",
            "Pending",
            "2025-05-11 08:00",
            "Info",
            "CPN1",
        ]
    )
    ws.append(
        [
            "2025-05-11",
            "ORD-002",
            "C002",
            "Bob",
            "456 Lane",
            "12345",
            "Delivery",
            "2025-05-15",
            None,
            3,
            2.0,
            900,
            "0",
            "0",
            "900",
            0,
            0,
            0,
            900,
            0,
            0,
            "Dana",
            None,
            None,
            "No",
            None,
            None,
            None,
            None,
            "No",
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
        ]
    )
    ws.append(["Total Order Summary", None, None, None, None, None, None, None, None, None, None])
    wb.save(path)
    return path


def _build_workbook_with_duplicate_bad_phones(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = list(_expected_headers())
    ws.append(headers)
    for idx in range(2):
        ws.append(
            [
                "2025-05-10 09:30",
                f"ORD-00{idx + 1}",
                "C001",
                "Alice",
                "123 Street",
                "12345",
                "Pickup",
                "",
                "2025-05-12 10:00",
                2,
                1.5,
                "1,200.00",
                "100",
                "50",
                "1050",
                None,
                500,
                0,
                550,
                None,
                None,
                "Bob",
                "Note A",
                "Order note",
                "Yes",
                "North",
                "Charlie",
                "GST123",
                "App",
                "POS1",
                "Yes",
                "TypeA",
                "Pack1",
                "Positive",
                "tag1",
                "Comment",
                "Dry Clean",
                "Extra",
                "Pending",
                "2025-05-11 08:00",
                "Info",
                "CPN1",
            ]
        )
    wb.save(path)
    return path


def _create_tables(database_url: str) -> None:
    metadata = sa.MetaData()
    _stg_td_orders_table(metadata)
    _orders_table(metadata)
    engine = sa.create_engine(database_url.replace("+aiosqlite", ""))
    metadata.create_all(engine)
    engine.dispose()


@pytest.mark.asyncio
async def test_ingest_td_orders_rows_api_payload(tmp_path: Path) -> None:
    db_path = tmp_path / "orders_rows.db"
    database_url = f"sqlite+aiosqlite:///{db_path}"
    _create_tables(database_url)

    tz = ZoneInfo("Asia/Kolkata")
    run_date = datetime(2025, 5, 20, 12, 0, tzinfo=tz)
    logger = get_logger(run_id="test_run")
    rows = [
        {
            "orderNo": "ORD-API-001",
            "orderDate": "2025-05-10 09:30",
            "customerCode": "C001",
            "customerName": "Alice",
            "mobileNumber": "+91 99999-88888",
            "grossAmount": "1200",
            "netAmount": "1100",
            "paid": "500",
        }
    ]

    result = await ingest_td_orders_rows(
        rows=rows,
        store_code="A668",
        cost_center="UN3668",
        run_id="test_run",
        run_date=run_date,
        database_url=database_url,
        logger=logger,
    )

    assert result.staging_rows == 1
    assert result.final_rows == 1
    assert result.rows_downloaded == 1
    assert result.warnings == []
    assert result.amount_metrics == {
        "explicit_zero_value_order_count": 0,
        "missing_amount_field_count": 0,
        "malformed_amount_field_count": 0,
        "canonical_zero_value_order_count": 0,
        "parsed_source_gross_amount_sum": Decimal("1200"),
        "parsed_source_net_amount_sum": Decimal("1100"),
    }

    metadata = sa.MetaData()
    stg_table = _stg_td_orders_table(metadata)
    orders_table = _orders_table(metadata)
    async with session_scope(database_url) as session:
        stg_row = (
            await session.execute(
                sa.select(stg_table.c.gross_amount, stg_table.c.net_amount).where(
                    stg_table.c.order_number == "ORD-API-001"
                )
            )
        ).one()
        order_row = (
            await session.execute(
                sa.select(orders_table.c.gross_amount, orders_table.c.net_amount).where(
                    orders_table.c.order_number == "ORD-API-001"
                )
            )
        ).one()
        await session.execute(
            sa.text(
                "CREATE VIEW vw_orders AS SELECT *, "
                "CASE WHEN net_amount IS NOT NULL AND net_amount <> 0 THEN net_amount ELSE gross_amount END "
                "AS order_amount FROM orders"
            )
        )
        canonical_amount = (
            await session.execute(
                sa.text("SELECT order_amount FROM vw_orders WHERE order_number = 'ORD-API-001'")
            )
        ).scalar_one()

    assert stg_row == (Decimal("1200.00"), Decimal("1100.00"))
    assert order_row == (Decimal("1200.00"), Decimal("1100.00"))
    assert canonical_amount == Decimal("1100")


@pytest.mark.asyncio
async def test_td_amount_provenance_preserves_explicit_zero_and_degrades_unusable_amounts(tmp_path: Path) -> None:
    db_path = tmp_path / "orders_amount_provenance.db"
    database_url = f"sqlite+aiosqlite:///{db_path}"
    _create_tables(database_url)

    result = await ingest_td_orders_rows(
        rows=[
            {"orderNo": "TD-ZERO", "orderDate": "2025-05-10", "customerPhone": "9999999999", "grossAmount": "0", "netAmount": 0},
            {"orderNo": "TD-MISSING", "orderDate": "2025-05-10", "customerPhone": "9999999999"},
            {"orderNo": "TD-MALFORMED", "orderDate": "2025-05-10", "customerPhone": "9999999999", "grossAmount": "bad", "netAmount": ""},
        ],
        store_code="A668",
        cost_center="UN3668",
        run_id="amount-provenance",
        run_date=datetime(2025, 5, 20, 12, 0, tzinfo=ZoneInfo("Asia/Kolkata")),
        database_url=database_url,
        logger=get_logger(run_id="amount-provenance"),
    )

    assert result.final_rows == 3
    assert result.amount_metrics == {
        "explicit_zero_value_order_count": 1,
        "missing_amount_field_count": 3,
        "malformed_amount_field_count": 1,
        "canonical_zero_value_order_count": 3,
        "parsed_source_gross_amount_sum": Decimal("0"),
        "parsed_source_net_amount_sum": Decimal("0"),
    }
    assert len(result.warning_rows) == 2
    assert any("provenance=absent" in warning for warning in result.warnings)
    assert any("provenance=malformed" in warning for warning in result.warnings)

    metadata = sa.MetaData()
    orders_table = _orders_table(metadata)
    async with session_scope(database_url) as session:
        persisted = (await session.execute(sa.select(orders_table.c.order_number, orders_table.c.gross_amount, orders_table.c.net_amount).order_by(orders_table.c.order_number))).all()
    assert persisted == [
        ("TD-MALFORMED", None, None),
        ("TD-MISSING", None, None),
        ("TD-ZERO", Decimal("0.00"), Decimal("0.00")),
    ]


@pytest.mark.asyncio
async def test_ingest_td_orders_workbook(tmp_path: Path) -> None:
    db_path = tmp_path / "orders.db"
    database_url = f"sqlite+aiosqlite:///{db_path}"
    _create_tables(database_url)

    workbook = _build_sample_workbook(tmp_path / "orders.xlsx")
    tz = ZoneInfo("Asia/Kolkata")
    run_date = datetime(2025, 5, 20, 12, 0, tzinfo=tz)
    logger = get_logger(run_id="test_run")

    result = await ingest_td_orders_workbook(
        workbook_path=workbook,
        store_code="A668",
        cost_center="UN3668",
        run_id="test_run",
        run_date=run_date,
        database_url=database_url,
        logger=logger,
    )

    assert isinstance(result, TdOrdersIngestResult)
    assert result.staging_rows == 2
    assert result.final_rows == 2
    assert len(result.warnings) == 1  # invalid phone in second row
    assert result.warnings == ["Invalid phone number fallback applied: 12345"]

    async with session_scope(database_url) as session:
        metadata = sa.MetaData()
        stg_table = _stg_td_orders_table(metadata)
        orders_table = _orders_table(metadata)
        staging_count = await session.scalar(sa.select(sa.func.count()).select_from(stg_table))
        final_count = await session.scalar(sa.select(sa.func.count()).select_from(orders_table))
        assert staging_count == 2
        assert final_count == 2

        stg_rows = (
            await session.execute(
                sa.select(stg_table.c.order_number, stg_table.c.adjustment).order_by(stg_table.c.order_number)
            )
        ).all()
        assert stg_rows[0].adjustment == Decimal("25.50")
        assert stg_rows[1].adjustment == Decimal("-10.25")

        orders_rows = (
            await session.execute(
                sa.select(
                    orders_table.c.order_number,
                    orders_table.c.mobile_number,
                    orders_table.c.adjustment,
                    orders_table.c.ingest_remarks,
                    orders_table.c.due_date,
                    orders_table.c.default_due_date,
                    orders_table.c.due_date_flag,
                ).order_by(orders_table.c.order_number)
            )
        ).all()
        ord1 = orders_rows[0]
        ord2 = orders_rows[1]
        assert ord1.order_number == "ORD-001"
        assert ord1.mobile_number == "9999988888"
        assert ord1.adjustment == Decimal("25.50")
        assert ord1.ingest_remarks is None
        assert ord1.due_date_flag == "Normal Delivery"
        assert ord2.mobile_number == "8888999762"  # invalid phone fallback applied
        assert ord2.adjustment == Decimal("-10.25")
        assert ord2.ingest_remarks == "MOBILE_FALLBACK_APPLIED"
        due_date = ord2.due_date if ord2.due_date.tzinfo else ord2.due_date.replace(tzinfo=tz)
        default_due_date = (
            ord2.default_due_date
            if ord2.default_due_date.tzinfo
            else ord2.default_due_date.replace(tzinfo=tz)
        )
        assert due_date == datetime(2025, 5, 15, tzinfo=tz)
        assert default_due_date == datetime(2025, 5, 14, tzinfo=tz)


@pytest.mark.asyncio
async def test_ingest_upsert_updates_existing(tmp_path: Path) -> None:
    db_path = tmp_path / "orders.db"
    database_url = f"sqlite+aiosqlite:///{db_path}"
    _create_tables(database_url)

    workbook = _build_sample_workbook(tmp_path / "orders.xlsx")
    tz = ZoneInfo("Asia/Kolkata")
    run_date = datetime(2025, 5, 20, 12, 0, tzinfo=tz)
    logger = get_logger(run_id="test_run")

    await ingest_td_orders_workbook(
        workbook_path=workbook,
        store_code="A668",
        cost_center="UN3668",
        run_id="test_run",
        run_date=run_date,
        database_url=database_url,
        logger=logger,
    )

    # modify workbook second row net amount to force update
    wb = openpyxl.load_workbook(workbook)
    ws = wb.active
    ws["O3"] = "950"  # Net Amount column
    wb.save(workbook)

    await ingest_td_orders_workbook(
        workbook_path=workbook,
        store_code="A668",
        cost_center="UN3668",
        run_id="test_run",
        run_date=run_date + timedelta(hours=1),
        database_url=database_url,
        logger=logger,
    )

    async with session_scope(database_url) as session:
        metadata = sa.MetaData()
        orders_table = _orders_table(metadata)
        net_amounts = (
            await session.execute(
                sa.select(orders_table.c.order_number, orders_table.c.net_amount).order_by(orders_table.c.order_number)
            )
        ).all()
        assert net_amounts[1].net_amount == 950


@pytest.mark.asyncio
async def test_footer_row_is_skipped(tmp_path: Path) -> None:
    db_path = tmp_path / "orders.db"
    database_url = f"sqlite+aiosqlite:///{db_path}"
    _create_tables(database_url)

    workbook = _build_sample_workbook_with_footer(tmp_path / "orders_with_footer.xlsx")

    tz = ZoneInfo("Asia/Kolkata")
    run_date = datetime(2025, 5, 20, 12, 0, tzinfo=tz)
    log_stream = io.StringIO()
    logger = JsonLogger(run_id="test_run", stream=log_stream, log_file_path=None)

    result = await ingest_td_orders_workbook(
        workbook_path=workbook,
        store_code="A668",
        cost_center="UN3668",
        run_id="test_run",
        run_date=run_date,
        database_url=database_url,
        logger=logger,
    )

    assert result.staging_rows == 2
    assert result.final_rows == 2
    # footer row should not add warnings beyond existing invalid phone warning and should not log warnings
    assert result.warnings == ["Invalid phone number fallback applied: 12345"]
    log_stream.seek(0)
    logs = log_stream.read().splitlines()
    assert logs == []


@pytest.mark.asyncio
async def test_ingest_remarks_populated_for_invalid_data(tmp_path: Path) -> None:
    db_path = tmp_path / "orders.db"
    database_url = f"sqlite+aiosqlite:///{db_path}"
    _create_tables(database_url)

    tz = ZoneInfo("Asia/Kolkata")
    run_date = datetime(2025, 5, 20, 12, 0, tzinfo=tz)
    log_stream = io.StringIO()
    logger = JsonLogger(run_id="test_run", stream=log_stream, log_file_path=None)

    workbook = _build_sample_workbook_with_invalid_data(tmp_path / "orders_invalid.xlsx")

    result = await ingest_td_orders_workbook(
        workbook_path=workbook,
        store_code="A668",
        cost_center="UN3668",
        run_id="test_run",
        run_date=run_date,
        database_url=database_url,
        logger=logger,
    )

    assert result.staging_rows == 2
    assert result.final_rows == 2

    async with session_scope(database_url) as session:
        metadata = sa.MetaData()
        stg_table = _stg_td_orders_table(metadata)
        orders_table = _orders_table(metadata)
        stg_rows = (
            await session.execute(
                sa.select(stg_table.c.order_number, stg_table.c.ingest_remarks).order_by(stg_table.c.order_number)
            )
        ).all()
        assert stg_rows[0].ingest_remarks == (
            "Field last_activity could not be parsed from value 'not-a-date' (field cleared); "
            "MOBILE_FALLBACK_APPLIED"
        )
        assert stg_rows[1].ingest_remarks is None
        orders_rows = (
            await session.execute(
                sa.select(orders_table.c.order_number, orders_table.c.ingest_remarks).order_by(orders_table.c.order_number)
            )
        ).all()
        assert orders_rows[0].ingest_remarks == (
            "Field last_activity could not be parsed from value 'not-a-date' (field cleared); "
            "MOBILE_FALLBACK_APPLIED"
        )
        assert orders_rows[1].ingest_remarks is None

    second_result = await ingest_td_orders_workbook(
        workbook_path=workbook,
        store_code="A668",
        cost_center="UN3668",
        run_id="test_run",
        run_date=run_date + timedelta(hours=1),
        database_url=database_url,
        logger=logger,
    )
    assert second_result.staging_rows == 2
    assert second_result.final_rows == 2
    assert second_result.ingest_remarks == [
        {
            "store_code": "A668",
            "order_number": "ORD-001",
            "ingest_remarks": (
                "Field last_activity could not be parsed from value 'not-a-date' (field cleared); "
                "MOBILE_FALLBACK_APPLIED"
            ),
        }
    ]

    async with session_scope(database_url) as session:
        metadata = sa.MetaData()
        orders_table = _orders_table(metadata)
        orders_rows = (
            await session.execute(
                sa.select(orders_table.c.order_number, orders_table.c.ingest_remarks).order_by(orders_table.c.order_number)
            )
        ).all()
        assert orders_rows[0].ingest_remarks == (
            "Field last_activity could not be parsed from value 'not-a-date' (field cleared); "
            "MOBILE_FALLBACK_APPLIED"
        )
        assert orders_rows[1].ingest_remarks is None


@pytest.mark.asyncio
async def test_duplicate_invalid_phone_warns_once(tmp_path: Path) -> None:
    db_path = tmp_path / "orders.db"
    database_url = f"sqlite+aiosqlite:///{db_path}"
    _create_tables(database_url)

    workbook = _build_workbook_with_duplicate_bad_phones(tmp_path / "orders_bad_phone.xlsx")

    tz = ZoneInfo("Asia/Kolkata")
    run_date = datetime(2025, 5, 20, 12, 0, tzinfo=tz)
    log_stream = io.StringIO()
    logger = JsonLogger(run_id="test_run", stream=log_stream, log_file_path=None)

    result = await ingest_td_orders_workbook(
        workbook_path=workbook,
        store_code="A668",
        cost_center="UN3668",
        run_id="test_run",
        run_date=run_date,
        database_url=database_url,
        logger=logger,
    )

    assert result.staging_rows == 2
    assert result.final_rows == 2
    assert result.warnings == ["Invalid phone number fallback applied: 12345"]
    log_stream.seek(0)
    assert log_stream.read().splitlines() == []
