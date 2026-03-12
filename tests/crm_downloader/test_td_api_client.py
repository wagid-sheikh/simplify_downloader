from __future__ import annotations

import asyncio
import json
from datetime import date
import logging
from datetime import datetime, timezone
from decimal import Decimal
import math
from pathlib import Path

import pytest

from app.crm_downloader.td_orders_sync.main import (
    _build_parser,
    _persist_compare_excel_artifact,
    _resolve_td_api_artifact_dir,
    _dataset_completion_health,
)
from app.crm_downloader.td_orders_sync.td_api_artifacts import (
    _validate_xlsx,
    _write_excel,
    persist_td_api_artifacts,
    persist_td_compare_artifacts,
)
from app.crm_downloader.td_orders_sync.td_api_client import (
    TdApiClient,
    TdApiClientConfig,
    _extract_rows,
    _filter_summary_rows,
)
from app.crm_downloader.td_orders_sync.td_api_compare import (
    COMPARE_KEY_FIELDS_BY_DATASET,
    compare_canonical_rows,
    project_api_rows_for_compare,
)


ORDERS_IGNORABLE_API_ONLY_FIELDS = {"bookingSlipUrl", "storeName", "deliveryDate"}
SALES_IGNORABLE_API_ONLY_FIELDS = {"printer", "storeName", "storeId"}


@pytest.fixture
def representative_orders_api_row() -> dict[str, str]:
    return {
        "orderNo": "O-1001",
        "orderDate": "2026-01-02 10:00:00",
        "amount": "120.50",
        "status": "Delivered",
        "bookingSlipUrl": "https://files.example/booking-slip/O-1001",
        "storeName": "Andheri",
        "deliveryDate": "2026-01-03",
    }


@pytest.fixture
def representative_sales_api_row() -> dict[str, str]:
    return {
        "orderNo": "S-2001",
        "paymentDate": "2026-01-02 11:00:00",
        "paymentMode": "UPI",
        "amount": "80.00",
        "status": "Collected",
        "printer": "FrontDesk-01",
        "storeName": "Andheri",
        "storeId": "A817",
    }


def test_source_mode_parser_accepts_api_modes() -> None:
    parser = _build_parser()
    args = parser.parse_args(["--source-mode", "api_shadow"])
    assert args.source_mode == "api_shadow"


def test_extract_rows_handles_common_response_shapes() -> None:
    assert _extract_rows([{"a": 1}, {"b": 2}]) == [{"a": 1}, {"b": 2}]
    assert _extract_rows({"items": [{"a": 1}], "x": 1}) == [{"a": 1}]
    assert _extract_rows({"data": {"rows": [{"x": 9}]}}) == [{"x": 9}]
    assert _extract_rows({"data": "bad"}) == []




def test_filter_summary_rows_removes_footer_like_rows() -> None:
    rows = [
        {"orderNumber": "A-1001", "customerName": "Alice", "totalAmount": "120.00"},
        {"orderNumber": None, "customerName": "Total Order", "totalAmount": "120.00"},
        {"orderNo": "S-2002", "paymentDate": "2026-01-02", "netAmount": "80.00"},
        {"orderNo": "", "description": "Grand Total", "netAmount": "200.00", "tax": "20.00"},
        {"orderNumber": 500, "orderDate": "Total Order", "netAmount": "200.00"},
    ]

    filtered_rows, summary_rows_filtered = _filter_summary_rows(rows)

    assert summary_rows_filtered == 3
    assert filtered_rows == [
        {"orderNumber": "A-1001", "customerName": "Alice", "totalAmount": "120.00"},
        {"orderNo": "S-2002", "paymentDate": "2026-01-02", "netAmount": "80.00"},
    ]


def test_api_client_reads_storage_state_artifact(tmp_path: Path) -> None:
    artifact = tmp_path / "store-state.json"
    artifact.write_text(json.dumps({"cookies": [{"name": "session"}], "origins": []}), encoding="utf-8")
    client = TdApiClient(store_code="a123", context=None, storage_state_path=artifact)  # type: ignore[arg-type]
    state = client.read_session_artifact()
    assert state["cookies"][0]["name"] == "session"


def test_storage_state_token_discovery_supports_json_wrapped_values_without_iframe(tmp_path: Path) -> None:
    artifact = tmp_path / "store-state.json"
    wrapped_token = "header.payload.signature"
    artifact.write_text(
        json.dumps(
            {
                "cookies": [{"name": "sessionid", "domain": ".quickdrycleaning.com"}],
                "origins": [
                    {
                        "origin": "https://reports.quickdrycleaning.com",
                        "localStorage": [
                            {
                                "name": "bootstrap_blob",
                                "value": json.dumps({"auth": {"reportingToken": wrapped_token}}),
                            }
                        ],
                    }
                ],
            }
        ),
        encoding="utf-8",
    )
    client = TdApiClient(store_code="a123", context=None, storage_state_path=artifact)  # type: ignore[arg-type]

    discovered = client._discover_token_from_storage_state()

    assert discovered.token == wrapped_token
    assert discovered.source == "storage_state"
    assert discovered.detail is not None


def test_compare_uses_identical_canonical_key_and_emits_mismatch_artifacts() -> None:
    ui_rows = [
        {
            "store_code": "A817",
            "order_number": "1001",
            "order_date": "2026-01-02 10:00:00",
            "amount": "12",
            "status": "Delivered",
        },
        {
            "store_code": "A817",
            "order_number": "1002",
            "order_date": "2026-01-02 11:00:00",
            "amount": "8",
            "status": "Delivered",
        },
    ]
    api_rows = [
        {
            "store_code": "A817",
            "order_number": "1001",
            "order_date": "2026-01-02T10:00:00+05:30",
            "amount": "12.00",
            "status": "delivered",
        },
        {
            "store_code": "A817",
            "order_number": "1003",
            "order_date": "2026-01-02T11:30:00+05:30",
            "amount": "8.00",
            "status": "delivered",
        },
    ]

    metrics = compare_canonical_rows(
        ui_rows=ui_rows,
        api_rows=api_rows,
        key_fields=COMPARE_KEY_FIELDS_BY_DATASET["orders"],
    ).as_dict()

    assert metrics["matched_rows"] == 1
    assert metrics["missing_in_api"] == 1
    assert metrics["missing_in_ui"] == 1
    assert metrics["amount_mismatches"] == 0
    assert metrics["status_mismatches"] == 0
    assert metrics["mismatch_artifacts"]["missing_in_api"][0]["key_components"]["order_number"] == "1002"
    assert metrics["mismatch_artifacts"]["missing_in_ui"][0]["key_components"]["order_number"] == "1003"


def test_compare_orders_missing_metrics_use_order_number_not_datetime() -> None:
    ui_rows = [
        {"store_code": "A817", "order_number": "1001", "order_date": "2026-01-02 10:00:00", "amount": "12"},
    ]
    api_rows = [
        {"store_code": "A817", "order_number": "1001", "order_date": "2026-01-02 10:05:00", "amount": "12"},
    ]

    metrics = compare_canonical_rows(
        ui_rows=ui_rows,
        api_rows=api_rows,
        key_fields=COMPARE_KEY_FIELDS_BY_DATASET["orders"],
    ).as_dict()

    assert metrics["missing_in_api"] == 0
    assert metrics["missing_in_ui"] == 0


def test_compare_sales_row_count_mismatch_does_not_report_missing_order() -> None:
    ui_rows = [
        {"store_code": "A817", "order_number": "2001", "payment_date": "2026-01-02 11:00:00", "payment_mode": "UPI"},
        {"store_code": "A817", "order_number": "2001", "payment_date": "2026-01-02 11:01:00", "payment_mode": "Card"},
    ]
    api_rows = [
        {"store_code": "A817", "order_number": "2001", "payment_date": "2026-01-02 11:02:00", "payment_mode": "UPI"},
    ]

    metrics = compare_canonical_rows(
        ui_rows=ui_rows,
        api_rows=api_rows,
        key_fields=COMPARE_KEY_FIELDS_BY_DATASET["sales"],
    ).as_dict()

    assert metrics["missing_in_api"] == 0
    assert metrics["missing_in_ui"] == 0
    assert metrics["mismatch_artifacts"]["sales_order_row_count_mismatches"] == [
        {"order_number": "2001", "ui_row_count": 2, "api_row_count": 1}
    ]








def test_compare_orders_aliases_resolve_api_native_amount_and_status_names() -> None:
    ui_rows = [
        {
            "store_code": "A817",
            "order_number": "5001",
            "order_date": "2026-01-02 10:00:00",
            "amount": "1200",
            "status": "Delivered",
        }
    ]
    api_rows = [
        {
            "store_code": "A817",
            "order_number": "5001",
            "order_date": "2026-01-02 10:00:00",
            "grossAmount": 1200.0,
            "orderStatus": "delivered",
        }
    ]

    metrics = compare_canonical_rows(
        ui_rows=ui_rows,
        api_rows=api_rows,
        key_fields=COMPARE_KEY_FIELDS_BY_DATASET["orders"],
    ).as_dict()

    assert metrics["amount_mismatches"] == 0
    assert metrics["status_mismatches"] == 0


def test_compare_sales_aliases_resolve_payment_received_and_order_status_names() -> None:
    ui_rows = [
        {
            "store_code": "A817",
            "order_number": "7001",
            "payment_date": "2026-01-02 11:00:00",
            "payment_mode": "UPI",
            "amount": "1200",
            "status": "Collected",
        }
    ]
    api_rows = [
        {
            "store_code": "A817",
            "order_number": "7001",
            "payment_date": "2026-01-02 11:00:00",
            "payment_mode": "UPI",
            "paymentReceived": 1200.0,
            "orderStatus": "collected",
        }
    ]

    metrics = compare_canonical_rows(
        ui_rows=ui_rows,
        api_rows=api_rows,
        key_fields=COMPARE_KEY_FIELDS_BY_DATASET["sales"],
    ).as_dict()

    assert metrics["amount_mismatches"] == 0
    assert metrics["status_mismatches"] == 0

def test_orders_projection_contract_required_fields_and_allow_list(
    representative_orders_api_row: dict[str, str],
) -> None:
    projected = project_api_rows_for_compare(
        dataset="orders",
        api_rows=[representative_orders_api_row],
        store_code="a817",
    )[0]

    required_fields = {"store_code", "order_number", "order_date", "amount", "status"}
    assert required_fields.issubset(projected.keys())
    assert projected["store_code"] == "A817"
    assert projected["order_number"] == representative_orders_api_row["orderNo"]
    assert projected["order_date"] == representative_orders_api_row["orderDate"]
    assert projected["amount"] == representative_orders_api_row["amount"]
    assert projected["status"] == representative_orders_api_row["status"]

    ui_equivalent_api_keys = {
        "orderNo",
        "orderDate",
        "amount",
        "status",
    }
    allow_list = ORDERS_IGNORABLE_API_ONLY_FIELDS
    unmapped_required = [
        key
        for key in representative_orders_api_row.keys()
        if key not in ui_equivalent_api_keys and key not in allow_list
    ]
    assert unmapped_required == [], f"Orders API keys require mapping or allow-listing: {unmapped_required}"


def test_sales_projection_contract_required_fields_and_allow_list(
    representative_sales_api_row: dict[str, str],
) -> None:
    projected = project_api_rows_for_compare(
        dataset="sales",
        api_rows=[representative_sales_api_row],
        store_code="a817",
    )[0]

    required_fields = {"store_code", "order_number", "payment_date", "payment_mode", "amount", "status"}
    assert required_fields.issubset(projected.keys())
    assert projected["store_code"] == "A817"
    assert projected["order_number"] == representative_sales_api_row["orderNo"]
    assert projected["payment_date"] == representative_sales_api_row["paymentDate"]
    assert projected["payment_mode"] == representative_sales_api_row["paymentMode"]
    assert projected["amount"] == representative_sales_api_row["amount"]
    assert projected["status"] == representative_sales_api_row["status"]

    ui_equivalent_api_keys = {
        "orderNo",
        "paymentDate",
        "paymentMode",
        "amount",
        "status",
    }
    allow_list = SALES_IGNORABLE_API_ONLY_FIELDS
    unmapped_required = [
        key
        for key in representative_sales_api_row.keys()
        if key not in ui_equivalent_api_keys and key not in allow_list
    ]
    assert unmapped_required == [], f"Sales API keys require mapping or allow-listing: {unmapped_required}"


def test_projection_round_trip_preserves_required_contract_fields(
    representative_orders_api_row: dict[str, str],
    representative_sales_api_row: dict[str, str],
) -> None:
    orders_projected = project_api_rows_for_compare(
        dataset="orders",
        api_rows=[representative_orders_api_row],
        store_code="a817",
    )[0]
    sales_projected = project_api_rows_for_compare(
        dataset="sales",
        api_rows=[representative_sales_api_row],
        store_code="a817",
    )[0]

    orders_api_path_output = {**representative_orders_api_row, **orders_projected}
    sales_api_path_output = {**representative_sales_api_row, **sales_projected}

    assert orders_api_path_output["order_number"] == representative_orders_api_row["orderNo"]
    assert orders_api_path_output["order_date"] == representative_orders_api_row["orderDate"]
    assert orders_api_path_output["amount"] == representative_orders_api_row["amount"]
    assert orders_api_path_output["status"] == representative_orders_api_row["status"]

    assert sales_api_path_output["order_number"] == representative_sales_api_row["orderNo"]
    assert sales_api_path_output["payment_date"] == representative_sales_api_row["paymentDate"]
    assert sales_api_path_output["payment_mode"] == representative_sales_api_row["paymentMode"]
    assert sales_api_path_output["amount"] == representative_sales_api_row["amount"]
    assert sales_api_path_output["status"] == representative_sales_api_row["status"]

def test_project_api_rows_for_compare_isolated_projection() -> None:
    projected = project_api_rows_for_compare(
        dataset="orders",
        api_rows=[{"orderNo": "1001", "orderDate": "2026-01-02 10:00:00", "amount": "12", "status": "Delivered", "sourceOnly": "keep"}],
        store_code="a817",
    )

    assert projected == [
        {
            "store_code": "A817",
            "order_number": "1001",
            "order_date": "2026-01-02 10:00:00",
            "amount": "12",
            "status": "Delivered",
        }
    ]

def test_resolve_td_api_artifact_dir_defaults_and_env_override(monkeypatch: pytest.MonkeyPatch, tmp_path: Path) -> None:
    monkeypatch.delenv("TD_API_ARTIFACT_DIR", raising=False)
    from app.crm_downloader.config import default_download_dir

    assert _resolve_td_api_artifact_dir() == default_download_dir().resolve()

    custom_dir = tmp_path / "custom-artifacts"
    monkeypatch.setenv("TD_API_ARTIFACT_DIR", str(custom_dir))
    assert _resolve_td_api_artifact_dir() == custom_dir.resolve()


def test_compare_excel_summary_excludes_nested_mismatch_artifacts(tmp_path: Path) -> None:
    artifact_path = tmp_path / "compare.xlsx"
    compare_metrics = {
        "total_rows": 2,
        "matched_rows": 1,
        "missing_in_api": 1,
        "missing_in_ui": 0,
        "amount_mismatches": 0,
        "status_mismatches": 0,
        "sample_mismatch_keys": "A817|1002",
        "mismatch_artifacts": {"missing_in_api": [{"order_number": "1002"}]},
    }

    _persist_compare_excel_artifact(
        artifact_path=artifact_path,
        compare_metrics=compare_metrics,
        api_request_metadata=[],
    )

    import openpyxl

    workbook = openpyxl.load_workbook(artifact_path)
    summary = workbook["summary"]
    headers = [cell.value for cell in next(summary.iter_rows(min_row=1, max_row=1))]
    assert "mismatch_artifacts" not in headers


def test_compare_excel_flattens_list_and_dict_cells(tmp_path: Path) -> None:
    artifact_path = tmp_path / "compare.xlsx"
    compare_metrics = {
        "total_rows": 3,
        "matched_rows": 1,
        "missing_in_api": 1,
        "missing_in_ui": 1,
        "amount_mismatches": 0,
        "status_mismatches": 1,
        "sample_mismatch_keys": ["A817|1002", "A817|1003"],
        "mismatch_artifacts": {
            "missing_in_api": [
                {
                    "key": "A817|1002",
                    "key_components": {"store_code": "A817", "order_number": "1002"},
                }
            ]
        },
    }

    _persist_compare_excel_artifact(
        artifact_path=artifact_path,
        compare_metrics=compare_metrics,
        api_request_metadata=[],
    )

    import openpyxl

    workbook = openpyxl.load_workbook(artifact_path)
    summary = workbook["summary"]
    summary_headers = [cell.value for cell in next(summary.iter_rows(min_row=1, max_row=1))]
    summary_row = [cell.value for cell in next(summary.iter_rows(min_row=2, max_row=2))]
    sample_keys_value = summary_row[summary_headers.index("sample_mismatch_keys")]
    assert sample_keys_value == '["A817|1002", "A817|1003"]'

    missing_in_api = workbook["missing_in_api"]
    row_headers = [cell.value for cell in next(missing_in_api.iter_rows(min_row=1, max_row=1))]
    row_values = [cell.value for cell in next(missing_in_api.iter_rows(min_row=2, max_row=2))]
    key_components_value = row_values[row_headers.index("key_components")]
    assert key_components_value == '{"order_number": "1002", "store_code": "A817"}'



def test_compare_excel_save_succeeds_with_non_empty_sample_mismatch_keys_list(tmp_path: Path) -> None:
    artifact_path = tmp_path / "compare.xlsx"
    compare_metrics = {
        "total_rows": 3,
        "matched_rows": 1,
        "missing_in_api": 1,
        "missing_in_ui": 1,
        "amount_mismatches": 1,
        "status_mismatches": 0,
        "sample_mismatch_keys": ["A817|1002", "A817|1003"],
        "mismatch_artifacts": {
            "missing_in_api": [
                {
                    "key": "A817|1002",
                    "context": {"store_code": "A817", "order_number": "1002"},
                }
            ],
            "missing_in_ui": [
                {
                    "key": "A817|1003",
                    "context": {"store_code": "A817", "order_number": "1003"},
                }
            ],
            "value_mismatches": [
                {
                    "key": "A817|1004",
                    "differences": ["amount", "status"],
                }
            ],
        },
    }

    _persist_compare_excel_artifact(
        artifact_path=artifact_path,
        compare_metrics=compare_metrics,
        api_request_metadata=[
            {
                "endpoint": "/reports/order-report",
                "request_payload": {"filters": ["orders", "delivery"]},
            }
        ],
    )

    import openpyxl

    assert artifact_path.exists()
    workbook = openpyxl.load_workbook(artifact_path)
    assert workbook["summary"].max_row >= 2






def test_persist_td_compare_artifacts_includes_endpoint_health_summary(tmp_path: Path) -> None:
    result = persist_td_compare_artifacts(
        download_dir=tmp_path,
        store_code="a817",
        from_date=date(2026, 1, 1),
        to_date=date(2026, 1, 2),
        orders_compare_metrics={"strict_verdict_ready": False, "dataset_health": {"ready": False}},
        sales_compare_metrics={"strict_verdict_ready": True, "dataset_health": {"ready": True}},
        endpoint_health_summary={"orders": {"ready": False, "degraded_reason": "http_401"}},
    )

    orders_payload = json.loads(Path(result.artifact_paths["orders_compare_mismatches"]).read_text(encoding="utf-8"))
    sales_payload = json.loads(Path(result.artifact_paths["sales_compare_mismatches"]).read_text(encoding="utf-8"))

    assert orders_payload["endpoint_health_summary"]["orders"]["degraded_reason"] == "http_401"
    assert sales_payload["endpoint_health_summary"]["orders"]["ready"] is False



def test_compare_excel_redacts_sensitive_api_request_metadata(tmp_path: Path) -> None:
    artifact_path = tmp_path / "compare.xlsx"
    _persist_compare_excel_artifact(
        artifact_path=artifact_path,
        compare_metrics={"total_rows": 1},
        api_request_metadata=[
            {
                "endpoint": "/reports/order-report",
                "query_params": {"token": ["secret-token"]},
                "headers": {"Authorization": "Bearer abc", "cookie": "sid=123"},
            }
        ],
    )

    import openpyxl

    workbook = openpyxl.load_workbook(artifact_path)
    metadata_sheet = workbook["api_request_metadata"]
    values = [cell.value for cell in next(metadata_sheet.iter_rows(min_row=2, max_row=2))]
    serialized = " ".join(str(value) for value in values if value is not None)
    assert "secret-token" not in serialized
    assert "Bearer abc" not in serialized
    assert "sid=123" not in serialized
    assert "***REDACTED***" in serialized


def test_persist_td_compare_artifacts_redacts_tokens_in_mismatch_payload(tmp_path: Path) -> None:
    result = persist_td_compare_artifacts(
        download_dir=tmp_path,
        store_code="a817",
        from_date=date(2026, 1, 1),
        to_date=date(2026, 1, 2),
        orders_compare_metrics={
            "strict_verdict_ready": False,
            "dataset_health": {"ready": False},
            "mismatch_artifacts": {
                "request_metadata": [
                    {
                        "query_params": {"token": ["visible-token"]},
                        "headers": {"authorization": "Bearer not-safe"},
                    }
                ]
            },
        },
        sales_compare_metrics={"strict_verdict_ready": True, "dataset_health": {"ready": True}},
        endpoint_health_summary={"diagnostics": {"set-cookie": "auth=clear"}},
    )

    orders_payload_text = Path(result.artifact_paths["orders_compare_mismatches"]).read_text(encoding="utf-8")
    assert "visible-token" not in orders_payload_text
    assert "not-safe" not in orders_payload_text
    assert "auth=clear" not in orders_payload_text
    assert "***REDACTED***" in orders_payload_text



def test_persist_td_api_artifacts_excel_rows_match_input_without_injected_columns(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("TD_API_HUMAN_READABLE_EXPORT", "true")
    order_rows = [
        {"order_number": "1001", "amount": "12.00", "nested": {"a": 1}, "notes": "hello\x00world"},
        {"order_number": "1002", "amount": "25.25", "nested": ["x", 2], "notes": "ok"},
    ]
    sale_rows = [{"sale_id": "S1", "amount": "10.00", "status": "Collected"}]
    garments_rows = [{"garment_id": "G1", "qty": 2, "meta": {"color": "blue"}}]

    result = persist_td_api_artifacts(
        download_dir=tmp_path,
        store_code="a817",
        from_date=date(2026, 1, 1),
        to_date=date(2026, 1, 2),
        raw_orders={"data": []},
        raw_sales={"data": []},
        raw_garments={"data": []},
        order_rows=order_rows,
        sale_rows=sale_rows,
        garments_rows=garments_rows,
    )

    import openpyxl

    workbook = openpyxl.load_workbook(result.artifact_paths["orders_excel"])
    sheet = workbook["orders"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    values_row1 = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]
    values_row2 = [cell.value for cell in next(sheet.iter_rows(min_row=3, max_row=3))]

    assert headers == ["order_number", "amount", "nested", "notes"]
    assert values_row1 == ["1001", "12.00", '{"a": 1}', "helloworld"]
    assert values_row2 == ["1002", "25.25", '["x", 2]', "ok"]


def test_persist_td_api_artifacts_rejects_compare_and_metadata_payload_shapes(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("TD_API_HUMAN_READABLE_EXPORT", "true")

    result_mapping_payload = persist_td_api_artifacts(
        download_dir=tmp_path,
        store_code="a817",
        from_date=date(2026, 1, 1),
        to_date=date(2026, 1, 2),
        raw_orders={"data": []},
        raw_sales={"data": []},
        raw_garments={"data": []},
        order_rows={"compare_metrics": {"matched_rows": 1}},  # type: ignore[arg-type]
        sale_rows=[{"sale_id": "S1"}],
        garments_rows=[{"garment_id": "G1"}],
    )

    assert "orders_excel" not in result_mapping_payload.artifact_paths
    assert any("API artifact purity violation" in warning for warning in result_mapping_payload.warnings)

    result_row_envelope = persist_td_api_artifacts(
        download_dir=tmp_path,
        store_code="a817",
        from_date=date(2026, 1, 1),
        to_date=date(2026, 1, 2),
        raw_orders={"data": []},
        raw_sales={"data": []},
        raw_garments={"data": []},
        order_rows=[{"api_request_metadata": [{"url": "x"}], "order_number": "1001"}],
        sale_rows=[{"sale_id": "S1"}],
        garments_rows=[{"garment_id": "G1"}],
    )

    assert "orders_excel" not in result_row_envelope.artifact_paths
    assert any("API artifact purity violation" in warning for warning in result_row_envelope.warnings)


def test_persist_td_api_artifacts_excel_contains_only_rows_sheet(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setenv("TD_API_HUMAN_READABLE_EXPORT", "true")

    result = persist_td_api_artifacts(
        download_dir=tmp_path,
        store_code="a817",
        from_date=date(2026, 1, 1),
        to_date=date(2026, 1, 2),
        raw_orders={"data": []},
        raw_sales={"data": []},
        raw_garments={"data": []},
        order_rows=[{"order_number": "1001"}],
        sale_rows=[{"sale_id": "S1"}],
        garments_rows=[{"garment_id": "G1"}],
    )

    import openpyxl

    orders_workbook = openpyxl.load_workbook(result.artifact_paths["orders_excel"])
    sales_workbook = openpyxl.load_workbook(result.artifact_paths["sales_excel"])

    assert orders_workbook.sheetnames == ["orders"]
    assert sales_workbook.sheetnames == ["sales"]
    assert "api_request_metadata" not in orders_workbook.sheetnames
    assert "compare_metrics" not in orders_workbook.sheetnames
def test_persist_td_api_artifacts_human_readable_export_toggle(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    base_kwargs = dict(
        download_dir=tmp_path,
        store_code="a817",
        from_date=date(2026, 1, 1),
        to_date=date(2026, 1, 2),
        raw_orders={"token": "secret", "data": []},
        raw_sales={"data": []},
        raw_garments={"data": []},
        order_rows=[{"order_number": "1001", "source_only": "x", "amount": "12.00"}],
        sale_rows=[{"order_number": "1001", "sales_only": "y", "amount": "12.00"}],
        garments_rows=[{"order_number": "1001", "garment_only": "z", "line_item_key": "L1"}],
    )

    monkeypatch.setenv("TD_API_HUMAN_READABLE_EXPORT", "false")
    disabled_result = persist_td_api_artifacts(**base_kwargs)

    assert disabled_result.human_readable_export_enabled is False
    assert disabled_result.human_readable_artifact_paths == []
    assert "orders_excel" not in disabled_result.artifact_paths
    assert "sales_excel" not in disabled_result.artifact_paths
    assert "garments_excel" not in disabled_result.artifact_paths

    orders_raw_path = Path(disabled_result.artifact_paths["orders_raw"])
    assert orders_raw_path.exists()
    orders_raw_text = orders_raw_path.read_text(encoding="utf-8")
    assert "***REDACTED***" in orders_raw_text
    assert "secret" not in orders_raw_text

    monkeypatch.setenv("TD_API_HUMAN_READABLE_EXPORT", "true")
    enabled_result = persist_td_api_artifacts(**base_kwargs)

    assert enabled_result.human_readable_export_enabled is True
    assert len(enabled_result.human_readable_artifact_paths) == 3
    assert "orders_excel" in enabled_result.artifact_paths
    assert "sales_excel" in enabled_result.artifact_paths
    assert "garments_excel" in enabled_result.artifact_paths
    assert Path(enabled_result.artifact_paths["orders_excel"]).exists()
    assert Path(enabled_result.artifact_paths["sales_excel"]).exists()
    assert Path(enabled_result.artifact_paths["garments_excel"]).exists()
    assert Path(enabled_result.artifact_paths["orders_raw"]).exists()


def test_td_api_artifact_write_excel_serializes_nested_non_scalar_values(tmp_path: Path) -> None:
    artifact_path = tmp_path / "td-artifact.xlsx"
    _write_excel(
        artifact_path,
        [
            {
                "plain": "ok",
                "nested": {1: "one", "items": ["a", {2: "two"}]},
                "tuple_value": ("x", {3: "three"}),
                "set_value": {"b", "a"},
            }
        ],
    )

    import openpyxl

    workbook = openpyxl.load_workbook(artifact_path)
    sheet = workbook["rows"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    values = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]

    nested_text = values[headers.index("nested")]
    tuple_text = values[headers.index("tuple_value")]
    set_text = values[headers.index("set_value")]

    assert nested_text == '{"1": "one", "items": ["a", {"2": "two"}]}'
    assert tuple_text == '["x", {"3": "three"}]'
    assert set_text == '["a", "b"]'


def test_td_api_artifact_write_excel_sanitizes_control_characters_and_reopens(tmp_path: Path) -> None:
    artifact_path = tmp_path / "td-control-chars.xlsx"
    _write_excel(
        artifact_path,
        [{"notes": "hello\x00world\x01!", "payload": {"text": "a\x0bb\x0cc"}}],
    )

    import openpyxl

    workbook = openpyxl.load_workbook(artifact_path)
    sheet = workbook["rows"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    values = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]

    assert values[headers.index("notes")] == "helloworld!"
    assert values[headers.index("payload")] == '{"text": "abc"}'


def test_td_api_artifact_write_excel_strips_xml_noncharacters(tmp_path: Path) -> None:
    artifact_path = tmp_path / "td-invalid-xml.xlsx"
    _write_excel(
        artifact_path,
        [{"notes": "prefixmiddle￿suffix", "payload": {"text": "ok￿done"}}],
    )

    import openpyxl

    workbook = openpyxl.load_workbook(artifact_path)
    sheet = workbook["rows"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    values = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]

    assert values[headers.index("notes")] == "prefixmiddlesuffix"
    assert values[headers.index("payload")] == '{"text": "okdone"}'

    empty_artifact_path = tmp_path / "td-empty.xlsx"
    _write_excel(empty_artifact_path, [])
    empty_workbook = openpyxl.load_workbook(empty_artifact_path)
    empty_sheet = empty_workbook["rows"]
    assert [cell.value for cell in next(empty_sheet.iter_rows(min_row=1, max_row=1))] == ["status"]
    assert [cell.value for cell in next(empty_sheet.iter_rows(min_row=2, max_row=2))] == ["no rows"]


def test_td_api_artifact_write_excel_coerces_problematic_scalar_types(tmp_path: Path) -> None:
    artifact_path = tmp_path / "td-coerced-scalars.xlsx"
    _write_excel(
        artifact_path,
        [
            {
                "tz_dt": datetime(2026, 1, 2, 3, 4, 5, tzinfo=timezone.utc),
                "nan_value": math.nan,
                "inf_value": math.inf,
                "neg_inf_value": -math.inf,
                "decimal_value": Decimal("123.4500"),
                "bytes_value": b"hello\xffworld",
                "unknown": Path("alpha\x00beta"),
            }
        ],
    )

    import openpyxl

    workbook = openpyxl.load_workbook(artifact_path)
    sheet = workbook["rows"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    values = [cell.value for cell in next(sheet.iter_rows(min_row=2, max_row=2))]

    assert values[headers.index("tz_dt")] == "2026-01-02T03:04:05+00:00"
    assert values[headers.index("nan_value")] == "NaN"
    assert values[headers.index("inf_value")] == "Infinity"
    assert values[headers.index("neg_inf_value")] == "-Infinity"
    assert values[headers.index("decimal_value")] == "123.4500"
    assert values[headers.index("bytes_value")] == "hello�world"
    assert values[headers.index("unknown")] == "alphabeta"


def test_td_api_artifact_write_excel_logs_coercion_type_summary(tmp_path: Path, caplog: pytest.LogCaptureFixture) -> None:
    artifact_path = tmp_path / "td-coercion-logging.xlsx"
    caplog.set_level(logging.DEBUG, logger="app.crm_downloader.td_orders_sync.td_api_artifacts")

    _write_excel(
        artifact_path,
        [
            {
                "d1": Decimal("1"),
                "d2": Decimal("2"),
                "b1": b"x",
                "dt": datetime(2026, 1, 2, 3, 4, 5, tzinfo=timezone.utc),
            }
        ],
    )

    messages = [record for record in caplog.records if "Excel cell serialization coerced" in record.message]
    assert len(messages) == 1
    assert "type counts={'Decimal': 2, 'bytes': 1, 'datetime': 1}" in messages[0].message




def test_td_api_artifact_write_excel_uses_atomic_replace(tmp_path: Path) -> None:
    artifact_path = tmp_path / "td-atomic.xlsx"
    _write_excel(artifact_path, [{"value": "first"}])

    original_bytes = artifact_path.read_bytes()

    from app.crm_downloader.td_orders_sync import td_api_artifacts

    real_replace = Path.replace

    def _failing_replace(self: Path, target: Path) -> Path:
        if self.name.endswith(".tmp.xlsx") and Path(target) == artifact_path:
            raise OSError("replace failed")
        return real_replace(self, target)

    monkeypatch = pytest.MonkeyPatch()
    monkeypatch.setattr(Path, "replace", _failing_replace)
    try:
        with pytest.raises(OSError, match="replace failed"):
            td_api_artifacts._write_excel(artifact_path, [{"value": "second"}])
    finally:
        monkeypatch.undo()

    assert artifact_path.exists()
    assert artifact_path.read_bytes() == original_bytes
    assert not list(tmp_path.glob("*.tmp.xlsx"))




def test_td_api_artifact_validate_xlsx_detects_openpyxl_roundtrip_failure(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    artifact_path = tmp_path / "td-roundtrip.xlsx"
    _write_excel(artifact_path, [{"value": "ok"}])

    import openpyxl

    real_load_workbook = openpyxl.load_workbook

    def _failing_load_workbook(*args, **kwargs):  # type: ignore[no-untyped-def]
        if Path(args[0]) == artifact_path:
            raise ValueError("simulated parser failure")
        return real_load_workbook(*args, **kwargs)

    monkeypatch.setattr(openpyxl, "load_workbook", _failing_load_workbook)

    with pytest.raises(ValueError, match="failed openpyxl round-trip validation"):
        _validate_xlsx(artifact_path)



def test_persist_td_api_artifacts_skips_invalid_excel_and_surfaces_validation_error(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    from app.crm_downloader.td_orders_sync import td_api_artifacts

    def _raise_validate(path: Path) -> None:
        raise ValueError(f"missing required OOXML entries: xl/workbook.xml ({path.name})")

    monkeypatch.setenv("TD_API_HUMAN_READABLE_EXPORT", "true")
    monkeypatch.setattr(td_api_artifacts, "_validate_xlsx", _raise_validate)

    result = persist_td_api_artifacts(
        download_dir=tmp_path,
        store_code="a817",
        from_date=date(2026, 1, 1),
        to_date=date(2026, 1, 2),
        raw_orders={"data": []},
        raw_sales={"data": []},
        raw_garments={"data": []},
        order_rows=[{"order_number": "1001"}],
        sale_rows=[{"order_number": "1001"}],
        garments_rows=[{"order_number": "1001"}],
    )

    assert "orders_excel" not in result.artifact_paths
    assert "sales_excel" not in result.artifact_paths
    assert "garments_excel" not in result.artifact_paths
    assert any("missing required OOXML entries" in warning for warning in result.warnings)

def test_persist_td_api_artifacts_unlinks_target_when_excel_write_fails(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    target = tmp_path / "A817_td_api_orders_20260101_20260102.xlsx"
    target.parent.mkdir(parents=True, exist_ok=True)
    target.write_text("stale", encoding="utf-8")

    from app.crm_downloader.td_orders_sync import td_api_artifacts

    def _raise_write_excel(path: Path, rows, *, sheet_name: str = "rows") -> None:  # type: ignore[no-untyped-def]
        raise RuntimeError("boom")

    monkeypatch.setenv("TD_API_HUMAN_READABLE_EXPORT", "true")
    monkeypatch.setattr(td_api_artifacts, "_write_excel", _raise_write_excel)

    result = persist_td_api_artifacts(
        download_dir=tmp_path,
        store_code="a817",
        from_date=date(2026, 1, 1),
        to_date=date(2026, 1, 2),
        raw_orders={"data": []},
        raw_sales={"data": []},
        raw_garments={"data": []},
        order_rows=[{"order_number": "1001"}],
        sale_rows=[{"order_number": "1001"}],
        garments_rows=[{"order_number": "1001"}],
    )

    assert not target.exists()
    assert any("orders_excel" in warning for warning in result.warnings)

class _StubResponse:
    def __init__(self, *, status: int, url: str, payload: dict[str, object]) -> None:
        self.status = status
        self.url = url
        self._payload = payload

    async def json(self) -> dict[str, object]:
        return self._payload


class _StubRequest:
    def __init__(self, responses: list[object]) -> None:
        self._responses = responses
        self.calls: list[dict[str, object]] = []

    async def get(self, url: str, params: object, headers: dict[str, str], timeout: int) -> _StubResponse:
        self.calls.append({"url": url, "params": params, "headers": dict(headers), "timeout": timeout})
        next_item = self._responses.pop(0)
        if isinstance(next_item, Exception):
            raise next_item
        return next_item




class _ReadTimeoutResponse(_StubResponse):
    async def json(self) -> dict[str, object]:
        raise asyncio.TimeoutError()


class _StubContext:
    def __init__(self, request: _StubRequest) -> None:
        self.request = request
        self.pages: list[object] = []


class _TokenRefreshingClient(TdApiClient):
    def __init__(self, **kwargs: object) -> None:
        super().__init__(**kwargs)
        self.discovery_calls: list[bool] = []

    async def _discover_reporting_token(self, *, force_refresh: bool = False):  # type: ignore[override]
        self.discovery_calls.append(force_refresh)
        if self._auth_state.token_discovery is not None and not force_refresh:
            return self._auth_state.token_discovery
        if force_refresh:
            token = type(self)._token_result("fresh-token")
            self._auth_state.token_discovery = token
            return token
        token = type(self)._token_result("stale-token")
        self._auth_state.token_discovery = token
        return token

    @staticmethod
    def _token_result(token: str):
        from app.crm_downloader.td_orders_sync.td_api_client import _TokenDiscoveryResult

        return _TokenDiscoveryResult(token=token, source="test", expiry=None)


@pytest.mark.asyncio
async def test_get_json_retries_once_after_401_with_refreshed_token(tmp_path: Path) -> None:
    request = _StubRequest(
        responses=[
            _StubResponse(status=401, url="https://reporting-api.quickdrycleaning.com/reports/order-report", payload={}),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/reports/order-report",
                payload={"data": [{"id": 1}]},
            ),
        ]
    )
    context = _StubContext(request=request)
    client = _TokenRefreshingClient(store_code="a123", context=context, storage_state_path=tmp_path / "s.json")

    metadata: list[dict[str, object]] = []
    result = await client._get_json(endpoint="/reports/order-report", params={"page": 1}, metadata=metadata)

    assert result.ok is True
    assert result.payload == {"data": [{"id": 1}]}
    assert len(request.calls) == 2
    assert request.calls[0]["headers"].get("Authorization") == "Bearer stale-token"
    assert request.calls[1]["headers"].get("Authorization") == "Bearer fresh-token"
    assert client.discovery_calls.count(True) == 1
    assert metadata[0]["retry_reason"] == "auth_refresh"
    assert metadata[0]["status"] == 401
    assert metadata[0]["retry_count"] == 0
    assert metadata[0]["token_refresh_attempted"] is True
    assert metadata[1]["status"] == 200
    assert metadata[1]["token_refresh_attempted"] is True




@pytest.mark.asyncio
async def test_fetch_reports_refreshes_auth_once_then_reuses_for_other_endpoints(tmp_path: Path) -> None:
    request = _StubRequest(
        responses=[
            _StubResponse(status=401, url="https://reporting-api.quickdrycleaning.com/reports/order-report", payload={}),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/reports/order-report",
                payload={"data": [{"orderNumber": "1001"}], "totalPages": 1},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales",
                payload={"data": [{"orderNo": "S-1"}], "totalPages": 1},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/garments/details",
                payload={"data": [{"orderNo": "G-1", "lineItemId": "L1"}], "totalPages": 1},
            ),
        ]
    )
    context = _StubContext(request=request)
    client = _TokenRefreshingClient(store_code="a123", context=context, storage_state_path=tmp_path / "s.json")

    result = await client.fetch_reports(from_date=date(2026, 1, 1), to_date=date(2026, 1, 2))

    assert result.endpoint_errors == {}
    assert result.orders_rows == [{"orderNumber": "1001"}]
    assert len(result.sales_rows) == 1
    assert len(result.garments_rows) == 1
    assert client.discovery_calls.count(True) == 1

    orders_calls = [call for call in request.calls if str(call["url"]).endswith("/reports/order-report")]
    sales_calls = [call for call in request.calls if str(call["url"]).endswith("/sales-and-deliveries/sales")]
    garments_calls = [call for call in request.calls if str(call["url"]).endswith("/garments/details")]
    assert len(orders_calls) == 2
    assert len(sales_calls) == 1
    assert len(garments_calls) == 1
    assert orders_calls[0]["headers"]["Authorization"] == "Bearer stale-token"
    assert orders_calls[1]["headers"]["Authorization"] == "Bearer fresh-token"
    assert sales_calls[0]["headers"]["Authorization"] == "Bearer fresh-token"
    assert garments_calls[0]["headers"]["Authorization"] == "Bearer fresh-token"
    assert sales_calls[0]["params"]["token"] == "fresh-token"
    assert garments_calls[0]["params"]["token"] == "fresh-token"

    auth_refresh_items = [
        item for item in result.request_metadata if item["endpoint"] == "/reports/order-report" and item["status"] == 401
    ]
    assert len(auth_refresh_items) == 1
    assert auth_refresh_items[0]["retry_reason"] == "auth_refresh"

    sales_metadata = next(item for item in result.request_metadata if item["endpoint"] == "/sales-and-deliveries/sales" and item["status"] == 200)
    assert sales_metadata["query_params"]["startDate"] == ["2026-01-01"]
    assert sales_metadata["query_params"]["endDate"] == ["2026-01-02"]
    assert sales_metadata["query_params"]["page"] == ["1"]
    assert sales_metadata["query_params"]["pageSize"] == ["500"]
    assert sales_metadata["query_params"]["token"] == ["fresh-token"]
    assert sales_metadata["auth_shape"] == "legacy"
    assert sales_metadata["primary_auth_shape"] == "legacy"
    assert sales_metadata["auth_shape_fallback_from_har_like"] is False

    garments_metadata = next(item for item in result.request_metadata if item["endpoint"] == "/garments/details" and item["status"] == 200)
    assert garments_metadata["query_params"]["startDate"] == ["2026-01-01"]
    assert garments_metadata["query_params"]["endDate"] == ["2026-01-02"]
    assert garments_metadata["query_params"]["page"] == ["1"]
    assert garments_metadata["query_params"]["pageSize"] == ["500"]
    assert garments_metadata["query_params"]["token"] == ["fresh-token"]
    assert garments_metadata["auth_shape"] == "legacy"
    assert garments_metadata["primary_auth_shape"] == "legacy"
    assert garments_metadata["auth_shape_fallback_from_har_like"] is False

@pytest.mark.asyncio
async def test_fetch_reports_captures_non_retriable_http_errors_per_endpoint(tmp_path: Path) -> None:
    request = _StubRequest(
        responses=[
            _StubResponse(status=403, url="https://reporting-api.quickdrycleaning.com/reports/order-report", payload={}),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales",
                payload={"data": [{"orderNo": "S-1", "paymentDate": "2026-01-02 10:00:00", "total": "10", "salesOnlyField": "survives"}], "totalPages": 1},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/garments/details",
                payload={"data": [{"orderNo": "G-1", "lineItemId": "L1", "garmentOnlyField": "survives"}], "totalPages": 1},
            ),
        ]
    )
    context = _StubContext(request=request)
    client = _TokenRefreshingClient(store_code="a123", context=context, storage_state_path=tmp_path / "s.json")

    result = await client.fetch_reports(from_date=date(2026, 1, 1), to_date=date(2026, 1, 2))

    assert result.raw_orders_payload["error"] == "http_403"
    assert result.raw_sales_payload["error"] is None
    assert result.raw_garments_payload["error"] is None
    assert result.orders_rows == []
    assert len(result.sales_rows) == 1
    assert len(result.garments_rows) == 1
    assert result.sales_rows[0]["salesOnlyField"] == "survives"
    assert result.garments_rows[0]["garmentOnlyField"] == "survives"


@pytest.mark.asyncio
async def test_fetch_reports_sets_expand_data_true_for_orders_and_sales(tmp_path: Path) -> None:
    request = _StubRequest(
        responses=[
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/reports/order-report",
                payload={"data": [], "totalPages": 1},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales",
                payload={"data": [], "totalPages": 1},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/garments/details",
                payload={"data": [], "totalPages": 1},
            ),
        ]
    )
    context = _StubContext(request=request)
    client = _TokenRefreshingClient(store_code="a123", context=context, storage_state_path=tmp_path / "s.json")

    await client.fetch_reports(from_date=date(2026, 1, 1), to_date=date(2026, 1, 2))

    orders_call = next(call for call in request.calls if call["url"].endswith("/reports/order-report"))
    sales_call = next(call for call in request.calls if call["url"].endswith("/sales-and-deliveries/sales"))
    garments_call = next(call for call in request.calls if call["url"].endswith("/garments/details"))

    assert orders_call["params"]["expandData"] == "true"
    assert sales_call["params"]["expandData"] == "true"
    assert "expandData" in orders_call["params"]
    assert "expandData" in sales_call["params"]
    assert "expandData" not in garments_call["params"]


@pytest.mark.asyncio
async def test_sales_and_garments_try_legacy_before_har_like_auth_shape(tmp_path: Path) -> None:
    request = _StubRequest(
        responses=[
            _StubResponse(status=200, url="https://reporting-api.quickdrycleaning.com/reports/order-report", payload={"data": [], "totalPages": 1}),
            _StubResponse(status=403, url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales", payload={}),
            _StubResponse(status=200, url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales", payload={"data": [], "totalPages": 1}),
            _StubResponse(status=403, url="https://reporting-api.quickdrycleaning.com/garments/details", payload={}),
            _StubResponse(status=200, url="https://reporting-api.quickdrycleaning.com/garments/details", payload={"data": [], "totalPages": 1}),
        ]
    )
    context = _StubContext(request=request)
    client = _TokenRefreshingClient(store_code="a123", context=context, storage_state_path=tmp_path / "s.json")

    result = await client.fetch_reports(from_date=date(2026, 1, 1), to_date=date(2026, 1, 2))

    sales_calls = [call for call in request.calls if call["url"].endswith("/sales-and-deliveries/sales")]
    garments_calls = [call for call in request.calls if call["url"].endswith("/garments/details")]
    assert sales_calls[0]["headers"]["Authorization"] == "Bearer stale-token"
    assert garments_calls[0]["headers"]["Authorization"] == "Bearer stale-token"
    assert sales_calls[0]["params"]["token"] == "stale-token"
    assert garments_calls[0]["params"]["token"] == "stale-token"
    assert sales_calls[1]["headers"].get("Authorization") is None
    assert garments_calls[1]["headers"].get("Authorization") is None
    assert "token" not in sales_calls[1]["params"]
    assert "token" not in garments_calls[1]["params"]

    fallback_entries = [item for item in result.request_metadata if item.get("auth_shape_fallback_from_har_like")]
    assert fallback_entries
    assert all(item["auth_shape"] == "har_like" for item in fallback_entries)
    assert all(item["primary_auth_shape"] == "legacy" for item in fallback_entries)
    assert all(item.get("auth_shape_fallback_used") is True for item in fallback_entries)
    assert all("auth_shape_status_delta" in item for item in fallback_entries)


@pytest.mark.asyncio
async def test_har_like_origin_and_referer_follow_report_iframe_context(tmp_path: Path) -> None:
    request = _StubRequest(
        responses=[
            _StubResponse(status=200, url="https://reporting-api.quickdrycleaning.com/reports/order-report", payload={"data": [], "totalPages": 1}),
            _StubResponse(status=200, url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales", payload={"data": [], "totalPages": 1}),
            _StubResponse(status=200, url="https://reporting-api.quickdrycleaning.com/garments/details", payload={"data": [], "totalPages": 1}),
        ]
    )
    context = _StubContext(request=request)
    iframe_src = "https://reports.quickdrycleaning.com/embed?view=sales&store=A123"
    client = _TokenRefreshingClient(store_code="a123", context=context, storage_state_path=tmp_path / "s.json", report_iframe_src=iframe_src)

    await client.fetch_reports(from_date=date(2026, 1, 1), to_date=date(2026, 1, 2))

    sales_call = next(call for call in request.calls if call["url"].endswith("/sales-and-deliveries/sales"))
    assert sales_call["headers"]["origin"] == "https://reports.quickdrycleaning.com"
    assert sales_call["headers"]["referer"] == iframe_src


@pytest.mark.asyncio
async def test_fetch_reports_preserves_full_fidelity_rows_for_artifacts(tmp_path: Path) -> None:
    request = _StubRequest(
        responses=[
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/reports/order-report",
                payload={"data": [{"orderNo": "O-1", "ordersOnlyField": "keep"}], "totalPages": 1},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales",
                payload={"data": [{"orderNo": "S-1", "salesOnlyField": "keep"}], "totalPages": 1},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/garments/details",
                payload={"data": [{"orderNo": "G-1", "garmentsOnlyField": "keep"}], "totalPages": 1},
            ),
        ]
    )
    context = _StubContext(request=request)
    client = _TokenRefreshingClient(store_code="a123", context=context, storage_state_path=tmp_path / "s.json")

    result = await client.fetch_reports(from_date=date(2026, 1, 1), to_date=date(2026, 1, 2))

    assert result.orders_rows[0]["ordersOnlyField"] == "keep"
    assert result.sales_rows[0]["salesOnlyField"] == "keep"
    assert result.garments_rows[0]["garmentsOnlyField"] == "keep"
    assert result.raw_orders_payload["data"][0]["ordersOnlyField"] == "keep"
    assert result.raw_sales_payload["data"][0]["salesOnlyField"] == "keep"
    assert result.raw_garments_payload["data"][0]["garmentsOnlyField"] == "keep"


@pytest.mark.asyncio
async def test_fetch_reports_filters_summary_rows_from_orders_and_sales(tmp_path: Path) -> None:
    request = _StubRequest(
        responses=[
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/reports/order-report",
                payload={
                    "data": [
                        {"orderNumber": "1001", "customer": "Alice", "amount": "10.00"},
                        {"orderNumber": "", "description": "Total", "amount": "10.00"},
                    ],
                    "totalPages": 1,
                },
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales",
                payload={
                    "data": [
                        {"orderNo": "S-1", "paymentDate": "2026-01-02", "netAmount": "7.00"},
                        {"orderNo": None, "label": "Grand Total", "netAmount": "7.00", "tax": "1.00"},
                    ],
                    "totalPages": 1,
                },
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/garments/details",
                payload={"data": [{"orderNo": "G-1", "lineItemId": "L1"}], "totalPages": 1},
            ),
        ]
    )
    context = _StubContext(request=request)
    client = _TokenRefreshingClient(store_code="a123", context=context, storage_state_path=tmp_path / "s.json")

    result = await client.fetch_reports(from_date=date(2026, 1, 1), to_date=date(2026, 1, 2))

    assert [row["orderNumber"] for row in result.orders_rows] == ["1001"]
    assert [row["orderNo"] for row in result.sales_rows] == ["S-1"]


@pytest.mark.asyncio
async def test_fetch_reports_exposes_summary_filter_counts(tmp_path: Path) -> None:
    request = _StubRequest(
        responses=[
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/reports/order-report",
                payload={
                    "data": [
                        {"orderNumber": "1001", "customer": "Alice", "amount": "10.00"},
                        {"orderNumber": "", "description": "Total", "amount": "10.00"},
                    ],
                    "totalPages": 1,
                },
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales",
                payload={
                    "data": [
                        {"orderNo": "S-1", "paymentDate": "2026-01-02", "netAmount": "7.00"},
                        {"orderNo": None, "label": "Grand Total", "netAmount": "7.00", "tax": "1.00"},
                    ],
                    "totalPages": 1,
                },
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/garments/details",
                payload={"data": [{"orderNo": "G-1", "lineItemId": "L1"}], "totalPages": 1},
            ),
        ]
    )
    context = _StubContext(request=request)
    client = _TokenRefreshingClient(store_code="a123", context=context, storage_state_path=tmp_path / "s.json")

    result = await client.fetch_reports(from_date=date(2026, 1, 1), to_date=date(2026, 1, 2))

    assert result.orders_summary_rows_filtered == 1
    assert result.sales_summary_rows_filtered == 1


@pytest.mark.asyncio
async def test_fetch_reports_honors_td_api_max_pages_cap(tmp_path: Path) -> None:
    request = _StubRequest(
        responses=[
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/reports/order-report",
                payload={"data": [{"orderNumber": "1001"}], "totalPages": 3, "total": 3},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales",
                payload={"data": [{"orderNo": "S-1"}], "totalPages": 3, "total": 3},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/garments/details",
                payload={"data": [{"orderNo": "G-1"}], "totalPages": 3, "total": 3},
            ),
        ]
    )
    context = _StubContext(request=request)
    client = _TokenRefreshingClient(store_code="a123", context=context, storage_state_path=tmp_path / "s.json")
    client.config = type(client.config)(max_pages=1)

    result = await client.fetch_reports(from_date=date(2026, 1, 1), to_date=date(2026, 1, 2))

    assert len(request.calls) == 3
    assert result.raw_orders_payload["pagination"]["pages_fetched"] == 1
    assert result.raw_sales_payload["pagination"]["pages_fetched"] == 1
    assert result.raw_garments_payload["pagination"]["pages_fetched"] == 1


class _NoTokenClient(TdApiClient):
    async def _discover_reporting_token(self, *, force_refresh: bool = False):  # type: ignore[override]
        from app.crm_downloader.td_orders_sync.td_api_client import _TokenDiscoveryResult

        return _TokenDiscoveryResult(token=None, source=None, expiry=None)


@pytest.mark.asyncio
async def test_fetch_reports_preflight_skips_endpoint_calls_when_auth_unavailable(tmp_path: Path) -> None:
    request = _StubRequest(responses=[])
    context = _StubContext(request=request)
    client = _NoTokenClient(store_code="a123", context=context, storage_state_path=tmp_path / "missing.json")

    result = await client.fetch_reports(from_date=date(2026, 1, 1), to_date=date(2026, 1, 2))

    assert request.calls == []
    assert result.endpoint_errors == {
        "/reports/order-report": "auth_unavailable",
        "/sales-and-deliveries/sales": "auth_unavailable",
        "/garments/details": "auth_unavailable",
    }
    assert result.request_metadata[0]["outcome"] == "auth_unavailable"




@pytest.mark.asyncio
async def test_fetch_reports_records_payload_error_as_endpoint_error(tmp_path: Path) -> None:
    request = _StubRequest(
        responses=[
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/reports/order-report",
                payload={"error": "read_timeout", "data": []},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales",
                payload={"data": [{"orderNo": "S-1"}], "totalPages": 1},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/garments/details",
                payload={"data": [{"orderNo": "G-1"}], "totalPages": 1},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/garments/details",
                payload={"data": [], "totalPages": 1},
            ),
        ]
    )
    context = _StubContext(request=request)
    client = _TokenRefreshingClient(store_code="a123", context=context, storage_state_path=tmp_path / "s.json")

    result = await client.fetch_reports(from_date=date(2026, 1, 1), to_date=date(2026, 1, 2))

    assert result.endpoint_errors["/reports/order-report"] == "read_timeout"
    assert result.raw_orders_payload["error"] == "read_timeout"
    assert result.endpoint_error_diagnostics["/reports/order-report"]["token_found"] is True

@pytest.mark.asyncio
async def test_fetch_reports_propagates_page_one_401_as_endpoint_error(tmp_path: Path) -> None:
    request = _StubRequest(
        responses=[
            _StubResponse(status=401, url="https://reporting-api.quickdrycleaning.com/reports/order-report", payload={}),
            _StubResponse(status=401, url="https://reporting-api.quickdrycleaning.com/reports/order-report", payload={}),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales",
                payload={"data": [{"orderNo": "S-1"}], "totalPages": 1},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/garments/details",
                payload={"data": [{"orderNo": "G-1"}], "totalPages": 1},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/garments/details",
                payload={"data": [], "totalPages": 1},
            ),
        ]
    )
    context = _StubContext(request=request)
    client = _TokenRefreshingClient(store_code="a123", context=context, storage_state_path=tmp_path / "s.json")

    result = await client.fetch_reports(from_date=date(2026, 1, 1), to_date=date(2026, 1, 2))

    assert result.endpoint_errors["/reports/order-report"] == "http_401"
    assert result.raw_orders_payload["error"] == "http_401"
    assert result.orders_rows == []
    assert result.endpoint_error_diagnostics["/reports/order-report"]["token_found"] is True


@pytest.mark.asyncio
async def test_fetch_reports_records_auth_error_diagnostics_when_rows_zero_due_to_auth(tmp_path: Path) -> None:
    request = _StubRequest(
        responses=[
            _StubResponse(status=401, url="https://reporting-api.quickdrycleaning.com/reports/order-report", payload={}),
            _StubResponse(status=401, url="https://reporting-api.quickdrycleaning.com/reports/order-report", payload={}),
            _StubResponse(status=401, url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales", payload={}),
            _StubResponse(status=401, url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales", payload={}),
            _StubResponse(status=401, url="https://reporting-api.quickdrycleaning.com/garments/details", payload={}),
            _StubResponse(status=401, url="https://reporting-api.quickdrycleaning.com/garments/details", payload={}),
        ]
    )
    context = _StubContext(request=request)
    client = _TokenRefreshingClient(store_code="a123", context=context, storage_state_path=tmp_path / "s.json")

    result = await client.fetch_reports(from_date=date(2026, 1, 1), to_date=date(2026, 1, 2))

    assert result.orders_rows == []
    assert result.sales_rows == []
    assert result.garments_rows == []
    assert result.endpoint_errors == {
        "/reports/order-report": "http_401",
        "/sales-and-deliveries/sales": "http_401",
        "/garments/details": "http_401",
    }
    assert set(result.endpoint_error_diagnostics.keys()) == set(result.endpoint_errors.keys())
    assert all(bool(diagnostics) for diagnostics in result.endpoint_error_diagnostics.values())


@pytest.mark.asyncio
async def test_get_json_classifies_total_timeout_and_bounds_retries(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    request = _StubRequest(responses=[asyncio.TimeoutError(), asyncio.TimeoutError(), asyncio.TimeoutError()])
    context = _StubContext(request=request)
    client = _TokenRefreshingClient(
        store_code="a123",
        context=context,
        storage_state_path=tmp_path / "s.json",
        config=TdApiClientConfig(max_retries=2, min_interval_seconds=0, backoff_base_seconds=0, backoff_jitter_seconds=0),
    )

    sleep_calls: list[float] = []

    async def _fake_sleep(seconds: float) -> None:
        sleep_calls.append(seconds)

    monkeypatch.setattr("app.crm_downloader.td_orders_sync.td_api_client.asyncio.sleep", _fake_sleep)

    async def _fake_wait_turn(_: str, __: float) -> None:
        return None

    monkeypatch.setattr("app.crm_downloader.td_orders_sync.td_api_client._StoreRateLimiter.wait_turn", _fake_wait_turn)

    metadata: list[dict[str, object]] = []
    result = await client._get_json(
        endpoint="/reports/order-report",
        params={"page": 1},
        metadata=metadata,
        connect_timeout_ms=10,
        read_timeout_ms=50,
    )

    assert result.ok is False
    assert result.error == "total_timeout"
    assert len(request.calls) == 3
    assert len(metadata) == 3
    assert {item["retry_reason"] for item in metadata} == {"total_timeout"}
    assert sleep_calls == [0.0, 0.0]


@pytest.mark.asyncio
async def test_fetch_reports_retries_timeout_then_falls_back_page_size(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    request = _StubRequest(
        responses=[
            _ReadTimeoutResponse(status=200, url="https://reporting-api.quickdrycleaning.com/reports/order-report", payload={}),
            _ReadTimeoutResponse(status=200, url="https://reporting-api.quickdrycleaning.com/reports/order-report", payload={}),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/reports/order-report",
                payload={"data": [{"orderNumber": "1001"}], "totalPages": 1},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales",
                payload={"data": [{"orderNo": "S-1"}], "totalPages": 1},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/garments/details",
                payload={"data": [{"orderNo": "G-1"}], "totalPages": 1},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/garments/details",
                payload={"data": [], "totalPages": 1},
            ),
        ]
    )
    context = _StubContext(request=request)
    client = _TokenRefreshingClient(
        store_code="a123",
        context=context,
        storage_state_path=tmp_path / "s.json",
        config=TdApiClientConfig(
            page_size=500,
            page_size_fallbacks=(250, 100),
            max_retries=1,
            orders_max_retries=1,
            min_interval_seconds=0,
            backoff_base_seconds=0,
            backoff_jitter_seconds=0,
            connect_timeout_ms=100,
            orders_read_timeout_ms=10,
            sales_read_timeout_ms=10,
            garments_read_timeout_ms=10,
        ),
    )

    async def _fake_sleep(_: float) -> None:
        return None

    monkeypatch.setattr("app.crm_downloader.td_orders_sync.td_api_client.asyncio.sleep", _fake_sleep)

    async def _fake_wait_turn(_: str, __: float) -> None:
        return None

    monkeypatch.setattr("app.crm_downloader.td_orders_sync.td_api_client._StoreRateLimiter.wait_turn", _fake_wait_turn)

    result = await client.fetch_reports(from_date=date(2026, 1, 1), to_date=date(2026, 1, 2))

    assert result.endpoint_errors == {}
    assert result.orders_rows == [{"orderNumber": "1001"}]

    order_calls = [call for call in request.calls if str(call["url"]).endswith("/reports/order-report")]
    assert len(order_calls) == 3
    assert [call["params"]["pageSize"] for call in order_calls] == [500, 500, 250]

    fallback_events = [
        item
        for item in result.request_metadata
        if item.get("endpoint") == "/reports/order-report" and item.get("retry_reason") == "page_size_fallback"
    ]
    assert len(fallback_events) == 1
    assert fallback_events[0]["fallback_page_size_from"] == 500
    assert fallback_events[0]["fallback_page_size_to"] == 250
    fallback_query_params = fallback_events[0]["query_params"]
    assert fallback_query_params["startDate"] == ["2026-01-01"]
    assert fallback_query_params["endDate"] == ["2026-01-02"]
    assert fallback_query_params["expandData"] == ["true"]

    assert result.metrics_counters["timeout_class|endpoint=/reports/order-report|timeout_class=read_timeout"] == 2
    assert result.metrics_counters["eventual_success_after_retry|endpoint=/reports/order-report"] == 1
    assert result.metrics_counters["fallback_page_size_attempts|endpoint=/reports/order-report"] == 1
    assert result.metrics_counters["fallback_page_size_successes|endpoint=/reports/order-report"] == 1


def test_endpoint_specific_retry_profile_defaults() -> None:
    client = TdApiClient(store_code="a123", context=None, storage_state_path=Path("/tmp/missing"))  # type: ignore[arg-type]
    assert client._retry_profile_for_endpoint("/reports/order-report")["max_retries"] == client.config.orders_max_retries
    assert client._retry_profile_for_endpoint("/sales-and-deliveries/sales")["max_retries"] == client.config.sales_max_retries
    assert client._retry_profile_for_endpoint("/garments/details")["max_retries"] == client.config.garments_max_retries




def test_endpoint_specific_total_timeout_defaults() -> None:
    client = TdApiClient(store_code="a123", context=None, storage_state_path=Path("/tmp/missing"))  # type: ignore[arg-type]
    assert client._total_timeout_ms_for_endpoint("/reports/order-report") == client.config.orders_total_timeout_ms
    assert client._total_timeout_ms_for_endpoint("/sales-and-deliveries/sales") == client.config.sales_total_timeout_ms
    assert client._total_timeout_ms_for_endpoint("/garments/details") == client.config.garments_total_timeout_ms
    assert client.config.sales_total_timeout_ms == 60000
    assert client.config.garments_total_timeout_ms == 60000


@pytest.mark.asyncio
async def test_fetch_reports_uses_endpoint_specific_effective_total_timeout(tmp_path: Path) -> None:
    request = _StubRequest(
        responses=[
            _StubResponse(status=200, url="https://reporting-api.quickdrycleaning.com/reports/order-report", payload={"data": [], "totalPages": 1}),
            _StubResponse(status=200, url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales", payload={"data": [], "totalPages": 1}),
            _StubResponse(status=200, url="https://reporting-api.quickdrycleaning.com/garments/details", payload={"data": [], "totalPages": 1}),
        ]
    )
    context = _StubContext(request=request)
    client = _TokenRefreshingClient(
        store_code="a123",
        context=context,
        storage_state_path=tmp_path / "s.json",
        config=TdApiClientConfig(connect_timeout_ms=5000, orders_read_timeout_ms=15000, sales_read_timeout_ms=45000, garments_read_timeout_ms=45000),
    )

    await client.fetch_reports(from_date=date(2026, 1, 1), to_date=date(2026, 1, 2))

    orders_call = next(call for call in request.calls if call["url"].endswith("/reports/order-report"))
    sales_call = next(call for call in request.calls if call["url"].endswith("/sales-and-deliveries/sales"))
    garments_call = next(call for call in request.calls if call["url"].endswith("/garments/details"))

    assert orders_call["timeout"] == client.config.orders_total_timeout_ms
    assert sales_call["timeout"] == client.config.sales_total_timeout_ms
    assert garments_call["timeout"] == client.config.garments_total_timeout_ms
    assert client.config.connect_timeout_ms + client.config.sales_read_timeout_ms == 50000
    assert sales_call["timeout"] > client.config.connect_timeout_ms

def test_endpoint_specific_read_timeout_defaults() -> None:
    client = TdApiClient(store_code="a123", context=None, storage_state_path=Path("/tmp/missing"))  # type: ignore[arg-type]
    assert client._read_timeout_ms_for_endpoint("/reports/order-report") == client.config.orders_read_timeout_ms
    assert client._read_timeout_ms_for_endpoint("/sales-and-deliveries/sales") == client.config.sales_read_timeout_ms
    assert client._read_timeout_ms_for_endpoint("/garments/details") == client.config.garments_read_timeout_ms
    assert client.config.sales_read_timeout_ms == 45000
    assert client.config.garments_read_timeout_ms == 45000


@pytest.mark.asyncio
async def test_get_json_timeout_retries_are_bounded_for_sales_endpoint(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    request = _StubRequest(
        responses=[
            _ReadTimeoutResponse(status=200, url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales", payload={}),
            _ReadTimeoutResponse(status=200, url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales", payload={}),
            _ReadTimeoutResponse(status=200, url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales", payload={}),
            _ReadTimeoutResponse(status=200, url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales", payload={}),
            _ReadTimeoutResponse(status=200, url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales", payload={}),
            _ReadTimeoutResponse(status=200, url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales", payload={}),
        ]
    )
    context = _StubContext(request=request)
    client = _TokenRefreshingClient(
        store_code="a123",
        context=context,
        storage_state_path=tmp_path / "s.json",
        config=TdApiClientConfig(
            min_interval_seconds=0,
            max_retries=5,
            sales_max_retries=5,
            sales_read_timeout_ms=15,
            timeout_retry_limit=2,
            backoff_base_seconds=0,
            backoff_jitter_seconds=0,
            sales_backoff_base_seconds=0,
            sales_max_backoff_seconds=0,
        ),
    )

    async def _fake_sleep(_: float) -> None:
        return None

    async def _fake_wait_turn(_: str, __: float) -> None:
        return None

    monkeypatch.setattr("app.crm_downloader.td_orders_sync.td_api_client.asyncio.sleep", _fake_sleep)
    monkeypatch.setattr("app.crm_downloader.td_orders_sync.td_api_client._StoreRateLimiter.wait_turn", _fake_wait_turn)

    metadata: list[dict[str, object]] = []
    retry_profile = client._retry_profile_for_endpoint("/sales-and-deliveries/sales")
    result = await client._get_json(
        endpoint="/sales-and-deliveries/sales",
        params={"startDate": "2026-01-01", "endDate": "2026-01-02", "page": 1, "pageSize": 500},
        metadata=metadata,
        connect_timeout_ms=client.config.connect_timeout_ms,
        read_timeout_ms=client.config.sales_read_timeout_ms,
        max_retries=int(retry_profile["max_retries"]),
        backoff_base_seconds=float(retry_profile["backoff_base_seconds"]),
        max_backoff_seconds=float(retry_profile["max_backoff_seconds"]),
    )

    assert result.ok is False
    assert result.error in {"read_timeout", "total_timeout"}
    assert len(request.calls) == (client.config.timeout_retry_limit + 1)
    timeout_metadata = [item for item in metadata if item.get("retry_reason") in {"read_timeout", "total_timeout"}]
    assert timeout_metadata
    assert max(item["retry_count"] for item in timeout_metadata) == client.config.timeout_retry_limit
    assert all(item["primary_auth_shape"] == "legacy" for item in timeout_metadata)
    assert all(item["auth_shape_fallback_from_har_like"] is False for item in timeout_metadata)
    assert all(item["timeout_diagnostics"]["configured"]["attempt_timeout_budget_ms"] == client.config.connect_timeout_ms + client.config.sales_read_timeout_ms for item in timeout_metadata)
    assert all(item["timeout_diagnostics"]["effective"]["attempt_timeout_budget_ms"] == client.config.sales_total_timeout_ms for item in timeout_metadata)


@pytest.mark.asyncio
async def test_get_json_fails_fast_when_sales_required_query_params_missing(tmp_path: Path) -> None:
    request = _StubRequest(responses=[])
    context = _StubContext(request=request)
    client = _TokenRefreshingClient(store_code="a123", context=context, storage_state_path=tmp_path / "s.json")

    metadata: list[dict[str, object]] = []
    result = await client._get_json(
        endpoint="/sales-and-deliveries/sales",
        params={"page": 1, "pageSize": 500},
        metadata=metadata,
    )

    assert result.ok is False
    assert result.error == "missing_required_query_params:startDate,endDate"
    assert len(request.calls) == 0
    assert metadata[0]["endpoint"] == "/sales-and-deliveries/sales"
    assert metadata[0]["retry_reason"] == "missing_required_query_params:startDate,endDate"


@pytest.mark.asyncio
async def test_fetch_reports_emits_endpoint_health_contract(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    request = _StubRequest(
        responses=[
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/reports/order-report",
                payload={"data": [{"orderNumber": "1001"}], "totalPages": 1},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/sales-and-deliveries/sales",
                payload={"data": [{"orderNo": "S-1"}], "totalPages": 1},
            ),
            _StubResponse(
                status=200,
                url="https://reporting-api.quickdrycleaning.com/garments/details",
                payload={"data": [{"orderNo": "G-1"}], "totalPages": 1},
            ),
        ]
    )
    context = _StubContext(request=request)
    client = _TokenRefreshingClient(
        store_code="a123",
        context=context,
        storage_state_path=tmp_path / "s.json",
        config=TdApiClientConfig(min_interval_seconds=0),
    )

    async def _fake_wait_turn(_: str, __: float) -> None:
        return None

    monkeypatch.setattr("app.crm_downloader.td_orders_sync.td_api_client._StoreRateLimiter.wait_turn", _fake_wait_turn)

    result = await client.fetch_reports(from_date=date(2026, 1, 1), to_date=date(2026, 1, 2))

    assert result.endpoint_health["/reports/order-report"]["success"] is True
    assert result.endpoint_health["/sales-and-deliveries/sales"]["success"] is True
    assert result.endpoint_health["/garments/details"]["attempts"] >= 1


def test_dataset_completion_health_includes_endpoint_health_fields() -> None:
    payload = {"pagination": {"pages_fetched": 2, "reported_total_pages": 2, "total_rows": 5}}
    health = _dataset_completion_health(
        payload,
        endpoint_error=None,
        endpoint_health={"success": True, "final_error_class": None, "attempts": 4},
    )
    assert health["ready"] is True
    assert health["endpoint_success"] is True
    assert health["endpoint_attempts"] == 4


@pytest.mark.asyncio
async def test_prepare_auth_context_dashboard_trial_cookie_only_marks_orders_not_ready(tmp_path: Path) -> None:
    artifact = tmp_path / "store-state.json"
    artifact.write_text(
        json.dumps({"cookies": [{"name": "sessionid", "domain": ".quickdrycleaning.com"}], "origins": []}),
        encoding="utf-8",
    )
    client = TdApiClient(store_code="a817", context=None, storage_state_path=artifact)  # type: ignore[arg-type]

    async def _tokenless_discovery() -> object:
        return type("TokenDiscovery", (), {"token": None, "source": None, "expiry": None})()

    client._discover_reporting_token = _tokenless_discovery  # type: ignore[method-assign]

    auth_result = await client.prepare_auth_context()

    assert auth_result.ready is False
    assert auth_result.cookies_present is True
    assert auth_result.token_present is False
    assert auth_result.auth_contract == "endpoint_specific"
    assert auth_result.endpoint_auth_requirements == {
        "/reports/order-report": ("token",),
        "/sales-and-deliveries/sales": ("cookie_session",),
        "/garments/details": ("cookie_session",),
    }
    assert auth_result.endpoint_readiness == {
        "/reports/order-report": False,
        "/sales-and-deliveries/sales": True,
        "/garments/details": True,
    }
    assert auth_result.endpoint_failure_reasons == {"/reports/order-report": "token_missing_or_expired"}
    assert auth_result.failure_reason == "/reports/order-report:token_missing_or_expired"
