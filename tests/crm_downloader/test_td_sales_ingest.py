from __future__ import annotations

import io
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from zoneinfo import ZoneInfo

import openpyxl
import pytest
import sqlalchemy as sa

from app.common.db import session_scope
from app.crm_downloader.td_orders_sync.sales_ingest import (
    TdSalesIngestResult,
    _expected_headers,
    _orders_table,
    _sales_table,
    _stg_td_sales_table,
    ingest_td_sales_workbook,
)
from app.dashboard_downloader.json_logger import JsonLogger, get_logger


async def _create_tables(database_url: str) -> None:
    metadata = sa.MetaData()
    _orders_table(metadata)
    _stg_td_sales_table(metadata)
    _sales_table(metadata)
    async with session_scope(database_url) as session:
        bind = session.bind
        async with bind.begin() as conn:
            await conn.run_sync(metadata.create_all)


def _build_sample_workbook(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = list(_expected_headers())
    ws.append(headers)
    ws.append(
        [
            "2025-05-10",
            "2025-05-11 09:30",
            "SAL-001",
            "C001",
            "Alice",
            "123 Street",
            "+91 99999-88888",
            "1,200.50",
            "10",
            "1190.50",
            "Bob",
            "UPI",
            "TX123",
            "Counter",
            "Regular",
        ]
    )
    ws.append(
        [
            "2025-05-11",
            "2025-05-12 10:00",
            "SAL-002",
            "C002",
            "Bob",
            "456 Lane",
            "08888877777",
            900,
            None,
            900,
            "Carol",
            "Cash",
            "",
            "",
            "Express",
        ]
    )
    wb.save(path)
    return path


def _build_duplicate_workbook(path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = list(_expected_headers())
    ws.append(headers)
    ws.append(
        [
            "2025-05-10",
            "2025-05-11 09:30",
            "DUP-001",
            "C001",
            "Alice",
            "123 Street",
            "12345",
            500,
            0,
            500,
            "Bob",
            "UPI",
            "TX123",
            "Counter",
            "Regular",
        ]
    )
    ws.append(
        [
            "2025-05-11",
            "2025-05-12 10:00",
            "DUP-001",
            "C002",
            "Bob",
            "456 Lane",
            "+91 99999-88888",
            900,
            None,
            900,
            "Carol",
            "UPI",
            "",
            "",
            "Express",
        ]
    )
    wb.save(path)
    return path


def _build_workbook_from_rows(path: Path, rows: list[dict[str, object]]) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = list(_expected_headers())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(header) for header in headers])
    wb.save(path)
    return path


async def _insert_order(
    database_url: str,
    *,
    store_code: str,
    cost_center: str,
    order_number: str,
    gross_amount: Decimal,
    order_date: datetime,
) -> None:
    async with session_scope(database_url) as session:
        metadata = sa.MetaData()
        orders_table = _orders_table(metadata)
        await session.execute(
            sa.insert(orders_table).values(
                run_id="test_run",
                run_date=order_date,
                cost_center=cost_center,
                store_code=store_code,
                source_system="TD",
                order_number=order_number,
                order_date=order_date,
                customer_name="Test Customer",
                mobile_number="9999999999",
                package_flag=False,
                due_date=order_date,
                default_due_date=order_date,
                complete_processing_by=order_date,
                gross_amount=gross_amount,
                order_edited_flag=False,
                created_at=order_date,
            )
        )
        await session.commit()


@pytest.mark.asyncio
async def test_sales_ingest_happy_path(tmp_path: Path) -> None:
    db_path = tmp_path / "sales.db"
    database_url = f"sqlite+aiosqlite:///{db_path}"
    await _create_tables(database_url)

    workbook = _build_sample_workbook(tmp_path / "sales.xlsx")
    tz = ZoneInfo("Asia/Kolkata")
    run_date = datetime(2025, 5, 20, 12, 0, tzinfo=tz)
    logger = get_logger(run_id="test_run")

    result = await ingest_td_sales_workbook(
        workbook_path=workbook,
        store_code="A668",
        cost_center="UN3668",
        run_id="test_run",
        run_date=run_date,
        database_url=database_url,
        logger=logger,
    )

    assert isinstance(result, TdSalesIngestResult)
    assert result.staging_rows == 2
    assert result.final_rows == 2
    assert result.warnings == []

    async with session_scope(database_url) as session:
        metadata = sa.MetaData()
        stg_table = _stg_td_sales_table(metadata)
        final_table = _sales_table(metadata)

        stg_rows = (
            await session.execute(
                sa.select(
                    stg_table.c.order_number,
                    stg_table.c.payment_received,
                    stg_table.c.is_duplicate,
                    stg_table.c.ingest_remarks,
                ).order_by(stg_table.c.order_number)
            )
        ).all()
        assert stg_rows[0].order_number == "SAL-001"
        assert float(stg_rows[0].payment_received) == 1200.5
        assert stg_rows[0].is_duplicate is False
        assert stg_rows[0].ingest_remarks is None
        assert stg_rows[1].is_duplicate is False

        final_rows = (
            await session.execute(
                sa.select(
                    final_table.c.order_number,
                    final_table.c.payment_mode,
                    final_table.c.is_duplicate,
                    final_table.c.ingest_remarks,
                ).order_by(final_table.c.order_number)
            )
        ).all()
        assert final_rows[0].payment_mode == "UPI"
        assert final_rows[0].is_duplicate is False
        assert final_rows[1].is_duplicate is False
        assert final_rows[0].ingest_remarks is None


@pytest.mark.asyncio
async def test_sales_ingest_marks_duplicates_and_warnings(tmp_path: Path) -> None:
    db_path = tmp_path / "sales_dupe.db"
    database_url = f"sqlite+aiosqlite:///{db_path}"
    await _create_tables(database_url)

    workbook = _build_duplicate_workbook(tmp_path / "sales_dup.xlsx")
    tz = ZoneInfo("Asia/Kolkata")
    run_date = datetime(2025, 5, 20, 12, 0, tzinfo=tz)
    log_stream = io.StringIO()
    logger = JsonLogger(run_id="test_run", stream=log_stream, log_file_path=None)

    result = await ingest_td_sales_workbook(
        workbook_path=workbook,
        store_code="A668",
        cost_center="UN3668",
        run_id="test_run",
        run_date=run_date,
        database_url=database_url,
        logger=logger,
    )

    assert result.staging_rows == 2
    assert result.final_rows == 2
    assert result.warnings == ["Invalid phone number fallback applied: 12345"]
    assert result.rows_duplicate == 2
    assert result.rows_edited == 2
    assert len(result.duplicate_rows) == 2
    assert len(result.edited_rows) == 2
    assert result.ingest_remarks == [
        {
            "store_code": "A668",
            "order_number": "DUP-001",
            "ingest_remarks": "MOBILE_FALLBACK_APPLIED; "
            "Duplicate order_number/payment_mode 'DUP-001'/'UPI' detected in sales data",
        },
        {
            "store_code": "A668",
            "order_number": "DUP-001",
            "ingest_remarks": "Duplicate order_number/payment_mode 'DUP-001'/'UPI' detected in sales data",
        },
    ]

    async with session_scope(database_url) as session:
        metadata = sa.MetaData()
        stg_table = _stg_td_sales_table(metadata)
        final_table = _sales_table(metadata)
        stg_rows = (
            await session.execute(
                sa.select(
                    stg_table.c.order_number,
                    stg_table.c.is_duplicate,
                    stg_table.c.ingest_remarks,
                ).order_by(stg_table.c.id)
            )
        ).all()
        assert stg_rows[0].is_duplicate is True
        assert "MOBILE_FALLBACK_APPLIED" in stg_rows[0].ingest_remarks
        assert "Duplicate order_number/payment_mode 'DUP-001'/'UPI' detected in sales data" in stg_rows[0].ingest_remarks
        assert stg_rows[1].is_duplicate is True
        assert stg_rows[1].ingest_remarks == "Duplicate order_number/payment_mode 'DUP-001'/'UPI' detected in sales data"

        final_rows = (
            await session.execute(
                sa.select(final_table.c.order_number, final_table.c.is_duplicate).order_by(final_table.c.id)
            )
        ).all()
        assert final_rows[0].is_duplicate is True
        assert final_rows[1].is_duplicate is True


@pytest.mark.asyncio
async def test_sales_ingest_split_payments_not_duplicate_or_edited(tmp_path: Path) -> None:
    db_path = tmp_path / "sales_split.db"
    database_url = f"sqlite+aiosqlite:///{db_path}"
    await _create_tables(database_url)

    tz = ZoneInfo("Asia/Kolkata")
    order_date = datetime(2025, 5, 10, 12, 0, tzinfo=tz)
    await _insert_order(
        database_url,
        store_code="A668",
        cost_center="UN3668",
        order_number="SPLIT-001",
        gross_amount=Decimal("1000"),
        order_date=order_date,
    )

    workbook = _build_workbook_from_rows(
        tmp_path / "sales_split.xlsx",
        [
            {
                "Order Date": "2025-05-10",
                "Payment Date": "2025-05-11 09:30",
                "Order Number": "SPLIT-001",
                "Customer Code": "C111",
                "Customer Name": "Split One",
                "Customer Address": "123 Split St",
                "Customer Mobile No.": "9999988888",
                "Payment Received": "600",
                "Adjustments": 0,
                "Balance": 0,
                "Accept By": "Agent",
                "Payment Mode": "Cash",
                "Online TransactionID": "S1",
                "Payment Made At": "Counter",
                "Type": "Regular",
            },
            {
                "Order Date": "2025-05-10",
                "Payment Date": "2025-05-11 10:30",
                "Order Number": "SPLIT-001",
                "Customer Code": "C111",
                "Customer Name": "Split One",
                "Customer Address": "123 Split St",
                "Customer Mobile No.": "9999988888",
                "Payment Received": "400",
                "Adjustments": 0,
                "Balance": 0,
                "Accept By": "Agent",
                "Payment Mode": "Card",
                "Online TransactionID": "S2",
                "Payment Made At": "Counter",
                "Type": "Regular",
            },
        ],
    )
    run_date = datetime(2025, 5, 20, 12, 0, tzinfo=tz)
    logger = get_logger(run_id="split_run")

    result = await ingest_td_sales_workbook(
        workbook_path=workbook,
        store_code="A668",
        cost_center="UN3668",
        run_id="split_run",
        run_date=run_date,
        database_url=database_url,
        logger=logger,
    )

    assert result.rows_duplicate == 0
    assert result.rows_edited == 0

    async with session_scope(database_url) as session:
        metadata = sa.MetaData()
        stg_table = _stg_td_sales_table(metadata)
        stg_rows = (
            await session.execute(
                sa.select(
                    stg_table.c.payment_mode,
                    stg_table.c.is_duplicate,
                    stg_table.c.is_edited_order,
                ).order_by(stg_table.c.payment_mode)
            )
        ).all()
        assert [row.payment_mode for row in stg_rows] == ["Card", "Cash"]
        assert all(row.is_duplicate is False for row in stg_rows)
        assert all(row.is_edited_order is False for row in stg_rows)


@pytest.mark.asyncio
async def test_sales_ingest_marks_edited_when_payments_below_gross(tmp_path: Path) -> None:
    db_path = tmp_path / "sales_shortfall.db"
    database_url = f"sqlite+aiosqlite:///{db_path}"
    await _create_tables(database_url)

    tz = ZoneInfo("Asia/Kolkata")
    order_date = datetime(2025, 5, 10, 12, 0, tzinfo=tz)
    await _insert_order(
        database_url,
        store_code="A668",
        cost_center="UN3668",
        order_number="EDIT-001",
        gross_amount=Decimal("1000"),
        order_date=order_date,
    )

    workbook = _build_workbook_from_rows(
        tmp_path / "sales_shortfall.xlsx",
        [
            {
                "Order Date": "2025-05-10",
                "Payment Date": "2025-05-11 09:30",
                "Order Number": "EDIT-001",
                "Customer Code": "C111",
                "Customer Name": "Short Fall",
                "Customer Address": "123 Short St",
                "Customer Mobile No.": "9999988888",
                "Payment Received": "400",
                "Adjustments": 0,
                "Balance": 0,
                "Accept By": "Agent",
                "Payment Mode": "Cash",
                "Online TransactionID": "E1",
                "Payment Made At": "Counter",
                "Type": "Regular",
            },
            {
                "Order Date": "2025-05-10",
                "Payment Date": "2025-05-11 10:30",
                "Order Number": "EDIT-001",
                "Customer Code": "C111",
                "Customer Name": "Short Fall",
                "Customer Address": "123 Short St",
                "Customer Mobile No.": "9999988888",
                "Payment Received": "300",
                "Adjustments": 0,
                "Balance": 0,
                "Accept By": "Agent",
                "Payment Mode": "Card",
                "Online TransactionID": "E2",
                "Payment Made At": "Counter",
                "Type": "Regular",
            },
        ],
    )
    run_date = datetime(2025, 5, 20, 12, 0, tzinfo=tz)
    logger = get_logger(run_id="shortfall_run")

    result = await ingest_td_sales_workbook(
        workbook_path=workbook,
        store_code="A668",
        cost_center="UN3668",
        run_id="shortfall_run",
        run_date=run_date,
        database_url=database_url,
        logger=logger,
    )

    assert result.rows_duplicate == 0
    assert result.rows_edited == 2

    shortfall_message = (
        "Total payment_received 700 is less than gross_amount 1000.00 for order 'EDIT-001'"
    )
    assert all(entry["ingest_remarks"] == shortfall_message for entry in result.ingest_remarks)

    async with session_scope(database_url) as session:
        metadata = sa.MetaData()
        stg_table = _stg_td_sales_table(metadata)
        stg_rows = (
            await session.execute(
                sa.select(
                    stg_table.c.payment_mode,
                    stg_table.c.is_duplicate,
                    stg_table.c.is_edited_order,
                    stg_table.c.ingest_remarks,
                ).order_by(stg_table.c.payment_mode)
            )
        ).all()
        assert [row.payment_mode for row in stg_rows] == ["Card", "Cash"]
        assert all(row.is_duplicate is False for row in stg_rows)
        assert all(row.is_edited_order is True for row in stg_rows)
        assert all(row.ingest_remarks == shortfall_message for row in stg_rows)


@pytest.mark.asyncio
async def test_sales_ingest_reingest_same_range_is_not_duplicate(tmp_path: Path) -> None:
    db_path = tmp_path / "sales_existing.db"
    database_url = f"sqlite+aiosqlite:///{db_path}"
    await _create_tables(database_url)

    base_workbook = _build_sample_workbook(tmp_path / "sales_base.xlsx")
    tz = ZoneInfo("Asia/Kolkata")
    run_date = datetime(2025, 5, 20, 12, 0, tzinfo=tz)
    logger = get_logger(run_id="test_run")

    await ingest_td_sales_workbook(
        workbook_path=base_workbook,
        store_code="A668",
        cost_center="UN3668",
        run_id="test_run",
        run_date=run_date,
        database_url=database_url,
        logger=logger,
    )

    # second workbook with same order_number but different payment_date should not be treated as duplicate
    wb = openpyxl.load_workbook(base_workbook)
    ws = wb.active
    ws["B2"] = "2025-05-15 09:30"
    ws["B3"] = "2025-05-16 10:00"
    wb.save(base_workbook)

    second_run_date = datetime(2025, 5, 21, 12, 0, tzinfo=tz)
    result = await ingest_td_sales_workbook(
        workbook_path=base_workbook,
        store_code="A668",
        cost_center="UN3668",
        run_id="test_run",
        run_date=second_run_date,
        database_url=database_url,
        logger=logger,
    )

    assert result.staging_rows == 2
    assert result.final_rows == 2
    assert result.rows_duplicate == 0
    assert result.rows_edited == 0
    assert result.duplicate_rows == []
    assert result.edited_rows == []

    async with session_scope(database_url) as session:
        metadata = sa.MetaData()
        stg_table = _stg_td_sales_table(metadata)
        dupe_flags = (
            await session.execute(
                sa.select(stg_table.c.order_number, stg_table.c.is_duplicate).order_by(stg_table.c.id)
            )
        ).all()
        assert all(flag.is_duplicate is False for flag in dupe_flags)


@pytest.mark.asyncio
async def test_sales_ingest_parses_dates_numbers_and_mobile(tmp_path: Path) -> None:
    db_path = tmp_path / "sales_parsing.db"
    database_url = f"sqlite+aiosqlite:///{db_path}"
    await _create_tables(database_url)

    workbook = _build_workbook_from_rows(
        tmp_path / "sales_parsing.xlsx",
        [
            {
                "Order Date": "10 May 2025",
                "Payment Date": datetime(2025, 5, 11, 9, 30),
                "Order Number": "PARSE-001",
                "Customer Code": "C123",
                "Customer Name": "Parsey",
                "Customer Address": "123 Parsing St",
                "Customer Mobile No.": "+91-99999 88888",
                "Payment Received": "1,000.75",
                "Adjustments": "oops",
                "Balance": 1000.75,
                "Accept By": "Checker",
                "Payment Mode": "Cash",
                "Online TransactionID": "TXP1",
                "Payment Made At": " Counter ",
                "Type": "Regular",
            }
        ],
    )
    tz = ZoneInfo("Asia/Kolkata")
    run_date = datetime(2025, 5, 22, 12, 0, tzinfo=tz)
    logger = get_logger(run_id="parse_run")

    result = await ingest_td_sales_workbook(
        workbook_path=workbook,
        store_code="A668",
        cost_center="UN3668",
        run_id="parse_run",
        run_date=run_date,
        database_url=database_url,
        logger=logger,
    )

    assert result.warnings == ["Non-numeric value for adjustments: oops"]
    assert result.staging_rows == 1
    assert result.final_rows == 1
    assert result.ingest_remarks == [
        {
            "store_code": "A668",
            "order_number": "PARSE-001",
            "ingest_remarks": "Field adjustments contained non-numeric value 'oops' (stored as 0)",
        }
    ]

    async with session_scope(database_url) as session:
        metadata = sa.MetaData()
        stg_table = _stg_td_sales_table(metadata)
        final_table = _sales_table(metadata)
        stg_row = (
            await session.execute(
                sa.select(
                    stg_table.c.order_date,
                    stg_table.c.payment_date,
                    stg_table.c.mobile_number,
                    stg_table.c.payment_received,
                    stg_table.c.adjustments,
                    stg_table.c.payment_made_at,
                    stg_table.c.ingest_remarks,
                )
            )
        ).one()
        assert stg_row.order_date.replace(tzinfo=tz) == datetime(2025, 5, 10, tzinfo=tz)
        assert stg_row.payment_date.replace(tzinfo=tz) == datetime(2025, 5, 11, 9, 30, tzinfo=tz)
        assert stg_row.mobile_number == "9999988888"
        assert float(stg_row.payment_received) == 1000.75
        assert float(stg_row.adjustments) == 0
        assert stg_row.payment_made_at == "Counter"
        assert stg_row.ingest_remarks == "Field adjustments contained non-numeric value 'oops' (stored as 0)"

        final_row = (
            await session.execute(
                sa.select(final_table.c.ingest_remarks, final_table.c.mobile_number, final_table.c.payment_mode)
            )
        ).one()
        assert final_row.ingest_remarks == stg_row.ingest_remarks
        assert final_row.mobile_number == "9999988888"
        assert final_row.payment_mode == "Cash"


@pytest.mark.asyncio
async def test_sales_upsert_respects_business_keys_and_propagates_remarks(tmp_path: Path) -> None:
    db_path = tmp_path / "sales_upsert.db"
    database_url = f"sqlite+aiosqlite:///{db_path}"
    await _create_tables(database_url)

    base_workbook = _build_workbook_from_rows(
        tmp_path / "sales_upsert.xlsx",
        [
            {
                "Order Date": "2025-05-10",
                "Payment Date": "2025-05-11 09:30",
                "Order Number": "UP-001",
                "Customer Code": "CU1",
                "Customer Name": "Upsert One",
                "Customer Address": "123 Up Lane",
                "Customer Mobile No.": "09876543210",
                "Payment Received": "500",
                "Adjustments": 0,
                "Balance": 500,
                "Accept By": "A",
                "Payment Mode": "UPI",
                "Online TransactionID": "T1",
                "Payment Made At": "Counter",
                "Type": "Regular",
            }
        ],
    )
    tz = ZoneInfo("Asia/Kolkata")
    run_date = datetime(2025, 5, 23, 12, 0, tzinfo=tz)
    logger = get_logger(run_id="upsert_run")

    await ingest_td_sales_workbook(
        workbook_path=base_workbook,
        store_code="A668",
        cost_center="UN3668",
        run_id="upsert_run",
        run_date=run_date,
        database_url=database_url,
        logger=logger,
    )

    updated_workbook = _build_workbook_from_rows(
        tmp_path / "sales_upsert_updated.xlsx",
        [
            {
                "Order Date": "2025-05-10",
                "Payment Date": "2025-05-11 09:30",
                "Order Number": "UP-001",
                "Customer Code": "CU1",
                "Customer Name": "Upsert One",
                "Customer Address": "123 Up Lane",
                "Customer Mobile No.": "09876543210",
                "Payment Received": "750",
                "Adjustments": 10,
                "Balance": 740,
                "Accept By": "A",
                "Payment Mode": "Card",
                "Online TransactionID": "T1",
                "Payment Made At": "Counter",
                "Type": "Regular",
            }
        ],
    )
    second_run_date = datetime(2025, 5, 24, 12, 0, tzinfo=tz)
    result = await ingest_td_sales_workbook(
        workbook_path=updated_workbook,
        store_code="A668",
        cost_center="UN3668",
        run_id="upsert_run",
        run_date=second_run_date,
        database_url=database_url,
        logger=logger,
    )

    assert result.staging_rows == 1
    assert result.final_rows == 1
    assert result.ingest_remarks == []

    async with session_scope(database_url) as session:
        metadata = sa.MetaData()
        stg_table = _stg_td_sales_table(metadata)
        final_table = _sales_table(metadata)

        stg_count = (await session.execute(sa.select(sa.func.count()).select_from(stg_table))).scalar_one()
        final_count = (await session.execute(sa.select(sa.func.count()).select_from(final_table))).scalar_one()
        assert stg_count == 1
        assert final_count == 1

        stg_row = (
            await session.execute(
                sa.select(
                    stg_table.c.payment_received,
                    stg_table.c.adjustments,
                    stg_table.c.is_duplicate,
                    stg_table.c.ingest_remarks,
                    stg_table.c.payment_mode,
                )
            )
        ).one()
        assert float(stg_row.payment_received) == 750
        assert float(stg_row.adjustments) == 10
        assert stg_row.is_duplicate is False
        assert stg_row.ingest_remarks is None
        assert stg_row.payment_mode == "Card"

        final_row = (
            await session.execute(
                sa.select(
                    final_table.c.payment_received,
                    final_table.c.adjustments,
                    final_table.c.is_duplicate,
                    final_table.c.ingest_remarks,
                    final_table.c.payment_mode,
                )
            )
        ).one()
        assert float(final_row.payment_received) == 750
        assert float(final_row.adjustments) == 10
        assert final_row.is_duplicate is False
        assert final_row.ingest_remarks is None
        assert final_row.payment_mode == "Card"
