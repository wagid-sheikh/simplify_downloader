from datetime import datetime
from decimal import Decimal

import openpyxl
from zoneinfo import ZoneInfo

from app.reports.pending_deliveries.data import (
    PendingDeliveriesBucket,
    PendingDeliveriesCostCenterSection,
    PendingDeliveriesReportData,
    PendingDeliveriesSummarySection,
    PendingDeliveryRow,
)
from app.reports.pending_deliveries.workbook import build_pending_deliveries_workbook


def _row(
    cost_center: str, order_number: str, order_date: datetime, *, age_days: int = 3
) -> PendingDeliveryRow:
    return PendingDeliveryRow(
        cost_center=cost_center,
        store_code="A1",
        order_number=order_number,
        customer_name="Alice",
        order_date=order_date,
        default_due_date=order_date,
        age_days=age_days,
        order_amount=Decimal("100"),
        paid_amount=Decimal("20"),
        pending_amount=Decimal("80"),
        adjustments=Decimal("0"),
        is_edited_order=False,
        is_duplicate=False,
        source_system="TumbleDry",
    )


def _report(
    rows_by_cc: dict[str, list[PendingDeliveryRow]],
) -> PendingDeliveriesReportData:
    sections = []
    summaries = []
    for cc, rows in rows_by_cc.items():
        bucket = PendingDeliveriesBucket(
            label="0-5 days",
            min_days=0,
            max_days=5,
            rows=rows,
            total_count=len(rows),
            total_pending_amount=Decimal("80"),
        )
        sections.append(
            PendingDeliveriesCostCenterSection(
                cost_center=cc,
                buckets=[bucket],
                total_pending_amount=Decimal("80"),
                total_count=len(rows),
            )
        )
        summaries.append(
            PendingDeliveriesSummarySection(
                cost_center=cc,
                buckets=[bucket],
                total_pending_amount=Decimal("80"),
                total_count=len(rows),
            )
        )
    return PendingDeliveriesReportData(
        report_date=datetime(2026, 4, 29).date(),
        summary_sections=summaries,
        cost_center_sections=sections,
        total_pending_amount=Decimal("160"),
        total_count=sum(len(v) for v in rows_by_cc.values()),
    )


def test_pending_deliveries_workbook_groups_sorts_and_normalizes_tz(tmp_path):
    output = tmp_path / "pending.xlsx"
    tz = ZoneInfo("Asia/Kolkata")
    rows = {
        "CC_A": [
            _row(
                "CC_A",
                "ORD-1",
                datetime(2026, 4, 29, 10, 0, tzinfo=ZoneInfo("UTC")),
                age_days=9,
            ),
            _row(
                "CC_A",
                "ORD-2",
                datetime(2026, 4, 30, 10, 0, tzinfo=ZoneInfo("UTC")),
                age_days=4,
            ),
        ],
        "CC_B": [
            _row("CC_B", "ORD-3", datetime(2026, 4, 28, 7, 0, tzinfo=ZoneInfo("UTC")))
        ],
    }

    row_count, sheet_count = build_pending_deliveries_workbook(
        data=_report(rows), output_path=output, business_timezone=tz
    )

    assert row_count == 3
    assert sheet_count == 2
    workbook = openpyxl.load_workbook(output, data_only=True)
    try:
        assert set(workbook.sheetnames) == {"CC_A", "CC_B"}
        cc_a = list(workbook["CC_A"].iter_rows(values_only=True))
        assert cc_a[0] == (
            "Cost Center",
            "Order Number",
            "Customer Name",
            "Order Date",
            "Due Date",
            "Age (Days)",
            "Order Amount",
        )
        assert cc_a[1][1] == "ORD-1"
        assert cc_a[2][1] == "ORD-2"
        assert cc_a[1][5] == 9
        assert cc_a[2][5] == 4
        written_dt = workbook["CC_A"].cell(row=2, column=4).value
        assert written_dt.tzinfo is None
        assert written_dt == datetime(2026, 4, 29, 15, 30)
        assert (
            workbook["CC_A"].cell(row=2, column=4).number_format
            == "yyyy-mm-dd hh:mm:ss"
        )
    finally:
        workbook.close()


def test_pending_deliveries_workbook_handles_zero_rows(tmp_path):
    output = tmp_path / "pending-empty.xlsx"
    data = _report({})
    row_count, sheet_count = build_pending_deliveries_workbook(
        data=data, output_path=output, business_timezone=ZoneInfo("Asia/Kolkata")
    )
    assert row_count == 0
    assert sheet_count == 1
    workbook = openpyxl.load_workbook(output, data_only=True)
    try:
        assert workbook.sheetnames == ["No Data"]
    finally:
        workbook.close()
