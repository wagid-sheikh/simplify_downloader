from __future__ import annotations

import json
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from types import SimpleNamespace
from zoneinfo import ZoneInfo

import openpyxl
import pytest
import sqlalchemy as sa
from jinja2 import Environment, FileSystemLoader, select_autoescape

from app.common.db import session_scope
from app.reports.daily_sales_report.data import RecoveryOrderRow
from app.reports.daily_sales_report.to_be_recovered import (
    build_context as build_to_be_recovered_context,
)
from app.reports.shared.short_payments import ShortPaymentRow
import app.reports.daily_sales_report.pipeline as pipeline
import app.reports.mtd_same_day_fulfillment.data as mtd_data


def test_daily_sales_template_excludes_apnf_section_markers() -> None:
    template = (
        Path("app")
        / "reports"
        / "daily_sales_report"
        / "templates"
        / "daily_sales_report.html"
    ).read_text(encoding="utf-8")

    assert "Actual Payments Not Found" not in template
    assert "{{ actual_payments_not_found_rows }}" not in template
    assert "actual_payments_not_found_rows" not in template


def test_mtd_template_excludes_apnf_section_markers() -> None:
    template = (
        Path("app")
        / "reports"
        / "mtd_same_day_fulfillment"
        / "templates"
        / "report.html"
    ).read_text(encoding="utf-8")

    assert "Actual Payments Not Found" not in template
    assert "missing_payment_rows" not in template


def test_build_context_includes_report_day_orders_by_cost_center() -> None:
    payload = SimpleNamespace(
        report_date=date(2026, 4, 29),
        rows=[],
        totals=SimpleNamespace(),
        edited_orders=[],
        edited_orders_summary=None,
        edited_orders_totals=None,
        missed_leads=[],
        cancelled_leads=[],
        lead_performance_summary=[],
        to_be_recovered=[],
        to_be_compensated=[],
        to_be_recovered_total_order_value=Decimal("0"),
        to_be_compensated_total_order_value=Decimal("0"),
        auto_cleared_order_numbers_text="",
        same_day_fulfillment_rows=[],
        missing_payment_rows=[],
        short_payment_rows=[],
        report_day_orders_by_cost_center=[SimpleNamespace(cost_center="CC1", order_numbers_text="ORD-1")],
    )
    context = pipeline._build_context(payload, "test")
    assert context["report_day_orders_by_cost_center"][0].order_numbers_text == "ORD-1"


def test_daily_sales_template_renders_same_day_cost_center_detail_values() -> None:
    payload = SimpleNamespace(
        report_date=date(2026, 4, 29),
        rows=[],
        totals=SimpleNamespace(),
        edited_orders=[],
        edited_orders_summary=None,
        edited_orders_totals=None,
        missed_leads=[],
        cancelled_leads=[],
        lead_performance_summary=[],
        to_be_recovered=[],
        to_be_compensated=[],
        to_be_recovered_total_order_value=Decimal("0"),
        to_be_compensated_total_order_value=Decimal("0"),
        auto_cleared_order_numbers_text="",
        same_day_fulfillment_rows=[
            SimpleNamespace(
                cost_center="CC77",
                store_code="S77",
                order_number="ORD-77",
                order_date=datetime(2026, 4, 29, 10, 30),
                delivery_or_payment_date=datetime(2026, 4, 29, 11, 10),
                customer_name="Alex",
                mobile_number="9999999999",
                line_items="Wash Shirt × 1",
                payment_mode="UPI",
                order_amount=Decimal("120"),
                payment_received=Decimal("120"),
                hours=Decimal("0.67"),
            )
        ],
        missing_payment_rows=[],
        short_payment_rows=[],
        report_day_orders_by_cost_center=[],
    )
    env = Environment(
        loader=FileSystemLoader(["app/reports/shared/templates"]),
        autoescape=select_autoescape(["html", "xml"]),
    )
    env.filters["format_amount"] = lambda value: str(value)
    env.filters["format_ddmmyyyy"] = lambda value: str(value)
    env.filters["format_hhmm_ampm"] = lambda value: str(value)
    template = env.get_template("same_day_fulfillment_table.html")
    html = template.render(
        same_day_title="Same-Day Fulfillment",
        same_day_grouped_rows_by_store=[("S77", payload.same_day_fulfillment_rows)],
        same_day_store_summary_rows=[],
        format_duration_hours=lambda value: str(value),
        format_duration_minutes=lambda value: str(value),
    )

    assert "Cost Center" in html
    assert "CC77</td>" in html and "ORD-77</td>" in html


def _create_tables(database_url: str) -> None:
    engine = sa.create_engine(database_url.replace("+aiosqlite", ""))
    with engine.begin() as conn:
        conn.execute(
            sa.text(
                """
                CREATE TABLE orders (
                    cost_center TEXT,
                    order_number TEXT,
                    order_date TIMESTAMP,
                    customer_name TEXT,
                    mobile_number TEXT,
                    net_amount NUMERIC,
                    recovery_status TEXT
                )
                """
            )
        )
        conn.execute(
            sa.text(
                """
                CREATE VIEW vw_orders AS
                SELECT *, COALESCE(net_amount, 0) AS order_amount
                FROM orders
                """
            )
        )
        conn.execute(
            sa.text(
                """
                CREATE TABLE order_line_items (
                    cost_center TEXT,
                    order_number TEXT,
                    service_name TEXT,
                    garment_name TEXT
                )
                """
            )
        )
        conn.execute(
            sa.text(
                """
                CREATE TABLE sales (
                    cost_center TEXT,
                    order_number TEXT,
                    payment_date TIMESTAMP,
                    payment_mode TEXT,
                    payment_received NUMERIC
                )
                """
            )
        )
        conn.execute(
            sa.text("CREATE TABLE store_master (cost_center TEXT, store_code TEXT)")
        )
        conn.execute(
            sa.text(
                "CREATE TABLE payment_collections (cost_center TEXT, order_number TEXT)"
            )
        )
        conn.execute(
            sa.text(
                """
                CREATE TABLE vw_orders_missing_in_payment_collections (
                    cost_center TEXT,
                    order_number TEXT,
                    order_date TIMESTAMP,
                    customer_name TEXT,
                    mobile_number TEXT,
                    net_amount NUMERIC,
                    recovery_status TEXT
                )
                """
            )
        )
        conn.execute(
            sa.text(
                """
                CREATE TABLE documents (
                    doc_type TEXT,
                    doc_subtype TEXT,
                    doc_date DATE,
                    reference_name_1 TEXT,
                    reference_id_1 TEXT,
                    reference_name_2 TEXT,
                    reference_id_2 TEXT,
                    reference_name_3 TEXT,
                    reference_id_3 TEXT,
                    file_name TEXT,
                    mime_type TEXT,
                    file_size_bytes INTEGER,
                    storage_backend TEXT,
                    file_path TEXT,
                    created_at TIMESTAMP,
                    created_by TEXT
                )
                """
            )
        )
    engine.dispose()


def test_to_be_recovered_template_renders_auto_cleared_empty_state() -> None:
    context = build_to_be_recovered_context(
        rows=[],
        report_date=date(2026, 4, 29),
        run_environment="test",
    )

    html = pipeline._render_html(
        context, template_name=pipeline.TO_BE_RECOVERED_TEMPLATE_NAME
    )

    assert (
        "Auto-cleared orders with payment proof in sales and payment collections"
        in html
    )
    assert "No unresolved TO_BE_RECOVERED orders" in html
    assert "Order Amount" in html
    assert "Order Value" not in html
    assert (
        "<strong>Auto-cleared orders with payment proof in sales and "
        "payment collections:</strong>" in html
    )
    assert "None" in html


def test_to_be_recovered_template_renders_auto_cleared_order_numbers() -> None:
    context = build_to_be_recovered_context(
        rows=[],
        report_date=date(2026, 4, 29),
        run_environment="test",
        auto_cleared_order_numbers_text="TD123, TD124, UC555",
    )

    html = pipeline._render_html(
        context, template_name=pipeline.TO_BE_RECOVERED_TEMPLATE_NAME
    )

    assert (
        "Auto-cleared orders with payment proof in sales and payment collections"
        in html
    )
    assert "TD123, TD124, UC555" in html


def test_to_be_recovered_template_renders_cost_center_summary() -> None:
    context = build_to_be_recovered_context(
        rows=[
            RecoveryOrderRow(
                cost_center="CC1",
                order_number="REC-1",
                order_date=date(2026, 4, 29),
                customer_name="Rhea",
                mobile_number="9999999990",
                order_value=Decimal("1250"),
            ),
            RecoveryOrderRow(
                cost_center="CC2",
                order_number="REC-2",
                order_date=date(2026, 4, 29),
                customer_name="Mina",
                mobile_number="9999999991",
                order_value=Decimal("2000"),
            ),
            RecoveryOrderRow(
                cost_center="CC1",
                order_number="REC-3",
                order_date=date(2026, 4, 29),
                customer_name="Nia",
                mobile_number="9999999992",
                order_value=Decimal("75"),
            ),
        ],
        report_date=date(2026, 4, 29),
        run_environment="test",
    )

    html = pipeline._render_html(
        context, template_name=pipeline.TO_BE_RECOVERED_TEMPLATE_NAME
    )

    assert 'class="micro-font summary-table"' in html
    assert "Total Order Amount" in html
    assert "Total Recoverable Amount" in html
    assert "Grand Total" in html
    assert "<td>1,325</td>" in html
    assert "<td>2,000</td>" in html
    assert "<td>3,325</td>" in html
    assert "REC-1" in html
    assert "REC-2" in html
    assert "REC-3" in html


def test_to_be_recovered_template_renders_summary_and_grouped_details() -> None:
    context = build_to_be_recovered_context(
        rows=[
            RecoveryOrderRow(
                cost_center="CC2",
                order_number="CC2-NEW",
                order_date=date(2026, 4, 29),
                customer_name="Newer",
                mobile_number="9999999993",
                order_value=Decimal("200"),
            ),
            RecoveryOrderRow(
                cost_center="CC1",
                order_number="CC1-LATE",
                order_date=date(2026, 4, 28),
                customer_name="Late First Store",
                mobile_number="9999999990",
                order_value=Decimal("50"),
            ),
            RecoveryOrderRow(
                cost_center="CC2",
                order_number="CC2-EARLY",
                order_date=date(2026, 4, 27),
                customer_name="Early Second Store",
                mobile_number="9999999992",
                order_value=Decimal("100"),
            ),
            RecoveryOrderRow(
                cost_center="CC1",
                order_number="CC1-EARLY",
                order_date=date(2026, 4, 27),
                customer_name="Early First Store",
                mobile_number="9999999991",
                order_value=Decimal("25"),
            ),
        ],
        report_date=date(2026, 4, 29),
        run_environment="test",
    )

    html = pipeline._render_html(
        context, template_name=pipeline.TO_BE_RECOVERED_TEMPLATE_NAME
    )

    assert "Order Amount" in html
    assert "Order Value" not in html
    assert "Cost Center" in html
    assert "Total Order Amount" in html
    assert "Total Recoverable Amount" in html

    summary_table = html[
        html.index('<table class="micro-font summary-table"') : html.index(
            '<table class="micro-font">'
        )
    ]
    assert summary_table.index('<td class="label">CC1</td>') < summary_table.index(
        '<td class="label">CC2</td>'
    )
    assert summary_table.count('<td class="label">CC1</td>') == 1
    assert summary_table.count('<td class="label">CC2</td>') == 1
    assert summary_table.count('<td class="label">Grand Total</td>') == 1
    assert "<td>75</td>" in summary_table
    assert "<td>300</td>" in summary_table
    assert "<td>375</td>" in summary_table

    detail_table = html[html.index('<table class="micro-font">') :]
    assert detail_table.index("Cost Center: CC1") < detail_table.index(
        "Cost Center: CC2"
    )

    cc1_section = detail_table[
        detail_table.index("Cost Center: CC1") : detail_table.index("CC1 Total")
    ]
    assert cc1_section.index("CC1-EARLY") < cc1_section.index("CC1-LATE")

    cc2_section = detail_table[
        detail_table.index("Cost Center: CC2") : detail_table.index("CC2 Total")
    ]
    assert cc2_section.index("CC2-EARLY") < cc2_section.index("CC2-NEW")

    assert (
        '<tr class="group-total-row">\n'
        '          <td class="label" colspan="5">CC1 Total</td>\n'
        '          <td>75</td>\n'
        '          <td>75</td>\n'
        '        </tr>'
    ) in detail_table
    assert (
        '<tr class="group-total-row">\n'
        '          <td class="label" colspan="5">CC2 Total</td>\n'
        '          <td>300</td>\n'
        '          <td>300</td>\n'
        '        </tr>'
    ) in detail_table
    assert detail_table.count('<tr class="group-total-row">') == 2
    assert (
        '<td class="label" colspan="6">Total Recoverable</td>\n'
        '          <td>375</td>'
    ) in detail_table


@pytest.mark.asyncio
async def test_daily_pipeline_writes_mtd_attachment_window_and_metadata(
    tmp_path, monkeypatch
) -> None:
    db_path = tmp_path / "daily_sales_pipeline.db"
    database_url = f"sqlite+aiosqlite:///{db_path}"
    _create_tables(database_url)

    report_date = date(2026, 4, 29)

    async with session_scope(database_url) as session:
        await session.execute(
            sa.text(
                "INSERT INTO store_master (cost_center, store_code) VALUES ('CC1', 'S1')"
            )
        )
        await session.execute(
            sa.text(
                """
                INSERT INTO orders (cost_center, order_number, order_date, customer_name, mobile_number, net_amount)
                VALUES
                    ('CC1', 'RPT-DATE-1', '2026-04-29T09:00:00+05:30', 'Alice', '9999999999', 900),
                    ('CC1', 'IN-MONTH-1', '2026-04-10T09:00:00+05:30', 'Bob', '8888888888', 800),
                    ('CC1', 'OUT-MONTH-1', '2026-03-31T09:00:00+05:30', 'Cora', '7777777777', 700)
                """
            )
        )
        await session.execute(
            sa.text(
                """
                INSERT INTO order_line_items (cost_center, order_number, service_name, garment_name)
                VALUES
                    ('CC1', 'RPT-DATE-1', 'Wash', 'Shirt'),
                    ('CC1', 'IN-MONTH-1', 'Iron', 'Pant'),
                    ('CC1', 'OUT-MONTH-1', 'Dry', 'Coat')
                """
            )
        )
        await session.execute(
            sa.text(
                """
                INSERT INTO sales (cost_center, order_number, payment_date, payment_mode, payment_received)
                VALUES
                    ('CC1', 'RPT-DATE-1', '2026-04-29T10:00:00+05:30', 'UPI', 900),
                    ('CC1', 'IN-MONTH-1', '2026-04-10T10:00:00+05:30', 'CARD', 800),
                    ('CC1', 'OUT-MONTH-1', '2026-04-29T10:00:00+05:30', 'CASH', 700)
                """
            )
        )
        await session.commit()

    monkeypatch.setattr(
        pipeline,
        "config",
        SimpleNamespace(database_url=database_url, pdf_render_timeout_seconds=30),
    )
    monkeypatch.setattr(pipeline, "get_timezone", lambda: ZoneInfo("Asia/Kolkata"))
    monkeypatch.setattr(pipeline, "resolve_run_env", lambda env: "test")
    monkeypatch.setattr(pipeline, "new_run_id", lambda: "run-mtd-1")

    async def _summary_noop(*args, **kwargs):
        return None

    monkeypatch.setattr(pipeline, "persist_summary_record", _summary_noop)
    monkeypatch.setattr(pipeline, "update_summary_record", _summary_noop)

    async def _fake_daily_data(*args, **kwargs):
        return SimpleNamespace(
            report_date=report_date,
            rows=[],
            totals=SimpleNamespace(),
            edited_orders=[],
            edited_orders_summary=SimpleNamespace(),
            edited_orders_totals=SimpleNamespace(),
            missed_leads=[],
            cancelled_leads=[],
            lead_performance_summary=SimpleNamespace(),
            to_be_recovered=[
                RecoveryOrderRow(
                    cost_center="CC1",
                    order_number="REC-1",
                    order_date=report_date,
                    customer_name="Rhea",
                    mobile_number="9999999990",
                    order_value=Decimal("125"),
                )
            ],
            to_be_compensated=[],
            to_be_recovered_total_order_value=0,
            to_be_compensated_total_order_value=0,
            same_day_fulfillment_rows=[SimpleNamespace(order_number="RPT-DATE-1")],
            missing_payment_rows=[],
            short_payment_rows=[],
        )

    monkeypatch.setattr(pipeline, "fetch_daily_sales_report", _fake_daily_data)

    monkeypatch.setattr(
        pipeline,
        "_render_html",
        lambda context, *args, **kwargs: (
            "DAILY SAME DAY: "
            + ",".join(row.order_number for row in context["same_day_fulfillment_rows"])
            if "same_day_fulfillment_rows" in context
            else "TO BE RECOVERED: "
            + ",".join(row.order_number for row in context["rows"])
        ),
    )

    rendered: dict[str, str] = {}

    async def _fake_render_pdf(html, output_path: Path, pdf_options=None, logger=None):
        rendered[str(output_path)] = html
        output_path.write_bytes(b"pdf")

    monkeypatch.setattr(
        pipeline, "render_pdf_with_configured_browser", _fake_render_pdf
    )

    async def _fake_notify(*args, **kwargs):
        return {"emails_planned": 1, "emails_sent": 1, "errors": []}

    monkeypatch.setattr(pipeline, "send_notifications_for_run", _fake_notify)

    await pipeline._run(report_date=report_date, env="test", force=True)

    daily_path = str(
        pipeline.OUTPUT_ROOT / f"{pipeline.PIPELINE_NAME}_{report_date.isoformat()}.pdf"
    )
    mtd_path = str(
        pipeline.OUTPUT_ROOT
        / f"reports.mtd_same_day_fulfillment_{report_date.isoformat()}.pdf"
    )
    recovered_path = str(
        pipeline.OUTPUT_ROOT / f"reports.to_be_recovered_{report_date.isoformat()}.pdf"
    )
    short_payments_path = str(
        pipeline.OUTPUT_ROOT
        / f"reports.daily_sales_report_short_payments_{report_date.isoformat()}.pdf"
    )
    apnf_path = str(
        pipeline.OUTPUT_ROOT
        / f"reports.daily_sales_report_actual_payments_not_found_{report_date.isoformat()}.pdf"
    )
    apnf_workbook_path = str(
        pipeline.OUTPUT_ROOT
        / f"reports.daily_sales_report_actual_payments_not_found_{report_date.isoformat()}.xlsx"
    )

    assert {daily_path, recovered_path, short_payments_path, apnf_path}.issubset(rendered)
    assert Path(apnf_workbook_path).exists()
    assert "RPT-DATE-1" in rendered[daily_path]
    assert "REC-1" in rendered[recovered_path]

    mtd_html = rendered[mtd_path]
    assert "RPT-DATE-1" in mtd_html
    assert "IN-MONTH-1" in mtd_html
    assert "OUT-MONTH-1" not in mtd_html
    assert "Wash Shirt" in mtd_html
    assert "Iron Pant" in mtd_html
    assert "Same-Day Fulfillment" in mtd_html
    assert "Window: 01-Apr-2026 to 29-Apr-2026" in mtd_html

    async with session_scope(database_url) as session:
        rows = (
            (
                await session.execute(
                    sa.text(
                        """
                    SELECT doc_type, file_name, reference_id_1, reference_id_3
                    FROM documents
                    ORDER BY file_name
                    """
                    )
                )
            )
            .mappings()
            .all()
        )

    assert len(rows) == 6
    by_type = {row["doc_type"]: row for row in rows}
    assert "daily_sales_actual_payments_not_found_pdf" in by_type
    assert "daily_sales_actual_payments_not_found_xlsx" in by_type
    assert "daily_sales_report_pdf" in by_type
    assert "daily_sales_short_payments_pdf" in by_type
    assert "mtd_same_day_fulfillment_pdf" in by_type
    assert "to_be_recovered_report_pdf" in by_type
    assert (
        by_type["daily_sales_short_payments_pdf"]["file_name"]
        == f"reports.daily_sales_report_short_payments_{report_date.isoformat()}.pdf"
    )
    assert by_type["daily_sales_short_payments_pdf"]["reference_id_1"] == pipeline.PIPELINE_NAME
    assert by_type["daily_sales_short_payments_pdf"]["reference_id_3"] == report_date.isoformat()
    assert (
        by_type["mtd_same_day_fulfillment_pdf"]["file_name"]
        == f"reports.mtd_same_day_fulfillment_{report_date.isoformat()}.pdf"
    )
    assert by_type["mtd_same_day_fulfillment_pdf"]["reference_id_1"] == pipeline.PIPELINE_NAME
    assert by_type["mtd_same_day_fulfillment_pdf"]["reference_id_3"] == report_date.isoformat()
    assert (
        by_type["to_be_recovered_report_pdf"]["file_name"]
        == f"reports.to_be_recovered_{report_date.isoformat()}.pdf"
    )
    assert by_type["to_be_recovered_report_pdf"]["reference_id_1"] == pipeline.PIPELINE_NAME
    assert by_type["to_be_recovered_report_pdf"]["reference_id_3"] == report_date.isoformat()
    workbook = openpyxl.load_workbook(Path(apnf_workbook_path), data_only=True)
    try:
        assert workbook.sheetnames == ["No Data"]
        rows = list(workbook["No Data"].iter_rows(values_only=True))
        assert rows[0] == (
            "Cost Center",
            "Order Number",
            "Order Date",
            "Customer Name",
            "Mobile Number",
            "Order Amount",
        )
    finally:
        workbook.close()


@pytest.mark.asyncio
async def test_daily_pipeline_metrics_are_json_safe_with_short_payment_rows(
    tmp_path, monkeypatch
) -> None:
    db_path = tmp_path / "daily_sales_pipeline_short_payment_metrics.db"
    database_url = f"sqlite+aiosqlite:///{db_path}"
    _create_tables(database_url)

    report_date = date(2026, 4, 29)
    short_payment_row = ShortPaymentRow(
        cost_center="CC1",
        order_number="SP-JSON-1",
        order_date=datetime(2026, 4, 29, 9, 30),
        customer_name="Alice",
        mobile_number="9999999999",
        order_amount=Decimal("100.50"),
        paid_amount=Decimal("80.25"),
        shortage_amount=Decimal("20.25"),
    )

    monkeypatch.setattr(
        pipeline,
        "config",
        SimpleNamespace(database_url=database_url, pdf_render_timeout_seconds=30),
    )
    monkeypatch.setattr(pipeline, "get_timezone", lambda: ZoneInfo("Asia/Kolkata"))
    monkeypatch.setattr(pipeline, "resolve_run_env", lambda env: "test")
    monkeypatch.setattr(pipeline, "new_run_id", lambda: "run-short-payment-json")

    captured_records: list[dict] = []

    async def _capture_persist(_database_url, record):
        captured_records.append(record)

    async def _capture_update(_database_url, _run_id, record):
        captured_records.append(record)

    monkeypatch.setattr(pipeline, "persist_summary_record", _capture_persist)
    monkeypatch.setattr(pipeline, "update_summary_record", _capture_update)

    async def _fake_daily_data(*args, **kwargs):
        return SimpleNamespace(
            report_date=report_date,
            rows=[],
            totals=SimpleNamespace(),
            edited_orders=[],
            edited_orders_summary=SimpleNamespace(),
            edited_orders_totals=SimpleNamespace(),
            missed_leads=[],
            cancelled_leads=[],
            lead_performance_summary=SimpleNamespace(),
            to_be_recovered=[],
            to_be_compensated=[],
            to_be_recovered_total_order_value=Decimal("0"),
            to_be_compensated_total_order_value=Decimal("0"),
            same_day_fulfillment_rows=[],
            missing_payment_rows=[],
            short_payment_rows=[short_payment_row],
        )

    async def _fake_mtd_rows(*args, **kwargs):
        return []

    async def _fake_render_pdf(html, output_path: Path, pdf_options=None, logger=None):
        output_path.write_bytes(b"pdf")

    async def _fake_notify(*args, **kwargs):
        return {"emails_planned": 1, "emails_sent": 1, "errors": []}

    monkeypatch.setattr(pipeline, "fetch_daily_sales_report", _fake_daily_data)
    monkeypatch.setattr(pipeline, "fetch_mtd_same_day_fulfillment", _fake_mtd_rows)
    monkeypatch.setattr(
        pipeline, "_render_html", lambda context, *args, **kwargs: "daily-html"
    )
    monkeypatch.setattr(
        pipeline, "render_pdf_with_configured_browser", _fake_render_pdf
    )
    monkeypatch.setattr(pipeline, "send_notifications_for_run", _fake_notify)

    await pipeline._run(
        report_date=report_date,
        env="test",
        force=True,
        orders_sync_upstream_status="failed",
        orders_sync_upstream_run_id="orders-run-1",
    )

    assert captured_records
    final_record = captured_records[-1]
    json.dumps(final_record["metrics_json"])
    for record in captured_records:
        json.dumps(record["metrics_json"])

    assert final_record["metrics_json"]["short_payment_rows"] == 1
    assert final_record["metrics_json"]["orders_sync_upstream_status"] == "failed"
    assert final_record["metrics_json"]["orders_sync_upstream_run_id"] == "orders-run-1"
    assert final_record["metrics_json"]["orders_sync_is_degraded"] is True
    assert "Orders sync failed before this report; data may be stale." in final_record["summary_text"]
    assert final_record["phases_json"]["upstream_orders_sync"]["warning"] == 1
    assert final_record["phases_json"]["send_email"]["ok"] == 1


@pytest.mark.asyncio
async def test_daily_pipeline_reaches_render_when_mtd_fetch_invoked_without_reflection(
    tmp_path, monkeypatch
) -> None:
    db_path = tmp_path / "daily_sales_pipeline_no_reflection.db"
    database_url = f"sqlite+aiosqlite:///{db_path}"
    _create_tables(database_url)

    report_date = date(2026, 4, 29)
    async with session_scope(database_url) as session:
        await session.execute(
            sa.text(
                "INSERT INTO store_master (cost_center, store_code) VALUES ('CC1', 'S1')"
            )
        )
        await session.execute(
            sa.text(
                "INSERT INTO orders (cost_center, order_number, order_date, customer_name, mobile_number, net_amount) VALUES ('CC1', 'RPT-1', '2026-04-29T09:00:00+05:30', 'Alice', '9999999999', 900)"
            )
        )
        await session.execute(
            sa.text(
                "INSERT INTO order_line_items (cost_center, order_number, service_name, garment_name) VALUES ('CC1', 'RPT-1', 'Wash', 'Shirt')"
            )
        )
        await session.execute(
            sa.text(
                "INSERT INTO sales (cost_center, order_number, payment_date, payment_mode, payment_received) VALUES ('CC1', 'RPT-1', '2026-04-29T10:00:00+05:30', 'UPI', 900)"
            )
        )
        await session.commit()

    monkeypatch.setattr(
        pipeline,
        "config",
        SimpleNamespace(database_url=database_url, pdf_render_timeout_seconds=30),
    )
    monkeypatch.setattr(pipeline, "get_timezone", lambda: ZoneInfo("Asia/Kolkata"))
    monkeypatch.setattr(pipeline, "resolve_run_env", lambda env: "test")
    monkeypatch.setattr(pipeline, "new_run_id", lambda: "run-mtd-2")
    monkeypatch.setattr(
        mtd_data.sa,
        "create_engine",
        lambda *a, **k: (_ for _ in ()).throw(
            AssertionError("unexpected create_engine")
        ),
    )

    async def _summary_noop(*args, **kwargs):
        return None

    monkeypatch.setattr(pipeline, "persist_summary_record", _summary_noop)
    monkeypatch.setattr(pipeline, "update_summary_record", _summary_noop)

    async def _fake_daily_data(*args, **kwargs):
        return SimpleNamespace(
            report_date=report_date,
            rows=[],
            totals=SimpleNamespace(),
            edited_orders=[],
            edited_orders_summary=SimpleNamespace(),
            edited_orders_totals=SimpleNamespace(),
            missed_leads=[],
            cancelled_leads=[],
            lead_performance_summary=SimpleNamespace(),
            to_be_recovered=[],
            to_be_compensated=[],
            to_be_recovered_total_order_value=0,
            to_be_compensated_total_order_value=0,
            same_day_fulfillment_rows=[SimpleNamespace(order_number="RPT-1")],
            missing_payment_rows=[],
            short_payment_rows=[],
        )

    monkeypatch.setattr(pipeline, "fetch_daily_sales_report", _fake_daily_data)
    monkeypatch.setattr(pipeline, "_render_html", lambda context, *args, **kwargs: "ok")

    async def _fake_render_pdf(*args, **kwargs):
        return None

    async def _fake_notify(*args, **kwargs):
        return {"emails_planned": 0, "emails_sent": 0, "errors": []}

    monkeypatch.setattr(
        pipeline, "render_pdf_with_configured_browser", _fake_render_pdf
    )
    monkeypatch.setattr(pipeline, "send_notifications_for_run", _fake_notify)

    await pipeline._run(report_date=report_date, env="test", force=True)


@pytest.mark.asyncio
async def test_daily_pipeline_continues_when_mtd_fetch_fails(
    tmp_path, monkeypatch
) -> None:
    db_path = tmp_path / "daily_sales_pipeline_mtd_failure.db"
    database_url = f"sqlite+aiosqlite:///{db_path}"
    _create_tables(database_url)

    report_date = date(2026, 4, 29)
    async with session_scope(database_url) as session:
        await session.execute(
            sa.text(
                "INSERT INTO store_master (cost_center, store_code) VALUES ('CC1', 'S1')"
            )
        )
        await session.execute(
            sa.text(
                "INSERT INTO orders (cost_center, order_number, order_date, customer_name, mobile_number, net_amount) VALUES ('CC1', 'RPT-2', '2026-04-29T09:00:00+05:30', 'Alice', '9999999999', 900)"
            )
        )
        await session.execute(
            sa.text(
                "INSERT INTO order_line_items (cost_center, order_number, service_name, garment_name) VALUES ('CC1', 'RPT-2', 'Wash', 'Shirt')"
            )
        )
        await session.execute(
            sa.text(
                "INSERT INTO sales (cost_center, order_number, payment_date, payment_mode, payment_received) VALUES ('CC1', 'RPT-2', '2026-04-29T10:00:00+05:30', 'UPI', 900)"
            )
        )
        await session.commit()

    monkeypatch.setattr(
        pipeline,
        "config",
        SimpleNamespace(database_url=database_url, pdf_render_timeout_seconds=30),
    )
    monkeypatch.setattr(pipeline, "get_timezone", lambda: ZoneInfo("Asia/Kolkata"))
    monkeypatch.setattr(pipeline, "resolve_run_env", lambda env: "test")
    monkeypatch.setattr(pipeline, "new_run_id", lambda: "run-mtd-failure")

    captured_records: list[dict] = []

    async def _capture_persist(_database_url, record):
        captured_records.append(record)

    async def _capture_update(_database_url, _run_id, record):
        captured_records.append(record)

    monkeypatch.setattr(pipeline, "persist_summary_record", _capture_persist)
    monkeypatch.setattr(pipeline, "update_summary_record", _capture_update)

    async def _fake_daily_data(*args, **kwargs):
        return SimpleNamespace(
            report_date=report_date,
            rows=[],
            totals=SimpleNamespace(),
            edited_orders=[],
            edited_orders_summary=SimpleNamespace(),
            edited_orders_totals=SimpleNamespace(),
            missed_leads=[],
            cancelled_leads=[],
            lead_performance_summary=SimpleNamespace(),
            to_be_recovered=[],
            to_be_compensated=[],
            to_be_recovered_total_order_value=0,
            to_be_compensated_total_order_value=0,
            same_day_fulfillment_rows=[SimpleNamespace(order_number="RPT-2")],
            missing_payment_rows=[],
            short_payment_rows=[],
        )

    async def _raise_mtd_fetch(*args, **kwargs):
        raise RuntimeError("mtd fetch boom")

    rendered_paths: list[str] = []

    async def _fake_render_pdf(html, output_path: Path, pdf_options=None, logger=None):
        rendered_paths.append(str(output_path))
        output_path.write_bytes(b"pdf")

    async def _fake_notify(*args, **kwargs):
        return {"emails_planned": 1, "emails_sent": 1, "errors": []}

    monkeypatch.setattr(pipeline, "fetch_daily_sales_report", _fake_daily_data)
    monkeypatch.setattr(pipeline, "fetch_mtd_same_day_fulfillment", _raise_mtd_fetch)
    monkeypatch.setattr(
        pipeline, "_render_html", lambda context, *args, **kwargs: "daily-html"
    )
    monkeypatch.setattr(
        pipeline, "render_pdf_with_configured_browser", _fake_render_pdf
    )
    monkeypatch.setattr(pipeline, "send_notifications_for_run", _fake_notify)

    await pipeline._run(report_date=report_date, env="test", force=True)

    daily_path = str(
        pipeline.OUTPUT_ROOT / f"{pipeline.PIPELINE_NAME}_{report_date.isoformat()}.pdf"
    )
    mtd_path = str(
        pipeline.OUTPUT_ROOT
        / f"reports.mtd_same_day_fulfillment_{report_date.isoformat()}.pdf"
    )

    short_payments_path = str(
        pipeline.OUTPUT_ROOT
        / f"reports.daily_sales_report_short_payments_{report_date.isoformat()}.pdf"
    )

    assert daily_path in rendered_paths
    assert short_payments_path in rendered_paths
    assert mtd_path not in rendered_paths

    async with session_scope(database_url) as session:
        rows = (
            (
                await session.execute(
                    sa.text(
                        """
                    SELECT doc_type, file_name
                    FROM documents
                    ORDER BY file_name
                    """
                    )
                )
            )
            .mappings()
            .all()
        )

    assert len(rows) == 5
    assert {row["doc_type"] for row in rows} == {
        "daily_sales_actual_payments_not_found_pdf",
        "daily_sales_actual_payments_not_found_xlsx",
        "daily_sales_report_pdf",
        "daily_sales_short_payments_pdf",
        "to_be_recovered_report_pdf",
    }

    assert captured_records
    final_record = captured_records[-1]
    assert final_record["overall_status"] == "warning"
    assert final_record["metrics_json"]["mtd_attachment_generated"] is False
    assert final_record["metrics_json"]["mtd_attachment_error"] == "mtd fetch boom"
    assert final_record["metrics_json"]["short_payment_rows"] == 0
    assert final_record["metrics_json"]["short_payments_pdf_generated"] is True
    assert final_record["metrics_json"]["actual_payments_not_found_rows"] == 0
    assert final_record["metrics_json"]["actual_payments_not_found_pdf_generated"] is True
    assert (
        "MTD same-day fulfillment attachment was not generated"
        in final_record["summary_text"]
    )


def test_actual_payments_not_found_workbook_groups_and_sorts_rows(tmp_path) -> None:
    output_path = tmp_path / "apnf.xlsx"
    rows = [
        SimpleNamespace(
            cost_center="CC/A",
            order_number="ORD-2",
            order_date=datetime(2026, 4, 28, 10, 0),
            customer_name="B",
            mobile_number="2",
            order_amount=Decimal("20"),
        ),
        SimpleNamespace(
            cost_center="CC/A",
            order_number="ORD-1",
            order_date=datetime(2026, 4, 29, 10, 0),
            customer_name="A",
            mobile_number="1",
            order_amount=Decimal("10"),
        ),
        SimpleNamespace(
            cost_center="CC:B",
            order_number="ORD-3",
            order_date=datetime(2026, 4, 27, 10, 0),
            customer_name="C",
            mobile_number="3",
            order_amount=Decimal("30"),
        ),
    ]

    pipeline._build_actual_payments_not_found_workbook(rows=rows, output_path=output_path, business_timezone=ZoneInfo("Asia/Kolkata"))

    workbook = openpyxl.load_workbook(output_path, data_only=True)
    try:
        assert set(workbook.sheetnames) == {"CC_A", "CC_B"}
        cc_a_rows = list(workbook["CC_A"].iter_rows(values_only=True))
        assert [row[1] for row in cc_a_rows[1:]] == ["ORD-1", "ORD-2"]
    finally:
        workbook.close()


def test_actual_payments_not_found_workbook_handles_zero_rows(tmp_path) -> None:
    output_path = tmp_path / "apnf-empty.xlsx"

    pipeline._build_actual_payments_not_found_workbook(rows=[], output_path=output_path, business_timezone=ZoneInfo("Asia/Kolkata"))

    workbook = openpyxl.load_workbook(output_path, data_only=True)
    try:
        assert workbook.sheetnames == ["No Data"]
        rows = list(workbook["No Data"].iter_rows(values_only=True))
        assert rows == [
            (
                "Cost Center",
                "Order Number",
                "Order Date",
                "Customer Name",
                "Mobile Number",
                "Order Amount",
            )
        ]
    finally:
        workbook.close()


def test_actual_payments_not_found_workbook_normalizes_tz_aware_datetimes(tmp_path) -> None:
    output_path = tmp_path / "apnf-tz-aware.xlsx"
    rows = [
        SimpleNamespace(
            cost_center="CC1",
            order_number="ORD-100",
            order_date=datetime(2026, 4, 29, 4, 30, tzinfo=ZoneInfo("UTC")),
            customer_name="Alice",
            mobile_number="123",
            order_amount=Decimal("100"),
        )
    ]

    pipeline._build_actual_payments_not_found_workbook(
        rows=rows,
        output_path=output_path,
        business_timezone=ZoneInfo("Asia/Kolkata"),
    )

    workbook = openpyxl.load_workbook(output_path, data_only=True)
    try:
        worksheet = workbook["CC1"]
        written_value = worksheet.cell(row=2, column=3).value
        assert isinstance(written_value, datetime)
        assert written_value == datetime(2026, 4, 29, 10, 0)
        assert written_value.tzinfo is None
        assert (
            worksheet.cell(row=2, column=3).number_format
            == pipeline.EXCEL_DATETIME_NUMBER_FORMAT
        )
    finally:
        workbook.close()


def test_daily_sales_context_and_template_mark_degraded_orders_sync() -> None:
    payload = SimpleNamespace(
        report_date=date(2026, 4, 29),
        rows=[],
        totals=SimpleNamespace(
            sales_ftd=Decimal("0"), sales_mtd=Decimal("0"), sales_lmtd=Decimal("0"),
            orders_count_ftd=0, orders_count_mtd=0, orders_count_lmtd=0,
            collections_ftd=Decimal("0"), collections_mtd=Decimal("0"), collections_lmtd=Decimal("0"),
            collections_count_ftd=0, collections_count_mtd=0, collections_count_lmtd=0,
            target=Decimal("0"), achieved=Decimal("0"), ttd=Decimal("0"),
            delta=Decimal("0"), reqd_per_day=Decimal("0"),
        ),
        edited_orders=[], edited_orders_summary=None, edited_orders_totals=None,
        missed_leads=[], cancelled_leads=[], lead_performance_summary=[],
        to_be_recovered=[], to_be_compensated=[],
        to_be_recovered_total_order_value=Decimal("0"),
        to_be_compensated_total_order_value=Decimal("0"),
        auto_cleared_order_numbers_text="", same_day_fulfillment_rows=[],
        missing_payment_rows=[], short_payment_rows=[], report_day_orders_by_cost_center=[],
    )
    upstream = pipeline.build_orders_sync_upstream_context(status="failed", run_id="orders-run-1")

    context = pipeline._build_context(payload, "prod", upstream)
    html = pipeline._render_html(context)

    assert context["orders_sync_is_degraded"] is True
    assert "Orders sync failed before this report; data may be stale." in html
    assert "Upstream orders sync status: failed; run ID: orders-run-1" in html
