from __future__ import annotations

import csv
from datetime import date, datetime, timezone
from pathlib import Path

import openpyxl
import pytest
import sqlalchemy as sa
from sqlalchemy.ext.asyncio import create_async_engine

from app.common.db import session_scope
from app.customer_retention.constants import SUPPRESSION_STATE_ACTIVE, SUPPRESSION_STATE_PENDING_APPROVAL, WORKBOOK_OUTCOME_SHIFTED_LOCATION
from app.customer_retention.db_tables import metadata, trx_customer_followup_history, trx_customer_followup_leads, trx_customer_suppression, trx_external_leads
from app.customer_retention.external_import import import_external_lead_file, parse_external_lead_file
from app.customer_retention.input_discovery import archive_processed_file, discover_external_lead_files, discover_returned_workbooks
from app.customer_retention.persistence import get_or_create_followup_lead
from app.customer_retention.mobile import MobileNormalizationStatus, normalize_mobile
from app.customer_retention.normalization import ValueNormalizer, normalize_value
from app.customer_retention.source_adapters import import_td_leads
from app.customer_retention.recovery_detection import detect_recoveries
from app.customer_retention.suppression import approve_suppression, check_active_suppression
from app.customer_retention.workbook_ingestor import ingest_returned_workbook


async def _prepare_db(tmp_path: Path) -> str:
    db = tmp_path / "phase2.db"
    url = f"sqlite+aiosqlite:///{db}"
    engine = create_async_engine(url, future=True)
    async with engine.begin() as conn:
        await conn.run_sync(metadata.create_all)
        await conn.execute(sa.text("CREATE TABLE store_master (cost_center TEXT PRIMARY KEY, store_code TEXT, sync_group TEXT, customer_retention_pipeline BOOLEAN NOT NULL DEFAULT 1)"))
        await conn.execute(sa.text("INSERT INTO store_master (cost_center, store_code, sync_group, customer_retention_pipeline) VALUES ('A100', 'SC-A', 'TD', 1), ('B200', 'SC-B', 'TD', 0), ('C300', 'SC-C', 'UC', 1)"))
        await conn.execute(sa.text("""
            CREATE TABLE crm_leads_current (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                lead_uid TEXT NOT NULL UNIQUE,
                store_code TEXT NOT NULL,
                pickup_no TEXT NOT NULL,
                status_bucket TEXT NOT NULL,
                customer_name TEXT,
                mobile TEXT,
                pickup_date TEXT,
                pickup_created_at TIMESTAMP,
                special_instruction TEXT,
                reason TEXT,
                source TEXT,
                customer_type TEXT,
                run_id TEXT,
                source_file TEXT,
                scraped_at TIMESTAMP
            )
        """))
    await engine.dispose()
    return url


def _write_external_csv(path: Path, rows: list[dict[str, object]]) -> None:
    with path.open("w", newline="", encoding="utf-8") as handle:
        writer = csv.DictWriter(handle, fieldnames=["cost_center", "customer_name", "mobile_number", "lead_source", "campaign_name", "lead_date", "remarks"])
        writer.writeheader()
        writer.writerows(rows)


def _write_workbook(path: Path, rows: list[dict[str, object]]) -> None:
    wb = openpyxl.Workbook()
    readme = wb.active
    readme.title = "READ_ME"
    readme.append(["Instructions"])
    ws = wb.create_sheet("FOLLOWUP_LEADS")
    headers = ["lead_id", "lead_source_type", "cost_center", "customer_name", "mobile_number", "generated_at", "Contact Attempted", "Contact Mode", "Customer Response", "Order Expected", "Next Follow-up Date", "Complaint", "Do Not Contact", "Handled By", "Staff Remarks", "Target Cost Center"]
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])
    wb.save(path)




async def _insert_workbook_lead(
    url: str,
    *,
    lead_id: int = 1,
    cost_center: str = "A100",
    mobile: str = "9876543210",
    lead_status: str = "OPEN",
    lead_date: date = date(2026, 6, 1),
) -> None:
    async with session_scope(url) as session:
        await session.execute(trx_customer_followup_leads.insert().values(
            lead_id=lead_id,
            lead_uuid=f"lead-{lead_id}",
            lead_source_type="EXTERNAL",
            source_system="test",
            source_table_name="source",
            source_record_id=str(lead_id),
            cost_center=cost_center,
            customer_name=f"Customer {lead_id}",
            mobile_number=mobile,
            normalized_mobile_number=mobile,
            lead_date=lead_date,
            lead_status=lead_status,
            contact_attempted=False,
            complaint_flag=False,
            do_not_contact_flag=False,
            is_closed=lead_status == "CLOSED",
            is_recovered=False,
            suppression_applied=False,
            created_at=datetime.now(timezone.utc),
            updated_at=datetime.now(timezone.utc),
        ))
        await session.commit()


def test_mobile_normalization_statuses_and_variants() -> None:
    assert normalize_mobile("98765 43210").normalized_mobile == "9876543210"
    assert normalize_mobile("+91-98765-43210").normalized_mobile == "9876543210"
    assert normalize_mobile("09876543210").normalized_mobile == "9876543210"
    assert normalize_mobile(" ").status == MobileNormalizationStatus.BLANK
    assert normalize_mobile("abc987").status == MobileNormalizationStatus.MALFORMED
    assert normalize_mobile("1111111111").status == MobileNormalizationStatus.UNNORMALIZABLE


def test_value_normalization_examples_and_unknowns() -> None:
    assert normalize_value(" whatsapp ").normalized_value == "WhatsApp Sent"
    assert normalize_value("watsapp").normalized_value == "WhatsApp Sent"
    assert normalize_value("WA sent").normalized_value == "WhatsApp Sent"
    assert normalize_value("no resp").normalized_value == "No Response"
    assert normalize_value("not interested").normalized_value == "Not Interested"
    assert normalize_value("dnd").normalized_value == "Do Not Contact"
    assert normalize_value("wrong no").normalized_value == "Wrong Number"
    assert normalize_value("invalid").normalized_value == "Invalid Number"
    assert normalize_value("pickup").normalized_value == "Pickup Requested"
    unknown = ValueNormalizer().normalize("mystery", allowed_values={"No Response"}, field_name="Customer Response")
    assert unknown.invalid is True
    assert unknown.warning_code == "value_unrecognized"


def test_external_xlsx_parser_reads_supported_phase2_columns(tmp_path: Path) -> None:
    path = tmp_path / "external.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["cost_center", "customer_name", "mobile_number", "lead_source", "campaign_name", "lead_date", "remarks"])
    ws.append(["A100", "Ada", "9876543210", "Meta", "June", "2026-06-01", "ok"])
    wb.save(path)
    rows = parse_external_lead_file(path)
    assert rows == [(2, {"cost_center": "A100", "customer_name": "Ada", "mobile_number": "9876543210", "lead_source": "Meta", "campaign_name": "June", "lead_date": "2026-06-01", "remarks": "ok"})]


@pytest.mark.asyncio
async def test_external_csv_import_idempotency_and_warnings(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    csv_path = tmp_path / "leads.csv"
    _write_external_csv(csv_path, [
        {"cost_center": "A100", "customer_name": "Ada", "mobile_number": "9876543210", "lead_source": "Meta", "campaign_name": "June", "lead_date": "2026-06-01", "remarks": "ok"},
        {"cost_center": "A100", "customer_name": "Bad", "mobile_number": "abc", "lead_source": "Meta", "campaign_name": "June", "lead_date": "2026-06-01", "remarks": "bad"},
        {"cost_center": "", "customer_name": "No Store", "mobile_number": "9876543211", "lead_source": "Meta", "campaign_name": "June", "lead_date": "2026-06-01", "remarks": "bad"},
    ])
    first = await import_external_lead_file(database_url=url, path=csv_path, pipeline_run_id="run1")
    second = await import_external_lead_file(database_url=url, path=csv_path, pipeline_run_id="run2")
    assert first.raw_rows_inserted == 3
    assert first.leads_created == 1
    assert {w.code for w in first.warnings} == {"mobile_malformed", "missing_required_field"}
    assert second.raw_rows_existing == 3
    assert second.leads_existing == 1
    async with session_scope(url) as session:
        assert (await session.execute(sa.select(sa.func.count()).select_from(trx_external_leads))).scalar_one() == 3
        assert (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_leads))).scalar_one() == 1


@pytest.mark.asyncio
async def test_external_import_blank_required_values_block_conversion_and_batch_continues(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    csv_path = tmp_path / "blank-required-values.csv"
    _write_external_csv(csv_path, [
        {"cost_center": "A100", "customer_name": "  ", "mobile_number": "9876543210", "lead_source": "", "campaign_name": "   ", "lead_date": "2026-06-01", "remarks": "  "},
        {"cost_center": "A100", "customer_name": "Blank Mobile", "mobile_number": " ", "lead_source": "Meta", "campaign_name": "June", "lead_date": "2026-06-01", "remarks": "bad"},
        {"cost_center": "A100", "customer_name": "Blank Date", "mobile_number": "9876543211", "lead_source": "Meta", "campaign_name": "June", "lead_date": " ", "remarks": "bad"},
        {"cost_center": " ", "customer_name": "Blank Store", "mobile_number": "9876543212", "lead_source": "Meta", "campaign_name": "June", "lead_date": "2026-06-01", "remarks": "bad"},
        {"cost_center": "A100", "customer_name": "Grace", "mobile_number": "9876543213", "lead_source": "Referral", "campaign_name": "June", "lead_date": "2026-06-02", "remarks": "ok"},
    ])

    result = await import_external_lead_file(database_url=url, path=csv_path, pipeline_run_id="run1")

    assert result.rows_seen == 5
    assert result.raw_rows_inserted == 4
    assert result.leads_created == 1
    assert result.rows_skipped == 4
    warning_pairs = [(warning.code, warning.field_name, warning.row_number) for warning in result.warnings]
    assert warning_pairs[:4] == [
        ("missing_required_field", "customer_name", 2),
        ("missing_required_field", "lead_source", 2),
        ("missing_required_field", "campaign_name", 2),
        ("missing_required_field", "remarks", 2),
    ]
    assert ("missing_required_field", "mobile_number", 3) in warning_pairs
    assert ("missing_required_field", "lead_date", 4) in warning_pairs
    assert ("invalid_lead_date", "lead_date", 4) in warning_pairs
    assert ("missing_required_field", "cost_center", 5) in warning_pairs
    async with session_scope(url) as session:
        external_rows = (
            await session.execute(
                sa.select(
                    trx_external_leads.c.customer_name,
                    trx_external_leads.c.lead_source,
                    trx_external_leads.c.campaign_name,
                    trx_external_leads.c.remarks,
                    trx_external_leads.c.converted_to_followup_lead,
                ).order_by(trx_external_leads.c.external_lead_id)
            )
        ).all()
        assert external_rows == [
            (None, "", None, None, False),
            ("Blank Mobile", "Meta", "June", "bad", False),
            ("Blank Store", "Meta", "June", "bad", False),
            ("Grace", "Referral", "June", "ok", True),
        ]
        assert (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_leads))).scalar_one() == 1


@pytest.mark.asyncio
@pytest.mark.parametrize("blank_field", ["lead_source", "campaign_name", "customer_name", "remarks"])
async def test_external_import_blank_descriptive_required_field_persists_raw_error_without_conversion(tmp_path: Path, blank_field: str) -> None:
    url = await _prepare_db(tmp_path)
    csv_path = tmp_path / f"blank-{blank_field}.csv"
    row = {
        "cost_center": "A100",
        "customer_name": "Ada",
        "mobile_number": "9876543210",
        "lead_source": "Meta",
        "campaign_name": "June",
        "lead_date": "2026-06-01",
        "remarks": "ok",
    }
    row[blank_field] = "  "
    _write_external_csv(csv_path, [row])

    result = await import_external_lead_file(database_url=url, path=csv_path, pipeline_run_id="run1")

    assert result.rows_seen == 1
    assert result.raw_rows_inserted == 1
    assert result.rows_skipped == 1
    assert result.leads_created == 0
    assert [(warning.code, warning.field_name, warning.row_number) for warning in result.warnings] == [("missing_required_field", blank_field, 2)]
    async with session_scope(url) as session:
        external = (
            await session.execute(
                sa.select(
                    trx_external_leads.c.lead_status,
                    trx_external_leads.c.converted_to_followup_lead,
                    trx_external_leads.c.converted_followup_lead_id,
                    trx_external_leads.c.raw_payload_json,
                )
            )
        ).one()
        assert external.lead_status == "ERROR"
        assert external.converted_to_followup_lead is False
        assert external.converted_followup_lead_id is None
        assert external.raw_payload_json[blank_field].strip() == ""
        assert (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_leads))).scalar_one() == 0


@pytest.mark.asyncio
async def test_external_import_persists_invalid_mobile_raw_without_conversion(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    csv_path = tmp_path / "invalid-mobile-external.csv"
    _write_external_csv(csv_path, [
        {"cost_center": "A100", "customer_name": "Bad Mobile", "mobile_number": "not-a-phone", "lead_source": "Meta", "campaign_name": "June", "lead_date": "2026-06-01", "remarks": "keep raw"},
    ])

    result = await import_external_lead_file(database_url=url, path=csv_path, pipeline_run_id="run1")

    assert result.rows_seen == 1
    assert result.raw_rows_inserted == 1
    assert result.rows_skipped == 1
    assert result.leads_created == 0
    assert [(warning.code, warning.field_name) for warning in result.warnings] == [("mobile_malformed", "mobile_number")]
    async with session_scope(url) as session:
        external = (
            await session.execute(
                sa.select(
                    trx_external_leads.c.raw_payload_json,
                    trx_external_leads.c.import_batch_id,
                    trx_external_leads.c.lead_source,
                    trx_external_leads.c.campaign_name,
                    trx_external_leads.c.mobile_number,
                    trx_external_leads.c.normalized_mobile_number,
                    trx_external_leads.c.lead_status,
                    trx_external_leads.c.converted_to_followup_lead,
                    trx_external_leads.c.converted_followup_lead_id,
                )
            )
        ).one()
        assert external.raw_payload_json["mobile_number"] == "not-a-phone"
        assert external.import_batch_id == result.import_batch_id
        assert external.lead_source == "Meta"
        assert external.campaign_name == "June"
        assert external.mobile_number == "not-a-phone"
        assert external.normalized_mobile_number == ""
        assert external.lead_status == "ERROR"
        assert external.converted_to_followup_lead is False
        assert external.converted_followup_lead_id is None
        assert (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_leads))).scalar_one() == 0


@pytest.mark.asyncio
async def test_td_pending_rows_same_store_and_mobile_share_unified_lead(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    async with session_scope(url) as session:
        await session.execute(sa.text("""
            INSERT INTO crm_leads_current (lead_uid, store_code, pickup_no, status_bucket, customer_name, mobile, pickup_created_at, run_id, scraped_at)
            VALUES ('TD-DUPE-1', 'SC-A', 'P1', 'pending', 'Ada One', '9876543210', '2026-06-01 00:00:00+00:00', 'r1', '2026-06-01 00:00:00+00:00'),
                   ('TD-DUPE-2', 'SC-A', 'P2', 'pending', 'Ada Two', '+91-98765-43210', '2026-06-01 00:00:00+00:00', 'r1', '2026-06-01 00:00:00+00:00')
        """))
        await session.commit()

    result = await import_td_leads(database_url=url, pipeline_run_id="run1")

    assert result.rows_seen == 2
    assert result.leads_created == 1
    assert result.leads_existing == 1
    async with session_scope(url) as session:
        assert (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_leads))).scalar_one() == 1


@pytest.mark.asyncio
async def test_external_import_batches_same_cost_center_and_mobile_share_unified_lead(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    first_path = tmp_path / "leads-first.csv"
    second_path = tmp_path / "leads-second.csv"
    _write_external_csv(first_path, [
        {"cost_center": "A100", "customer_name": "Ada One", "mobile_number": "9876543210", "lead_source": "Meta", "campaign_name": "June", "lead_date": "2026-06-01", "remarks": "first"},
    ])
    _write_external_csv(second_path, [
        {"cost_center": "A100", "customer_name": "Ada Two", "mobile_number": "+91-98765-43210", "lead_source": "Google", "campaign_name": "July", "lead_date": "2026-06-02", "remarks": "second"},
    ])

    first = await import_external_lead_file(database_url=url, path=first_path, pipeline_run_id="run1")
    second = await import_external_lead_file(database_url=url, path=second_path, pipeline_run_id="run2")

    assert first.raw_rows_inserted == 1
    assert first.leads_created == 1
    assert second.raw_rows_inserted == 1
    assert second.leads_existing == 1
    async with session_scope(url) as session:
        assert (await session.execute(sa.select(sa.func.count()).select_from(trx_external_leads))).scalar_one() == 2
        assert (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_leads))).scalar_one() == 1
        converted_ids = (await session.execute(sa.select(trx_external_leads.c.converted_followup_lead_id).order_by(trx_external_leads.c.external_lead_id))).scalars().all()
        assert converted_ids == [1, 1]


@pytest.mark.asyncio
async def test_source_record_match_wins_before_customer_identity_dedupe(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    async with session_scope(url) as session:
        await session.execute(trx_customer_followup_leads.insert().values(
            lead_id=1,
            lead_uuid="lead-source-exact",
            lead_source_type="TD",
            source_system="TD_CRM_LEADS_SYNC",
            source_table_name="crm_leads_current",
            source_record_id="TD-SAME",
            cost_center="A100",
            customer_name="Closed Source Lead",
            mobile_number="9876543210",
            normalized_mobile_number="9876543210",
            lead_date=date(2026, 6, 1),
            lead_status="CLOSED",
            is_closed=True,
            contact_attempted=False,
            complaint_flag=False,
            do_not_contact_flag=False,
            created_at=datetime.now(timezone.utc),
            updated_at=datetime.now(timezone.utc),
        ))
        await session.execute(trx_customer_followup_leads.insert().values(
            lead_id=2,
            lead_uuid="lead-open-customer",
            lead_source_type="EXTERNAL",
            source_system="CUSTOMER_FOLLOWUP_EXTERNAL_IMPORT",
            source_table_name="trx_external_leads",
            source_record_id="external-other",
            cost_center="A100",
            customer_name="Open Customer Lead",
            mobile_number="9876543210",
            normalized_mobile_number="9876543210",
            lead_date=date(2026, 6, 2),
            lead_status="OPEN",
            is_closed=False,
            contact_attempted=False,
            complaint_flag=False,
            do_not_contact_flag=False,
            created_at=datetime.now(timezone.utc),
            updated_at=datetime.now(timezone.utc),
        ))
        lead_id, created = await get_or_create_followup_lead(
            session,
            lead_source_type="TD",
            source_system="TD_CRM_LEADS_SYNC",
            source_table_name="crm_leads_current",
            source_record_id="TD-SAME",
            source_reference="P1",
            cost_center="A100",
            customer_name="Reprocessed",
            mobile_number="9876543210",
            normalized_mobile_number="9876543210",
            lead_date=date(2026, 6, 3),
            pipeline_run_id="run2",
            dedupe_by_customer_identity=True,
        )

    assert lead_id == 1
    assert created is False


@pytest.mark.asyncio
async def test_td_adapter_import_idempotency_and_invalid_mobile_warning(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    async with session_scope(url) as session:
        await session.execute(sa.text("""
            INSERT INTO crm_leads_current (lead_uid, store_code, pickup_no, status_bucket, customer_name, mobile, pickup_created_at, run_id, scraped_at)
            VALUES ('TD1', 'SC-A', 'P1', 'pending', 'Ada', '9876543210', '2026-06-01 00:00:00+00:00', 'r1', '2026-06-01 00:00:00+00:00'),
                   ('TD2', 'SC-A', 'P2', 'pending', 'Bad', 'abc', '2026-06-01 00:00:00+00:00', 'r1', '2026-06-01 00:00:00+00:00'),
                   ('TD3', 'SC-A', 'P3', 'completed', 'Done', '9876543211', '2026-06-01 00:00:00+00:00', 'r1', '2026-06-01 00:00:00+00:00')
        """))
        await session.commit()
    first = await import_td_leads(database_url=url, pipeline_run_id="run1")
    second = await import_td_leads(database_url=url, pipeline_run_id="run2")
    assert first.rows_seen == 2
    assert first.leads_created == 1
    assert [w.code for w in first.warnings] == ["mobile_malformed"]
    assert second.leads_existing == 1
    async with session_scope(url) as session:
        assert (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_leads))).scalar_one() == 1


@pytest.mark.asyncio
async def test_td_adapter_resolves_store_code_to_active_retention_cost_center(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    async with session_scope(url) as session:
        await session.execute(sa.text("""
            INSERT INTO crm_leads_current (lead_uid, store_code, pickup_no, status_bucket, customer_name, mobile, pickup_created_at, run_id, scraped_at)
            VALUES ('TD-ACTIVE', 'SC-A', 'P-ACTIVE', 'pending', 'Ada', '9876543210', '2026-06-01 00:00:00+00:00', 'r1', '2026-06-01 00:00:00+00:00'),
                   ('TD-INACTIVE', 'SC-B', 'P-INACTIVE', 'pending', 'Bea', '9876543211', '2026-06-01 00:00:00+00:00', 'r1', '2026-06-01 00:00:00+00:00'),
                   ('TD-WRONG-GROUP', 'SC-C', 'P-WRONG-GROUP', 'pending', 'Cara', '9876543212', '2026-06-01 00:00:00+00:00', 'r1', '2026-06-01 00:00:00+00:00'),
                   ('TD-UNMAPPED', 'SC-X', 'P-UNMAPPED', 'pending', 'Dee', '9876543213', '2026-06-01 00:00:00+00:00', 'r1', '2026-06-01 00:00:00+00:00')
        """))
        await session.commit()

    result = await import_td_leads(database_url=url, pipeline_run_id="run1")

    assert result.rows_seen == 4
    assert result.leads_created == 1
    assert result.rows_skipped == 3
    assert [warning.code for warning in result.warnings] == ["td_store_inactive", "td_store_inactive", "td_store_unmapped"]
    async with session_scope(url) as session:
        lead = (await session.execute(sa.select(trx_customer_followup_leads))).mappings().one()
        assert lead["cost_center"] == "A100"
        assert lead["assigned_store"] == "A100"
        assert lead["source_record_id"] == "TD-ACTIVE"
        assert lead["source_reference"] == "P-ACTIVE"


@pytest.mark.asyncio
async def test_td_cost_center_mapping_keeps_suppression_and_recovery_store_scoped(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    async with session_scope(url) as session:
        await session.execute(sa.text("""
            INSERT INTO crm_leads_current (lead_uid, store_code, pickup_no, status_bucket, customer_name, mobile, pickup_created_at, run_id, scraped_at)
            VALUES ('TD-A100', 'SC-A', 'P-A100', 'pending', 'Ada', '9876543210', '2026-06-01 00:00:00+00:00', 'r1', '2026-06-01 00:00:00+00:00')
        """))
        await session.execute(trx_customer_suppression.insert().values(
            suppression_id=1,
            cost_center="SC-A",
            mobile_number="9876543210",
            normalized_mobile_number="9876543210",
            suppression_reason="Not Interested",
            suppression_state=SUPPRESSION_STATE_ACTIVE,
            suppression_start_date=date(2026, 6, 1),
            suppression_until=date(2026, 7, 1),
            is_permanent=False,
            approval_required=False,
            created_at=datetime.now(timezone.utc),
        ))
        await session.execute(trx_customer_suppression.insert().values(
            suppression_id=2,
            cost_center="A100",
            mobile_number="9876543210",
            normalized_mobile_number="9876543210",
            suppression_reason="Not Interested",
            suppression_state=SUPPRESSION_STATE_ACTIVE,
            suppression_start_date=date(2026, 6, 1),
            suppression_until=date(2026, 7, 1),
            is_permanent=False,
            approval_required=False,
            created_at=datetime.now(timezone.utc),
        ))
        await session.execute(sa.text("""
            CREATE TABLE vw_orders (
                cost_center TEXT,
                order_number TEXT,
                order_date TIMESTAMP,
                customer_name TEXT,
                mobile_number TEXT,
                order_amount NUMERIC
            )
        """))
        await session.execute(sa.text("""
            INSERT INTO vw_orders (cost_center, order_number, order_date, customer_name, mobile_number, order_amount)
            VALUES ('SC-A', 'ORD-STORE-CODE', '2026-06-02 00:00:00+00:00', 'Ada', '9876543210', 100),
                   ('A100', 'ORD-COST-CENTER', '2026-06-03 00:00:00+00:00', 'Ada', '9876543210', 200)
        """))
        await session.commit()

    import_result = await import_td_leads(database_url=url, pipeline_run_id="run1")
    assert import_result.leads_created == 1

    async with session_scope(url) as session:
        lead = (await session.execute(sa.select(trx_customer_followup_leads))).mappings().one()
        assert lead["cost_center"] == "A100"
        store_code_suppression = await check_active_suppression(session, cost_center="SC-A", normalized_mobile_number="9876543210", as_of_date=date(2026, 6, 2))
        cost_center_suppression = await check_active_suppression(session, cost_center="A100", normalized_mobile_number="9876543210", as_of_date=date(2026, 6, 2))
        assert store_code_suppression.is_suppressed is True
        assert cost_center_suppression.is_suppressed is True

        recovery = await detect_recoveries(session, as_of_date=date(2026, 6, 4), pipeline_run_id="run2")
        await session.commit()

        assert recovery.leads_recovered == 1
        assert [match.recovered_order_id for match in recovery.matches] == ["ORD-COST-CENTER"]
        recovered = (await session.execute(sa.select(trx_customer_followup_leads.c.recovered_order_id))).scalar_one()
        assert recovered == "ORD-COST-CENTER"


@pytest.mark.asyncio
async def test_workbook_duplicate_upload_protected_edits_invalid_mobile_and_required_blank(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    async with session_scope(url) as session:
        await session.execute(trx_customer_followup_leads.insert().values(
            lead_id=1,
            lead_uuid="lead-1",
            lead_source_type="EXTERNAL",
            source_system="test",
            source_table_name="source",
            source_record_id="1",
            cost_center="A100",
            customer_name="Ada",
            mobile_number="9876543210",
            normalized_mobile_number="9876543210",
            lead_date=date(2026, 6, 1),
            lead_status="OPEN",
            contact_attempted=False,
            complaint_flag=False,
            do_not_contact_flag=False,
            created_at=datetime.now(timezone.utc),
            updated_at=datetime.now(timezone.utc),
        ))
        await session.commit()
    wb_path = tmp_path / "returned.xlsx"
    _write_workbook(wb_path, [
        {"lead_id": 1, "lead_source_type": "TD", "cost_center": "A100", "customer_name": "Edited", "mobile_number": "9876543210", "Contact Attempted": "Yes", "Customer Response": "wa sent", "Complaint": "No", "Do Not Contact": "No", "Handled By": "Staff"},
        {"lead_id": 1, "lead_source_type": "EXTERNAL", "cost_center": "A100", "customer_name": "Ada", "mobile_number": "1111111111", "Contact Attempted": "Yes", "Customer Response": "No Response", "Complaint": "No", "Do Not Contact": "No"},
        {"lead_id": 1, "lead_source_type": "EXTERNAL", "cost_center": "A100", "customer_name": "Ada", "mobile_number": "9876543210", "Contact Attempted": "", "Customer Response": "", "Complaint": "No", "Do Not Contact": "No"},
    ])
    first = await ingest_returned_workbook(database_url=url, path=wb_path, pipeline_run_id="run1")
    second = await ingest_returned_workbook(database_url=url, path=wb_path, pipeline_run_id="run2")
    assert first.history_inserted == 2
    assert first.protected_edits_ignored >= 1
    assert "mobile_unnormalizable" in {w.code for w in first.warnings}
    assert first.rows_pending_not_updated == 1
    assert second.history_existing == 2
    async with session_scope(url) as session:
        assert (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_history))).scalar_one() == 2
        lead_row = (
            await session.execute(
                sa.select(
                    trx_customer_followup_leads.c.customer_name,
                    trx_customer_followup_leads.c.lead_status,
                    trx_customer_followup_leads.c.lead_stage,
                    trx_customer_followup_leads.c.staff_remarks,
                    trx_customer_followup_leads.c.is_closed,
                    trx_customer_followup_leads.c.is_recovered,
                    trx_customer_followup_leads.c.suppression_applied,
                ).where(trx_customer_followup_leads.c.lead_id == 1)
            )
        ).one()
        assert lead_row.customer_name == "Ada"
        assert lead_row.lead_status == "WORKED"
        assert lead_row.lead_stage is None
        assert lead_row.staff_remarks is None
        assert lead_row.is_closed is False
        assert lead_row.is_recovered is False
        assert lead_row.suppression_applied is False
        pending_event = (
            await session.execute(
                sa.select(
                    trx_customer_followup_history.c.event_type,
                    trx_customer_followup_history.c.previous_status,
                    trx_customer_followup_history.c.new_status,
                ).where(trx_customer_followup_history.c.event_type.like("Pending_Not_Updated:%"))
            )
        ).one()
        assert pending_event.previous_status == "WORKED"
        assert pending_event.new_status == "WORKED"


@pytest.mark.asyncio
async def test_workbook_blank_handled_by_warns_but_ingests_worked_row(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    await _insert_workbook_lead(url)
    wb_path = tmp_path / "blank_handler.xlsx"
    _write_workbook(wb_path, [
        {
            "lead_id": 1,
            "lead_source_type": "EXTERNAL",
            "cost_center": "A100",
            "customer_name": "Customer 1",
            "mobile_number": "9876543210",
            "Contact Attempted": "Yes",
            "Customer Response": "No Response",
            "Complaint": "No",
            "Do Not Contact": "No",
            "Handled By": "  ",
        },
    ])

    result = await ingest_returned_workbook(database_url=url, path=wb_path, pipeline_run_id="run-blank-handler")

    handled_by_warnings = [warning for warning in result.warnings if warning.code == "handled_by_blank"]
    assert len(handled_by_warnings) == 1
    assert handled_by_warnings[0].field_name == "handled_by"
    assert result.history_inserted == 1
    assert result.rows_pending_not_updated == 0
    async with session_scope(url) as session:
        lead = (await session.execute(sa.select(trx_customer_followup_leads.c.lead_status, trx_customer_followup_leads.c.handled_by).where(trx_customer_followup_leads.c.lead_id == 1))).one()
        history = (await session.execute(sa.select(trx_customer_followup_history.c.handled_by, trx_customer_followup_history.c.customer_response).where(trx_customer_followup_history.c.pipeline_run_id == "run-blank-handler"))).one()
    assert lead.lead_status == "WORKED"
    assert lead.handled_by is None
    assert history.handled_by is None
    assert history.customer_response == "No Response"


@pytest.mark.asyncio
async def test_workbook_lifecycle_transition_updates_leads_and_suppressions(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    async with session_scope(url) as session:
        for lead_id, mobile in ((1, "9876543210"), (2, "9876543211"), (3, "9876543212"), (4, "9876543213")):
            await session.execute(trx_customer_followup_leads.insert().values(
                lead_id=lead_id,
                lead_uuid=f"lead-{lead_id}",
                lead_source_type="EXTERNAL",
                source_system="test",
                source_table_name="source",
                source_record_id=str(lead_id),
                cost_center="A100",
                customer_name=f"Customer {lead_id}",
                mobile_number=mobile,
                normalized_mobile_number=mobile,
                lead_date=date(2026, 6, 1),
                lead_status="OPEN",
                contact_attempted=False,
                complaint_flag=False,
                do_not_contact_flag=False,
                created_at=datetime.now(timezone.utc),
                updated_at=datetime.now(timezone.utc),
            ))
        await session.commit()

    wb_path = tmp_path / "lifecycle_returned.xlsx"
    _write_workbook(wb_path, [
        {"lead_id": 1, "lead_source_type": "EXTERNAL", "cost_center": "A100", "customer_name": "Customer 1", "mobile_number": "9876543210", "Contact Attempted": "Yes", "Contact Mode": "Call", "Customer Response": "Pickup Requested", "Order Expected": "Yes", "Next Follow-up Date": "2026-06-20", "Complaint": "No", "Do Not Contact": "No", "Handled By": "Staff", "Staff Remarks": "Will order"},
        {"lead_id": 2, "lead_source_type": "EXTERNAL", "cost_center": "A100", "customer_name": "Customer 2", "mobile_number": "9876543211", "generated_at": "2026-06-10", "Contact Attempted": "Yes", "Customer Response": "Not Interested", "Complaint": "No", "Do Not Contact": "No"},
        {"lead_id": 3, "lead_source_type": "EXTERNAL", "cost_center": "A100", "customer_name": "Customer 3", "mobile_number": "9876543212", "generated_at": "2026-06-10", "Contact Attempted": "Yes", "Customer Response": "Do Not Contact", "Complaint": "No", "Do Not Contact": "Yes"},
        {"lead_id": 4, "lead_source_type": "EXTERNAL", "cost_center": "A100", "customer_name": "Customer 4", "mobile_number": "9876543213", "Contact Attempted": "", "Customer Response": "", "Complaint": "No", "Do Not Contact": "No"},
    ])

    result = await ingest_returned_workbook(database_url=url, path=wb_path, pipeline_run_id="run-lifecycle")

    assert result.history_inserted == 4
    assert result.rows_pending_not_updated == 1
    async with session_scope(url) as session:
        leads = (
            await session.execute(
                sa.select(
                    trx_customer_followup_leads.c.lead_id,
                    trx_customer_followup_leads.c.lead_status,
                    trx_customer_followup_leads.c.customer_response,
                    trx_customer_followup_leads.c.next_followup_date,
                    trx_customer_followup_leads.c.staff_remarks,
                    trx_customer_followup_leads.c.is_closed,
                ).order_by(trx_customer_followup_leads.c.lead_id)
            )
        ).all()
        by_id = {row.lead_id: row for row in leads}
        assert by_id[1].lead_status == "DUE_FOLLOWUP"
        assert by_id[1].customer_response == "Pickup Requested"
        assert by_id[1].next_followup_date == date(2026, 6, 20)
        assert by_id[1].staff_remarks == "Will order"
        assert by_id[2].lead_status == "CLOSED"
        assert by_id[2].is_closed is True
        assert by_id[3].lead_status == "CLOSED"
        assert by_id[3].is_closed is True
        assert by_id[4].lead_status == "OPEN"
        assert by_id[4].customer_response is None

        suppressions = (
            await session.execute(
                sa.select(
                    trx_customer_suppression.c.source_lead_id,
                    trx_customer_suppression.c.suppression_state,
                    trx_customer_suppression.c.is_permanent,
                    trx_customer_suppression.c.suppression_until,
                ).order_by(trx_customer_suppression.c.source_lead_id)
            )
        ).all()
        assert len(suppressions) == 2
        assert suppressions[0].source_lead_id == 2
        assert suppressions[0].suppression_state == SUPPRESSION_STATE_ACTIVE
        assert suppressions[0].is_permanent is False
        assert suppressions[0].suppression_until is not None
        assert suppressions[0].suppression_until == date(2026, 9, 8)
        assert suppressions[1].source_lead_id == 3
        assert suppressions[1].suppression_state == SUPPRESSION_STATE_PENDING_APPROVAL
        assert suppressions[1].is_permanent is True
        assert suppressions[1].suppression_until is None

        assert (
            await session.execute(
                sa.select(sa.func.count()).select_from(trx_customer_suppression).where(
                    trx_customer_suppression.c.is_permanent.is_(True),
                    trx_customer_suppression.c.suppression_state == SUPPRESSION_STATE_ACTIVE,
                )
            )
        ).scalar_one() == 0
        pending_event = (
            await session.execute(
                sa.select(trx_customer_followup_history.c.previous_status, trx_customer_followup_history.c.new_status).where(
                    trx_customer_followup_history.c.lead_id == 4,
                    trx_customer_followup_history.c.event_type.like("Pending_Not_Updated:%"),
                )
            )
        ).one()
        assert pending_event.previous_status == "OPEN"
        assert pending_event.new_status == "OPEN"

@pytest.mark.asyncio
async def test_workbook_suppression_uses_deterministic_lead_date_when_process_date_changes(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    async def ingest_with_fake_today(case_name: str, fake_today: date) -> tuple[date, date, int, int, str]:
        case_dir = tmp_path / case_name
        case_dir.mkdir()
        url = await _prepare_db(case_dir)
        await _insert_workbook_lead(url, lead_id=30, mobile="9876543230", lead_date=date(2026, 5, 15))
        wb_path = case_dir / "returned_without_generated_at.xlsx"
        _write_workbook(wb_path, [
            {
                "lead_id": 30,
                "lead_source_type": "EXTERNAL",
                "cost_center": "A100",
                "customer_name": "Customer 30",
                "mobile_number": "9876543230",
                "Contact Attempted": "Yes",
                "Customer Response": "Not Interested",
                "Complaint": "No",
                "Do Not Contact": "No",
            },
        ])

        class FakeDate(date):
            @classmethod
            def today(cls) -> date:
                return fake_today

        import app.customer_retention.workbook_ingestor as workbook_ingestor

        monkeypatch.setattr(workbook_ingestor, "date", FakeDate)
        first = await ingest_returned_workbook(database_url=url, path=wb_path, pipeline_run_id=f"{case_name}-run-1")
        monkeypatch.setattr(workbook_ingestor, "date", FakeDate)
        replay = await ingest_returned_workbook(database_url=url, path=wb_path, pipeline_run_id=f"{case_name}-run-2")

        async with session_scope(url) as session:
            suppression = (await session.execute(sa.select(trx_customer_suppression))).mappings().one()
            event_type = (
                await session.execute(
                    sa.select(trx_customer_followup_history.c.event_type).where(
                        trx_customer_followup_history.c.event_type.like("LIFECYCLE_TRANSITION:%")
                    )
                )
            ).scalar_one()
        return suppression["suppression_start_date"], suppression["suppression_until"], first.history_inserted, replay.history_existing, event_type

    early = await ingest_with_fake_today("early", date(2026, 6, 13))
    late = await ingest_with_fake_today("late", date(2026, 7, 20))

    assert early[0] == late[0] == date(2026, 5, 15)
    assert early[1] == late[1] == date(2026, 8, 13)
    assert early[2:4] == late[2:4] == (1, 1)
    assert early[4].endswith(":2:Not Interested")
    assert late[4].endswith(":2:Not Interested")


@pytest.mark.asyncio
async def test_workbook_suppression_prefers_explicit_run_date_over_process_date(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    url = await _prepare_db(tmp_path)
    await _insert_workbook_lead(url, lead_id=31, mobile="9876543231", lead_date=date(2026, 5, 15))
    wb_path = tmp_path / "returned_with_generated_at.xlsx"
    _write_workbook(wb_path, [
        {
            "lead_id": 31,
            "lead_source_type": "EXTERNAL",
            "cost_center": "A100",
            "customer_name": "Customer 31",
            "mobile_number": "9876543231",
            "generated_at": "2026-06-10",
            "Contact Attempted": "Yes",
            "Customer Response": "Not Interested",
            "Complaint": "No",
            "Do Not Contact": "No",
        },
    ])

    class FakeDate(date):
        @classmethod
        def today(cls) -> date:
            return date(2030, 1, 1)

    import app.customer_retention.workbook_ingestor as workbook_ingestor

    monkeypatch.setattr(workbook_ingestor, "date", FakeDate)
    first = await ingest_returned_workbook(
        database_url=url,
        path=wb_path,
        pipeline_run_id="run-date-1",
        run_date=date(2026, 6, 13),
    )
    monkeypatch.setattr(workbook_ingestor, "date", FakeDate)
    replay = await ingest_returned_workbook(
        database_url=url,
        path=wb_path,
        pipeline_run_id="run-date-2",
        run_date=date(2026, 6, 13),
    )

    assert first.history_inserted == 1
    assert replay.history_existing == 1
    async with session_scope(url) as session:
        suppression = (await session.execute(sa.select(trx_customer_suppression))).mappings().one()
    assert suppression["suppression_start_date"] == date(2026, 6, 13)
    assert suppression["suppression_until"] == date(2026, 9, 11)


@pytest.mark.asyncio
async def test_shifted_location_valid_target_creates_one_idempotent_external_destination(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    await _insert_workbook_lead(url, lead_id=10, cost_center="A100", mobile="9876543210")
    wb_path = tmp_path / "shifted_valid.xlsx"
    _write_workbook(wb_path, [
        {"lead_id": 10, "lead_source_type": "EXTERNAL", "cost_center": "A100", "customer_name": "Customer 10", "mobile_number": "9876543210", "generated_at": "2026-06-10", "Contact Attempted": "Yes", "Customer Response": WORKBOOK_OUTCOME_SHIFTED_LOCATION, "Complaint": "No", "Do Not Contact": "No", "Target Cost Center": "c300"},
    ])

    first = await ingest_returned_workbook(database_url=url, path=wb_path, pipeline_run_id="run-shift-1")
    replay = await ingest_returned_workbook(database_url=url, path=wb_path, pipeline_run_id="run-shift-2")

    assert first.history_inserted == 1
    assert replay.history_existing == 1
    assert not first.warnings
    async with session_scope(url) as session:
        leads = (await session.execute(sa.select(trx_customer_followup_leads).order_by(trx_customer_followup_leads.c.lead_id))).mappings().all()
        assert len(leads) == 2
        source = next(row for row in leads if row["lead_id"] == 10)
        destination = next(row for row in leads if row["lead_id"] != 10)
        assert source["lead_status"] == "CLOSED"
        assert source["is_closed"] is True
        assert source["closed_reason"] == WORKBOOK_OUTCOME_SHIFTED_LOCATION
        assert source["target_cost_center"] == "C300"
        assert destination["lead_source_type"] == "EXTERNAL"
        assert destination["cost_center"] == "C300"
        assert destination["assigned_store"] == "C300"
        assert destination["target_cost_center"] == "C300"
        assert destination["shifted_from_lead_id"] == 10
        assert destination["shifted_from_cost_center"] == "A100"
        assert destination["source_system"] == "CUSTOMER_FOLLOWUP_SHIFTED_LOCATION"
        assert destination["source_record_id"] == "10:C300"
        assert (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_leads).where(trx_customer_followup_leads.c.shifted_from_lead_id == 10))).scalar_one() == 1


@pytest.mark.asyncio
async def test_shifted_location_blank_target_no_destination_and_normal_pending_suppression(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    await _insert_workbook_lead(url, lead_id=11, cost_center="A100", mobile="9876543211")
    wb_path = tmp_path / "shifted_blank.xlsx"
    _write_workbook(wb_path, [
        {"lead_id": 11, "lead_source_type": "EXTERNAL", "cost_center": "A100", "customer_name": "Customer 11", "mobile_number": "9876543211", "generated_at": "2026-06-10", "Contact Attempted": "Yes", "Customer Response": WORKBOOK_OUTCOME_SHIFTED_LOCATION, "Complaint": "No", "Do Not Contact": "No", "Target Cost Center": ""},
    ])

    result = await ingest_returned_workbook(database_url=url, path=wb_path, pipeline_run_id="run-shift-blank")

    assert {warning.code for warning in result.warnings} == {"target_cost_center_blank"}
    async with session_scope(url) as session:
        assert (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_leads))).scalar_one() == 1
        suppression = (await session.execute(sa.select(trx_customer_suppression).where(trx_customer_suppression.c.source_lead_id == 11))).mappings().one()
        assert suppression["suppression_state"] == SUPPRESSION_STATE_PENDING_APPROVAL
        assert suppression["is_permanent"] is True
        source = (await session.execute(sa.select(trx_customer_followup_leads).where(trx_customer_followup_leads.c.lead_id == 11))).mappings().one()
        assert source["lead_status"] == "CLOSED"
        assert source["target_cost_center"] is None


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("target_cost_center", "warning_code"),
    [("B200", "target_cost_center_invalid"), ("A100", "target_cost_center_same_store")],
)
async def test_shifted_location_invalid_or_same_store_target_warns_and_creates_no_destination(tmp_path: Path, target_cost_center: str, warning_code: str) -> None:
    url = await _prepare_db(tmp_path)
    await _insert_workbook_lead(url, lead_id=12, cost_center="A100", mobile="9876543212")
    wb_path = tmp_path / f"shifted_{target_cost_center}.xlsx"
    _write_workbook(wb_path, [
        {"lead_id": 12, "lead_source_type": "EXTERNAL", "cost_center": "A100", "customer_name": "Customer 12", "mobile_number": "9876543212", "generated_at": "2026-06-10", "Contact Attempted": "Yes", "Customer Response": WORKBOOK_OUTCOME_SHIFTED_LOCATION, "Complaint": "No", "Do Not Contact": "No", "Target Cost Center": target_cost_center},
    ])

    result = await ingest_returned_workbook(database_url=url, path=wb_path, pipeline_run_id=f"run-shift-{target_cost_center}")

    assert {warning.code for warning in result.warnings} == {warning_code}
    async with session_scope(url) as session:
        assert (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_leads))).scalar_one() == 1
        assert (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_leads).where(trx_customer_followup_leads.c.shifted_from_lead_id == 12))).scalar_one() == 0


@pytest.mark.asyncio
async def test_shifted_location_source_suppression_records_remain_intact_and_target_identity_is_isolated(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    await _insert_workbook_lead(url, lead_id=13, cost_center="A100", mobile="9876543213")
    wb_path = tmp_path / "shifted_suppression_isolation.xlsx"
    _write_workbook(wb_path, [
        {"lead_id": 13, "lead_source_type": "EXTERNAL", "cost_center": "A100", "customer_name": "Customer 13", "mobile_number": "9876543213", "generated_at": "2026-06-10", "Contact Attempted": "Yes", "Customer Response": WORKBOOK_OUTCOME_SHIFTED_LOCATION, "Complaint": "No", "Do Not Contact": "No", "Target Cost Center": "C300"},
    ])

    await ingest_returned_workbook(database_url=url, path=wb_path, pipeline_run_id="run-shift-isolation")

    async with session_scope(url) as session:
        suppression = (await session.execute(sa.select(trx_customer_suppression).where(trx_customer_suppression.c.source_lead_id == 13))).mappings().one()
        assert suppression["cost_center"] == "A100"
        assert suppression["suppression_state"] == SUPPRESSION_STATE_PENDING_APPROVAL
        assert suppression["is_permanent"] is True
        history_before = (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_history).where(trx_customer_followup_history.c.lead_id == 13))).scalar_one()

        approval = await approve_suppression(session, suppression_id=int(suppression["suppression_id"]), approved_by="manager", pipeline_run_id="approval-run")
        source_decision = await check_active_suppression(session, cost_center="A100", normalized_mobile_number="9876543213", as_of_date=date(2026, 6, 12))
        target_decision = await check_active_suppression(session, cost_center="C300", normalized_mobile_number="9876543213", as_of_date=date(2026, 6, 12))
        await session.commit()

        assert approval.changed is True
        assert source_decision.is_suppressed is True
        assert target_decision.is_suppressed is False
        assert (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_leads).where(trx_customer_followup_leads.c.cost_center == "C300", trx_customer_followup_leads.c.normalized_mobile_number == "9876543213"))).scalar_one() == 1
        assert (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_history).where(trx_customer_followup_history.c.lead_id == 13))).scalar_one() == history_before + 1


def test_input_discovery_and_archive_move_processed_external_file_out_of_input_scope(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs" / "customer_followup"
    external_dir = input_dir / "external_leads"
    external_dir.mkdir(parents=True)
    (input_dir / ".hidden.xlsx").write_text("x")
    (input_dir / "~$temp.xlsx").write_text("x")
    workbook = input_dir / "returned.xlsx"
    workbook.write_text("workbook")
    external = external_dir / "leads.csv"
    external.write_text("cost_center,customer_name,mobile_number,lead_source,campaign_name,lead_date,remarks\n")

    assert [f.file_name for f in discover_returned_workbooks(input_dir=input_dir)] == ["returned.xlsx"]
    assert [f.file_name for f in discover_external_lead_files(external_input_dir=external_dir)] == ["leads.csv"]

    archive1 = archive_processed_file(external, archive_dir=tmp_path / "archive" / "customer_followup", run_id="run1", result_metadata={"status": "ok"})

    assert archive1.exists()
    assert archive1.read_text() == "cost_center,customer_name,mobile_number,lead_source,campaign_name,lead_date,remarks\n"
    assert not external.exists()
    assert archive1.with_suffix(archive1.suffix + ".json").exists()
    assert discover_external_lead_files(external_input_dir=external_dir) == []

    external.write_text("changed")
    archive2 = archive_processed_file(external, archive_dir=tmp_path / "archive" / "customer_followup", run_id="run1")

    assert archive2 != archive1
    assert archive2.exists()
    assert not external.exists()
    assert discover_external_lead_files(external_input_dir=external_dir) == []


def test_archive_move_prevents_repeated_workbook_discovery_after_processed_run(tmp_path: Path) -> None:
    input_dir = tmp_path / "inputs" / "customer_followup"
    input_dir.mkdir(parents=True)
    workbook = input_dir / "returned.xlsx"
    workbook.write_text("workbook")

    first_discovery = discover_returned_workbooks(input_dir=input_dir)
    assert [f.file_name for f in first_discovery] == ["returned.xlsx"]

    archived = archive_processed_file(first_discovery[0].path, archive_dir=tmp_path / "archive" / "customer_followup", run_id="run-workbook")

    assert archived.exists()
    assert not workbook.exists()
    assert discover_returned_workbooks(input_dir=input_dir) == []
