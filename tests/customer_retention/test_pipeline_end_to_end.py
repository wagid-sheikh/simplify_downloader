from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import openpyxl
import pytest
import pytest_asyncio
import sqlalchemy as sa
from sqlalchemy.ext.asyncio import create_async_engine

from app.common.db import session_scope
from app.customer_retention.constants import (
    CAP_WORK_SECTION_EXTERNAL_LEAD,
    CAP_WORK_SECTION_FRESH_RETENTION,
    LEAD_SOURCE_EXTERNAL,
    LEAD_SOURCE_RETENTION,
    LEAD_SOURCE_TD,
    LEAD_STATUS_CLOSED,
    LEAD_STATUS_DUE_FOLLOWUP,
    LEAD_STATUS_OPEN,
    LEAD_STATUS_RECOVERED,
    SUPPRESSION_STATE_PENDING_APPROVAL,
    WORKBOOK_OUTCOME_DO_NOT_CONTACT,
    WORKBOOK_OUTCOME_LEAD_STALE,
    WORKBOOK_OUTCOME_PICKUP_REQUESTED,
)
from app.customer_retention.db_tables import (
    customer_followup_cap_config,
    metadata as retention_metadata,
    trx_customer_followup_history,
    trx_customer_followup_leads,
    trx_customer_suppression,
)
from app.customer_retention.workbook_ingestor import FOLLOWUP_SHEET

RUN_DATE = date(2026, 6, 12)


@pytest_asyncio.fixture()
async def retention_db(tmp_path: Path) -> str:
    url = f"sqlite+aiosqlite:///{tmp_path / 'retention_e2e.db'}"
    engine = create_async_engine(url, future=True)
    async with engine.begin() as conn:
        await conn.run_sync(retention_metadata.create_all)
        await conn.execute(sa.text("""
            CREATE TABLE store_master (
                store_code TEXT PRIMARY KEY,
                cost_center TEXT NOT NULL,
                sync_group TEXT NOT NULL,
                customer_retention_pipeline BOOLEAN NOT NULL DEFAULT 1
            )
        """))
        await conn.execute(sa.text("""
            CREATE TABLE crm_leads_current (
                id INTEGER PRIMARY KEY,
                lead_uid TEXT,
                store_code TEXT,
                pickup_no TEXT,
                status_bucket TEXT,
                customer_name TEXT,
                mobile TEXT,
                pickup_date TEXT,
                pickup_created_at TIMESTAMP,
                special_instruction TEXT,
                reason TEXT,
                source TEXT,
                customer_type TEXT,
                run_id TEXT,
                source_file TEXT,
                scraped_at TIMESTAMP
            )
        """))
        await conn.execute(sa.text("""
            CREATE TABLE vw_orders (
                cost_center TEXT NOT NULL,
                order_number TEXT NOT NULL,
                order_date TIMESTAMP NOT NULL,
                customer_name TEXT,
                mobile_number TEXT,
                order_amount NUMERIC(12, 2) NOT NULL
            )
        """))
        await conn.execute(sa.text("""
            CREATE TABLE pipeline_run_summaries (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                pipeline_name TEXT,
                run_id TEXT UNIQUE,
                run_env TEXT,
                started_at TIMESTAMP,
                finished_at TIMESTAMP,
                total_time_taken TEXT,
                report_date DATE,
                overall_status TEXT,
                summary_text TEXT,
                phases_json JSON,
                metrics_json JSON,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """))
        await conn.execute(sa.text("INSERT INTO store_master VALUES ('TD01','A100','TD',1), ('TD02','B200','TD',1)"))
    await engine.dispose()
    async with session_scope(url) as session:
        await _cap(session, LEAD_SOURCE_TD, 'TD_LEAD', None, uncapped=True)
        await _cap(session, LEAD_SOURCE_EXTERNAL, CAP_WORK_SECTION_EXTERNAL_LEAD, 'A100', daily_cap=1)
        await _cap(session, LEAD_SOURCE_RETENTION, CAP_WORK_SECTION_FRESH_RETENTION, 'A100', daily_cap=2)
        await _cap(session, LEAD_SOURCE_RETENTION, CAP_WORK_SECTION_FRESH_RETENTION, 'B200', daily_cap=2)
        await session.commit()
    return url


@pytest.fixture()
def retention_dirs(tmp_path: Path):
    input_dir = tmp_path / 'input'; external = input_dir / 'external_leads'; output = tmp_path / 'output'; archive = tmp_path / 'archive'
    for p in (input_dir, external, output, archive):
        p.mkdir(parents=True)
    return input_dir, external, output, archive


@pytest.fixture(autouse=True)
def no_email(monkeypatch):
    async def fake_send(*_args, **_kwargs):
        from app.customer_retention.pipeline import NotificationResult
        return NotificationResult(planned=1, sent=0, skipped=True, reason='test')
    monkeypatch.setattr('app.customer_retention.pipeline.send_owner_summary', fake_send)


@pytest.fixture()
def configured(monkeypatch, retention_db: str, retention_dirs):
    input_dir, external, output, archive = retention_dirs
    cfg = type('Cfg', (), {
        'database_url': retention_db,
        'customer_followup_input_dir': str(input_dir),
        'customer_followup_external_input_dir': str(external),
        'customer_followup_output_dir': str(output),
        'customer_followup_archive_dir': str(archive),
        'customer_followup_backlog_warning_threshold': 3,
        'run_env': 'test',
        'environment': 'test',
    })()
    monkeypatch.setattr('app.customer_retention.pipeline.config', cfg)
    monkeypatch.setattr('app.config.config', cfg)
    return cfg


async def _cap(session, source, section, cost_center, *, daily_cap=None, uncapped=False):
    next_id = int((await session.execute(sa.select(sa.func.coalesce(sa.func.max(customer_followup_cap_config.c.cap_config_id), 0) + 1))).scalar_one())
    await session.execute(customer_followup_cap_config.insert().values(
        cap_config_id=next_id, lead_source_type=source, work_section=section, cost_center=cost_center,
        daily_cap=daily_cap, is_uncapped=uncapped, enabled=True,
        effective_from=date(2026, 1, 1), created_at=datetime.now(timezone.utc), updated_at=datetime.now(timezone.utc),
    ))


def _external_csv(path: Path, rows: list[tuple[str, str, str]]):
    path.write_text('cost_center,customer_name,mobile_number,lead_source,campaign_name,lead_date,remarks\n' + ''.join(
        f'A100,{name},{mobile},META,June,{RUN_DATE.isoformat()},{remark}\n' for name, mobile, remark in rows
    ), encoding='utf-8')


def _returned_workbook(generated: Path, destination: Path, *, outcome: str, bad_mobile: bool = False):
    wb = openpyxl.load_workbook(generated)
    ws = wb[FOLLOWUP_SHEET]
    headers = [c.value for c in ws[1]]
    idx = {h: i + 1 for i, h in enumerate(headers)}
    for r in range(2, ws.max_row + 1):
        ws.cell(r, idx['Handled By']).value = 'Riya'
        ws.cell(r, idx['Contact Attempted']).value = 'Yes'
        ws.cell(r, idx['Contact Mode']).value = 'Call'
        ws.cell(r, idx['Customer Response']).value = outcome
        ws.cell(r, idx['Order Expected']).value = 'Yes'
        ws.cell(r, idx['Complaint']).value = 'No'
        ws.cell(r, idx['Do Not Contact']).value = 'No'
        if outcome == WORKBOOK_OUTCOME_PICKUP_REQUESTED:
            ws.cell(r, idx['Next Follow-up Date']).value = (RUN_DATE + timedelta(days=2)).isoformat()
        if bad_mobile:
            ws.cell(r, idx['mobile_number']).value = '11111'
            break
    wb.save(destination)


def _workbook_rows(path: Path) -> int:
    return openpyxl.load_workbook(path, read_only=True, data_only=True)[FOLLOWUP_SHEET].max_row - 1


@pytest.mark.asyncio
async def test_td_lifecycle_end_to_end_generates_ingests_and_is_idempotent(configured, retention_db: str, retention_dirs):
    from app.customer_retention.pipeline import run_customer_retention_pipeline
    input_dir, _external, _output, _archive = retention_dirs
    async with session_scope(retention_db) as session:
        await session.execute(sa.text("""INSERT INTO crm_leads_current
            (lead_uid, store_code, pickup_no, status_bucket, customer_name, mobile, pickup_created_at, scraped_at)
            VALUES ('td-1','TD01','PU1','pending','TD Customer','9876543210',:ts,:ts)"""), {'ts': datetime(2026, 6, 11, tzinfo=timezone.utc)})
        await session.commit()

    first = await run_customer_retention_pipeline(run_date=RUN_DATE, run_id='td-gen', database_url=retention_db)
    workbook = Path(first.generated_files[0])
    assert _workbook_rows(workbook) == 1
    async with session_scope(retention_db) as session:
        lead = (await session.execute(sa.select(trx_customer_followup_leads).where(trx_customer_followup_leads.c.lead_source_type == LEAD_SOURCE_TD))).mappings().one()
        assert lead['lead_status'] == LEAD_STATUS_OPEN

    returned = input_dir / 'td_returned.xlsx'
    _returned_workbook(workbook, returned, outcome=WORKBOOK_OUTCOME_PICKUP_REQUESTED)
    returned_bytes = returned.read_bytes()
    await run_customer_retention_pipeline(run_date=RUN_DATE, run_id='td-return-1', database_url=retention_db)
    returned.write_bytes(returned_bytes)
    await run_customer_retention_pipeline(run_date=RUN_DATE, run_id='td-return-2', database_url=retention_db)

    async with session_scope(retention_db) as session:
        status = (await session.execute(sa.select(trx_customer_followup_leads.c.lead_status).where(trx_customer_followup_leads.c.lead_source_type == LEAD_SOURCE_TD))).scalar_one()
        history_count = (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_history))).scalar_one()
        assert status == LEAD_STATUS_DUE_FOLLOWUP
        assert history_count == 1


@pytest.mark.asyncio
async def test_external_lifecycle_respects_cap_and_updates_reporting(configured, retention_db: str, retention_dirs):
    from app.customer_retention.pipeline import run_customer_retention_pipeline
    input_dir, external, _output, _archive = retention_dirs
    _external_csv(external / 'external.csv', [('Ext One', '9876543211', 'first'), ('Ext Two', '9876543212', 'second')])
    result = await run_customer_retention_pipeline(run_date=RUN_DATE, run_id='external-gen', database_url=retention_db)
    assert result.counts['external_leads_created'] == 1
    assert _workbook_rows(Path(result.generated_files[0])) == 1
    _returned_workbook(Path(result.generated_files[0]), input_dir / 'external_returned.xlsx', outcome=WORKBOOK_OUTCOME_LEAD_STALE)
    after = await run_customer_retention_pipeline(run_date=RUN_DATE, run_id='external-return', database_url=retention_db)
    async with session_scope(retention_db) as session:
        closed = (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_leads).where(trx_customer_followup_leads.c.lead_source_type == LEAD_SOURCE_EXTERNAL, trx_customer_followup_leads.c.lead_status == LEAD_STATUS_CLOSED))).scalar_one()
        suppressions = (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_suppression))).scalar_one()
        assert closed == 1
        assert suppressions == 1
    assert after.summary_payload['warning_error_summary']['returned_files_processed'] == 1
    assert after.summary_payload['source_wise_summary']


@pytest.mark.asyncio
async def test_retention_lifecycle_freeze_suppression_recovery_and_notification_payload(configured, retention_db: str, retention_dirs):
    from app.customer_retention.pipeline import run_customer_retention_pipeline
    input_dir, _external, _output, _archive = retention_dirs
    async with session_scope(retention_db) as session:
        # B200 has open backlog > threshold so fresh RETENTION generation is frozen there.
        
        for backlog_id in range(900, 904):
            await session.execute(trx_customer_followup_leads.insert().values(lead_id=backlog_id, lead_uuid=f'backlog-{backlog_id}', lead_source_type=LEAD_SOURCE_RETENTION, source_system='x', cost_center='B200', customer_name='Backlog', mobile_number=f'9876500{backlog_id}', normalized_mobile_number=f'9876500{backlog_id}', lead_date=RUN_DATE - timedelta(days=2), lead_status=LEAD_STATUS_OPEN, lifecycle_bucket='WARM', contact_attempted=False, complaint_flag=False, do_not_contact_flag=False, is_closed=False, is_recovered=False, suppression_applied=False, created_by_pipeline_run_id='old', updated_at=datetime.now(timezone.utc), created_at=datetime.now(timezone.utc)))
        for i, mobile in enumerate(['9876543220', '9876543221', '9876543222']):
            await session.execute(sa.text("INSERT INTO vw_orders VALUES ('A100', :order_no, :order_date, :name, :mobile, :amount)"), {'order_no': f'O{i}', 'order_date': datetime(2026, 4, 1), 'name': f'Ret {i}', 'mobile': mobile, 'amount': 100 + i})
        await session.execute(sa.text("INSERT INTO vw_orders VALUES ('B200', 'OB', :order_date, 'Frozen', '9876543299', 150)"), {'order_date': datetime(2026, 4, 1)})
        await session.commit()
    result = await run_customer_retention_pipeline(run_date=RUN_DATE, run_id='ret-gen', database_url=retention_db)
    assert result.counts['retention_leads_created'] == 2
    assert 'B200' in result.summary_payload['warning_error_summary']['frozen_stores']
    _returned_workbook(Path(result.generated_files[0]), input_dir / 'ret_returned.xlsx', outcome=WORKBOOK_OUTCOME_DO_NOT_CONTACT)
    closed = await run_customer_retention_pipeline(run_date=RUN_DATE, run_id='ret-dead', database_url=retention_db)
    assert closed.summary_payload['warning_error_summary']['pending_suppression_approval_count'] == 2
    async with session_scope(retention_db) as session:
        pending_approval = (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_suppression).where(trx_customer_suppression.c.suppression_state == SUPPRESSION_STATE_PENDING_APPROVAL))).scalar_one()
        assert pending_approval == 2
        # Recovery uses vw_orders.order_amount for revenue after a later order appears.
        mobile = (await session.execute(sa.select(trx_customer_followup_leads.c.normalized_mobile_number).where(trx_customer_followup_leads.c.lead_source_type == LEAD_SOURCE_RETENTION, trx_customer_followup_leads.c.cost_center == 'A100', trx_customer_followup_leads.c.is_closed.is_(False)).limit(1))).scalar_one()
        await session.execute(sa.text("INSERT INTO vw_orders VALUES ('A100', 'REC1', :order_date, 'Recovered', :mobile, 432.10)"), {'order_date': datetime(2026, 6, 13), 'mobile': mobile})
        await session.commit()
    recovered = await run_customer_retention_pipeline(run_date=date(2026, 6, 14), run_id='ret-recover', database_url=retention_db)
    assert recovered.summary_payload['run_summary']['pipeline_run_id'] == 'ret-recover'
    assert recovered.summary_payload['aging_actionable_workload']
    assert recovered.summary_payload['staff_productivity']
    assert recovered.summary_payload['source_wise_summary']
    assert any(float(value) == 432.10 for value in recovered.summary_payload['recovered_revenue_by_lead'].values())
    async with session_scope(retention_db) as session:
        assert (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_leads).where(trx_customer_followup_leads.c.is_recovered.is_(True), trx_customer_followup_leads.c.lead_status == LEAD_STATUS_RECOVERED))).scalar_one() == 1


@pytest.mark.asyncio
async def test_invalid_mobile_cross_source_does_not_create_or_mutate(configured, retention_db: str, retention_dirs):
    from app.customer_retention.pipeline import run_customer_retention_pipeline
    input_dir, external, _output, _archive = retention_dirs
    async with session_scope(retention_db) as session:
        await session.execute(sa.text("INSERT INTO crm_leads_current (lead_uid, store_code, status_bucket, customer_name, mobile, pickup_created_at, scraped_at) VALUES ('bad-td','TD01','pending','Bad TD','11111',:ts,:ts)"), {'ts': datetime(2026, 6, 11, tzinfo=timezone.utc)})
        await session.execute(sa.text("INSERT INTO vw_orders VALUES ('A100','BADORDER',:order_date,'Bad Order','11111',999)"), {'order_date': datetime(2026, 4, 1)})
        await session.commit()
    _external_csv(external / 'bad_external.csv', [('Bad Ext', '11111', 'bad')])
    bad = await run_customer_retention_pipeline(run_date=RUN_DATE, run_id='invalid', database_url=retention_db)
    assert bad.counts.get('td_leads_created', 0) == 0
    assert bad.counts.get('external_leads_created', 0) == 0
    assert bad.counts['snapshot_invalid_mobile_rows'] == 1
    # A valid row then proves an invalid returned workbook mobile cannot mutate state.
    async with session_scope(retention_db) as session:
        await session.execute(sa.text("INSERT INTO crm_leads_current (lead_uid, store_code, status_bucket, customer_name, mobile, pickup_created_at, scraped_at) VALUES ('good-td','TD01','pending','Good TD','9876543230',:ts,:ts)"), {'ts': datetime(2026, 6, 11, tzinfo=timezone.utc)})
        await session.commit()
    good = await run_customer_retention_pipeline(run_date=RUN_DATE, run_id='valid-for-bad-workbook', database_url=retention_db)
    _returned_workbook(Path(good.generated_files[0]), input_dir / 'bad_mobile_return.xlsx', outcome=WORKBOOK_OUTCOME_PICKUP_REQUESTED, bad_mobile=True)
    await run_customer_retention_pipeline(run_date=RUN_DATE, run_id='bad-workbook-mobile', database_url=retention_db)
    async with session_scope(retention_db) as session:
        assert (await session.execute(sa.select(sa.func.count()).select_from(trx_customer_followup_history))).scalar_one() == 0


def test_cli_dispatch_customer_retention_dry_run(monkeypatch):
    from app.__main__ import main
    seen = {}
    async def fake_run(**kwargs):
        seen.update(kwargs)
        return type('Result', (), {'run_id': 'cli', 'status': 'success'})()
    monkeypatch.setattr('app.customer_retention.pipeline.run_customer_retention_pipeline', fake_run)
    assert main(['customer-retention', 'run', '--dry-run', '--run-date', '2026-06-12', '--run-id', 'cli']) == 0
    assert seen['dry_run'] is True
    assert seen['run_date'] == RUN_DATE


def test_customer_retention_business_logic_does_not_read_raw_financial_columns():
    forbidden = ('orders.net_amount', 'orders.gross_amount', 'orders.adjustment')
    offenders = []
    for path in Path('app/customer_retention').rglob('*.py'):
        text = path.read_text(encoding='utf-8')
        if path.name == 'analytics.py':
            assert 'vw_orders.order_amount' in text
        for needle in forbidden:
            if needle in text:
                offenders.append(f'{path}:{needle}')
    assert offenders == []
