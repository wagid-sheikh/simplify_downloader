from __future__ import annotations

from datetime import date, datetime, timedelta, timezone
import io
from contextlib import asynccontextmanager
import json
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest
import sqlalchemy as sa
from sqlalchemy.ext.asyncio import create_async_engine

from app.common.db import session_scope
from app.dashboard_downloader.json_logger import JsonLogger
from app.customer_retention.caps import resolve_active_cap
from app.customer_retention.retention_generation import allocate_and_generate_retention_leads
from app.customer_retention.snapshot import CustomerRetentionSnapshotRow, SnapshotResult
from app.customer_retention.constants import (
    CAP_WORK_SECTION_EXTERNAL_LEAD,
    CAP_WORK_SECTION_FRESH_RETENTION,
    CAP_WORK_SECTION_PENDING_CARRY_FORWARD,
    CAP_WORK_SECTION_TD_LEAD,
    CAP_WORK_SECTION_DUE_FOLLOWUP,
    LEAD_SOURCE_EXTERNAL,
    LEAD_SOURCE_RETENTION,
    LEAD_SOURCE_TD,
    LEAD_STATUS_CLOSED,
    LEAD_STATUS_OPEN,
    LEAD_STATUS_PENDING,
    LEAD_STATUS_RECOVERED,
    SUPPRESSION_STATE_ACTIVE,
    WORKBOOK_OUTCOME_LABELS,
    WORKBOOK_OUTCOME_SHIFTED_LOCATION,
)
from app.customer_retention.db_tables import customer_followup_cap_config, metadata, trx_customer_followup_history, trx_customer_followup_leads, trx_customer_suppression
from app.customer_retention.workbook_generator import (
    DROPDOWN_VALUES_BY_HEADER,
    FOLLOWUP_HEADERS,
    default_customer_followup_output_root,
    generate_store_workbook,
)
from app.customer_retention.workbook_ingestor import EDITABLE_COLUMNS, FOLLOWUP_SHEET, READ_ME_SHEET, ingest_returned_workbook
from app.customer_retention.workbook_selection import (
    StoreWorkbookSelectionResult,
    load_active_retention_stores,
    select_workbook_leads_for_store,
)
from app.customer_retention.workload import evaluate_retention_workload_freeze


@pytest.mark.asyncio
async def test_pipeline_passes_configured_backlog_threshold(monkeypatch) -> None:
    from app.customer_retention import pipeline

    observed: dict[str, int] = {}

    @asynccontextmanager
    async def fake_session_scope(_database_url):
        class FakeSession:
            async def commit(self):
                return None

            async def flush(self):
                return None

        yield FakeSession()

    async def fake_import_td_leads(*_args, **_kwargs):
        return type("TdResult", (), {"rows_seen": 0, "leads_created": 0, "warnings": []})()

    async def fake_detect_recoveries(*_args, **_kwargs):
        return type("Recovery", (), {"leads_recovered": 0, "leads_closed": 0})()

    async def fake_snapshot(*_args, **_kwargs):
        return type("Snapshot", (), {"rows": (), "rows_invalid_mobile": 0})()

    async def fake_generate(*_args, **_kwargs):
        return type("Generation", (), {"rows_seen": 0, "leads_created": 0, "leads_reused": 0, "rows_skipped": 0, "warnings": []})()

    async def fake_load_active_retention_stores(*_args, **_kwargs):
        return ("A100",)

    async def fake_select(_session, *, backlog_threshold, **_kwargs):
        observed["backlog_threshold"] = backlog_threshold
        return ()

    async def fake_summary(*_args, **_kwargs):
        return {}

    async def fake_send(*_args, **_kwargs):
        return pipeline.NotificationResult(planned=0, sent=0, skipped=True, reason="test")

    monkeypatch.setattr(pipeline, "config", type("Cfg", (), {"database_url": "sqlite+aiosqlite:///unused.db", "customer_followup_output_dir": "/tmp/customer_followup", "customer_followup_backlog_warning_threshold": 7})())
    monkeypatch.setattr(pipeline, "session_scope", fake_session_scope)
    monkeypatch.setattr(pipeline, "get_customer_followup_paths", lambda: type("Paths", (), {"archive_dir": Path("/tmp/archive")})())
    monkeypatch.setattr(pipeline, "discover_returned_workbooks", lambda logger=None: [])
    monkeypatch.setattr(pipeline, "discover_external_lead_files", lambda logger=None: [])
    monkeypatch.setattr(pipeline, "load_active_retention_stores", fake_load_active_retention_stores)
    monkeypatch.setattr(pipeline, "_import_td_leads", fake_import_td_leads)
    monkeypatch.setattr(pipeline, "detect_recoveries", fake_detect_recoveries)
    monkeypatch.setattr(pipeline, "build_customer_retention_snapshot", fake_snapshot)
    monkeypatch.setattr(pipeline, "allocate_and_generate_retention_leads", fake_generate)
    monkeypatch.setattr(pipeline, "select_workbook_leads_for_active_stores", fake_select)
    monkeypatch.setattr(pipeline, "generate_workbooks", lambda **kwargs: type("WorkbookResult", (), {"outputs": []})())
    monkeypatch.setattr(pipeline, "build_management_summary_payload", fake_summary)
    monkeypatch.setattr(pipeline, "send_owner_summary", fake_send)

    await pipeline.run_customer_retention_pipeline(run_date=date(2026, 6, 12), run_id="threshold-run")

    assert observed["backlog_threshold"] == 7


def test_default_customer_followup_output_root_uses_config(monkeypatch, tmp_path: Path):
    configured_root = tmp_path / "configured_followup_output"
    monkeypatch.setattr("app.config.config", type("Cfg", (), {"customer_followup_output_dir": str(configured_root)})())

    assert default_customer_followup_output_root() == configured_root


def test_generate_store_workbook_uses_configured_default_output_root(monkeypatch, tmp_path: Path):
    configured_root = tmp_path / "configured_followup_output"
    monkeypatch.setattr("app.config.config", type("Cfg", (), {"customer_followup_output_dir": str(configured_root)})())
    selection = StoreWorkbookSelectionResult(
        cost_center="A100",
        run_date=date(2026, 6, 12),
        rows=(),
        counts_by_category={},
        warnings=(),
    )

    output = generate_store_workbook(
        selection=selection,
        active_cost_centers=("A100",),
        generated_at=datetime(2026, 6, 12, 8, 0, tzinfo=timezone.utc),
    )

    assert output.output_path == (
        configured_root / "2026-06" / "customer_followup_A100_2026-06-12.xlsx"
    )
    assert output.output_path.exists()


async def _prepare_db(tmp_path: Path, *, relaxed_caps: bool = False) -> str:
    url = f"sqlite+aiosqlite:///{tmp_path / ('phase4_relaxed.db' if relaxed_caps else 'phase4.db')}"
    engine = create_async_engine(url, future=True)
    async with engine.begin() as conn:
        if relaxed_caps:
            await conn.execute(sa.text("""
                CREATE TABLE customer_followup_cap_config (
                    cap_config_id INTEGER PRIMARY KEY AUTOINCREMENT,
                    cost_center TEXT NULL,
                    lead_source_type TEXT NOT NULL,
                    work_section TEXT NOT NULL,
                    daily_cap INTEGER NULL,
                    is_uncapped BOOLEAN NOT NULL DEFAULT 0,
                    enabled BOOLEAN NOT NULL DEFAULT 1,
                    effective_from DATE NOT NULL,
                    effective_until DATE NULL,
                    created_at TIMESTAMP,
                    updated_at TIMESTAMP
                )
            """))
        else:
            await conn.run_sync(metadata.create_all)
            await conn.execute(sa.text("CREATE TABLE store_master (cost_center TEXT PRIMARY KEY, customer_retention_pipeline BOOLEAN NOT NULL DEFAULT 1)"))
            await conn.execute(sa.text("INSERT INTO store_master (cost_center, customer_retention_pipeline) VALUES ('A100', 1), ('B200', 0), ('C300', 1), ('D400', 1)"))
    await engine.dispose()
    return url


async def _insert_cap(session, *, source=LEAD_SOURCE_RETENTION, section=CAP_WORK_SECTION_FRESH_RETENTION, cost_center=None, daily_cap=13, uncapped=False, enabled=True, effective_from=date(2026, 1, 1), effective_until=None):
    next_id = int((await session.execute(sa.select(sa.func.coalesce(sa.func.max(customer_followup_cap_config.c.cap_config_id), 0) + 1))).scalar_one())
    await session.execute(customer_followup_cap_config.insert().values(
        cap_config_id=next_id,
        cost_center=cost_center,
        lead_source_type=source,
        work_section=section,
        daily_cap=daily_cap,
        is_uncapped=uncapped,
        enabled=enabled,
        effective_from=effective_from,
        effective_until=effective_until,
        created_at=datetime.now(timezone.utc),
        updated_at=datetime.now(timezone.utc),
    ))


async def _insert_lead(session, *, lead_id: int, cost_center="A100", source=LEAD_SOURCE_RETENTION, status=LEAD_STATUS_OPEN, lead_date=date(2026, 6, 12), next_followup_date=None, mobile=None, priority=Decimal("10"), closed=False, recovered=False, suppression=False):
    normalized = mobile if mobile is not None else f"98765432{lead_id:02d}"[-10:]
    await session.execute(trx_customer_followup_leads.insert().values(
        lead_id=lead_id,
        lead_uuid=f"lead-{lead_id}",
        lead_source_type=source,
        source_system="test",
        source_table_name="source" if source in {LEAD_SOURCE_TD, LEAD_SOURCE_EXTERNAL} else None,
        source_record_id=str(lead_id) if source in {LEAD_SOURCE_TD, LEAD_SOURCE_EXTERNAL} else None,
        cost_center=cost_center,
        customer_name=f"Customer {lead_id}",
        mobile_number=normalized,
        normalized_mobile_number=normalized,
        lead_date=lead_date,
        lead_status=status,
        lifecycle_bucket="WARM" if source == LEAD_SOURCE_RETENTION else None,
        last_order_date=lead_date - timedelta(days=35),
        days_since_last_order=35,
        total_orders=3,
        lifetime_spend=Decimal("900.00"),
        average_order_value=Decimal("300.00"),
        last_order_amount=Decimal("300.00"),
        priority_score=priority,
        recommended_strategy="Call promptly" if source != LEAD_SOURCE_RETENTION else "Gentle reminder call",
        next_followup_date=next_followup_date,
        contact_attempted=False,
        complaint_flag=False,
        do_not_contact_flag=False,
        is_closed=closed or status == LEAD_STATUS_CLOSED,
        is_recovered=recovered or status == LEAD_STATUS_RECOVERED,
        suppression_applied=suppression,
        created_at=datetime.now(timezone.utc),
        created_by_pipeline_run_id="run-retention" if source == LEAD_SOURCE_RETENTION else None,
        updated_by_pipeline_run_id="run-retention" if source == LEAD_SOURCE_RETENTION else None,
        updated_at=datetime.now(timezone.utc),
    ))


@pytest.mark.asyncio
async def test_cap_resolver_selection_rules_and_missing_disabled_uncapped(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    async with session_scope(url) as session:
        await _insert_cap(session, daily_cap=13, effective_from=date(2026, 1, 1))
        await _insert_cap(session, cost_center="A100", daily_cap=7, effective_from=date(2026, 1, 1))
        await _insert_cap(session, cost_center="A100", daily_cap=5, enabled=False, effective_from=date(2026, 6, 1))
        await _insert_cap(session, cost_center="C300", daily_cap=9, effective_from=date(2026, 1, 1), effective_until=date(2026, 5, 31))
        await _insert_cap(session, cost_center="D400", daily_cap=8, effective_from=date(2026, 1, 1))
        await _insert_cap(session, cost_center="D400", daily_cap=6, effective_from=date(2026, 6, 1))
        await _insert_cap(session, source=LEAD_SOURCE_TD, section=CAP_WORK_SECTION_TD_LEAD, daily_cap=None, uncapped=True)
        await session.commit()

        assert (await resolve_active_cap(session, lead_source_type=LEAD_SOURCE_RETENTION, work_section=CAP_WORK_SECTION_FRESH_RETENTION, cost_center="Z999", run_date=date(2026, 6, 12))).daily_cap == 13
        assert (await resolve_active_cap(session, lead_source_type=LEAD_SOURCE_RETENTION, work_section=CAP_WORK_SECTION_FRESH_RETENTION, cost_center="A100", run_date=date(2026, 6, 12))).daily_cap == 7
        assert (await resolve_active_cap(session, lead_source_type=LEAD_SOURCE_RETENTION, work_section=CAP_WORK_SECTION_FRESH_RETENTION, cost_center="C300", run_date=date(2026, 6, 12))).daily_cap == 13
        assert (await resolve_active_cap(session, lead_source_type=LEAD_SOURCE_RETENTION, work_section=CAP_WORK_SECTION_FRESH_RETENTION, cost_center="D400", run_date=date(2026, 6, 12))).daily_cap == 6
        assert (await resolve_active_cap(session, lead_source_type=LEAD_SOURCE_TD, work_section=CAP_WORK_SECTION_TD_LEAD, cost_center="A100", run_date=date(2026, 6, 12))).is_uncapped is True
        assert (await resolve_active_cap(session, lead_source_type=LEAD_SOURCE_EXTERNAL, work_section=CAP_WORK_SECTION_EXTERNAL_LEAD, cost_center="A100", run_date=date(2026, 6, 12))).missing is True


@pytest.mark.asyncio
async def test_cap_resolver_detects_invalid_td_and_ambiguous_overlap(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path, relaxed_caps=True)
    async with session_scope(url) as session:
        await _insert_cap(session, source=LEAD_SOURCE_TD, section=CAP_WORK_SECTION_TD_LEAD, daily_cap=1, uncapped=False)
        invalid = await resolve_active_cap(session, lead_source_type=LEAD_SOURCE_TD, work_section=CAP_WORK_SECTION_TD_LEAD, cost_center="A100", run_date=date(2026, 6, 12))
        assert invalid.valid is False
        assert invalid.warnings[0].code == "td_cap_contract_violation"
        await _insert_cap(session, daily_cap=13, effective_from=date(2026, 1, 1))
        await _insert_cap(session, daily_cap=12, effective_from=date(2026, 1, 1))
        ambiguous = await resolve_active_cap(session, lead_source_type=LEAD_SOURCE_RETENTION, work_section=CAP_WORK_SECTION_FRESH_RETENTION, cost_center="A100", run_date=date(2026, 6, 12))
        assert ambiguous.valid is False
        assert ambiguous.warnings[0].code == "ambiguous_active_cap"


@pytest.mark.asyncio
async def test_workload_freeze_threshold_and_old_carry_forward_exclusion(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    run_date = date(2026, 6, 12)
    async with session_scope(url) as session:
        for i in range(1, 4):
            await _insert_lead(session, lead_id=i, lead_date=run_date - timedelta(days=i))
        await _insert_lead(session, lead_id=99, lead_date=run_date - timedelta(days=30))
        await session.commit()
        below = await evaluate_retention_workload_freeze(session, cost_center="A100", run_date=run_date, threshold=4)
        at_threshold_allowed = await evaluate_retention_workload_freeze(session, cost_center="A100", run_date=run_date, threshold=3)
        above = await evaluate_retention_workload_freeze(session, cost_center="A100", run_date=run_date, threshold=2)
        assert below.frozen is False
        assert at_threshold_allowed.frozen is False
        assert above.frozen is True
        assert above.incomplete_recent_retention_count == 3
        assert above.older_carry_forward_count == 1


@pytest.mark.asyncio
async def test_workbook_selection_categories_caps_freeze_suppression_and_ordering(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    run_date = date(2026, 6, 12)
    async with session_scope(url) as session:
        await _insert_cap(session, daily_cap=1)
        await _insert_cap(session, source=LEAD_SOURCE_EXTERNAL, section=CAP_WORK_SECTION_EXTERNAL_LEAD, daily_cap=1)
        await _insert_lead(session, lead_id=1, next_followup_date=run_date, priority=Decimal("1"))
        await _insert_lead(session, lead_id=2, lead_date=run_date - timedelta(days=1), priority=Decimal("100"))
        await _insert_lead(session, lead_id=3, source=LEAD_SOURCE_TD, priority=Decimal("100"))
        await _insert_lead(session, lead_id=4, source=LEAD_SOURCE_TD, priority=Decimal("90"))
        await _insert_lead(session, lead_id=5, source=LEAD_SOURCE_EXTERNAL, priority=Decimal("80"))
        await _insert_lead(session, lead_id=6, source=LEAD_SOURCE_EXTERNAL, priority=Decimal("70"))
        await _insert_lead(session, lead_id=7, source=LEAD_SOURCE_RETENTION, priority=Decimal("60"))
        await _insert_lead(session, lead_id=8, source=LEAD_SOURCE_RETENTION, priority=Decimal("50"))
        await _insert_lead(session, lead_id=9, source=LEAD_SOURCE_RETENTION, mobile="1111111111")
        await _insert_lead(session, lead_id=10, source=LEAD_SOURCE_RETENTION, closed=True, status=LEAD_STATUS_CLOSED)
        await _insert_lead(session, lead_id=11, source=LEAD_SOURCE_RETENTION, recovered=True, status=LEAD_STATUS_RECOVERED)
        await _insert_lead(session, lead_id=12, source=LEAD_SOURCE_EXTERNAL, mobile="9876543299")
        await session.execute(trx_customer_suppression.insert().values(suppression_id=1, cost_center="A100", mobile_number="9876543299", normalized_mobile_number="9876543299", suppression_reason="Not Interested", suppression_state=SUPPRESSION_STATE_ACTIVE, suppression_start_date=run_date, suppression_until=run_date + timedelta(days=90), is_permanent=False, approval_required=False, source_lead_id=12, created_at=datetime.now(timezone.utc)))
        await session.commit()

        assert await load_active_retention_stores(session) == ("A100", "C300", "D400")
        result = await select_workbook_leads_for_store(session, cost_center="A100", run_date=run_date, backlog_threshold=99)
        assert [row.work_section for row in result.rows] == [CAP_WORK_SECTION_DUE_FOLLOWUP, CAP_WORK_SECTION_PENDING_CARRY_FORWARD, CAP_WORK_SECTION_TD_LEAD, CAP_WORK_SECTION_TD_LEAD, CAP_WORK_SECTION_EXTERNAL_LEAD, CAP_WORK_SECTION_FRESH_RETENTION]
        assert [row.lead_id for row in result.rows] == [1, 2, 3, 4, 5, 7]
        assert result.counts_by_category[CAP_WORK_SECTION_TD_LEAD] == 2
        frozen = await select_workbook_leads_for_store(session, cost_center="A100", run_date=run_date, backlog_threshold=0)
        assert CAP_WORK_SECTION_FRESH_RETENTION not in {row.work_section for row in frozen.rows}
        assert {CAP_WORK_SECTION_DUE_FOLLOWUP, CAP_WORK_SECTION_PENDING_CARRY_FORWARD, CAP_WORK_SECTION_TD_LEAD, CAP_WORK_SECTION_EXTERNAL_LEAD}.issubset({row.work_section for row in frozen.rows})


@pytest.mark.asyncio
async def test_workbook_selection_emits_phase4_observability_events(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    run_date = date(2026, 6, 12)
    log_stream = io.StringIO()
    logger = JsonLogger(run_id="logger-run", stream=log_stream, log_file_path=None)
    async with session_scope(url) as session:
        await _insert_cap(session, daily_cap=1)
        await _insert_cap(session, source=LEAD_SOURCE_EXTERNAL, section=CAP_WORK_SECTION_EXTERNAL_LEAD, daily_cap=1)
        await _insert_lead(session, lead_id=101, next_followup_date=run_date, priority=Decimal("100"))
        await _insert_lead(session, lead_id=102, lead_date=run_date - timedelta(days=1), priority=Decimal("90"))
        await _insert_lead(session, lead_id=103, source=LEAD_SOURCE_TD, priority=Decimal("80"))
        await _insert_lead(session, lead_id=104, source=LEAD_SOURCE_EXTERNAL, priority=Decimal("70"))
        await _insert_lead(session, lead_id=105, source=LEAD_SOURCE_EXTERNAL, priority=Decimal("60"))
        await _insert_lead(session, lead_id=106, source=LEAD_SOURCE_RETENTION, priority=Decimal("50"))
        await _insert_lead(session, lead_id=107, source=LEAD_SOURCE_RETENTION, priority=Decimal("40"))
        await _insert_lead(session, lead_id=108, source=LEAD_SOURCE_RETENTION, mobile="123", priority=Decimal("30"))
        await session.commit()

        stores = await load_active_retention_stores(
            session, logger=logger, run_id="phase4-run", phase="phase4", pipeline="customer_retention", run_date=run_date
        )
        result = await select_workbook_leads_for_store(
            session,
            cost_center="A100",
            run_date=run_date,
            backlog_threshold=99,
            logger=logger,
            run_id="phase4-run",
            phase="phase4",
            pipeline="customer_retention",
        )

    events = [json.loads(line) for line in log_stream.getvalue().splitlines()]
    by_message = {event["message"]: event for event in events}

    assert stores == ("A100", "C300", "D400")
    assert by_message["active_retention_stores_loaded"]["active_store_count"] == 3
    assert by_message["active_retention_stores_loaded"]["run_id"] == "phase4-run"
    cap_events = [event for event in events if event["message"] == "cap_resolution_result"]
    assert {event["lead_source_type"] for event in cap_events} == {LEAD_SOURCE_RETENTION, LEAD_SOURCE_EXTERNAL}
    assert all(event["phase"] == "phase4" and event["pipeline"] == "customer_retention" for event in cap_events)
    assert by_message["workload_freeze_result"]["frozen"] is False
    assert by_message["due_followup_rows_selected"]["selected_count"] == 1
    assert by_message["pending_carry_forward_rows_selected"]["selected_count"] == 1
    assert by_message["td_rows_selected"]["selected_count"] == 1
    assert by_message["external_rows_selected_capped"]["selected_count"] == 1
    assert by_message["external_rows_selected_capped"]["capped_count"] == 1
    assert by_message["fresh_retention_rows_selected_frozen_capped"]["selected_count"] == 1
    assert by_message["fresh_retention_rows_selected_frozen_capped"]["capped_count"] == 1
    assert by_message["invalid_normalized_mobile_identity_exclusions"]["excluded_count"] == 1
    assert by_message["invalid_normalized_mobile_identity_exclusions"]["status"] == "warning"
    assert by_message["final_workbook_row_counts"]["row_count"] == len(result.rows) == 5
    assert by_message["final_workbook_row_counts"]["counts_by_category"][CAP_WORK_SECTION_FRESH_RETENTION] == 1


@pytest.mark.asyncio
async def test_workbook_generation_structure_validation_and_ingestion_identity(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    run_date = date(2026, 6, 12)
    async with session_scope(url) as session:
        await _insert_cap(session, daily_cap=2)
        await _insert_lead(session, lead_id=1, priority=Decimal("20"))
        await session.commit()
        selection = await select_workbook_leads_for_store(
            session, cost_center="A100", run_date=run_date, backlog_threshold=99
        )

    output = generate_store_workbook(selection=selection, active_cost_centers=("A100", "C300", "D400"), output_root=tmp_path / "outputs" / "customer_followup", generated_at=datetime(2026, 6, 12, 8, 0, tzinfo=timezone.utc))
    assert output.output_path == tmp_path / "outputs" / "customer_followup" / "2026-06" / "customer_followup_A100_2026-06-12.xlsx"

    workbook = openpyxl.load_workbook(output.output_path)
    assert workbook.sheetnames == [READ_ME_SHEET, FOLLOWUP_SHEET]
    readme_text = "\n".join(str(workbook[READ_ME_SHEET].cell(row=i, column=1).value or "") for i in range(1, 7))
    assert "Only edit allowed columns" in readme_text
    assert "Target Cost Center" in readme_text
    assert "invalid entries become warnings" in readme_text
    sheet = workbook[FOLLOWUP_SHEET]
    headers = [cell.value for cell in sheet[1]]
    assert headers == FOLLOWUP_HEADERS
    assert sheet.max_row == 2
    assert sheet.cell(row=2, column=headers.index("lead_source_type") + 1).value == LEAD_SOURCE_RETENTION
    assert sheet.cell(row=2, column=headers.index("recommended_strategy") + 1).value == "Gentle reminder call"
    assert sheet.cell(row=2, column=headers.index("lead_id") + 1).value == 1
    assert sheet.cell(row=2, column=headers.index("normalized_mobile_number") + 1).value
    assert sheet.cell(row=2, column=headers.index("normalized_mobile_number") + 1).protection.locked is True
    assert sheet.cell(row=2, column=headers.index("lead_id") + 1).protection.locked is True
    assert sheet.cell(row=2, column=headers.index("Contact Attempted") + 1).protection.locked is False

    validations = list(sheet.data_validations.dataValidation)
    validation_by_range = {str(rng): dv for dv in validations for rng in dv.cells.ranges}
    for header in DROPDOWN_VALUES_BY_HEADER:
        col = openpyxl.utils.get_column_letter(headers.index(header) + 1)
        assert col + "2" in validation_by_range
    target_col = openpyxl.utils.get_column_letter(headers.index("Target Cost Center") + 1)
    target_validation = validation_by_range[f"{target_col}2"]
    assert target_validation.type == "list"
    assert target_validation.allow_blank is True
    assert "A100,C300,D400" in target_validation.formula1
    assert target_validation.showErrorMessage is True
    response_col = openpyxl.utils.get_column_letter(headers.index("Customer Response") + 1)
    assert all(value in validation_by_range[f"{response_col}2"].formula1 for value in WORKBOOK_OUTCOME_LABELS)

    # The generated row can be returned through Phase 2 ingestion using its stable
    # lead_id + normalized mobile identity without creating an identity mismatch.
    sheet.cell(row=2, column=headers.index("Contact Attempted") + 1, value="Yes")
    sheet.cell(row=2, column=headers.index("Customer Response") + 1, value="Interested")
    sheet.cell(row=2, column=headers.index("Complaint") + 1, value="No")
    sheet.cell(row=2, column=headers.index("Do Not Contact") + 1, value="No")
    sheet.cell(row=2, column=headers.index("Handled By") + 1, value="Staff A")
    workbook.save(output.output_path)

    ingest_result = await ingest_returned_workbook(database_url=url, path=output.output_path, pipeline_run_id="run-phase4")
    assert ingest_result.rows_seen == 1
    assert ingest_result.history_inserted == 1
    assert ingest_result.warning_count == 0


@pytest.mark.asyncio
async def test_target_cost_center_ingestion_warnings_and_destination_creation_contract(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    run_date = date(2026, 6, 12)
    async with session_scope(url) as session:
        await _insert_cap(session, daily_cap=5)
        for lead_id in range(31, 36):
            await _insert_lead(session, lead_id=lead_id, priority=Decimal(str(100 - lead_id)))
        await session.commit()
        selection = await select_workbook_leads_for_store(
            session, cost_center="A100", run_date=run_date, backlog_threshold=99
        )

    output = generate_store_workbook(
        selection=selection,
        active_cost_centers=("A100", "C300", "D400"),
        output_root=tmp_path / "outputs",
        generated_at=datetime(2026, 6, 12, 8, 0, tzinfo=timezone.utc),
    )

    workbook = openpyxl.load_workbook(output.output_path)
    sheet = workbook[FOLLOWUP_SHEET]
    headers = [cell.value for cell in sheet[1]]
    col = {header: headers.index(header) + 1 for header in headers}
    rows_by_lead_id = {sheet.cell(row=row_number, column=col["lead_id"]).value: row_number for row_number in range(2, sheet.max_row + 1)}
    cases = {
        31: (WORKBOOK_OUTCOME_SHIFTED_LOCATION, None),
        32: (WORKBOOK_OUTCOME_SHIFTED_LOCATION, "B200"),
        33: (WORKBOOK_OUTCOME_SHIFTED_LOCATION, "A100"),
        34: ("Interested", "C300"),
        35: (WORKBOOK_OUTCOME_SHIFTED_LOCATION, "D400"),
    }
    for lead_id, (response, target_cost_center) in cases.items():
        row_number = rows_by_lead_id[lead_id]
        sheet.cell(row=row_number, column=col["Contact Attempted"], value="Yes")
        sheet.cell(row=row_number, column=col["Customer Response"], value=response)
        sheet.cell(row=row_number, column=col["Complaint"], value="No")
        sheet.cell(row=row_number, column=col["Do Not Contact"], value="No")
        sheet.cell(row=row_number, column=col["Target Cost Center"], value=target_cost_center)
    workbook.save(output.output_path)

    result = await ingest_returned_workbook(database_url=url, path=output.output_path, pipeline_run_id="run-target-contract")

    warning_codes_by_lead_id = {warning.lead_id: warning.code for warning in result.warnings}
    assert warning_codes_by_lead_id == {
        31: "target_cost_center_blank",
        32: "target_cost_center_invalid",
        33: "target_cost_center_same_store",
        34: "target_cost_center_ignored",
    }
    async with session_scope(url) as session:
        destination_leads = (
            await session.execute(
                sa.select(trx_customer_followup_leads).where(
                    trx_customer_followup_leads.c.source_system == "CUSTOMER_FOLLOWUP_SHIFTED_LOCATION"
                )
            )
        ).mappings().all()
        assert len(destination_leads) == 1
        assert destination_leads[0]["lead_source_type"] == LEAD_SOURCE_EXTERNAL
        assert destination_leads[0]["cost_center"] == "D400"
        assert destination_leads[0]["shifted_from_lead_id"] == 35
        assert (
            await session.execute(
                sa.select(sa.func.count()).select_from(trx_customer_followup_leads).where(
                    trx_customer_followup_leads.c.shifted_from_lead_id.in_([31, 32, 33, 34])
                )
            )
        ).scalar_one() == 0


@pytest.mark.asyncio
async def test_target_cost_center_dropdown_uses_inline_validation_for_small_active_store_lists(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    run_date = date(2026, 6, 12)
    async with session_scope(url) as session:
        await _insert_cap(session, daily_cap=2)
        await _insert_lead(session, lead_id=21, priority=Decimal("20"))
        await session.commit()
        selection = await select_workbook_leads_for_store(
            session, cost_center="A100", run_date=run_date, backlog_threshold=99
        )

    output = generate_store_workbook(
        selection=selection,
        active_cost_centers=("A100", "C300", "D400"),
        output_root=tmp_path / "outputs",
        generated_at=datetime(2026, 6, 12, 8, 0, tzinfo=timezone.utc),
    )

    workbook = openpyxl.load_workbook(output.output_path)
    sheet = workbook[FOLLOWUP_SHEET]
    headers = [cell.value for cell in sheet[1]]
    target_col = openpyxl.utils.get_column_letter(headers.index("Target Cost Center") + 1)
    validation_by_range = {
        str(rng): dv for dv in sheet.data_validations.dataValidation for rng in dv.cells.ranges
    }
    target_validation = validation_by_range[f"{target_col}2"]
    assert target_validation.type == "list"
    assert target_validation.formula1 == '"A100,C300,D400"'
    assert workbook.sheetnames == [READ_ME_SHEET, FOLLOWUP_SHEET]


@pytest.mark.asyncio
async def test_target_cost_center_dropdown_uses_hidden_lookup_range_for_large_active_store_lists(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    run_date = date(2026, 6, 12)
    active_cost_centers = tuple(f"S{i:04d}" for i in range(80))
    async with session_scope(url) as session:
        await _insert_cap(session, daily_cap=2)
        await _insert_lead(session, lead_id=22, priority=Decimal("20"))
        await session.commit()
        selection = await select_workbook_leads_for_store(
            session, cost_center="A100", run_date=run_date, backlog_threshold=99
        )

    output = generate_store_workbook(
        selection=selection,
        active_cost_centers=active_cost_centers,
        output_root=tmp_path / "outputs",
        generated_at=datetime(2026, 6, 12, 8, 0, tzinfo=timezone.utc),
    )

    workbook = openpyxl.load_workbook(output.output_path)
    sheet = workbook[FOLLOWUP_SHEET]
    headers = [cell.value for cell in sheet[1]]
    target_col = openpyxl.utils.get_column_letter(headers.index("Target Cost Center") + 1)
    validation_by_range = {
        str(rng): dv for dv in sheet.data_validations.dataValidation for rng in dv.cells.ranges
    }
    target_validation = validation_by_range[f"{target_col}2"]
    assert target_validation.type == "list"
    assert target_validation.formula1 == "=ActiveCostCenters"
    assert "ActiveCostCenters" in workbook.defined_names
    destinations = list(workbook.defined_names["ActiveCostCenters"].destinations)
    assert destinations == [("_ACTIVE_COST_CENTER_LOOKUP", "$A$2:$A$81")]
    lookup = workbook["_ACTIVE_COST_CENTER_LOOKUP"]
    assert lookup.sheet_state == "veryHidden"
    assert lookup.protection.sheet is True
    assert lookup.cell(row=2, column=1).value == "S0000"
    assert lookup.cell(row=81, column=1).value == "S0079"


@pytest.mark.asyncio
async def test_large_target_cost_center_lookup_sheet_is_not_user_facing_tab(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    run_date = date(2026, 6, 12)
    async with session_scope(url) as session:
        await _insert_cap(session, daily_cap=2)
        await _insert_lead(session, lead_id=23, priority=Decimal("20"))
        await session.commit()
        selection = await select_workbook_leads_for_store(
            session, cost_center="A100", run_date=run_date, backlog_threshold=99
        )

    output = generate_store_workbook(
        selection=selection,
        active_cost_centers=tuple(f"STORE-{i:04d}" for i in range(80)),
        output_root=tmp_path / "outputs",
        generated_at=datetime(2026, 6, 12, 8, 0, tzinfo=timezone.utc),
    )

    workbook = openpyxl.load_workbook(output.output_path)
    visible_tabs = [sheet.title for sheet in workbook.worksheets if sheet.sheet_state == "visible"]
    assert visible_tabs == [READ_ME_SHEET, FOLLOWUP_SHEET]
    assert workbook["_ACTIVE_COST_CENTER_LOOKUP"].sheet_state in {"hidden", "veryHidden"}


@pytest.mark.asyncio
async def test_active_store_dropdown_is_loaded_from_enabled_store_master_only(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    run_date = date(2026, 6, 12)
    async with session_scope(url) as session:
        await _insert_cap(session, daily_cap=2)
        await _insert_lead(session, lead_id=501, priority=Decimal("20"))
        await session.commit()
        active_cost_centers = await load_active_retention_stores(session)
        selection = await select_workbook_leads_for_store(
            session, cost_center="A100", run_date=run_date, backlog_threshold=99
        )

    output = generate_store_workbook(
        selection=selection,
        active_cost_centers=active_cost_centers,
        output_root=tmp_path / "outputs",
        generated_at=datetime(2026, 6, 12, 8, 0, tzinfo=timezone.utc),
    )

    workbook = openpyxl.load_workbook(output.output_path)
    sheet = workbook[FOLLOWUP_SHEET]
    headers = [cell.value for cell in sheet[1]]
    target_col = openpyxl.utils.get_column_letter(headers.index("Target Cost Center") + 1)
    validation_by_range = {
        str(rng): dv for dv in sheet.data_validations.dataValidation for rng in dv.cells.ranges
    }

    assert active_cost_centers == ("A100", "C300", "D400")
    assert validation_by_range[f"{target_col}2"].formula1 == '"A100,C300,D400"'
    assert "B200" not in validation_by_range[f"{target_col}2"].formula1


@pytest.mark.asyncio
async def test_controlled_editable_dropdowns_are_unlocked_and_have_expected_validations(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    run_date = date(2026, 6, 12)
    async with session_scope(url) as session:
        await _insert_cap(session, daily_cap=2)
        await _insert_lead(session, lead_id=502, priority=Decimal("20"))
        await session.commit()
        selection = await select_workbook_leads_for_store(
            session, cost_center="A100", run_date=run_date, backlog_threshold=99
        )

    output = generate_store_workbook(
        selection=selection,
        active_cost_centers=("A100", "C300", "D400"),
        output_root=tmp_path / "outputs",
        generated_at=datetime(2026, 6, 12, 8, 0, tzinfo=timezone.utc),
    )

    workbook = openpyxl.load_workbook(output.output_path)
    sheet = workbook[FOLLOWUP_SHEET]
    headers = [cell.value for cell in sheet[1]]
    validation_by_range = {
        str(rng): dv for dv in sheet.data_validations.dataValidation for rng in dv.cells.ranges
    }
    expected_dropdowns = {
        **{header: values for header, values in DROPDOWN_VALUES_BY_HEADER.items()},
        "Target Cost Center": ("A100", "C300", "D400"),
    }

    for header, values in expected_dropdowns.items():
        col_idx = headers.index(header) + 1
        col_letter = openpyxl.utils.get_column_letter(col_idx)
        validation = validation_by_range[f"{col_letter}2"]
        assert header in EDITABLE_COLUMNS.values()
        assert sheet.cell(row=2, column=col_idx).protection.locked is False
        assert validation.type == "list"
        assert validation.showErrorMessage is True
        assert validation.allow_blank is (header == "Target Cost Center")
        for value in values:
            assert value in validation.formula1


@pytest.mark.asyncio
async def test_generated_workbook_reingestion_ignores_protected_edits_and_preserves_identity_history(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    run_date = date(2026, 6, 12)
    async with session_scope(url) as session:
        await _insert_cap(session, daily_cap=2)
        await _insert_lead(session, lead_id=503, priority=Decimal("20"), mobile="9876500503")
        await session.commit()
        selection = await select_workbook_leads_for_store(
            session, cost_center="A100", run_date=run_date, backlog_threshold=99
        )

    output = generate_store_workbook(
        selection=selection,
        active_cost_centers=("A100", "C300", "D400"),
        output_root=tmp_path / "outputs",
        generated_at=datetime(2026, 6, 12, 8, 0, tzinfo=timezone.utc),
    )

    workbook = openpyxl.load_workbook(output.output_path)
    sheet = workbook[FOLLOWUP_SHEET]
    headers = [cell.value for cell in sheet[1]]
    col = {header: headers.index(header) + 1 for header in headers}

    assert sheet.cell(row=2, column=col["lead_id"]).value == 503
    assert sheet.cell(row=2, column=col["mobile_number"]).value == "9876500503"
    assert sheet.cell(row=2, column=col["normalized_mobile_number"]).value == "9876500503"

    sheet.cell(row=2, column=col["customer_name"], value="Tampered Name")
    sheet.cell(row=2, column=col["cost_center"], value="C300")
    sheet.cell(row=2, column=col["recommended_strategy"], value="Tampered Strategy")
    sheet.cell(row=2, column=col["normalized_mobile_number"], value="9999999999")
    sheet.cell(row=2, column=col["Contact Attempted"], value="Yes")
    sheet.cell(row=2, column=col["Contact Mode"], value="Call")
    sheet.cell(row=2, column=col["Customer Response"], value="Interested")
    sheet.cell(row=2, column=col["Order Expected"], value="Maybe")
    sheet.cell(row=2, column=col["Complaint"], value="No")
    sheet.cell(row=2, column=col["Do Not Contact"], value="No")
    sheet.cell(row=2, column=col["Handled By"], value="Staff Phase4")
    sheet.cell(row=2, column=col["Staff Remarks"], value="Valid editable values")
    workbook.save(output.output_path)

    result = await ingest_returned_workbook(database_url=url, path=output.output_path, pipeline_run_id="run-phase4-return")

    assert result.rows_seen == 1
    assert result.protected_edits_ignored == 4
    assert result.history_inserted == 1
    assert result.rows_skipped == 0
    assert {warning.code for warning in result.warnings} == {"protected_column_edit"}
    assert {warning.field_name for warning in result.warnings} >= {"customer_name", "cost_center", "recommended_strategy", "normalized_mobile_number"}

    async with session_scope(url) as session:
        lead = (
            await session.execute(
                sa.select(trx_customer_followup_leads).where(trx_customer_followup_leads.c.lead_id == 503)
            )
        ).mappings().one()
        history = (
            await session.execute(
                sa.select(trx_customer_followup_history).where(trx_customer_followup_history.c.lead_id == 503)
            )
        ).mappings().one()

    assert lead["cost_center"] == "A100"
    assert lead["customer_name"] == "Customer 503"
    assert lead["recommended_strategy"] == "Gentle reminder call"
    assert lead["normalized_mobile_number"] == "9876500503"
    assert history["pipeline_run_id"] == "run-phase4-return"
    assert history["handled_by"] == "Staff Phase4"
    assert history["contact_attempted"] is True
    assert history["contact_mode"] == "Call"
    assert history["customer_response"] == "Interested"
    assert history["order_expected"] == "Maybe"
    assert history["complaint_flag"] is False
    assert history["do_not_contact_flag"] is False
    assert history["staff_remarks"] == "Valid editable values"


def _snapshot_row(index: int, *, cost_center: str = "A100", snapshot_date: date = date(2026, 6, 12), priority: Decimal | None = None) -> CustomerRetentionSnapshotRow:
    mobile = f"987650{index:04d}"
    return CustomerRetentionSnapshotRow(
        snapshot_date=snapshot_date,
        cost_center=cost_center,
        customer_name=f"Snapshot Customer {index}",
        mobile_number=mobile,
        normalized_mobile_number=mobile,
        last_order_date=snapshot_date - timedelta(days=45),
        days_since_last_order=45,
        lifecycle_bucket="WARM",
        total_orders=2,
        lifetime_spend=Decimal("500.00"),
        average_order_value=Decimal("250.00"),
        last_order_amount=Decimal("250.00"),
        last_followup_date=None,
        last_followup_status=None,
        suppression_status=None,
        eligible_for_retention=True,
        priority_score=priority if priority is not None else Decimal(str(1000 - index)),
        recommended_strategy="Gentle reminder call",
    )


def _snapshot(rows) -> SnapshotResult:
    rows = tuple(rows)
    return SnapshotResult(
        snapshot_date=date(2026, 6, 12),
        rows=rows,
        rows_seen=len(rows),
        rows_invalid_mobile=0,
        rows_suppressed=0,
        rows_existing_open_lead=0,
    )


@pytest.mark.asyncio
async def test_retention_allocation_freeze_prevents_fresh_db_inserts(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    run_date = date(2026, 6, 12)
    async with session_scope(url) as session:
        await _insert_cap(session, daily_cap=13)
        await _insert_lead(session, lead_id=900, lead_date=run_date - timedelta(days=1), mobile="9999999900")
        result = await allocate_and_generate_retention_leads(
            session,
            snapshot=_snapshot(_snapshot_row(i) for i in range(1, 6)),
            active_stores=("A100",),
            run_date=run_date,
            backlog_threshold=0,
            pipeline_run_id="freeze-run",
        )
        await session.commit()

        fresh_created = (
            await session.execute(
                sa.select(sa.func.count()).select_from(trx_customer_followup_leads).where(
                    trx_customer_followup_leads.c.created_by_pipeline_run_id == "freeze-run",
                    trx_customer_followup_leads.c.lead_source_type == LEAD_SOURCE_RETENTION,
                )
            )
        ).scalar_one()

    assert result.leads_created == 0
    assert result.skipped_frozen == 5
    assert fresh_created == 0


@pytest.mark.asyncio
async def test_retention_allocation_caps_db_inserts_without_consuming_due_or_carry_forward(tmp_path: Path) -> None:
    url = await _prepare_db(tmp_path)
    run_date = date(2026, 6, 12)
    async with session_scope(url) as session:
        await _insert_cap(session, daily_cap=13)
        await _insert_lead(session, lead_id=901, lead_date=run_date - timedelta(days=30), mobile="9999999901", status=LEAD_STATUS_PENDING)
        await _insert_lead(session, lead_id=902, lead_date=run_date - timedelta(days=2), next_followup_date=run_date, mobile="9999999902")
        result = await allocate_and_generate_retention_leads(
            session,
            snapshot=_snapshot(_snapshot_row(i) for i in range(1, 21)),
            active_stores=("A100",),
            run_date=run_date,
            backlog_threshold=99,
            pipeline_run_id="cap-run",
        )
        await session.flush()
        selection = await select_workbook_leads_for_store(session, cost_center="A100", run_date=run_date, backlog_threshold=99)
        fresh_created = (
            await session.execute(
                sa.select(sa.func.count()).select_from(trx_customer_followup_leads).where(
                    trx_customer_followup_leads.c.created_by_pipeline_run_id == "cap-run",
                    trx_customer_followup_leads.c.lead_source_type == LEAD_SOURCE_RETENTION,
                    trx_customer_followup_leads.c.lead_date == run_date,
                )
            )
        ).scalar_one()
        await session.commit()

    assert result.leads_created == 13
    assert result.skipped_cap == 7
    assert fresh_created == 13
    assert selection.counts_by_category[CAP_WORK_SECTION_FRESH_RETENTION] == 13
    assert selection.counts_by_category[CAP_WORK_SECTION_PENDING_CARRY_FORWARD] == 1
    assert selection.counts_by_category[CAP_WORK_SECTION_DUE_FOLLOWUP] == 1
