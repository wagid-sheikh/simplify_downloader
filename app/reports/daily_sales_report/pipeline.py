from __future__ import annotations

import asyncio
from datetime import date, datetime, timezone
from time import perf_counter
from decimal import Decimal
from pathlib import Path
from typing import Mapping

import openpyxl
from jinja2 import Environment, FileSystemLoader, select_autoescape
from openpyxl.workbook import Workbook

from app.common.date_utils import aware_now, get_timezone
from app.common.db import session_scope
from app.config import config
from app.dashboard_downloader.db_tables import documents
from app.dashboard_downloader.json_logger import get_logger, log_event, new_run_id
from app.dashboard_downloader.notifications import send_notifications_for_run
from app.dashboard_downloader.pipelines.base import (
    PipelinePhaseTracker,
    persist_summary_record,
    resolve_run_env,
    update_summary_record,
)
from app.dashboard_downloader.report_generator import render_pdf_with_configured_browser

from app.reports.mtd_same_day_fulfillment.data import fetch_mtd_same_day_fulfillment
from app.reports.mtd_same_day_fulfillment.render import (
    render_html as render_mtd_same_day_html,
)
from app.reports.shared.formatters import (
    format_amount,
    format_ddmmyyyy,
    format_hhmm_ampm,
)
from app.reports.shared.same_day_fulfillment import (
    build_store_summary,
    format_duration_hours,
    format_duration_minutes,
    group_rows_by_store,
)

from app.reports.upstream import (
    DEGRADED_ORDERS_SYNC_MESSAGE,
    OrdersSyncUpstreamContext,
    build_orders_sync_upstream_context,
)

from .data import DailySalesReportData, fetch_daily_sales_report
from .to_be_recovered import (
    DOCUMENT_TYPE as TO_BE_RECOVERED_DOCUMENT_TYPE,
    PIPELINE_OUTPUT_PREFIX as TO_BE_RECOVERED_OUTPUT_PREFIX,
    TEMPLATE_NAME as TO_BE_RECOVERED_TEMPLATE_NAME,
    build_context as build_to_be_recovered_context,
)

PIPELINE_NAME = "reports.daily_sales_report"
TEMPLATE_NAME = "daily_sales_report.html"
SHORT_PAYMENTS_TEMPLATE_NAME = "short_payments_report.html"
SHORT_PAYMENTS_DOCUMENT_TYPE = "daily_sales_short_payments_pdf"
SHORT_PAYMENTS_OUTPUT_PREFIX = "reports.daily_sales_report_short_payments"
ACTUAL_PAYMENTS_NOT_FOUND_TEMPLATE_NAME = "actual_payments_not_found_report.html"
ACTUAL_PAYMENTS_NOT_FOUND_DOCUMENT_TYPE = "daily_sales_actual_payments_not_found_pdf"
ACTUAL_PAYMENTS_NOT_FOUND_OUTPUT_PREFIX = (
    "reports.daily_sales_report_actual_payments_not_found"
)
ACTUAL_PAYMENTS_NOT_FOUND_WORKBOOK_DOCUMENT_TYPE = (
    "daily_sales_actual_payments_not_found_xlsx"
)
ACTUAL_PAYMENTS_NOT_FOUND_WORKBOOK_OUTPUT_PREFIX = (
    "reports.daily_sales_report_actual_payments_not_found"
)
TEMPLATE_DIR = Path("app") / "reports" / "daily_sales_report" / "templates"
SHARED_TEMPLATE_DIR = Path("app") / "reports" / "shared" / "templates"
OUTPUT_ROOT = Path("app") / "reports" / "output_files"
EXCEL_DATETIME_NUMBER_FORMAT = "yyyy-mm-dd hh:mm:ss"


def _elapsed_ms(start: float) -> int:
    return int((perf_counter() - start) * 1000)


def _log_timing(
    *,
    logger: object,
    phase: str,
    message: str,
    report_date: date,
    started_at: float,
    **counts: object,
) -> None:
    log_event(
        logger=logger,
        phase=phase,
        status="ok",
        message=message,
        report_date=report_date.isoformat(),
        elapsed_ms=_elapsed_ms(started_at),
        **counts,
    )


def _format_amount(value: Decimal | int | float | None) -> str:
    return format_amount(value)


def _format_ddmmyyyy(value: object | None) -> str:
    return format_ddmmyyyy(value)


def _format_hhmm_ampm(value: object | None) -> str:
    return format_hhmm_ampm(value)


def _render_html(
    context: Mapping[str, object], template_name: str = TEMPLATE_NAME
) -> str:
    env = Environment(
        loader=FileSystemLoader([str(TEMPLATE_DIR), str(SHARED_TEMPLATE_DIR)]),
        autoescape=select_autoescape(["html", "xml"]),
    )
    env.filters["format_amount"] = _format_amount
    env.filters["format_ddmmyyyy"] = _format_ddmmyyyy
    env.filters["format_hhmm_ampm"] = _format_hhmm_ampm
    template = env.get_template(template_name)
    return template.render(**context)


async def _persist_document(
    *,
    database_url: str,
    run_id: str,
    report_date: date,
    file_path: Path,
    doc_type: str,
) -> None:
    async with session_scope(database_url) as session:
        await session.execute(
            documents.insert().values(
                doc_type=doc_type,
                doc_subtype="pipeline_report",
                doc_date=report_date,
                reference_name_1="pipeline",
                reference_id_1=PIPELINE_NAME,
                reference_name_2="run_id",
                reference_id_2=run_id,
                reference_name_3="report_date",
                reference_id_3=report_date.isoformat(),
                file_name=file_path.name,
                mime_type=(
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    if file_path.suffix.lower() == ".xlsx"
                    else "application/pdf"
                ),
                file_size_bytes=(
                    file_path.stat().st_size if file_path.exists() else None
                ),
                storage_backend="fs",
                file_path=str(file_path),
                created_at=datetime.now(timezone.utc),
                created_by="pipeline",
            )
        )
        await session.commit()


def _build_context(
    data: DailySalesReportData,
    run_environment: str,
    orders_sync_upstream: OrdersSyncUpstreamContext | None = None,
) -> dict[str, object]:
    report_date_display = data.report_date.strftime("%d-%b-%Y")
    missing_payment_rows = list(data.missing_payment_rows)
    missing_payment_summary_by_cost_center: list[dict[str, object]] = []
    missing_payment_grouped_by_cost_center: list[dict[str, object]] = []
    missing_payment_by_cost_center: dict[str, list[object]] = {}
    for row in sorted(
        missing_payment_rows,
        key=lambda item: (
            str(getattr(item, "cost_center", "") or ""),
            getattr(item, "order_date", None) or datetime.min,
            str(getattr(item, "order_number", "") or ""),
        ),
        reverse=True,
    ):
        cost_center = str(getattr(row, "cost_center", "") or "")
        missing_payment_by_cost_center.setdefault(cost_center, []).append(row)
    for cost_center in sorted(missing_payment_by_cost_center):
        rows = sorted(
            missing_payment_by_cost_center[cost_center],
            key=lambda item: (
                getattr(item, "order_date", None) or datetime.min,
                str(getattr(item, "order_number", "") or ""),
            ),
            reverse=True,
        )
        total_order_amount = sum(
            (
                getattr(item, "order_amount", Decimal("0")) or Decimal("0")
                for item in rows
            ),
            Decimal("0"),
        )
        missing_payment_summary_by_cost_center.append(
            {
                "cost_center": cost_center,
                "count": len(rows),
                "total_order_amount": total_order_amount,
            }
        )
        missing_payment_grouped_by_cost_center.append(
            {
                "cost_center": cost_center,
                "rows": rows,
                "group_count": len(rows),
                "group_total_order_amount": total_order_amount,
            }
        )
    orders_sync_upstream = orders_sync_upstream or build_orders_sync_upstream_context()
    integrity_findings = list(getattr(data, "integrity_findings", []))
    integrity_has_errors = any(finding.severity == "error" for finding in integrity_findings)
    return {
        "company_name": "The Shaw Ventures",
        "report_date_display": report_date_display,
        "run_environment": run_environment,
        "orders_sync_upstream_status": orders_sync_upstream.status or "",
        "orders_sync_upstream_run_id": orders_sync_upstream.run_id or "",
        "orders_sync_is_degraded": orders_sync_upstream.is_degraded,
        "orders_sync_warning_text": orders_sync_upstream.warning_text,
        "integrity_findings": integrity_findings,
        "integrity_has_errors": integrity_has_errors,
        "rows": data.rows,
        "totals": data.totals,
        "edited_orders": data.edited_orders,
        "edited_orders_summary": data.edited_orders_summary,
        "edited_orders_totals": data.edited_orders_totals,
        "missed_leads": data.missed_leads,
        "cancelled_leads": data.cancelled_leads,
        "lead_performance_summary": data.lead_performance_summary,
        "to_be_recovered": data.to_be_recovered,
        "to_be_compensated": data.to_be_compensated,
        "to_be_recovered_total_order_value": data.to_be_recovered_total_order_value,
        "to_be_compensated_total_order_value": data.to_be_compensated_total_order_value,
        "auto_cleared_order_numbers_text": getattr(
            data, "auto_cleared_order_numbers_text", ""
        ),
        "same_day_fulfillment_rows": data.same_day_fulfillment_rows,
        "missing_payment_rows": missing_payment_rows,
        "missing_payment_summary_by_cost_center": missing_payment_summary_by_cost_center,
        "missing_payment_grouped_by_cost_center": missing_payment_grouped_by_cost_center,
        "short_payment_rows": getattr(data, "short_payment_rows", []),
        "report_day_orders_by_cost_center": getattr(
            data, "report_day_orders_by_cost_center", []
        ),
        "same_day_grouped_rows_by_store": group_rows_by_store(
            data.same_day_fulfillment_rows
        ),
        "same_day_store_summary_rows": build_store_summary(
            data.same_day_fulfillment_rows
        ),
        "format_duration_hours": format_duration_hours,
        "format_duration_minutes": format_duration_minutes,
    }


def _short_payments_output_path(report_date: date) -> Path:
    return OUTPUT_ROOT / f"{SHORT_PAYMENTS_OUTPUT_PREFIX}_{report_date.isoformat()}.pdf"


def _actual_payments_not_found_output_path(report_date: date) -> Path:
    return (
        OUTPUT_ROOT
        / f"{ACTUAL_PAYMENTS_NOT_FOUND_OUTPUT_PREFIX}_{report_date.isoformat()}.pdf"
    )


def _actual_payments_not_found_workbook_output_path(report_date: date) -> Path:
    return OUTPUT_ROOT / (
        f"{ACTUAL_PAYMENTS_NOT_FOUND_WORKBOOK_OUTPUT_PREFIX}_{report_date.isoformat()}.xlsx"
    )


def _sanitize_worksheet_name(value: str, used_names: set[str]) -> str:
    normalized = "".join("_" if char in r"[]:*?/\\" else char for char in value).strip()
    base_name = normalized[:31] if normalized else "Unspecified"
    candidate = base_name
    suffix = 1
    while candidate in used_names:
        suffix_text = f"_{suffix}"
        candidate = f"{base_name[: 31 - len(suffix_text)]}{suffix_text}"
        suffix += 1
    used_names.add(candidate)
    return candidate


def _build_actual_payments_not_found_workbook(
    *, rows: list[object], output_path: Path, business_timezone: object
) -> tuple[Path, int, int]:
    def _coerce_datetime_like(value: object) -> datetime | None:
        if isinstance(value, datetime):
            return value
        to_pydatetime = getattr(value, "to_pydatetime", None)
        if callable(to_pydatetime):
            coerced = to_pydatetime()
            if isinstance(coerced, datetime):
                return coerced
        return None

    def _normalize_datetime_for_excel(value: object) -> tuple[object, bool]:
        dt_value = _coerce_datetime_like(value)
        if dt_value is None:
            return value, False
        if dt_value.tzinfo is None:
            return dt_value, True
        localized = dt_value.astimezone(business_timezone)
        return localized.replace(tzinfo=None), True

    workbook: Workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    headers = [
        "Cost Center",
        "Order Number",
        "Order Date",
        "Customer Name",
        "Mobile Number",
        "Order Amount",
    ]
    rows_by_cost_center: dict[str, list[object]] = {}
    for row in rows:
        cost_center = str(getattr(row, "cost_center", "") or "")
        rows_by_cost_center.setdefault(cost_center, []).append(row)

    used_sheet_names: set[str] = set()
    rows_processed = 0
    datetime_fields_normalized = 0
    for cost_center in sorted(rows_by_cost_center):
        sheet_name = _sanitize_worksheet_name(cost_center, used_sheet_names)
        worksheet = workbook.create_sheet(title=sheet_name)
        worksheet.append(headers)
        sorted_rows = sorted(
            rows_by_cost_center[cost_center],
            key=lambda item: (
                getattr(item, "order_date", None) or datetime.min,
                str(getattr(item, "order_number", "") or ""),
            ),
            reverse=True,
        )
        for row in sorted_rows:
            rows_processed += 1
            raw_row_values = [
                getattr(row, "cost_center", ""),
                getattr(row, "order_number", ""),
                getattr(row, "order_date", None),
                getattr(row, "customer_name", ""),
                getattr(row, "mobile_number", ""),
                getattr(row, "order_amount", None),
            ]
            normalized_row_values: list[object] = []
            datetime_columns: set[int] = set()
            for column_index, raw_value in enumerate(raw_row_values, start=1):
                field_name = headers[column_index - 1]
                try:
                    normalized_value, is_datetime = _normalize_datetime_for_excel(
                        raw_value
                    )
                except Exception as exc:
                    raise RuntimeError(
                        "Failed to normalize datetime for APNF workbook "
                        f"(sheet={sheet_name}, field={field_name}, value={raw_value!r}): {exc}"
                    ) from exc
                normalized_row_values.append(normalized_value)
                if is_datetime:
                    datetime_fields_normalized += 1
                    datetime_columns.add(column_index)

            worksheet.append(normalized_row_values)
            row_number = worksheet.max_row
            for column_index in datetime_columns:
                worksheet.cell(row=row_number, column=column_index).number_format = (
                    EXCEL_DATETIME_NUMBER_FORMAT
                )
    if not workbook.sheetnames:
        worksheet = workbook.create_sheet(title="No Data")
        worksheet.append(headers)
    workbook.save(output_path)
    return output_path, rows_processed, datetime_fields_normalized


async def _generate_short_payments_pdf(
    *,
    context: Mapping[str, object],
    report_date: date,
    logger: object,
) -> Path:
    """Render the Short Payments PDF independently from the main report PDF.

    Keep this helper outside any Daily Sales existing-run skip path: Short
    Payments is a current action list and must refresh on every report run,
    even when the main Daily Sales snapshot for the date already exists.
    """

    short_payments_html = _render_html(
        context, template_name=SHORT_PAYMENTS_TEMPLATE_NAME
    )
    short_payments_output_path = _short_payments_output_path(report_date)
    if short_payments_output_path.exists():
        short_payments_output_path.unlink()
    await render_pdf_with_configured_browser(
        short_payments_html,
        short_payments_output_path,
        pdf_options={"format": "A4", "landscape": True},
        logger=logger,
    )
    return short_payments_output_path


async def _run(
    report_date: date | None,
    env: str | None,
    force: bool,
    orders_sync_upstream_status: str | None = None,
    orders_sync_upstream_run_id: str | None = None,
) -> None:
    run_env = resolve_run_env(env)
    run_id = new_run_id()
    logger = get_logger(run_id=run_id)
    tracker = PipelinePhaseTracker(
        pipeline_name=PIPELINE_NAME, env=run_env, run_id=run_id
    )
    database_url = config.database_url
    orders_sync_upstream = build_orders_sync_upstream_context(
        status=orders_sync_upstream_status, run_id=orders_sync_upstream_run_id
    )

    try:
        if not database_url:
            tracker.mark_phase("load_data", "error")
            tracker.add_summary(
                "Database URL is missing; cannot generate daily sales report."
            )
            tracker.overall = "error"
            log_event(
                logger=logger,
                phase="load_data",
                status="error",
                message="Database URL is missing; cannot generate daily sales report.",
            )
            return

        tz = get_timezone()
        resolved_date = report_date or aware_now(tz).date()
        tracker.set_report_date(resolved_date)
        if orders_sync_upstream.status or orders_sync_upstream.run_id:
            status_text = orders_sync_upstream.status or "unknown"
            run_id_text = orders_sync_upstream.run_id or "unknown"
            tracker.add_summary(
                f"Upstream orders sync: status={status_text}, run_id={run_id_text}."
            )
        if orders_sync_upstream.is_degraded:
            tracker.mark_phase("upstream_orders_sync", "warning")
            tracker.add_summary(DEGRADED_ORDERS_SYNC_MESSAGE)
        log_event(
            logger=logger,
            phase="orchestrator",
            message="starting daily sales report pipeline",
            report_date=resolved_date.isoformat(),
            run_env=run_env,
            force=force,
            orders_sync_upstream_status=orders_sync_upstream.status,
            orders_sync_upstream_run_id=orders_sync_upstream.run_id,
            orders_sync_is_degraded=orders_sync_upstream.is_degraded,
        )

        data_fetch_started_at = perf_counter()
        data = await fetch_daily_sales_report(
            database_url=database_url,
            report_date=resolved_date,
        )
        tracker.mark_phase("load_data", "ok")
        _log_timing(
            logger=logger,
            phase="load_data",
            message="daily sales data fetched",
            report_date=resolved_date,
            started_at=data_fetch_started_at,
            rows=len(data.rows),
            edited_orders=len(data.edited_orders),
            missing_payment_rows=len(data.missing_payment_rows),
            short_payment_rows=len(getattr(data, "short_payment_rows", [])),
            to_be_recovered_rows=len(getattr(data, "to_be_recovered", [])),
            integrity_findings=len(getattr(data, "integrity_findings", [])),
        )
        log_event(
            logger=logger,
            phase="load_data",
            message="daily sales report data loaded",
            report_date=resolved_date.isoformat(),
            rows=len(data.rows),
            edited_orders=len(data.edited_orders),
            missing_payment_rows=len(data.missing_payment_rows),
            short_payment_rows=len(getattr(data, "short_payment_rows", [])),
            integrity_findings=len(getattr(data, "integrity_findings", [])),
        )
        integrity_findings = list(getattr(data, "integrity_findings", []))
        integrity_errors = [finding for finding in integrity_findings if finding.severity == "error"]
        integrity_warnings = [finding for finding in integrity_findings if finding.severity == "warning"]
        if integrity_findings:
            tracker.mark_phase("validate_integrity", "warning")
            tracker.add_summary(
                f"Daily Sales integrity findings: {len(integrity_errors)} hard error(s), "
                f"{len(integrity_warnings)} warning(s). "
                + " ".join(finding.message for finding in integrity_findings)
            )
        else:
            tracker.mark_phase("validate_integrity", "ok")

        context_started_at = perf_counter()
        context = _build_context(data, run_env, orders_sync_upstream)
        _log_timing(
            logger=logger,
            phase="build_context",
            message="daily sales context built",
            report_date=resolved_date,
            started_at=context_started_at,
            rows=len(data.rows),
            edited_orders=len(data.edited_orders),
            missing_payment_rows=len(data.missing_payment_rows),
            short_payment_rows=len(getattr(data, "short_payment_rows", [])),
            same_day_fulfillment_rows=len(
                getattr(data, "same_day_fulfillment_rows", [])
            ),
        )
        html_started_at = perf_counter()
        html = _render_html(context)
        _log_timing(
            logger=logger,
            phase="render_html",
            message="daily sales html rendered",
            report_date=resolved_date,
            started_at=html_started_at,
            rows=len(data.rows),
            missing_payment_rows=len(data.missing_payment_rows),
            short_payment_rows=len(getattr(data, "short_payment_rows", [])),
        )
        short_payments_pdf_generated = True
        actual_payments_not_found_pdf_generated = True
        to_be_recovered_context_started_at = perf_counter()
        to_be_recovered_context = build_to_be_recovered_context(
            rows=data.to_be_recovered,
            report_date=resolved_date,
            run_environment=run_env,
            auto_cleared_order_numbers_text=getattr(
                data, "auto_cleared_order_numbers_text", ""
            ),
        )
        _log_timing(
            logger=logger,
            phase="build_context",
            message="to-be-recovered context built",
            report_date=resolved_date,
            started_at=to_be_recovered_context_started_at,
            to_be_recovered_rows=len(data.to_be_recovered),
            auto_cleared_orders=len(getattr(data, "auto_cleared_order_numbers", [])),
        )
        to_be_recovered_html_started_at = perf_counter()
        to_be_recovered_html = _render_html(
            to_be_recovered_context,
            template_name=TO_BE_RECOVERED_TEMPLATE_NAME,
        )
        _log_timing(
            logger=logger,
            phase="render_html",
            message="to-be-recovered html rendered",
            report_date=resolved_date,
            started_at=to_be_recovered_html_started_at,
            to_be_recovered_rows=len(data.to_be_recovered),
        )
        mtd_attachment_generated = True
        mtd_rows = []
        same_day_html: str | None = None
        mtd_attachment_error: str | None = None
        mtd_fetch_started_at = perf_counter()
        try:
            mtd_rows = await fetch_mtd_same_day_fulfillment(
                database_url=database_url, report_date=resolved_date
            )
            _log_timing(
                logger=logger,
                phase="load_data",
                message="mtd same-day fulfillment data fetched",
                report_date=resolved_date,
                started_at=mtd_fetch_started_at,
                mtd_row_count=len(mtd_rows),
            )
        except Exception as exc:
            mtd_attachment_generated = False
            mtd_attachment_error = str(exc)
            tracker.mark_phase("render_html", "warning")
            tracker.add_summary(
                f"MTD attachment generation failed during data load ({exc})."
            )
            tracker.overall = "warning"
            log_event(
                logger=logger,
                phase="render_html",
                status="warning",
                message="failed to fetch mtd same-day fulfillment data; continuing daily report generation",
                report_date=resolved_date.isoformat(),
                database_backend=database_url.split("://", 1)[0],
                function_name="fetch_mtd_same_day_fulfillment",
                error=str(exc),
            )
        mtd_start = resolved_date.replace(day=1)
        mtd_end = resolved_date
        if mtd_attachment_generated:
            mtd_html_started_at = perf_counter()
            same_day_html = render_mtd_same_day_html(
                rows=mtd_rows,
                report_date_display=resolved_date.strftime("%d-%b-%Y"),
                mtd_start_display=mtd_start.strftime("%d-%b-%Y"),
                mtd_end_display=mtd_end.strftime("%d-%b-%Y"),
            )
            _log_timing(
                logger=logger,
                phase="render_html",
                message="mtd same-day fulfillment html rendered",
                report_date=resolved_date,
                started_at=mtd_html_started_at,
                mtd_row_count=len(mtd_rows),
            )
            tracker.mark_phase("render_html", "ok")
        log_event(
            logger=logger,
            phase="render_html",
            status="warning" if not mtd_attachment_generated else "ok",
            message=(
                "rendered daily sales report html; skipped MTD attachment"
                if not mtd_attachment_generated
                else "rendered daily sales report html"
            ),
            report_date=resolved_date.isoformat(),
            pipeline_name=PIPELINE_NAME,
            mtd_start=mtd_start.isoformat(),
            mtd_end=mtd_end.isoformat(),
            mtd_row_count=len(mtd_rows),
            short_payment_rows=len(getattr(data, "short_payment_rows", [])),
            short_payments_pdf_generated=short_payments_pdf_generated,
            actual_payments_not_found_rows=len(data.missing_payment_rows),
            actual_payments_not_found_pdf_generated=actual_payments_not_found_pdf_generated,
        )

        OUTPUT_ROOT.mkdir(parents=True, exist_ok=True)
        output_path = OUTPUT_ROOT / f"{PIPELINE_NAME}_{resolved_date.isoformat()}.pdf"
        same_day_output_path = (
            OUTPUT_ROOT
            / f"reports.mtd_same_day_fulfillment_{resolved_date.isoformat()}.pdf"
        )
        short_payments_output_path = _short_payments_output_path(resolved_date)
        actual_payments_not_found_output_path = _actual_payments_not_found_output_path(
            resolved_date
        )
        actual_payments_not_found_workbook_output_path = (
            _actual_payments_not_found_workbook_output_path(resolved_date)
        )
        to_be_recovered_output_path = (
            OUTPUT_ROOT
            / f"{TO_BE_RECOVERED_OUTPUT_PREFIX}_{resolved_date.isoformat()}.pdf"
        )
        if output_path.exists():
            output_path.unlink()
        if same_day_output_path.exists():
            same_day_output_path.unlink()
        if to_be_recovered_output_path.exists():
            to_be_recovered_output_path.unlink()
        if actual_payments_not_found_output_path.exists():
            actual_payments_not_found_output_path.unlink()
        if actual_payments_not_found_workbook_output_path.exists():
            actual_payments_not_found_workbook_output_path.unlink()
        try:
            pdf_started_at = perf_counter()
            await render_pdf_with_configured_browser(
                html,
                output_path,
                pdf_options={"format": "A4", "landscape": True},
                logger=logger,
            )
            _log_timing(
                logger=logger,
                phase="render_pdf",
                message="daily sales pdf rendered",
                report_date=resolved_date,
                started_at=pdf_started_at,
                rows=len(data.rows),
                file_path=str(output_path),
            )

            pdf_started_at = perf_counter()
            await render_pdf_with_configured_browser(
                to_be_recovered_html,
                to_be_recovered_output_path,
                pdf_options={"format": "A4", "landscape": True},
                logger=logger,
            )
            _log_timing(
                logger=logger,
                phase="render_pdf",
                message="to-be-recovered pdf rendered",
                report_date=resolved_date,
                started_at=pdf_started_at,
                to_be_recovered_rows=len(data.to_be_recovered),
                file_path=str(to_be_recovered_output_path),
            )

            pdf_started_at = perf_counter()
            short_payments_output_path = await _generate_short_payments_pdf(
                context=context,
                report_date=resolved_date,
                logger=logger,
            )
            _log_timing(
                logger=logger,
                phase="render_pdf",
                message="short payments pdf rendered",
                report_date=resolved_date,
                started_at=pdf_started_at,
                short_payment_rows=len(getattr(data, "short_payment_rows", [])),
                file_path=str(short_payments_output_path),
            )

            actual_payments_not_found_html = _render_html(
                context,
                template_name=ACTUAL_PAYMENTS_NOT_FOUND_TEMPLATE_NAME,
            )
            pdf_started_at = perf_counter()
            await render_pdf_with_configured_browser(
                actual_payments_not_found_html,
                actual_payments_not_found_output_path,
                pdf_options={"format": "A4", "landscape": True},
                logger=logger,
            )
            _log_timing(
                logger=logger,
                phase="render_pdf",
                message="actual payments not found pdf rendered",
                report_date=resolved_date,
                started_at=pdf_started_at,
                actual_payments_not_found_rows=len(data.missing_payment_rows),
                file_path=str(actual_payments_not_found_output_path),
            )

            log_event(
                logger=logger,
                phase="render_xlsx",
                status="ok",
                message="generating actual payments not found workbook",
                report_date=resolved_date.isoformat(),
                run_id=run_id,
                rows=len(data.missing_payment_rows),
            )
            workbook_started_at = perf_counter()
            _, rows_processed, datetime_fields_normalized = (
                _build_actual_payments_not_found_workbook(
                    rows=list(data.missing_payment_rows),
                    output_path=actual_payments_not_found_workbook_output_path,
                    business_timezone=tz,
                )
            )
            _log_timing(
                logger=logger,
                phase="render_xlsx",
                message="actual payments not found workbook generated",
                report_date=resolved_date,
                started_at=workbook_started_at,
                run_id=run_id,
                rows_processed=rows_processed,
                datetime_fields_normalized=datetime_fields_normalized,
                timezone=str(getattr(tz, "key", tz)),
                file_path=str(actual_payments_not_found_workbook_output_path),
            )
            if mtd_attachment_generated and same_day_html is not None:
                pdf_started_at = perf_counter()
                await render_pdf_with_configured_browser(
                    same_day_html,
                    same_day_output_path,
                    pdf_options={"format": "A4", "landscape": True},
                    logger=logger,
                )
                _log_timing(
                    logger=logger,
                    phase="render_pdf",
                    message="mtd same-day fulfillment pdf rendered",
                    report_date=resolved_date,
                    started_at=pdf_started_at,
                    mtd_row_count=len(mtd_rows),
                    file_path=str(same_day_output_path),
                )
        except asyncio.TimeoutError as exc:
            tracker.mark_phase("render_pdf", "error")
            tracker.add_summary(
                f"PDF rendering timed out after {config.pdf_render_timeout_seconds}s."
            )
            tracker.overall = "error"
            log_event(
                logger=logger,
                phase="render_pdf",
                status="error",
                message="daily sales report pdf render timed out",
                report_date=resolved_date.isoformat(),
                file_path=str(output_path),
                same_day_file_path=str(same_day_output_path),
                to_be_recovered_file_path=str(to_be_recovered_output_path),
                short_payments_file_path=str(short_payments_output_path),
                actual_payments_not_found_file_path=str(
                    actual_payments_not_found_output_path
                ),
                actual_payments_not_found_workbook_file_path=str(
                    actual_payments_not_found_workbook_output_path
                ),
                pipeline_name=PIPELINE_NAME,
                mtd_start=mtd_start.isoformat(),
                mtd_end=mtd_end.isoformat(),
                mtd_row_count=len(mtd_rows),
                short_payment_rows=len(getattr(data, "short_payment_rows", [])),
                short_payments_pdf_generated=False,
                actual_payments_not_found_rows=len(data.missing_payment_rows),
                actual_payments_not_found_pdf_generated=False,
                error=str(exc),
            )
            finished_at = datetime.now(timezone.utc)
            record = tracker.build_record(finished_at)
            await persist_summary_record(database_url, record)
            return
        tracker.mark_phase(
            "render_pdf", "warning" if not mtd_attachment_generated else "ok"
        )
        log_event(
            logger=logger,
            phase="render_pdf",
            status="warning" if not mtd_attachment_generated else "ok",
            message=(
                "rendered daily sales report pdf; skipped MTD attachment pdf"
                if not mtd_attachment_generated
                else "rendered daily sales report pdf"
            ),
            report_date=resolved_date.isoformat(),
            file_path=str(output_path),
            same_day_file_path=str(same_day_output_path),
            to_be_recovered_file_path=str(to_be_recovered_output_path),
            short_payments_file_path=str(short_payments_output_path),
            actual_payments_not_found_file_path=str(
                actual_payments_not_found_output_path
            ),
            actual_payments_not_found_workbook_file_path=str(
                actual_payments_not_found_workbook_output_path
            ),
            pipeline_name=PIPELINE_NAME,
            mtd_start=mtd_start.isoformat(),
            mtd_end=mtd_end.isoformat(),
            mtd_row_count=len(mtd_rows),
            short_payment_rows=len(getattr(data, "short_payment_rows", [])),
            short_payments_pdf_generated=short_payments_pdf_generated,
            actual_payments_not_found_rows=len(data.missing_payment_rows),
            actual_payments_not_found_pdf_generated=actual_payments_not_found_pdf_generated,
        )

        persist_documents_started_at = perf_counter()
        documents_persisted = 0
        await _persist_document(
            database_url=database_url,
            run_id=run_id,
            report_date=resolved_date,
            file_path=output_path,
            doc_type="daily_sales_report_pdf",
        )
        documents_persisted += 1
        await _persist_document(
            database_url=database_url,
            run_id=run_id,
            report_date=resolved_date,
            file_path=to_be_recovered_output_path,
            doc_type=TO_BE_RECOVERED_DOCUMENT_TYPE,
        )
        documents_persisted += 1
        await _persist_document(
            database_url=database_url,
            run_id=run_id,
            report_date=resolved_date,
            file_path=short_payments_output_path,
            doc_type=SHORT_PAYMENTS_DOCUMENT_TYPE,
        )
        documents_persisted += 1
        await _persist_document(
            database_url=database_url,
            run_id=run_id,
            report_date=resolved_date,
            file_path=actual_payments_not_found_output_path,
            doc_type=ACTUAL_PAYMENTS_NOT_FOUND_DOCUMENT_TYPE,
        )
        documents_persisted += 1
        await _persist_document(
            database_url=database_url,
            run_id=run_id,
            report_date=resolved_date,
            file_path=actual_payments_not_found_workbook_output_path,
            doc_type=ACTUAL_PAYMENTS_NOT_FOUND_WORKBOOK_DOCUMENT_TYPE,
        )
        documents_persisted += 1
        if mtd_attachment_generated:
            await _persist_document(
                database_url=database_url,
                run_id=run_id,
                report_date=resolved_date,
                file_path=same_day_output_path,
                doc_type="mtd_same_day_fulfillment_pdf",
            )
            documents_persisted += 1
        _log_timing(
            logger=logger,
            phase="persist_documents",
            message="daily sales documents persisted",
            report_date=resolved_date,
            started_at=persist_documents_started_at,
            documents_persisted=documents_persisted,
            mtd_attachment_generated=mtd_attachment_generated,
        )
        tracker.mark_phase(
            "persist_documents", "warning" if not mtd_attachment_generated else "ok"
        )
        log_event(
            logger=logger,
            phase="persist_documents",
            status="warning" if not mtd_attachment_generated else "ok",
            message=(
                "saved daily sales report document; MTD attachment document skipped"
                if not mtd_attachment_generated
                else "saved daily sales report document"
            ),
            report_date=resolved_date.isoformat(),
            file_path=str(output_path),
            same_day_file_path=str(same_day_output_path),
            to_be_recovered_file_path=str(to_be_recovered_output_path),
            short_payments_file_path=str(short_payments_output_path),
            actual_payments_not_found_file_path=str(
                actual_payments_not_found_output_path
            ),
            actual_payments_not_found_workbook_file_path=str(
                actual_payments_not_found_workbook_output_path
            ),
            pipeline_name=PIPELINE_NAME,
            mtd_start=mtd_start.isoformat(),
            mtd_end=mtd_end.isoformat(),
            mtd_row_count=len(mtd_rows),
            short_payment_rows=len(getattr(data, "short_payment_rows", [])),
            short_payments_pdf_generated=short_payments_pdf_generated,
            actual_payments_not_found_rows=len(data.missing_payment_rows),
            actual_payments_not_found_pdf_generated=actual_payments_not_found_pdf_generated,
        )

        tracker.metrics = {
            "report_date": resolved_date.isoformat(),
            "orders_sync_upstream": orders_sync_upstream.as_metrics(),
            "orders_sync_upstream_status": orders_sync_upstream.status,
            "orders_sync_upstream_run_id": orders_sync_upstream.run_id,
            "orders_sync_is_degraded": orders_sync_upstream.is_degraded,
            "rows": len(data.rows),
            "edited_orders": len(data.edited_orders),
            "integrity_findings": [finding.as_metrics_payload() for finding in integrity_findings],
            "integrity_error_count": len(integrity_errors),
            "integrity_warning_count": len(integrity_warnings),
            "integrity_invalid_data_report": bool(integrity_errors),
            "mtd_attachment_generated": mtd_attachment_generated,
            "mtd_attachment_row_count": len(mtd_rows),
            "mtd_attachment_error": mtd_attachment_error,
            "short_payment_rows": len(getattr(data, "short_payment_rows", [])),
            "short_payments_pdf_file_path": str(short_payments_output_path),
            "short_payments_pdf_generated": short_payments_pdf_generated,
            "actual_payments_not_found_rows": len(data.missing_payment_rows),
            "actual_payments_not_found_pdf_file_path": str(
                actual_payments_not_found_output_path
            ),
            "actual_payments_not_found_pdf_generated": actual_payments_not_found_pdf_generated,
            "actual_payments_not_found_workbook_file_path": str(
                actual_payments_not_found_workbook_output_path
            ),
            "to_be_recovered_orders": len(data.to_be_recovered),
            "to_be_recovered_total_order_value": str(
                data.to_be_recovered_total_order_value
            ),
        }
        tracker.add_summary(
            f"Daily sales report generated for {resolved_date.isoformat()} with "
            f"{len(data.rows)} cost centers and {len(data.edited_orders)} edited orders."
        )
        if not mtd_attachment_generated:
            tracker.add_summary(
                "MTD same-day fulfillment attachment was not generated; daily sales report PDF was still produced."
            )

        pre_finished_at = datetime.now(timezone.utc)
        pre_record = tracker.build_record(pre_finished_at)
        await persist_summary_record(database_url, pre_record)

        notification_started_at = perf_counter()
        try:
            notification_result = await send_notifications_for_run(
                PIPELINE_NAME, run_id
            )
            emails_planned = int(notification_result.get("emails_planned") or 0)
            emails_sent = int(notification_result.get("emails_sent") or 0)
            notification_errors = notification_result.get("errors") or []
            _log_timing(
                logger=logger,
                phase="send_email",
                message="daily sales notification sent",
                report_date=resolved_date,
                started_at=notification_started_at,
                emails_planned=emails_planned,
                emails_sent=emails_sent,
                notification_error_count=len(notification_errors),
            )
            if emails_sent > 0:
                tracker.mark_phase("send_email", "ok")
                log_event(
                    logger=logger,
                    phase="send_email",
                    message="notification sent",
                    report_date=resolved_date.isoformat(),
                    emails_planned=emails_planned,
                    emails_sent=emails_sent,
                    notification_errors=notification_errors,
                )
            else:
                tracker.mark_phase("send_email", "warning")
                tracker.add_summary(
                    "Notification dispatch completed but no emails were sent; see logs."
                )
                tracker.overall = "warning"
                log_event(
                    logger=logger,
                    phase="send_email",
                    status="warning",
                    message="notification dispatch completed with zero emails sent",
                    report_date=resolved_date.isoformat(),
                    emails_planned=emails_planned,
                    emails_sent=emails_sent,
                    notification_errors=notification_errors,
                )
        except Exception as exc:  # pragma: no cover - defensive guardrail
            tracker.mark_phase("send_email", "warning")
            tracker.add_summary(
                f"Notification dispatch failed; see logs for details ({exc})."
            )
            tracker.overall = "warning"
            log_event(
                logger=logger,
                phase="send_email",
                status="warning",
                message="notification dispatch failed",
                report_date=resolved_date.isoformat(),
                error=str(exc),
            )

        final_finished_at = datetime.now(timezone.utc)
        final_record = tracker.build_record(final_finished_at)
        await update_summary_record(database_url, run_id, final_record)
        log_event(
            logger=logger,
            phase="orchestrator",
            message="daily sales report pipeline complete",
            report_date=resolved_date.isoformat(),
            status=tracker.overall or "ok",
        )
    finally:
        logger.close()


def run_pipeline(
    report_date: date | None = None,
    env: str | None = None,
    force: bool = False,
    orders_sync_upstream_status: str | None = None,
    orders_sync_upstream_run_id: str | None = None,
) -> None:
    asyncio.run(
        _run(
            report_date,
            env,
            force,
            orders_sync_upstream_status=orders_sync_upstream_status,
            orders_sync_upstream_run_id=orders_sync_upstream_run_id,
        )
    )


__all__ = ["run_pipeline"]
