from __future__ import annotations

from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl.workbook import Workbook

from .data import PendingDeliveriesReportData, PendingDeliveryRow

DATETIME_FORMAT = "yyyy-mm-dd hh:mm:ss"


def workbook_output_path(
    *, output_root: Path, pipeline_name: str, report_date: date
) -> Path:
    return output_root / f"{pipeline_name}_{report_date.isoformat()}.xlsx"


def _sanitize_worksheet_name(value: str, used_names: set[str]) -> str:
    trimmed = (value or "Unknown").strip() or "Unknown"
    sanitized = "".join(
        "_" if ch in {"[", "]", ":", "*", "?", "/", "\\"} else ch for ch in trimmed
    )
    candidate = sanitized[:31] or "Sheet"
    if candidate not in used_names:
        used_names.add(candidate)
        return candidate
    index = 2
    while True:
        suffix = f"_{index}"
        base = candidate[: max(1, 31 - len(suffix))]
        next_candidate = f"{base}{suffix}"
        if next_candidate not in used_names:
            used_names.add(next_candidate)
            return next_candidate
        index += 1


def _normalize_excel_datetime(value: object, *, business_timezone) -> datetime | object:
    if not isinstance(value, datetime):
        return value
    normalized = value
    if normalized.tzinfo is None:
        normalized = normalized.replace(tzinfo=business_timezone)
    else:
        normalized = normalized.astimezone(business_timezone)
    return normalized.replace(tzinfo=None)


def _sorted_rows(rows: Iterable[PendingDeliveryRow]) -> list[PendingDeliveryRow]:
    return sorted(
        rows,
        key=lambda row: (
            row.age_days,
            row.order_date,
            row.order_number,
        ),
        reverse=True,
    )


def build_pending_deliveries_workbook(
    *,
    data: PendingDeliveriesReportData,
    output_path: Path,
    business_timezone,
) -> tuple[int, int]:
    workbook: Workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)

    headers = [
        "Cost Center",
        "Order Number",
        "Customer Name",
        "Order Date",
        "Due Date",
        "Age (Days)",
        "Order Amount",
    ]

    grouped: dict[str, list[PendingDeliveryRow]] = defaultdict(list)
    for cost_center_section in data.cost_center_sections:
        for bucket in cost_center_section.buckets:
            grouped[cost_center_section.cost_center].extend(bucket.rows)

    used_sheet_names: set[str] = set()
    sheet_count = 0
    row_count = 0
    for cost_center in sorted(grouped):
        rows = _sorted_rows(grouped[cost_center])
        if not rows:
            continue
        worksheet = workbook.create_sheet(
            title=_sanitize_worksheet_name(cost_center, used_sheet_names)
        )
        worksheet.append(headers)
        sheet_count += 1
        for row in rows:
            values = [
                row.cost_center,
                row.order_number,
                row.customer_name,
                (
                    row.order_date
                    if isinstance(row.order_date, datetime)
                    else datetime.combine(row.order_date, datetime.min.time())
                ),
                (
                    row.default_due_date
                    if isinstance(row.default_due_date, datetime)
                    else datetime.combine(row.default_due_date, datetime.min.time())
                ),
                row.age_days,
                Decimal(row.order_amount),
            ]
            normalized = [
                _normalize_excel_datetime(value, business_timezone=business_timezone)
                for value in values
            ]
            worksheet.append(normalized)
            row_count += 1
            current_row = worksheet.max_row
            for datetime_col in (4, 5):
                worksheet.cell(row=current_row, column=datetime_col).number_format = (
                    DATETIME_FORMAT
                )

    if not workbook.sheetnames:
        ws = workbook.create_sheet(title="No Data")
        ws.append(headers)
        sheet_count = 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return row_count, sheet_count
