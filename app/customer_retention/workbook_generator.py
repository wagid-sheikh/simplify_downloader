"""Generate operator workbooks for customer retention follow-up."""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl.styles import Font, PatternFill, Protection
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

from app.dashboard_downloader.json_logger import JsonLogger, log_event

from .constants import WORKBOOK_OUTCOME_LABELS
from .types import RowWarning
from .workbook_ingestor import EDITABLE_COLUMNS, FOLLOWUP_SHEET, PROTECTED_COLUMNS, READ_ME_SHEET
from .workbook_selection import StoreWorkbookSelectionResult, WorkbookLeadRow

CONTACT_ATTEMPTED_VALUES = ("Yes", "No")
CONTACT_MODE_VALUES = ("Call", "WhatsApp", "Both", "Not Contacted")
ORDER_EXPECTED_VALUES = ("Yes", "No", "Maybe", "Not Applicable")
YES_NO_VALUES = ("Yes", "No")

FOLLOWUP_HEADERS = [
    "lead_id",
    "lead_source_type",
    "work_section",
    "cost_center",
    "customer_name",
    "mobile_number",
    "normalized_mobile_number",
    "lifecycle_bucket",
    "last_order_date",
    "days_since_last_order",
    "total_orders",
    "lifetime_spend",
    "average_order_value",
    "last_order_amount",
    "priority_score",
    "recommended_strategy",
    "generated_at",
    "Handled By",
    "Contact Attempted",
    "Contact Mode",
    "Customer Response",
    "Order Expected",
    "Next Follow-up Date",
    "Complaint",
    "Do Not Contact",
    "Staff Remarks",
    "Target Cost Center",
]
EDITABLE_HEADER_SET = set(EDITABLE_COLUMNS.values())
EXCEL_INLINE_LIST_FORMULA_LIMIT = 255
ACTIVE_COST_CENTER_LOOKUP_SHEET = "_ACTIVE_COST_CENTER_LOOKUP"
ACTIVE_COST_CENTER_RANGE_NAME = "ActiveCostCenters"

DROPDOWN_VALUES_BY_HEADER = {
    "Contact Attempted": CONTACT_ATTEMPTED_VALUES,
    "Contact Mode": CONTACT_MODE_VALUES,
    "Customer Response": WORKBOOK_OUTCOME_LABELS,
    "Order Expected": ORDER_EXPECTED_VALUES,
    "Complaint": YES_NO_VALUES,
    "Do Not Contact": YES_NO_VALUES,
}


@dataclass(frozen=True)
class StoreWorkbookOutputResult:
    cost_center: str
    output_path: Path
    row_count: int
    warning_count: int = 0
    warnings: tuple[RowWarning, ...] = field(default_factory=tuple)


@dataclass(frozen=True)
class WorkbookGenerationResult:
    run_date: date
    outputs: tuple[StoreWorkbookOutputResult, ...]
    warnings: tuple[RowWarning, ...] = field(default_factory=tuple)


def customer_followup_output_path(*, output_root: Path, cost_center: str, run_date: date) -> Path:
    month_dir = output_root / f"{run_date:%Y-%m}"
    return month_dir / f"customer_followup_{cost_center}_{run_date.isoformat()}.xlsx"


def default_customer_followup_output_root() -> Path:
    from app.config import config

    return Path(config.customer_followup_output_dir).expanduser()


def generate_store_workbook(
    *,
    selection: StoreWorkbookSelectionResult,
    active_cost_centers: Iterable[str],
    output_root: Path | None = None,
    generated_at: datetime | None = None,
    logger: JsonLogger | None = None,
) -> StoreWorkbookOutputResult:
    output_path = customer_followup_output_path(
        output_root=output_root or default_customer_followup_output_root(),
        cost_center=selection.cost_center,
        run_date=selection.run_date,
    )
    workbook = openpyxl.Workbook()
    readme = workbook.active
    readme.title = READ_ME_SHEET
    followup = workbook.create_sheet(FOLLOWUP_SHEET)
    _populate_readme(readme)
    _populate_followup_sheet(followup, selection.rows, sorted(set(active_cost_centers)), generated_at or datetime.now(timezone.utc))
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    if logger:
        log_event(
            logger=logger,
            phase="workbook",
            message="customer_followup_workbook_generated",
            cost_center=selection.cost_center,
            file_path=str(output_path),
            row_count=len(selection.rows),
            warnings=len(selection.warnings),
        )
    return StoreWorkbookOutputResult(selection.cost_center, output_path, len(selection.rows), len(selection.warnings), selection.warnings)


def generate_workbooks(
    *,
    selections: Iterable[StoreWorkbookSelectionResult],
    active_cost_centers: Iterable[str],
    output_root: Path | None = None,
    generated_at: datetime | None = None,
    logger: JsonLogger | None = None,
) -> WorkbookGenerationResult:
    selection_list = list(selections)
    outputs = [generate_store_workbook(selection=selection, active_cost_centers=active_cost_centers, output_root=output_root, generated_at=generated_at, logger=logger) for selection in selection_list]
    warnings = tuple(warning for output in outputs for warning in output.warnings)
    run_date = selection_list[0].run_date if selection_list else date.today()
    return WorkbookGenerationResult(run_date=run_date, outputs=tuple(outputs), warnings=warnings)


def _populate_readme(sheet) -> None:
    rows = [
        ("Customer Follow-Up Workbook Instructions",),
        ("Only edit allowed columns: Handled By, Contact Attempted, Contact Mode, Customer Response, Order Expected, Next Follow-up Date, Complaint, Do Not Contact, Staff Remarks, Target Cost Center.",),
        ("Do not change system columns; protected columns are ignored during ingestion because DB values are authoritative.",),
        ("Select dropdown values only for controlled fields; invalid entries become warnings and may keep rows pending.",),
        ("Target Cost Center is only for Shifted Location and must be selected from the active-store dropdown. Excel shows the active-store dropdown on every row; ingestion ignores populated targets on other responses with a warning.",),
        ("Pickup Requested means urgent action. Use Lead Stale only when no further action is useful.",),
    ]
    for row in rows:
        sheet.append(row)
    sheet.column_dimensions["A"].width = 120
    sheet.protection.sheet = True


def _populate_followup_sheet(sheet, rows: tuple[WorkbookLeadRow, ...], active_cost_centers: list[str], generated_at: datetime) -> None:
    sheet.append(FOLLOWUP_HEADERS)
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    editable_fill = PatternFill("solid", fgColor="FFF2CC")
    for col_idx, header in enumerate(FOLLOWUP_HEADERS, start=1):
        cell = sheet.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)
        cell.fill = editable_fill if header in EDITABLE_HEADER_SET else header_fill
        width = max(14, min(42, len(header) + 3))
        sheet.column_dimensions[cell.column_letter].width = width
    for lead in rows:
        sheet.append(_row_values(lead, generated_at))
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    max_row = max(sheet.max_row, 2)
    editable_columns = {idx for idx, header in enumerate(FOLLOWUP_HEADERS, start=1) if header in EDITABLE_HEADER_SET}
    for row in sheet.iter_rows(min_row=1, max_row=max_row, max_col=len(FOLLOWUP_HEADERS)):
        for cell in row:
            unlocked = cell.column in editable_columns and cell.row > 1
            cell.protection = Protection(locked=not unlocked)
            if unlocked:
                cell.fill = editable_fill
    _add_dropdown_validations(sheet, active_cost_centers, max_row=max_row)
    _add_date_validation(sheet, max_row=max_row)
    sheet.protection.sheet = True
    sheet.protection.enable()


def _row_values(row: WorkbookLeadRow, generated_at: datetime) -> list[object | None]:
    return [
        row.lead_id,
        row.lead_source_type,
        row.work_section,
        row.cost_center,
        row.customer_name,
        row.mobile_number,
        row.normalized_mobile_number,
        row.lifecycle_bucket,
        row.last_order_date.isoformat() if row.last_order_date else None,
        row.days_since_last_order,
        row.total_orders,
        _decimal(row.lifetime_spend),
        _decimal(row.average_order_value),
        _decimal(row.last_order_amount),
        _decimal(row.priority_score),
        row.recommended_strategy,
        generated_at.replace(tzinfo=None),
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
        None,
    ]


def _decimal(value: Decimal | None) -> str | None:
    return str(value.quantize(Decimal("0.01"))) if value is not None else None


def _quote_list(values: Iterable[str]) -> str:
    escaped = [str(value).replace('"', '""') for value in values]
    return '"' + ",".join(escaped) + '"'


def _column_letter(header: str) -> str:
    idx = FOLLOWUP_HEADERS.index(header) + 1
    return openpyxl.utils.get_column_letter(idx)


def _add_dropdown_validations(sheet, active_cost_centers: list[str], *, max_row: int) -> None:
    for header, values in DROPDOWN_VALUES_BY_HEADER.items():
        validation = DataValidation(type="list", formula1=_quote_list(values), allow_blank=False, showErrorMessage=True)
        validation.error = "Select a value from the dropdown list."
        validation.errorTitle = "Invalid dropdown value"
        column = _column_letter(header)
        validation.add(f"{column}2:{column}{max_row}")
        sheet.add_data_validation(validation)
    target_formula = _target_cost_center_validation_formula(sheet, active_cost_centers)
    target_validation = DataValidation(type="list", formula1=target_formula, allow_blank=True, showErrorMessage=True)
    target_validation.error = "Select an active target cost center from the dropdown list."
    target_validation.errorTitle = "Invalid target cost center"
    target_column = _column_letter("Target Cost Center")
    target_validation.add(f"{target_column}2:{target_column}{max_row}")
    sheet.add_data_validation(target_validation)


def _target_cost_center_validation_formula(sheet, active_cost_centers: list[str]) -> str:
    inline_formula = _quote_list(active_cost_centers)
    if len(inline_formula) < EXCEL_INLINE_LIST_FORMULA_LIMIT:
        return inline_formula
    return f"={_ensure_active_cost_center_lookup_range(sheet, active_cost_centers)}"


def _ensure_active_cost_center_lookup_range(sheet, active_cost_centers: list[str]) -> str:
    workbook = sheet.parent
    if ACTIVE_COST_CENTER_LOOKUP_SHEET in workbook.sheetnames:
        lookup = workbook[ACTIVE_COST_CENTER_LOOKUP_SHEET]
        lookup.delete_rows(1, lookup.max_row)
    else:
        lookup = workbook.create_sheet(ACTIVE_COST_CENTER_LOOKUP_SHEET)

    lookup.sheet_state = "veryHidden"
    lookup.protection.sheet = True
    lookup.protection.enable()
    lookup.cell(row=1, column=1, value="active_cost_center")
    for row_idx, cost_center in enumerate(active_cost_centers, start=2):
        lookup.cell(row=row_idx, column=1, value=cost_center)

    # Excel list validations cannot safely carry long inline CSV formulas. A
    # workbook-scoped name keeps the user-facing tab list clean while allowing
    # arbitrarily many active stores.
    last_row = max(len(active_cost_centers) + 1, 2)
    range_reference = f"\'{ACTIVE_COST_CENTER_LOOKUP_SHEET}\'!$A$2:$A${last_row}"
    if ACTIVE_COST_CENTER_RANGE_NAME in workbook.defined_names:
        del workbook.defined_names[ACTIVE_COST_CENTER_RANGE_NAME]
    workbook.defined_names.add(DefinedName(ACTIVE_COST_CENTER_RANGE_NAME, attr_text=range_reference))
    return ACTIVE_COST_CENTER_RANGE_NAME


def _add_date_validation(sheet, *, max_row: int) -> None:
    column = _column_letter("Next Follow-up Date")
    validation = DataValidation(type="date", operator="greaterThanOrEqual", formula1="DATE(2000,1,1)", allow_blank=True, showErrorMessage=True)
    validation.error = "Enter a valid follow-up date."
    validation.errorTitle = "Invalid date"
    validation.add(f"{column}2:{column}{max_row}")
    sheet.add_data_validation(validation)
