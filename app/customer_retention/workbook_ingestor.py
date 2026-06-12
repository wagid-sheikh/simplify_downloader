"""Returned workbook ingestion shell for Customer Retention Phase 2."""

from __future__ import annotations

import hashlib
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl
import sqlalchemy as sa

from app.common.db import session_scope
from app.dashboard_downloader.json_logger import JsonLogger, log_event

from .constants import WORKBOOK_OUTCOME_LABELS
from .db_tables import trx_customer_followup_leads
from .lifecycle import apply_lifecycle_transition
from .mobile import normalize_mobile
from .normalization import ValueNormalizer
from .persistence import insert_history_once
from .types import RowWarning, WorkbookIngestionResult

FOLLOWUP_SHEET = "FOLLOWUP_LEADS"
READ_ME_SHEET = "READ_ME"
PROTECTED_COLUMNS = {
    "lead_id",
    "lead_source_type",
    "work_section",
    "cost_center",
    "customer_name",
    "mobile_number",
    "lifecycle_bucket",
    "last_order_date",
    "days_since_last_order",
    "total_orders",
    "lifetime_spend",
    "average_order_value",
    "last_order_amount",
    "priority_score",
    "recommended_strategy",
    "generated_at",
}
EDITABLE_COLUMNS = {
    "handled_by": "Handled By",
    "contact_attempted": "Contact Attempted",
    "contact_mode": "Contact Mode",
    "customer_response": "Customer Response",
    "order_expected": "Order Expected",
    "next_followup_date": "Next Follow-up Date",
    "complaint": "Complaint",
    "do_not_contact": "Do Not Contact",
    "staff_remarks": "Staff Remarks",
    "target_cost_center": "Target Cost Center",
}
REQUIRED_EDITABLE = ("contact_attempted", "customer_response", "complaint", "do_not_contact")


def _key(value: Any) -> str:
    return "_".join(str(value or "").strip().lower().replace("-", " ").split())


def _row_dict(headers: list[str], values: tuple[Any, ...]) -> dict[str, Any]:
    return {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}


def _changed(a: Any, b: Any) -> bool:
    return str(a or "").strip() != str(b or "").strip()


def _normalized_bool(value: Any) -> bool | None:
    if value is None:
        return None
    text = str(value).strip().lower()
    if text == "yes":
        return True
    if text == "no":
        return False
    return None


def _normalized_date(value: Any) -> date | None:
    if value is None or not str(value).strip():
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _normalized_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _suppression_start_date_from_row(row: dict[str, Any]) -> date:
    generated_at = _normalized_date(row.get("generated_at"))
    if generated_at is not None:
        return generated_at
    # Boundary fallback for legacy/manual workbooks that do not carry the
    # generated_at business date. Suppression helpers deliberately do not use
    # process dates internally so deterministic callers can pass an explicit date.
    return date.today()


async def ingest_returned_workbook(*, database_url: str, path: Path, pipeline_run_id: str, logger: JsonLogger | None = None, normalizer: ValueNormalizer | None = None) -> WorkbookIngestionResult:
    file_digest = hashlib.sha256(path.read_bytes()).hexdigest()
    result = WorkbookIngestionResult(source_file=path.name, file_identity=file_digest)
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if READ_ME_SHEET not in workbook.sheetnames or FOLLOWUP_SHEET not in workbook.sheetnames:
        result.warnings.append(RowWarning("workbook_sheet_missing", "Workbook must include READ_ME and FOLLOWUP_LEADS sheets", source_file=path.name))
        return result
    sheet = workbook[FOLLOWUP_SHEET]
    rows_iter = sheet.iter_rows(values_only=True)
    try:
        headers = [_key(cell) for cell in next(rows_iter)]
    except StopIteration:
        return result
    value_normalizer = normalizer or ValueNormalizer()
    async with session_scope(database_url) as session:
        for row_number, values in enumerate(rows_iter, start=2):
            row = _row_dict(headers, values)
            if "next_follow_up_date" in row and "next_followup_date" not in row:
                row["next_followup_date"] = row["next_follow_up_date"]
            if not any(value is not None and str(value).strip() for value in row.values()):
                continue
            result.rows_seen += 1
            lead = None
            lead_id_raw = row.get("lead_id")
            if lead_id_raw not in (None, ""):
                try:
                    lead_id = int(lead_id_raw)
                except (TypeError, ValueError):
                    lead_id = None
                if lead_id is not None:
                    lead = (await session.execute(sa.select(trx_customer_followup_leads).where(trx_customer_followup_leads.c.lead_id == lead_id))).mappings().first()
            mobile_result = normalize_mobile(row.get("mobile_number"))
            if lead is None and mobile_result.is_valid:
                cost_center = str(row.get("cost_center") or "").strip().upper()
                lead = (await session.execute(sa.select(trx_customer_followup_leads).where(trx_customer_followup_leads.c.cost_center == cost_center, trx_customer_followup_leads.c.normalized_mobile_number == mobile_result.normalized_mobile))).mappings().first()
            if lead is None:
                result.rows_skipped += 1
                result.warnings.append(RowWarning("lead_not_found", "Workbook row could not be matched to an existing lead", row_number, path.name))
                continue
            lead_map = dict(lead)
            if not mobile_result.is_valid:
                result.rows_skipped += 1
                result.warnings.append(RowWarning(mobile_result.warning_code or "invalid_mobile", mobile_result.warning_message or "Invalid mobile number", row_number, path.name, "mobile_number", lead_id=int(lead_map["lead_id"]), cost_center=lead_map.get("cost_center")))
                continue
            if mobile_result.normalized_mobile != lead_map.get("normalized_mobile_number"):
                result.rows_skipped += 1
                result.warnings.append(RowWarning("mobile_identity_conflict", "Workbook mobile number conflicts with protected lead identity", row_number, path.name, "mobile_number", lead_id=int(lead_map["lead_id"]), cost_center=lead_map.get("cost_center")))
                continue
            for column in PROTECTED_COLUMNS:
                if column in row and column in lead_map and _changed(row.get(column), lead_map.get(column)):
                    result.protected_edits_ignored += 1
                    result.warnings.append(RowWarning("protected_column_edit", "Protected workbook column edit ignored", row_number, path.name, column, lead_id=int(lead_map["lead_id"]), cost_center=lead_map.get("cost_center")))
            normalized_values: dict[str, Any] = {}
            row_has_required_blank = False
            for field in REQUIRED_EDITABLE:
                normalized = value_normalizer.normalize(row.get(field), field_name=EDITABLE_COLUMNS[field], required=True)
                normalized_values[field] = normalized.normalized_value
                if normalized.invalid:
                    row_has_required_blank = True
                    result.warnings.append(RowWarning(normalized.warning_code or "invalid_value", normalized.warning_message or "Required editable field is invalid", row_number, path.name, field, lead_id=int(lead_map["lead_id"]), cost_center=lead_map.get("cost_center")))
            response = value_normalizer.normalize(row.get("customer_response"), allowed_values=WORKBOOK_OUTCOME_LABELS, field_name="Customer Response", required=True)
            normalized_values["customer_response"] = response.normalized_value
            if response.invalid and response.warning_code != "required_blank":
                row_has_required_blank = True
                result.warnings.append(RowWarning(response.warning_code or "invalid_response", response.warning_message or "Customer response is invalid", row_number, path.name, "customer_response", lead_id=int(lead_map["lead_id"]), cost_center=lead_map.get("cost_center")))
            for field in ("contact_mode", "order_expected"):
                normalized = value_normalizer.normalize(row.get(field), field_name=EDITABLE_COLUMNS[field])
                normalized_values[field] = normalized.normalized_value
            event_suffix = f"{file_digest[:16]}:{row_number}"
            lead_id = int(lead_map["lead_id"])
            if row_has_required_blank:
                result.rows_pending_not_updated += 1
                inserted = await insert_history_once(
                    session,
                    lead_id=lead_id,
                    pipeline_run_id=pipeline_run_id,
                    event_type=f"Pending_Not_Updated:{event_suffix}",
                    previous_status=str(lead_map.get("lead_status")) if lead_map.get("lead_status") is not None else None,
                    new_status=str(lead_map.get("lead_status")) if lead_map.get("lead_status") is not None else None,
                    raw_excel_value_json={key: (str(value) if value is not None else None) for key, value in row.items() if key in EDITABLE_COLUMNS or key in PROTECTED_COLUMNS},
                    normalized_value_json=normalized_values,
                )
            else:
                # Successful rows must pass through the lifecycle engine so status,
                # closure, suppression, and history contracts stay centralized.
                transition = await apply_lifecycle_transition(
                    session,
                    lead_id=lead_id,
                    customer_response=normalized_values.get("customer_response"),
                    contact_attempted=_normalized_bool(normalized_values.get("contact_attempted")),
                    contact_mode=_normalized_text(normalized_values.get("contact_mode")),
                    order_expected=_normalized_text(normalized_values.get("order_expected")),
                    next_followup_date=_normalized_date(row.get("next_followup_date")),
                    complaint_flag=_normalized_bool(normalized_values.get("complaint")),
                    do_not_contact_flag=_normalized_bool(normalized_values.get("do_not_contact")),
                    staff_remarks=_normalized_text(row.get("staff_remarks")),
                    handled_by=_normalized_text(row.get("handled_by")),
                    pipeline_run_id=pipeline_run_id,
                    event_key=event_suffix,
                    suppression_start_date=_suppression_start_date_from_row(row),
                )
                inserted = transition.history_inserted
            if inserted:
                result.history_inserted += 1
            else:
                result.history_existing += 1
        await session.commit()
    if logger:
        log_event(logger=logger, phase="history_update", message="workbook_ingestion_complete", source_file=path.name, rows_seen=result.rows_seen, history_inserted=result.history_inserted, warnings=result.warning_count)
    return result
