from __future__ import annotations

import json
import logging
import math
import os
import tempfile
import zipfile

import openpyxl
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import Any, Mapping, Sequence

logger = logging.getLogger(__name__)

_SENSITIVE_FIELD_NAMES = {"token", "authorization", "cookie", "set-cookie"}
_REDACTED = "***REDACTED***"
_FORBIDDEN_API_ARTIFACT_ENVELOPE_KEYS = {
    "mismatch_artifacts",
    "api_request_metadata",
    "compare_metrics",
    "summary",
    "endpoint_health",
}


def _validate_artifact_rows_payload(rows: Any, *, artifact_key: str) -> Sequence[Mapping[str, Any]]:
    """Validate API artifact row payload purity before Excel serialization.

    API snapshot Excel artifacts must receive only row lists (`Sequence[Mapping[str, Any]]`) and
    must not receive compare/diagnostic wrapper structures.
    """

    if isinstance(rows, (str, bytes, bytearray)) or not isinstance(rows, Sequence):
        raise TypeError(
            f"API artifact purity violation for '{artifact_key}': expected Sequence[Mapping[str, Any]], "
            f"got {type(rows).__name__}."
        )
    if isinstance(rows, Mapping):
        overlapping_keys = sorted(_FORBIDDEN_API_ARTIFACT_ENVELOPE_KEYS.intersection({str(key) for key in rows.keys()}))
        if overlapping_keys:
            raise ValueError(
                f"API artifact purity violation for '{artifact_key}': wrapper/envelope payload keys are not allowed "
                f"({', '.join(overlapping_keys)}). Pass raw row dicts only."
            )
        raise TypeError(
            f"API artifact purity violation for '{artifact_key}': expected Sequence[Mapping[str, Any]], got Mapping."
        )

    validated_rows: list[Mapping[str, Any]] = []
    for index, row in enumerate(rows):
        if not isinstance(row, Mapping):
            raise TypeError(
                f"API artifact purity violation for '{artifact_key}': row at index {index} must be Mapping[str, Any], "
                f"got {type(row).__name__}."
            )
        forbidden_keys = sorted(_FORBIDDEN_API_ARTIFACT_ENVELOPE_KEYS.intersection({str(key) for key in row.keys()}))
        if forbidden_keys:
            raise ValueError(
                f"API artifact purity violation for '{artifact_key}': row at index {index} contains envelope keys "
                f"({', '.join(forbidden_keys)}). Pass raw API rows only."
            )
        validated_rows.append(row)
    return validated_rows


def redact_sensitive_fields(payload: Any) -> Any:
    def _is_sensitive_field(key: Any) -> bool:
        return str(key).strip().lower() in _SENSITIVE_FIELD_NAMES

    if isinstance(payload, Mapping):
        redacted: dict[str, Any] = {}
        for key, value in payload.items():
            key_text = str(key)
            redacted[key_text] = _REDACTED if _is_sensitive_field(key) else redact_sensitive_fields(value)
        return redacted
    if isinstance(payload, list):
        return [redact_sensitive_fields(item) for item in payload]
    if isinstance(payload, tuple):
        return tuple(redact_sensitive_fields(item) for item in payload)
    if isinstance(payload, set):
        return {redact_sensitive_fields(item) for item in payload}
    return payload


@dataclass
class TdApiArtifactPersistResult:
    artifact_paths: dict[str, str] = field(default_factory=dict)
    human_readable_export_enabled: bool = False
    human_readable_artifact_paths: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)



def _window_token(value: date) -> str:
    return value.strftime("%Y%m%d")



def _raw_filename(store_code: str, dataset: str, from_date: date, to_date: date) -> str:
    return f"{store_code}_td_api_{dataset}_{_window_token(from_date)}_{_window_token(to_date)}.json"



def _write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(redact_sensitive_fields(payload), ensure_ascii=False, indent=2), encoding="utf-8")



def _write_jsonl(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(redact_sensitive_fields(dict(row)), ensure_ascii=False, sort_keys=True))
            handle.write("\n")


def _validate_xlsx(path: Path) -> None:
    required_entries = {"[Content_Types].xml", "xl/workbook.xml"}
    try:
        with zipfile.ZipFile(path, "r") as zip_handle:
            first_corrupt_entry = zip_handle.testzip()
            if first_corrupt_entry is not None:
                raise ValueError(f"zip integrity check failed at entry '{first_corrupt_entry}'")
            available_entries = set(zip_handle.namelist())
    except zipfile.BadZipFile as exc:
        raise ValueError(f"not a valid ZIP-based OOXML workbook: {exc}") from exc

    missing_entries = sorted(required_entries - available_entries)
    if missing_entries:
        raise ValueError(
            "missing required OOXML entries: " + ", ".join(missing_entries)
        )

    # Excel is stricter than a simple ZIP/XML integrity check. Re-opening with openpyxl
    # gives us an additional sanity gate before we publish artifacts to users.
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"failed openpyxl round-trip validation: {exc}") from exc

    try:
        if not workbook.sheetnames:
            raise ValueError("workbook has no worksheets")
        # Touch at least one row in each sheet to force XML parsing of worksheet payloads.
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            _ = next(sheet.iter_rows(min_row=1, max_row=1), None)
    finally:
        workbook.close()


def _write_excel(path: Path, rows: Sequence[Mapping[str, Any]], *, sheet_name: str = "rows") -> None:
    """Write a single-sheet API snapshot workbook from raw API row dictionaries only.

    Contract:
    - input must already be a pure `Sequence[Mapping[str, Any]]`
    - row/column structure is preserved from source row dicts (no additional computed columns)
    - only serialization-safe coercions are applied (nested JSON stringification, XML char
      sanitization, and scalar coercions required for Excel compatibility)
    """

    rows = _validate_artifact_rows_payload(rows, artifact_key=sheet_name)
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    max_coercion_debug_entries = 20
    coerced_type_counts: dict[str, int] = {}
    coerced_type_order: list[str] = []
    total_coercions = 0

    def _record_coercion(value: Any) -> None:
        nonlocal total_coercions
        total_coercions += 1
        if len(coerced_type_order) >= max_coercion_debug_entries:
            return
        type_name = type(value).__name__
        if type_name not in coerced_type_counts:
            coerced_type_order.append(type_name)
            coerced_type_counts[type_name] = 0
        coerced_type_counts[type_name] += 1

    def _sanitize_excel_string(value: str) -> str:
        def _is_valid_xml_char(char: str) -> bool:
            codepoint = ord(char)
            if codepoint in {0x9, 0xA, 0xD}:
                return True
            if codepoint < 0x20:
                return False
            if 0x7F <= codepoint <= 0x9F:
                return False
            if 0xD800 <= codepoint <= 0xDFFF:
                return False
            if 0xFDD0 <= codepoint <= 0xFDEF:
                return False
            if codepoint & 0xFFFF in {0xFFFE, 0xFFFF}:
                return False
            return codepoint <= 0x10FFFF

        return "".join(char for char in value if _is_valid_xml_char(char))

    def _json_compatible(value: Any) -> Any:
        if isinstance(value, Mapping):
            return {str(key): _json_compatible(nested_value) for key, nested_value in value.items()}
        if isinstance(value, (list, tuple)):
            return [_json_compatible(item) for item in value]
        if isinstance(value, set):
            normalized_items = [_json_compatible(item) for item in value]
            return sorted(normalized_items, key=lambda item: json.dumps(item, ensure_ascii=False, sort_keys=True))
        if isinstance(value, str):
            return _sanitize_excel_string(value)
        return _coerce_scalar_for_excel(value)

    def _coerce_scalar_for_excel(value: Any) -> Any:
        if isinstance(value, datetime) and value.tzinfo is not None and value.utcoffset() is not None:
            _record_coercion(value)
            return value.isoformat()
        if isinstance(value, float) and (math.isnan(value) or math.isinf(value)):
            _record_coercion(value)
            if math.isnan(value):
                return "NaN"
            if value > 0:
                return "Infinity"
            return "-Infinity"
        if isinstance(value, Decimal):
            _record_coercion(value)
            return str(value)
        if isinstance(value, bytes):
            _record_coercion(value)
            return value.decode("utf-8", errors="replace")
        if value is None or isinstance(value, (bool, int, float, date, str)):
            return value
        _record_coercion(value)
        return _sanitize_excel_string(str(value))

    def _serialize_excel_cell_value(value: Any) -> Any:
        if isinstance(value, (Mapping, list, tuple, set)):
            value = json.dumps(_json_compatible(value), ensure_ascii=False, sort_keys=True)
        value = _coerce_scalar_for_excel(value)
        if isinstance(value, str):
            return _sanitize_excel_string(value)
        return value

    def _append_serialized_row(values: Sequence[Any]) -> None:
        worksheet.append([_serialize_excel_cell_value(value) for value in values])

    if not rows:
        _append_serialized_row(["status"])
        _append_serialized_row(["no rows"])
    else:
        columns: list[str] = []
        for row in rows:
            for key in row.keys():
                key_text = str(key)
                if key_text not in columns:
                    columns.append(key_text)
        _append_serialized_row(columns)
        for row in rows:
            _append_serialized_row([row.get(column) for column in columns])
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(
        mode="wb",
        suffix=f".{path.name}.tmp.xlsx",
        dir=path.parent,
        delete=False,
    ) as temp_handle:
        temp_path = Path(temp_handle.name)
    try:
        workbook.save(temp_path)
        if temp_path.stat().st_size == 0:
            raise ValueError(f"Temporary Excel artifact is empty: {temp_path}")
        _validate_xlsx(temp_path)
        temp_path.replace(path)
    except Exception:
        if temp_path.exists():
            temp_path.unlink()
        raise
    if total_coercions:
        logger.debug(
            "Excel cell serialization coerced %s values; first %s type counts=%s",
            total_coercions,
            max_coercion_debug_entries,
            {type_name: coerced_type_counts[type_name] for type_name in coerced_type_order},
        )


def _excel_filename(store_code: str, dataset: str, from_date: date, to_date: date) -> str:
    return f"{store_code}_td_api_{dataset}_{_window_token(from_date)}_{_window_token(to_date)}.xlsx"


def _human_readable_export_enabled() -> bool:
    return (os.environ.get("TD_API_HUMAN_READABLE_EXPORT") or "true").strip().lower() in {"1", "true", "yes", "on"}



def persist_td_api_artifacts(
    *,
    download_dir: Path,
    store_code: str,
    from_date: date,
    to_date: date,
    raw_orders: Any,
    raw_sales: Any,
    raw_garments: Any,
    order_rows: Sequence[Mapping[str, Any]],
    sale_rows: Sequence[Mapping[str, Any]],
    garments_rows: Sequence[Mapping[str, Any]],
) -> TdApiArtifactPersistResult:
    """Persist TD API artifacts.

    Purity contract for human-readable Excel snapshots:
    - `orders_excel` must be written only from `order_rows`
    - `sales_excel` must be written only from `sale_rows`
    - `garments_excel` must be written only from `garments_rows`
    - compare metrics, mismatch artifacts, request metadata, endpoint diagnostics, and any
      other UI-derived or wrapped payloads must never be passed into these row parameters.
    """

    human_readable_export_enabled = _human_readable_export_enabled()
    result = TdApiArtifactPersistResult(human_readable_export_enabled=human_readable_export_enabled)
    store = (store_code or "").strip().upper() or "UNKNOWN"
    artifact_targets: list[tuple[str, Path, Any, str]] = [
        ("orders_raw", download_dir / _raw_filename(store, "orders", from_date, to_date), raw_orders, "json"),
        ("sales_raw", download_dir / _raw_filename(store, "sales", from_date, to_date), raw_sales, "json"),
        ("garments_raw", download_dir / _raw_filename(store, "garments", from_date, to_date), raw_garments, "json"),
    ]

    for key, path, payload, kind in artifact_targets:
        try:
            if kind == "jsonl":
                _write_jsonl(path, payload)
            elif kind == "xlsx":
                _write_excel(path, payload, sheet_name=key)
            else:
                _write_json(path, payload)
            result.artifact_paths[key] = str(path)
        except Exception as exc:  # pragma: no cover - defensive guard
            if kind == "xlsx" and path.exists():
                path.unlink()
            warning = f"Failed to persist TD API artifact '{key}' at {path}: {exc}"
            result.warnings.append(warning)
            logger.warning(warning)

    if human_readable_export_enabled:
        diagnostic_targets: list[tuple[str, Path, Sequence[Mapping[str, Any]], str]] = [
            ("orders_excel", download_dir / _excel_filename(store, "orders", from_date, to_date), order_rows, "orders"),
            ("sales_excel", download_dir / _excel_filename(store, "sales", from_date, to_date), sale_rows, "sales"),
            (
                "garments_excel",
                download_dir / _excel_filename(store, "garments", from_date, to_date),
                garments_rows,
                "garments",
            ),
        ]
        for key, path, rows, sheet_name in diagnostic_targets:
            try:
                pure_rows = _validate_artifact_rows_payload(rows, artifact_key=key)
                _write_excel(path, pure_rows, sheet_name=sheet_name)
                _validate_xlsx(path)
                result.artifact_paths[key] = str(path)
                result.human_readable_artifact_paths.append(str(path))
            except Exception as exc:  # pragma: no cover - defensive guard
                if path.exists():
                    path.unlink()
                warning = f"Failed to persist TD API artifact '{key}' at {path}: {exc}"
                result.warnings.append(warning)
                logger.warning(warning)

    return result


def persist_td_compare_artifacts(
    *,
    download_dir: Path,
    store_code: str,
    from_date: date,
    to_date: date,
    orders_compare_metrics: Mapping[str, Any],
    sales_compare_metrics: Mapping[str, Any],
    endpoint_health_summary: Mapping[str, Any] | None = None,
) -> TdApiArtifactPersistResult:
    _ = (
        download_dir,
        store_code,
        from_date,
        to_date,
        orders_compare_metrics,
        sales_compare_metrics,
        endpoint_health_summary,
    )
    return TdApiArtifactPersistResult()
