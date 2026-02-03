from __future__ import annotations

from collections import Counter
import re
from dataclasses import dataclass, field
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping, Sequence
from zoneinfo import ZoneInfo

import openpyxl
import sqlalchemy as sa
from dateutil import parser
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.dialects.sqlite import insert as sqlite_insert
from sqlalchemy.ext.asyncio import AsyncConnection, AsyncEngine

from app.common.db import session_scope
from app.common.date_utils import get_timezone
from app.crm_downloader.td_orders_sync.ingest import _orders_table
from app.dashboard_downloader.json_logger import JsonLogger, log_event

STG_TD_SALES_COLUMNS = [
    "order_date",
    "payment_date",
    "order_number",
    "customer_code",
    "customer_name",
    "customer_address",
    "mobile_number",
    "payment_received",
    "adjustments",
    "balance",
    "accepted_by",
    "payment_mode",
    "transaction_id",
    "payment_made_at",
    "order_type",
    "is_duplicate",
    "is_edited_order",
    "ingest_remarks",
]

SALES_COLUMNS = [
    "order_date",
    "payment_date",
    "order_number",
    "customer_code",
    "customer_name",
    "customer_address",
    "mobile_number",
    "payment_received",
    "adjustments",
    "balance",
    "accepted_by",
    "payment_mode",
    "transaction_id",
    "payment_made_at",
    "order_type",
    "is_duplicate",
    "is_edited_order",
    "ingest_remarks",
]

HEADER_MAP: Mapping[str, str] = {
    "Order Date": "order_date",
    "Payment Date": "payment_date",
    "Order Number": "order_number",
    "Customer Code": "customer_code",
    "Customer Name": "customer_name",
    "Customer Address": "customer_address",
    "Customer Mobile No.": "mobile_number",
    "Payment Received": "payment_received",
    "Adjustments": "adjustments",
    "Balance": "balance",
    "Accept By": "accepted_by",
    "Payment Mode": "payment_mode",
    "Online TransactionID": "transaction_id",
    "Payment Made At": "payment_made_at",
    "Type": "order_type",
}

NUMERIC_FIELDS = {"payment_received", "adjustments", "balance"}
DATE_FIELDS = {"order_date", "payment_date"}
REQUIRED_HEADERS = set(HEADER_MAP.keys())


def _stringify_value(value: Any) -> str:
    try:
        if hasattr(value, "isoformat"):
            return value.isoformat()  # type: ignore[call-arg]
    except Exception:
        pass
    return str(value) if value is not None else ""


@dataclass
class TdSalesIngestResult:
    staging_rows: int
    staging_inserted: int
    staging_updated: int
    final_rows: int
    final_inserted: int
    final_updated: int
    warnings: list[str]
    ingest_remarks: list[dict[str, str]] = field(default_factory=list)
    rows_downloaded: int = 0
    dropped_rows: list[dict[str, Any]] = field(default_factory=list)
    warning_rows: list[dict[str, Any]] = field(default_factory=list)
    edited_rows: list[dict[str, Any]] = field(default_factory=list)
    duplicate_rows: list[dict[str, Any]] = field(default_factory=list)
    rows_edited: int = 0
    rows_duplicate: int = 0


def _chunked(values: Sequence[tuple[Any, ...]], chunk_size: int = 500) -> Iterable[list[tuple[Any, ...]]]:
    for index in range(0, len(values), chunk_size):
        yield list(values[index : index + chunk_size])


async def _fetch_existing_keys(
    session: sa.ext.asyncio.AsyncSession,
    table: sa.Table,
    key_columns: Sequence[sa.Column],
    keys: Sequence[tuple[Any, ...]],
) -> set[tuple[Any, ...]]:
    if not keys:
        return set()
    existing: set[tuple[Any, ...]] = set()
    for chunk in _chunked(list(keys)):
        stmt = sa.select(*key_columns).where(sa.tuple_(*key_columns).in_(chunk))
        result = await session.execute(stmt)
        existing.update(tuple(row) for row in result.fetchall())
    return existing


def _stg_td_sales_table(metadata: sa.MetaData) -> sa.Table:
    return sa.Table(
        "stg_td_sales",
        metadata,
        sa.Column(
            "id",
            sa.BigInteger().with_variant(sa.Integer(), "sqlite"),
            primary_key=True,
            autoincrement=True,
        ),
        sa.Column("run_id", sa.Text()),
        sa.Column("run_date", sa.DateTime(timezone=True)),
        sa.Column("cost_center", sa.String(length=8)),
        sa.Column("store_code", sa.String(length=16)),
        sa.Column("order_date", sa.DateTime(timezone=True)),
        sa.Column("payment_date", sa.DateTime(timezone=True)),
        sa.Column("order_number", sa.String(length=16)),
        sa.Column("customer_code", sa.String(length=16)),
        sa.Column("customer_name", sa.String(length=128)),
        sa.Column("customer_address", sa.String(length=256)),
        sa.Column("mobile_number", sa.String(length=16)),
        sa.Column("payment_received", sa.Numeric(12, 2)),
        sa.Column("adjustments", sa.Numeric(12, 2)),
        sa.Column("balance", sa.Numeric(12, 2)),
        sa.Column("accepted_by", sa.String(length=64)),
        sa.Column("payment_mode", sa.String(length=32)),
        sa.Column("transaction_id", sa.String(length=64)),
        sa.Column("payment_made_at", sa.String(length=128)),
        sa.Column("order_type", sa.String(length=32)),
        sa.Column("is_duplicate", sa.Boolean()),
        sa.Column("is_edited_order", sa.Boolean()),
        sa.Column("ingest_remarks", sa.Text()),
        sa.UniqueConstraint(
            "store_code",
            "order_number",
            "payment_date",
            name="uq_stg_td_sales_store_order_payment_date",
        ),
        sqlite_autoincrement=True,
    )


def _sales_table(metadata: sa.MetaData) -> sa.Table:
    return sa.Table(
        "sales",
        metadata,
        sa.Column(
            "id",
            sa.BigInteger().with_variant(sa.Integer(), "sqlite"),
            primary_key=True,
            autoincrement=True,
        ),
        sa.Column("run_id", sa.Text()),
        sa.Column("run_date", sa.DateTime(timezone=True)),
        sa.Column("cost_center", sa.String(length=8)),
        sa.Column("store_code", sa.String(length=16)),
        sa.Column("order_date", sa.DateTime(timezone=True)),
        sa.Column("payment_date", sa.DateTime(timezone=True)),
        sa.Column("order_number", sa.String(length=16)),
        sa.Column("customer_code", sa.String(length=16)),
        sa.Column("customer_name", sa.String(length=128)),
        sa.Column("customer_address", sa.String(length=256)),
        sa.Column("mobile_number", sa.String(length=16)),
        sa.Column("payment_received", sa.Numeric(12, 2)),
        sa.Column("adjustments", sa.Numeric(12, 2)),
        sa.Column("balance", sa.Numeric(12, 2)),
        sa.Column("accepted_by", sa.String(length=64)),
        sa.Column("payment_mode", sa.String(length=32)),
        sa.Column("transaction_id", sa.String(length=64)),
        sa.Column("payment_made_at", sa.String(length=128)),
        sa.Column("order_type", sa.String(length=32)),
        sa.Column("is_duplicate", sa.Boolean()),
        sa.Column("is_edited_order", sa.Boolean()),
        sa.Column("ingest_remarks", sa.Text()),
        sa.UniqueConstraint(
            "cost_center",
            "order_number",
            "payment_date",
            name="uq_sales_cost_center_order_number_payment_date",
        ),
        sqlite_autoincrement=True,
    )


MOBILE_FALLBACK_NUMBER = "8888999762"


def _normalize_phone(
    value: str | None, *, warnings: list[str], invalid_phone_numbers: set[str], row_remarks: list[str]
) -> str | None:
    value_str = "" if value is None else str(value)
    if value_str.strip() == "":
        row_remarks.append("MOBILE_FALLBACK_APPLIED")
        if value_str not in invalid_phone_numbers:
            invalid_phone_numbers.add(value_str)
            warnings.append("Invalid phone number fallback applied: <missing>")
        return MOBILE_FALLBACK_NUMBER
    sanitized = value_str
    if re.search(r"[oO]", sanitized):
        sanitized = re.sub(r"[oO]", "0", sanitized)
        row_remarks.append("MOBILE_SANITIZED_O_TO_0")
    digits = re.sub(r"\D", "", sanitized)
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[-10:]
    if len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    if len(digits) == 10:
        return digits
    row_remarks.append("MOBILE_FALLBACK_APPLIED")
    if value_str not in invalid_phone_numbers:
        invalid_phone_numbers.add(value_str)
        warnings.append(f"Invalid phone number fallback applied: {value_str}")
    return MOBILE_FALLBACK_NUMBER


def _parse_numeric(value: Any, *, warnings: list[str], field: str, row_remarks: list[str]) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    try:
        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value))
        cleaned = str(value).replace(",", "")
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        warnings.append(f"Non-numeric value for {field}: {value}")
        row_remarks.append(f"Field {field} contained non-numeric value '{value}' (stored as 0)")
        return Decimal("0")


def _parse_datetime(
    value: Any, *, tz: ZoneInfo, field: str, warnings: list[str], row_remarks: list[str]
) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=tz)
    try:
        parsed = parser.parse(str(value))
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=tz)
    except Exception:
        warnings.append(f"Could not parse datetime for {field}: {value}")
        row_remarks.append(f"Field {field} could not be parsed from value '{value}' (field cleared)")
        return None


def _is_footer_row(values: Sequence[Any]) -> bool:
    non_empty = [value for value in values if value not in (None, "")]
    if not non_empty:
        return False
    combined = " ".join(str(value).strip().lower() for value in non_empty)
    footer_markers = (
        "total records",
        "total order",
        "report generated",
        "this is a computer generated",
        "this is a system generated",
        "powered by quick",
    )
    if any(marker in combined for marker in footer_markers):
        return True
    first_value = non_empty[0]
    return isinstance(first_value, str) and first_value.strip().lower().startswith("total")


def _coerce_row(
    raw: Mapping[str, Any], *, tz: ZoneInfo, warnings: list[str], invalid_phone_numbers: set[str]
) -> tuple[Dict[str, Any], list[str], str | None]:
    row: Dict[str, Any] = {}
    row_remarks: list[str] = []
    drop_reason: str | None = None
    for header, field in HEADER_MAP.items():
        row[field] = raw.get(header)

    for field in DATE_FIELDS:
        row[field] = _parse_datetime(row[field], tz=tz, field=field, warnings=warnings, row_remarks=row_remarks)

    for field in NUMERIC_FIELDS:
        row[field] = _parse_numeric(row[field], warnings=warnings, field=field, row_remarks=row_remarks)

    row["mobile_number"] = _normalize_phone(
        row.get("mobile_number"), warnings=warnings, invalid_phone_numbers=invalid_phone_numbers, row_remarks=row_remarks
    )

    if row.get("order_number") in (None, ""):
        drop_reason = "Skipping row with blank order_number"
        warnings.append(drop_reason)
        return {}, [], drop_reason
    if row.get("payment_date") is None:
        drop_reason = f"Skipping row with missing payment_date for order {row.get('order_number')}"
        warnings.append(drop_reason)
        return {}, [], drop_reason

    if isinstance(row.get("payment_made_at"), str):
        row["payment_made_at"] = row["payment_made_at"].strip()

    return row, row_remarks, drop_reason


def _read_workbook_rows(
    workbook_path: Path, *, tz: ZoneInfo, warnings: list[str], store_code: str
) -> tuple[list[Dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], int]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = wb.active
    header_cells = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    headers = [cell for cell in header_cells if cell]
    missing = REQUIRED_HEADERS - set(headers)
    if missing:
        raise ValueError(f"TD Sales workbook missing expected columns: {sorted(missing)}")

    rows: list[Dict[str, Any]] = []
    warning_rows: list[dict[str, Any]] = []
    dropped_rows: list[dict[str, Any]] = []
    data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    while data_rows and _is_footer_row(data_rows[-1]):
        data_rows.pop()
    rows_downloaded = len(data_rows)

    invalid_phone_numbers: set[str] = set()

    for values in data_rows:
        raw_row = {header: values[idx] if idx < len(values) else None for idx, header in enumerate(headers)}
        normalized, row_remarks, drop_reason = _coerce_row(
            raw_row, tz=tz, warnings=warnings, invalid_phone_numbers=invalid_phone_numbers
        )
        order_number = _stringify_value(raw_row.get("Order Number"))
        if normalized:
            normalized["_remarks"] = row_remarks
            if row_remarks:
                warning_rows.append(
                    {
                        "store_code": store_code,
                        "order_number": order_number,
                        "headers": headers,
                        "values": {header: _stringify_value(raw_row.get(header)) for header in headers},
                        "remarks": "; ".join(row_remarks),
                        "ingest_remarks": "; ".join(row_remarks),
                    }
                )
            rows.append(normalized)
        else:
            dropped_rows.append(
                {
                    "store_code": store_code,
                    "order_number": order_number,
                    "headers": headers,
                    "values": {header: _stringify_value(raw_row.get(header)) for header in headers},
                    "remarks": drop_reason or "; ".join(row_remarks) or "Row dropped due to missing required values",
                    "ingest_remarks": drop_reason or "; ".join(row_remarks) or "Row dropped due to missing required values",
                }
            )
    return rows, warning_rows, dropped_rows, rows_downloaded


def _make_insert(table: sa.Table, values: Mapping[str, Any], *, use_sqlite: bool) -> sa.sql.dml.Insert:
    insert_fn = sqlite_insert if use_sqlite else pg_insert
    insert = insert_fn(table).values(**values)
    conflict_cols = [col.name for col in table.primary_key.columns] if table.primary_key else []
    for constraint in table.constraints:
        if isinstance(constraint, sa.UniqueConstraint):
            conflict_cols = [col.name for col in constraint.columns]
            break
    if not conflict_cols:
        return insert
    return insert.on_conflict_do_update(index_elements=conflict_cols, set_={key: insert.excluded[key] for key in values})


async def ingest_td_sales_workbook(
    *,
    workbook_path: Path,
    store_code: str,
    cost_center: str,
    run_id: str,
    run_date: datetime,
    database_url: str,
    logger: JsonLogger,
) -> TdSalesIngestResult:
    tz = get_timezone()
    warnings: list[str] = []
    rows, warning_rows, dropped_rows, rows_downloaded = _read_workbook_rows(
        workbook_path, tz=tz, warnings=warnings, store_code=store_code
    )
    header_labels = list(HEADER_MAP.keys())

    if not rows:
        log_event(
            logger=logger,
            phase="sales_ingest",
            status="warn",
            message="No rows parsed from TD Sales workbook",
            store_code=store_code,
            workbook=str(workbook_path),
        )
        return TdSalesIngestResult(
            staging_rows=0,
            staging_inserted=0,
            staging_updated=0,
            final_rows=0,
            final_inserted=0,
            final_updated=0,
            warnings=warnings,
            ingest_remarks=[],
            rows_downloaded=rows_downloaded,
            dropped_rows=dropped_rows,
            warning_rows=warning_rows,
        )

    metadata = sa.MetaData()
    stg_table = _stg_td_sales_table(metadata)
    final_table = _sales_table(metadata)
    orders_table = _orders_table(metadata)
    use_sqlite = database_url.startswith("sqlite")

    async with session_scope(database_url) as session:
        bind = session.bind
        if isinstance(bind, AsyncEngine):
            async with bind.begin() as conn:
                await conn.run_sync(metadata.create_all)
        elif isinstance(bind, AsyncConnection):
            await bind.run_sync(metadata.create_all)
        else:  # pragma: no cover - safety fallback
            raise TypeError(f"Unsupported SQLAlchemy bind for TD Sales ingest: {type(bind)!r}")

        staging_keys = [
            (store_code, row.get("order_number"), row.get("payment_date")) for row in rows
        ]
        final_keys = [
            (cost_center, row.get("order_number"), row.get("payment_date")) for row in rows
        ]
        existing_staging = await _fetch_existing_keys(
            session,
            stg_table,
            (stg_table.c.store_code, stg_table.c.order_number, stg_table.c.payment_date),
            staging_keys,
        )
        existing_final = await _fetch_existing_keys(
            session,
            final_table,
            (final_table.c.cost_center, final_table.c.order_number, final_table.c.payment_date),
            final_keys,
        )
        staging_key_set = set(staging_keys)
        final_key_set = set(final_keys)
        staging_inserted = len(staging_key_set - existing_staging)
        staging_updated = len(staging_key_set & existing_staging)
        final_inserted = len(final_key_set - existing_final)
        final_updated = len(final_key_set & existing_final)

        duplicate_counts = Counter(
            (store_code, str(row.get("order_number")), row.get("payment_mode")) for row in rows
        )
        duplicates_set = {key for key, count in duplicate_counts.items() if count > 1}

        payment_totals: dict[tuple[str, str], Decimal] = {}
        payment_dates_by_order: dict[tuple[str, str], set[datetime]] = {}
        for row in rows:
            order_key = (store_code, str(row.get("order_number")))
            payment_received = row.get("payment_received") or Decimal("0")
            payment_totals[order_key] = payment_totals.get(order_key, Decimal("0")) + payment_received
            payment_date = row.get("payment_date")
            if payment_date is not None:
                payment_dates_by_order.setdefault(order_key, set()).add(payment_date)

        order_numbers = {str(row.get("order_number")) for row in rows}
        net_amounts: dict[tuple[str, str], Decimal] = {}
        if order_numbers:
            order_rows = await session.execute(
                sa.select(orders_table.c.store_code, orders_table.c.order_number, orders_table.c.net_amount).where(
                    sa.and_(orders_table.c.store_code == store_code, orders_table.c.order_number.in_(order_numbers))
                )
            )
            for row in order_rows:
                if row.net_amount is not None:
                    net_amounts[(row.store_code, str(row.order_number))] = row.net_amount

        existing_payment_totals: dict[tuple[str, str], Decimal] = {}
        if order_numbers:
            existing_rows = await session.execute(
                sa.select(
                    final_table.c.store_code,
                    final_table.c.order_number,
                    final_table.c.payment_date,
                    final_table.c.payment_received,
                ).where(
                    sa.and_(final_table.c.store_code == store_code, final_table.c.order_number.in_(order_numbers))
                )
            )
            for row in existing_rows:
                order_key = (row.store_code, str(row.order_number))
                payment_date = row.payment_date
                if payment_date and payment_date in payment_dates_by_order.get(order_key, set()):
                    continue
                payment_received = row.payment_received or Decimal("0")
                existing_payment_totals[order_key] = (
                    existing_payment_totals.get(order_key, Decimal("0")) + payment_received
                )

        # Compare net_amount using the cumulative payment totals (existing payments + current workbook).
        edited_shortfall_keys = {
            key
            for key, total in payment_totals.items()
            if key in net_amounts
            and (total + existing_payment_totals.get(key, Decimal("0"))) < net_amounts[key]
        }

        staging_count = 0
        final_count = 0
        remark_entries: list[dict[str, str]] = []
        edited_rows: list[dict[str, Any]] = []
        duplicate_rows: list[dict[str, Any]] = []

        if duplicates_set:
            await session.execute(
                sa.update(stg_table)
                .where(
                    sa.tuple_(
                        stg_table.c.store_code, stg_table.c.order_number, stg_table.c.payment_mode
                    ).in_(duplicates_set)
                )
                .values(is_duplicate=True, is_edited_order=True)
            )
            await session.execute(
                sa.update(final_table)
                .where(
                    sa.tuple_(
                        final_table.c.store_code, final_table.c.order_number, final_table.c.payment_mode
                    ).in_(duplicates_set)
                )
                .values(is_duplicate=True, is_edited_order=True)
            )

        if edited_shortfall_keys:
            await session.execute(
                sa.update(stg_table)
                .where(
                    sa.tuple_(stg_table.c.store_code, stg_table.c.order_number).in_(edited_shortfall_keys)
                )
                .values(is_edited_order=True)
            )
            await session.execute(
                sa.update(final_table)
                .where(
                    sa.tuple_(final_table.c.store_code, final_table.c.order_number).in_(edited_shortfall_keys)
                )
                .values(is_edited_order=True)
            )

        for row in rows:
            order_number = str(row.get("order_number"))
            duplicate_key = (store_code, order_number, row.get("payment_mode"))
            order_key = (store_code, order_number)
            remarks = row.pop("_remarks", [])
            is_duplicate = duplicate_key in duplicates_set
            is_shortfall = order_key in edited_shortfall_keys
            adjustments = row.get("adjustments") or Decimal("0")
            is_adjusted = adjustments > 0
            if is_duplicate:
                remarks.append(
                    "Duplicate order_number/payment_mode "
                    f"'{order_number}'/'{_stringify_value(row.get('payment_mode'))}' detected in sales data"
                )
            if is_shortfall:
                remarks.append(
                    "Total payment_received "
                    f"{_stringify_value(payment_totals[order_key])} is less than net_amount "
                    f"{_stringify_value(net_amounts[order_key])} for order '{order_number}'"
                )
            if is_adjusted:
                remarks.append("Orders Value was adjusted")
            row["is_duplicate"] = is_duplicate
            row["is_edited_order"] = is_duplicate or is_shortfall or is_adjusted
            row["ingest_remarks"] = "; ".join(remarks) if remarks else None
            if row["is_edited_order"]:
                edited_rows.append(
                    {
                        "store_code": store_code,
                        "order_number": order_number,
                        "headers": header_labels,
                        "values": {label: _stringify_value(row.get(field)) for label, field in HEADER_MAP.items()},
                        "ingest_remarks": row.get("ingest_remarks"),
                    }
                )
            if row["is_duplicate"]:
                duplicate_rows.append(
                    {
                        "store_code": store_code,
                        "order_number": order_number,
                        "headers": header_labels,
                        "values": {label: _stringify_value(row.get(field)) for label, field in HEADER_MAP.items()},
                        "ingest_remarks": row.get("ingest_remarks"),
                    }
                )

            if row.get("ingest_remarks"):
                remark_entries.append(
                    {
                        "store_code": store_code,
                        "order_number": order_number,
                        "ingest_remarks": str(row["ingest_remarks"]),
                    }
                )

            stg_values = {
                **{field: row.get(field) for field in STG_TD_SALES_COLUMNS},
                "run_id": run_id,
                "run_date": run_date,
                "cost_center": cost_center,
                "store_code": store_code,
            }
            stg_stmt = _make_insert(stg_table, stg_values, use_sqlite=use_sqlite)
            await session.execute(stg_stmt)
            staging_count += 1

            final_values = {
                **{field: row.get(field) for field in SALES_COLUMNS},
                "run_id": run_id,
                "run_date": run_date,
                "cost_center": cost_center,
                "store_code": store_code,
            }
            final_stmt = _make_insert(final_table, final_values, use_sqlite=use_sqlite)
            await session.execute(final_stmt)
            final_count += 1

        await session.commit()

    return TdSalesIngestResult(
        staging_rows=staging_count,
        staging_inserted=staging_inserted,
        staging_updated=staging_updated,
        final_rows=final_count,
        final_inserted=final_inserted,
        final_updated=final_updated,
        warnings=warnings,
        ingest_remarks=remark_entries,
        rows_downloaded=rows_downloaded,
        dropped_rows=dropped_rows,
        warning_rows=warning_rows,
        edited_rows=edited_rows,
        duplicate_rows=duplicate_rows,
        rows_edited=len(edited_rows),
        rows_duplicate=len(duplicate_rows),
    )


def _expected_headers() -> Sequence[str]:
    return list(HEADER_MAP.keys())


__all__ = [
    "TdSalesIngestResult",
    "ingest_td_sales_workbook",
    "_orders_table",
    "_stg_td_sales_table",
    "_sales_table",
    "_expected_headers",
]
