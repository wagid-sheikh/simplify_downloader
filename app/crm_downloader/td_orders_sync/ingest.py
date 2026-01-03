from __future__ import annotations

import re
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Sequence
from zoneinfo import ZoneInfo

import openpyxl
import sqlalchemy as sa
from dateutil import parser
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.dialects.sqlite import insert as sqlite_insert

from app.common.db import session_scope
from app.common.date_utils import get_timezone
from app.dashboard_downloader.json_logger import JsonLogger, log_event

STG_TD_ORDERS_COLUMNS = [
    "order_date",
    "order_number",
    "customer_code",
    "customer_name",
    "customer_address",
    "mobile_number",
    "preference",
    "due_date",
    "last_activity",
    "pieces",
    "weight",
    "gross_amount",
    "discount",
    "tax_amount",
    "net_amount",
    "advance",
    "paid",
    "adjustment",
    "balance",
    "advance_received",
    "advance_used",
    "booked_by",
    "workshop_note",
    "order_note",
    "home_delivery",
    "area_location",
    "garments_inspected_by",
    "customer_gstin",
    "registration_source",
    "order_from_pos",
    "package",
    "package_type",
    "package_name",
    "feedback",
    "tags",
    "comment",
    "primary_service",
    "topup_service",
    "order_status",
    "last_payment_activity",
    "package_payment_info",
    "coupon_code",
    "ingest_remarks",
]

ORDERS_COLUMNS = [
    "cost_center",
    "store_code",
    "source_system",
    "order_number",
    "invoice_number",
    "order_date",
    "customer_code",
    "customer_name",
    "mobile_number",
    "customer_gstin",
    "customer_source",
    "package_flag",
    "service_type",
    "customer_address",
    "pieces",
    "weight",
    "due_date",
    "default_due_date",
    "due_days_delta",
    "due_date_flag",
    "complete_processing_by",
    "gross_amount",
    "discount_amount",
    "tax_amount",
    "net_amount",
    "payment_status",
    "order_status",
    "payment_mode",
    "payment_date",
    "payment_amount",
    "order_edited_flag",
    "system_order_status",
    "google_maps_url",
    "latitude",
    "longitude",
    "created_by",
    "created_at",
    "updated_by",
    "updated_at",
    "run_id",
    "run_date",
]

HEADER_MAP: Mapping[str, str] = {
    "Order Date / Time": "order_date",
    "Order No.": "order_number",
    "Customer Code": "customer_code",
    "Name": "customer_name",
    "Address": "customer_address",
    "Phone": "mobile_number",
    "Preference": "preference",
    "Due Date": "due_date",
    "Last Activity": "last_activity",
    "Pcs.": "pieces",
    "Weight": "weight",
    "Gross Amount": "gross_amount",
    "Discount": "discount",
    "Tax": "tax_amount",
    "Net Amount": "net_amount",
    "Advance": "advance",
    "Paid": "paid",
    "Adjustment": "adjustment",
    "Balance": "balance",
    "Advance Received": "advance_received",
    "Advance Used": "advance_used",
    "Booked By": "booked_by",
    "Workshop Note": "workshop_note",
    "Order Note": "order_note",
    "Home Delivery": "home_delivery",
    "Area Location": "area_location",
    "Garments Inspected By": "garments_inspected_by",
    "Customer GSTIN": "customer_gstin",
    "Registration Source": "registration_source",
    "Order From POS": "order_from_pos",
    "Package": "package",
    "Package Type": "package_type",
    "Package Name": "package_name",
    "Feedback": "feedback",
    "Tags": "tags",
    "Comment": "comment",
    "Primary Services": "primary_service",
    "Top Up/Extra Service": "topup_service",
    "Order Status": "order_status",
    "Last Payment Activity": "last_payment_activity",
    "Package Payment Info": "package_payment_info",
    "Coupon Code": "coupon_code",
}

NUMERIC_FIELDS = {
    "pieces",
    "weight",
    "gross_amount",
    "discount",
    "tax_amount",
    "net_amount",
    "advance",
    "paid",
    "adjustment",
    "balance",
    "advance_received",
    "advance_used",
}

DATE_FIELDS = {
    "order_date",
    "due_date",
    "last_activity",
    "last_payment_activity",
}

REQUIRED_HEADERS = set(HEADER_MAP.keys())


@dataclass
class TdOrdersIngestResult:
    staging_rows: int
    final_rows: int
    warnings: list[str]
    ingest_remarks: list[dict[str, str]] = field(default_factory=list)


def _stg_td_orders_table(metadata: sa.MetaData) -> sa.Table:
    return sa.Table(
        "stg_td_orders",
        metadata,
        sa.Column(
            "id",
            sa.BigInteger().with_variant(sa.Integer(), "sqlite"),
            primary_key=True,
            autoincrement=True,
        ),
        sa.Column("run_id", sa.Text()),
        sa.Column("run_date", sa.DateTime(timezone=True)),
        sa.Column("cost_center", sa.String(length=8)),
        sa.Column("store_code", sa.String(length=8)),
        sa.Column("order_date", sa.DateTime(timezone=True)),
        sa.Column("order_number", sa.String(length=12)),
        sa.Column("customer_code", sa.String(length=12)),
        sa.Column("customer_name", sa.String(length=128)),
        sa.Column("customer_address", sa.String(length=256)),
        sa.Column("mobile_number", sa.String(length=16)),
        sa.Column("preference", sa.String(length=128)),
        sa.Column("due_date", sa.DateTime(timezone=True)),
        sa.Column("last_activity", sa.DateTime(timezone=True)),
        sa.Column("pieces", sa.Numeric(12, 0)),
        sa.Column("weight", sa.Numeric(12, 2)),
        sa.Column("gross_amount", sa.Numeric(12, 2)),
        sa.Column("discount", sa.Numeric(12, 2)),
        sa.Column("tax_amount", sa.Numeric(12, 2)),
        sa.Column("net_amount", sa.Numeric(12, 2)),
        sa.Column("advance", sa.Numeric(12, 2)),
        sa.Column("paid", sa.Numeric(12, 2)),
        sa.Column("adjustment", sa.Numeric(12, 2)),
        sa.Column("balance", sa.Numeric(12, 2)),
        sa.Column("advance_received", sa.Numeric(12, 2)),
        sa.Column("advance_used", sa.Numeric(12, 2)),
        sa.Column("booked_by", sa.String(length=32)),
        sa.Column("workshop_note", sa.Text()),
        sa.Column("order_note", sa.Text()),
        sa.Column("home_delivery", sa.String(length=32)),
        sa.Column("area_location", sa.Text()),
        sa.Column("garments_inspected_by", sa.String(length=32)),
        sa.Column("customer_gstin", sa.String(length=32)),
        sa.Column("registration_source", sa.String(length=24)),
        sa.Column("order_from_pos", sa.String(length=32)),
        sa.Column("package", sa.String(length=32)),
        sa.Column("package_type", sa.String(length=32)),
        sa.Column("package_name", sa.String(length=32)),
        sa.Column("feedback", sa.String(length=32)),
        sa.Column("tags", sa.String(length=32)),
        sa.Column("comment", sa.Text()),
        sa.Column("primary_service", sa.String(length=24)),
        sa.Column("topup_service", sa.String(length=32)),
        sa.Column("order_status", sa.String(length=32)),
        sa.Column("last_payment_activity", sa.DateTime(timezone=True)),
        sa.Column("package_payment_info", sa.String(length=32)),
        sa.Column("coupon_code", sa.String(length=32)),
        sa.Column("ingest_remarks", sa.Text()),
        sa.UniqueConstraint("store_code", "order_number", "order_date", name="uq_stg_td_orders_store_order_date"),
        sqlite_autoincrement=True,
    )


def _orders_table(metadata: sa.MetaData) -> sa.Table:
    return sa.Table(
        "orders",
        metadata,
        sa.Column(
            "id",
            sa.BigInteger().with_variant(sa.Integer(), "sqlite"),
            primary_key=True,
            autoincrement=True,
        ),
        sa.Column("run_id", sa.Text()),
        sa.Column("run_date", sa.DateTime(timezone=True)),
        sa.Column("cost_center", sa.String(length=8), nullable=False),
        sa.Column("store_code", sa.String(length=8), nullable=False),
        sa.Column("source_system", sa.String(length=12), nullable=False),
        sa.Column("order_number", sa.String(length=12), nullable=False),
        sa.Column("invoice_number", sa.String(length=12)),
        sa.Column("order_date", sa.DateTime(timezone=True), nullable=False),
        sa.Column("customer_code", sa.String(length=12)),
        sa.Column("customer_name", sa.String(length=128), nullable=False),
        sa.Column("mobile_number", sa.String(length=16), nullable=False),
        sa.Column("customer_gstin", sa.String(length=32)),
        sa.Column("customer_source", sa.String(length=24)),
        sa.Column("package_flag", sa.Boolean(), nullable=False, server_default=sa.text("false")),
        sa.Column("service_type", sa.String(length=24)),
        sa.Column("customer_address", sa.Text()),
        sa.Column("pieces", sa.Numeric(12, 0)),
        sa.Column("weight", sa.Numeric(12, 2)),
        sa.Column("due_date", sa.DateTime(timezone=True), nullable=False),
        sa.Column("default_due_date", sa.DateTime(timezone=True), nullable=False),
        sa.Column("due_days_delta", sa.Numeric(10, 0)),
        sa.Column("due_date_flag", sa.String(length=24)),
        sa.Column("complete_processing_by", sa.DateTime(timezone=True), nullable=False),
        sa.Column("gross_amount", sa.Numeric(12, 2)),
        sa.Column("discount_amount", sa.Numeric(12, 2)),
        sa.Column("tax_amount", sa.Numeric(12, 2)),
        sa.Column("net_amount", sa.Numeric(12, 2)),
        sa.Column("payment_status", sa.String(length=24)),
        sa.Column("order_status", sa.String(length=24)),
        sa.Column("payment_mode", sa.String(length=24)),
        sa.Column("payment_date", sa.DateTime(timezone=True)),
        sa.Column("payment_amount", sa.Numeric(12, 2)),
        sa.Column("order_edited_flag", sa.Boolean(), nullable=False, server_default=sa.text("false")),
        sa.Column("system_order_status", sa.String(length=24), server_default=sa.text("'Active'")),
        sa.Column("google_maps_url", sa.String(length=256)),
        sa.Column("latitude", sa.Float()),
        sa.Column("longitude", sa.Float()),
        sa.Column("created_by", sa.BigInteger()),
        sa.Column("created_at", sa.DateTime(timezone=True), nullable=False),
        sa.Column("updated_by", sa.BigInteger()),
        sa.Column("updated_at", sa.DateTime(timezone=True)),
        sa.UniqueConstraint("cost_center", "order_number", "order_date", name="uq_orders_cost_center_order_number_order_date"),
        sqlite_autoincrement=True,
    )


def _normalize_phone(
    value: str | None, *, warnings: list[str], invalid_phone_numbers: set[str], row_remarks: list[str]
) -> str | None:
    if value is None:
        return None
    digits = re.sub(r"\D", "", str(value))
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[-10:]
    if len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    if len(digits) == 10:
        return digits
    value_str = str(value)
    row_remarks.append(f"phone: {value_str}")
    if value_str not in invalid_phone_numbers:
        invalid_phone_numbers.add(value_str)
        warnings.append(f"Invalid phone number dropped: {value_str}")
    return None


def _parse_numeric(value: Any, *, warnings: list[str], field: str, row_remarks: list[str]) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    try:
        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value))
        cleaned = str(value).replace(",", "")
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        warnings.append(f"Non-numeric value for {field}: {value}")
        row_remarks.append(f"{field}: {value}")
        return Decimal("0")


def _parse_datetime(
    value: Any, *, tz: ZoneInfo, field: str, warnings: list[str], row_remarks: list[str]
) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=tz)
    try:
        parsed = parser.parse(str(value))
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=tz)
    except Exception:
        warnings.append(f"Could not parse datetime for {field}: {value}")
        row_remarks.append(f"{field}: {value}")
        return None


def _coerce_row(raw: Mapping[str, Any], *, tz: ZoneInfo, warnings: list[str], invalid_phone_numbers: set[str]) -> Dict[str, Any]:
    row: Dict[str, Any] = {}
    row_remarks: list[str] = []
    for header, field in HEADER_MAP.items():
        row[field] = raw.get(header)

    row["order_date"] = _parse_datetime(
        row["order_date"], tz=tz, field="order_date", warnings=warnings, row_remarks=row_remarks
    )
    row["due_date"] = _parse_datetime(row["due_date"], tz=tz, field="due_date", warnings=warnings, row_remarks=row_remarks)
    row["last_activity"] = _parse_datetime(
        row["last_activity"], tz=tz, field="last_activity", warnings=warnings, row_remarks=row_remarks
    )
    row["last_payment_activity"] = _parse_datetime(
        row["last_payment_activity"], tz=tz, field="last_payment_activity", warnings=warnings, row_remarks=row_remarks
    )

    for field in NUMERIC_FIELDS:
        row[field] = _parse_numeric(row[field], warnings=warnings, field=field, row_remarks=row_remarks)

    row["mobile_number"] = _normalize_phone(
        row.get("mobile_number"),
        warnings=warnings,
        invalid_phone_numbers=invalid_phone_numbers,
        row_remarks=row_remarks,
    )
    if row["order_number"] in (None, ""):
        warnings.append("Skipping row with blank order_number")
        return {}
    if row["order_date"] is None:
        warnings.append(f"Skipping row with missing order_date for order {row['order_number']}")
        return {}
    if row["due_date"] is None and row["order_date"] is not None:
        row["due_date"] = row["order_date"] + timedelta(days=3)

    row["ingest_remarks"] = "; ".join(row_remarks) if row_remarks else None

    return row


def _is_footer_row(values: Sequence[Any]) -> bool:
    non_empty = [value for value in values if value not in (None, "")]
    if not non_empty:
        return False
    combined = " ".join(str(value).strip().lower() for value in non_empty)
    footer_markers = (
        "total records",
        "total order",
        "report generated",
        "this is a computer generated",
        "this is a system generated",
        "powered by quick",
    )
    if any(marker in combined for marker in footer_markers):
        return True
    first_value = non_empty[0]
    return isinstance(first_value, str) and first_value.strip().lower().startswith("total")


def _read_workbook_rows(
    workbook_path: Path, *, tz: ZoneInfo, warnings: list[str], logger: JsonLogger
) -> list[Dict[str, Any]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = wb.active
    header_cells = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    headers = [cell for cell in header_cells if cell]
    missing = REQUIRED_HEADERS - set(headers)
    if missing:
        raise ValueError(f"TD Orders workbook missing expected columns: {sorted(missing)}")

    rows: list[Dict[str, Any]] = []
    data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    while data_rows and _is_footer_row(data_rows[-1]):
        data_rows.pop()

    invalid_phone_numbers: set[str] = set()

    for values in data_rows:
        raw_row = {header: values[idx] if idx < len(values) else None for idx, header in enumerate(headers)}
        normalized = _coerce_row(raw_row, tz=tz, warnings=warnings, invalid_phone_numbers=invalid_phone_numbers)
        if normalized:
            rows.append(normalized)
    return rows


def _make_insert(table: sa.Table, values: Mapping[str, Any], *, use_sqlite: bool) -> sa.sql.dml.Insert:
    insert_fn = sqlite_insert if use_sqlite else pg_insert
    insert = insert_fn(table).values(**values)
    conflict_cols = [col.name for col in table.primary_key.columns] if table.primary_key else []
    for constraint in table.constraints:
        if isinstance(constraint, sa.UniqueConstraint):
            conflict_cols = [col.name for col in constraint.columns]
            break
    if not conflict_cols:
        return insert
    return insert.on_conflict_do_update(index_elements=conflict_cols, set_={key: insert.excluded[key] for key in values})


async def ingest_td_orders_workbook(
    *,
    workbook_path: Path,
    store_code: str,
    cost_center: str,
    run_id: str,
    run_date: datetime,
    database_url: str,
    logger: JsonLogger,
) -> TdOrdersIngestResult:
    tz = get_timezone()
    warnings: list[str] = []
    rows = _read_workbook_rows(workbook_path, tz=tz, warnings=warnings, logger=logger)
    remark_entries = [
        {
            "store_code": store_code,
            "order_number": str(row.get("order_number")) if row.get("order_number") is not None else "",
            "ingest_remarks": str(row["ingest_remarks"]),
        }
        for row in rows
        if row.get("ingest_remarks")
    ]
    if not rows:
        log_event(
            logger=logger,
            phase="ingest",
            status="warn",
            message="No rows parsed from TD Orders workbook",
            store_code=store_code,
            workbook=str(workbook_path),
        )
        return TdOrdersIngestResult(
            staging_rows=0, final_rows=0, warnings=warnings, ingest_remarks=remark_entries
        )

    metadata = sa.MetaData()
    stg_table = _stg_td_orders_table(metadata)
    final_table = _orders_table(metadata)
    use_sqlite = database_url.startswith("sqlite")

    async with session_scope(database_url) as session:
        bind = session.bind
        if isinstance(bind, sa.ext.asyncio.AsyncEngine):
            async with bind.begin() as conn:
                await conn.run_sync(metadata.create_all)
        elif isinstance(bind, sa.ext.asyncio.AsyncConnection):
            await bind.run_sync(metadata.create_all)
        else:
            raise TypeError(f"Unsupported SQLAlchemy bind for TD Orders ingest: {type(bind)!r}")

        staging_count = 0
        final_count = 0
        for row in rows:
            stg_values = {
                **{field: row.get(field) for field in STG_TD_ORDERS_COLUMNS},
                "run_id": run_id,
                "run_date": run_date,
                "cost_center": cost_center,
                "store_code": store_code,
            }
            stmt = _make_insert(stg_table, stg_values, use_sqlite=use_sqlite)
            await session.execute(stmt)
            staging_count += 1

            order_date = stg_values["order_date"]
            due_date = stg_values["due_date"]
            default_due_date = order_date + timedelta(days=3)
            due_days_delta = (due_date - default_due_date).days if due_date and default_due_date else 0
            if due_days_delta == 0:
                due_flag = "Normal Delivery"
            elif due_days_delta > 0:
                due_flag = "Date Extended"
            else:
                due_flag = "Express Delivery"
            complete_processing_by = default_due_date - timedelta(days=1)
            package_flag = False if (stg_values.get("package") or "").strip().lower() == "no" else True

            final_values = {
                "cost_center": cost_center,
                "store_code": store_code,
                "source_system": "TumbleDry",
                "order_number": stg_values["order_number"],
                "invoice_number": None,
                "order_date": order_date,
                "customer_code": stg_values.get("customer_code"),
                "customer_name": stg_values.get("customer_name") or "",
                "mobile_number": stg_values.get("mobile_number") or "",
                "customer_gstin": stg_values.get("customer_gstin"),
                "customer_source": stg_values.get("registration_source"),
                "package_flag": package_flag,
                "service_type": stg_values.get("primary_service"),
                "customer_address": stg_values.get("customer_address"),
                "pieces": stg_values.get("pieces"),
                "weight": stg_values.get("weight"),
                "due_date": due_date or default_due_date,
                "default_due_date": default_due_date,
                "due_days_delta": due_days_delta,
                "due_date_flag": due_flag,
                "complete_processing_by": complete_processing_by,
                "gross_amount": stg_values.get("gross_amount"),
                "discount_amount": stg_values.get("discount"),
                "tax_amount": stg_values.get("tax_amount"),
                "net_amount": stg_values.get("net_amount"),
                "payment_status": "Pending",
                "order_status": "Pending",
                "payment_mode": None,
                "payment_date": None,
                "payment_amount": None,
                "order_edited_flag": False,
                "system_order_status": "Active",
                "google_maps_url": None,
                "latitude": None,
                "longitude": None,
                "created_by": 1,
                "created_at": run_date,
                "updated_by": None,
                "updated_at": None,
                "run_id": run_id,
                "run_date": run_date,
            }
            final_stmt = _make_insert(final_table, final_values, use_sqlite=use_sqlite)
            await session.execute(final_stmt)
            final_count += 1

        await session.commit()

    return TdOrdersIngestResult(
        staging_rows=staging_count,
        final_rows=final_count,
        warnings=warnings,
        ingest_remarks=remark_entries,
    )


def _expected_headers() -> Sequence[str]:
    return list(HEADER_MAP.keys())


__all__ = [
    "TdOrdersIngestResult",
    "ingest_td_orders_workbook",
    "_stg_td_orders_table",
    "_orders_table",
    "_expected_headers",
]
