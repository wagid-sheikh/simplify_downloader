from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from datetime import datetime, timedelta
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Dict, Mapping, Sequence

import openpyxl
import sqlalchemy as sa
from dateutil import parser
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.dialects.sqlite import insert as sqlite_insert

from app.common.date_utils import get_timezone
from app.common.db import session_scope
from app.dashboard_downloader.json_logger import JsonLogger, log_event

STG_UC_ORDERS_COLUMNS = [
    "s_no",
    "order_number",
    "invoice_number",
    "invoice_date",
    "customer_name",
    "mobile_number",
    "payment_status",
    "customer_gstin",
    "place_of_supply",
    "net_amount",
    "cgst",
    "sgst",
    "gross_amount",
    "ingest_remarks",
]

ORDERS_COLUMNS = [
    "cost_center",
    "store_code",
    "source_system",
    "order_number",
    "invoice_number",
    "order_date",
    "customer_code",
    "customer_name",
    "mobile_number",
    "customer_gstin",
    "customer_source",
    "package_flag",
    "service_type",
    "customer_address",
    "pieces",
    "weight",
    "due_date",
    "default_due_date",
    "due_days_delta",
    "due_date_flag",
    "complete_processing_by",
    "gross_amount",
    "discount_amount",
    "tax_amount",
    "net_amount",
    "payment_status",
    "order_status",
    "payment_mode",
    "payment_date",
    "payment_amount",
    "order_edited_flag",
    "system_order_status",
    "ingest_remarks",
    "google_maps_url",
    "latitude",
    "longitude",
    "created_by",
    "created_at",
    "updated_by",
    "updated_at",
    "run_id",
    "run_date",
]

HEADER_MAP: Mapping[str, str] = {
    "S.No.": "s_no",
    "Booking ID": "order_number",
    "Invoice No.": "invoice_number",
    "Invoice Date": "invoice_date",
    "Customer Name": "customer_name",
    "Customer Ph. No.": "mobile_number",
    "Payment Status": "payment_status",
    "Customer GSTIN": "customer_gstin",
    "Place of Supply": "place_of_supply",
    "Taxable Value": "net_amount",
    "CGST": "cgst",
    "SGST": "sgst",
    "Total Invoice Value": "gross_amount",
}

NUMERIC_FIELDS = {"net_amount", "cgst", "sgst", "gross_amount"}

REQUIRED_HEADERS = set(HEADER_MAP.keys())

GSTIN_PATTERN = re.compile(r"^[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z]{1}[1-9A-Z]{1}Z[0-9A-Z]{1}$")


def _stringify_value(value: Any) -> str:
    try:
        if hasattr(value, "isoformat"):
            return value.isoformat()  # type: ignore[call-arg]
    except Exception:
        pass
    return str(value) if value is not None else ""


@dataclass
class UcOrdersIngestResult:
    staging_rows: int
    final_rows: int
    staging_inserted: int
    staging_updated: int
    final_inserted: int
    final_updated: int
    warnings: list[str]
    ingest_remarks: list[dict[str, str]] = field(default_factory=list)
    rows_downloaded: int = 0
    dropped_rows: list[dict[str, Any]] = field(default_factory=list)
    warning_rows: list[dict[str, Any]] = field(default_factory=list)


def _stg_uc_orders_table(metadata: sa.MetaData) -> sa.Table:
    return sa.Table(
        "stg_uc_orders",
        metadata,
        sa.Column(
            "id",
            sa.BigInteger().with_variant(sa.Integer(), "sqlite"),
            primary_key=True,
            autoincrement=True,
        ),
        sa.Column("run_id", sa.Text()),
        sa.Column("run_date", sa.DateTime(timezone=True)),
        sa.Column("cost_center", sa.String(length=8)),
        sa.Column("store_code", sa.String(length=8)),
        sa.Column("s_no", sa.Integer()),
        sa.Column("order_number", sa.String(length=24)),
        sa.Column("invoice_number", sa.String(length=24)),
        sa.Column("invoice_date", sa.DateTime(timezone=True)),
        sa.Column("customer_name", sa.String(length=128)),
        sa.Column("mobile_number", sa.String(length=16)),
        sa.Column("payment_status", sa.String(length=24)),
        sa.Column("customer_gstin", sa.String(length=32)),
        sa.Column("place_of_supply", sa.String(length=64)),
        sa.Column("net_amount", sa.Numeric(12, 2)),
        sa.Column("cgst", sa.Numeric(12, 2)),
        sa.Column("sgst", sa.Numeric(12, 2)),
        sa.Column("gross_amount", sa.Numeric(12, 2)),
        sa.Column("ingest_remarks", sa.Text()),
        sa.UniqueConstraint("store_code", "order_number", "invoice_date", name="uq_stg_uc_orders"),
        sqlite_autoincrement=True,
    )


def _orders_table(metadata: sa.MetaData) -> sa.Table:
    return sa.Table(
        "orders",
        metadata,
        sa.Column(
            "id",
            sa.BigInteger().with_variant(sa.Integer(), "sqlite"),
            primary_key=True,
            autoincrement=True,
        ),
        sa.Column("cost_center", sa.String(length=8)),
        sa.Column("store_code", sa.String(length=8)),
        sa.Column("source_system", sa.String(length=24)),
        sa.Column("order_number", sa.String(length=24)),
        sa.Column("invoice_number", sa.String(length=24)),
        sa.Column("order_date", sa.DateTime(timezone=True), nullable=False),
        sa.Column("customer_code", sa.String(length=16)),
        sa.Column("customer_name", sa.String(length=128), nullable=False),
        sa.Column("mobile_number", sa.String(length=16), nullable=False),
        sa.Column("customer_gstin", sa.String(length=32)),
        sa.Column("customer_source", sa.String(length=64)),
        sa.Column("package_flag", sa.Boolean(), nullable=False, server_default=sa.text("false")),
        sa.Column("service_type", sa.String(length=64)),
        sa.Column("customer_address", sa.String(length=256)),
        sa.Column("pieces", sa.Numeric(12, 0)),
        sa.Column("weight", sa.Numeric(12, 2)),
        sa.Column("due_date", sa.DateTime(timezone=True)),
        sa.Column("default_due_date", sa.DateTime(timezone=True)),
        sa.Column("due_days_delta", sa.Integer()),
        sa.Column("due_date_flag", sa.String(length=24)),
        sa.Column("complete_processing_by", sa.DateTime(timezone=True)),
        sa.Column("gross_amount", sa.Numeric(12, 2)),
        sa.Column("discount_amount", sa.Numeric(12, 2)),
        sa.Column("tax_amount", sa.Numeric(12, 2)),
        sa.Column("net_amount", sa.Numeric(12, 2)),
        sa.Column("payment_status", sa.String(length=24)),
        sa.Column("order_status", sa.String(length=24)),
        sa.Column("payment_mode", sa.String(length=24)),
        sa.Column("payment_date", sa.DateTime(timezone=True)),
        sa.Column("payment_amount", sa.Numeric(12, 2)),
        sa.Column("order_edited_flag", sa.Boolean(), nullable=False, server_default=sa.text("false")),
        sa.Column("system_order_status", sa.String(length=24), server_default=sa.text("'Active'")),
        sa.Column("ingest_remarks", sa.Text()),
        sa.Column("google_maps_url", sa.String(length=256)),
        sa.Column("latitude", sa.Float()),
        sa.Column("longitude", sa.Float()),
        sa.Column("created_by", sa.BigInteger()),
        sa.Column("created_at", sa.DateTime(timezone=True), nullable=False),
        sa.Column("updated_by", sa.BigInteger()),
        sa.Column("updated_at", sa.DateTime(timezone=True)),
        sa.Column("run_id", sa.Text()),
        sa.Column("run_date", sa.DateTime(timezone=True)),
        sa.UniqueConstraint("cost_center", "order_number", "order_date", name="uq_orders_cost_center_order_number_order_date"),
        sqlite_autoincrement=True,
    )


def _normalize_phone(
    value: str | None, *, warnings: list[str], invalid_phone_numbers: set[str], row_remarks: list[str]
) -> str | None:
    if value is None:
        return None
    digits = re.sub(r"\D", "", str(value))
    if len(digits) == 12 and digits.startswith("91"):
        digits = digits[-10:]
    if len(digits) == 11 and digits.startswith("0"):
        digits = digits[1:]
    if len(digits) == 10:
        return digits
    value_str = str(value)
    row_remarks.append(f"Phone value '{value_str}' is invalid and was dropped")
    if value_str not in invalid_phone_numbers:
        invalid_phone_numbers.add(value_str)
        warnings.append(f"Invalid phone number dropped: {value_str}")
    return None


def _parse_numeric(value: Any, *, warnings: list[str], field: str, row_remarks: list[str]) -> Decimal:
    if value is None or value == "":
        return Decimal("0")
    try:
        if isinstance(value, (int, float, Decimal)):
            return Decimal(str(value))
        cleaned = str(value).replace(",", "")
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        warnings.append(f"Non-numeric value for {field}: {value}")
        row_remarks.append(f"Field {field} contained non-numeric value '{value}' (stored as 0)")
        return Decimal("0")


def _parse_s_no(value: Any, *, warnings: list[str], row_remarks: list[str]) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
    try:
        if isinstance(value, bool):
            raise ValueError
        if isinstance(value, int):
            return value
        if isinstance(value, float):
            if value.is_integer():
                return int(value)
            raise ValueError
        cleaned = str(value).replace(",", "").strip()
        if not cleaned:
            return None
        parsed = Decimal(cleaned)
        if parsed != parsed.to_integral_value():
            raise InvalidOperation
        return int(parsed)
    except (InvalidOperation, ValueError):
        warnings.append(f"Non-numeric S.No. value dropped: {value}")
        row_remarks.append(f"S.No. value '{value}' could not be parsed and was cleared")
        return None


def _parse_datetime(value: Any, *, field: str, warnings: list[str], row_remarks: list[str]) -> datetime | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value if value.tzinfo else value.replace(tzinfo=get_timezone())
    try:
        parsed = parser.parse(str(value))
        return parsed if parsed.tzinfo else parsed.replace(tzinfo=get_timezone())
    except Exception:
        warnings.append(f"Could not parse datetime for {field}: {value}")
        row_remarks.append(f"Field {field} could not be parsed from value '{value}' (field cleared)")
        return None


def _normalize_order_number(value: Any, *, warnings: list[str], row_remarks: list[str]) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, float) and value.is_integer():
        normalized = str(int(value))
    else:
        normalized = str(value).strip()
    if not normalized:
        return None
    if normalized != str(value).strip():
        row_remarks.append(f"Order number normalized from '{value}' to '{normalized}'")
    return normalized


def _normalize_invoice_number(value: Any, *, warnings: list[str], row_remarks: list[str]) -> str | None:
    if value is None or value == "":
        return None
    if isinstance(value, float) and value.is_integer():
        normalized = str(int(value))
    else:
        normalized = re.sub(r"\s+", "", str(value).strip())
    if not normalized:
        return None
    if normalized != str(value).strip():
        row_remarks.append(f"Invoice number normalized from '{value}' to '{normalized}'")
    return normalized


def _normalize_gstin(value: Any, *, warnings: list[str], row_remarks: list[str]) -> str | None:
    if value is None or value == "":
        return None
    normalized = re.sub(r"\s+", "", str(value)).upper()
    if GSTIN_PATTERN.fullmatch(normalized):
        return normalized
    row_remarks.append(f"Customer GSTIN '{value}' is invalid and was cleared")
    warnings.append(f"Invalid GSTIN dropped: {value}")
    return None


def _build_ingest_remarks_payload(*, warnings: list[str], failures: list[str] | None = None) -> str | None:
    if not warnings and not failures:
        return None
    payload = {
        "warnings": list(warnings),
        "failures": list(failures or []),
    }
    return json.dumps(payload, ensure_ascii=False)


def _coerce_row(
    raw: Mapping[str, Any], *, warnings: list[str], invalid_phone_numbers: set[str]
) -> tuple[Dict[str, Any], list[str], str | None]:
    row: Dict[str, Any] = {}
    row_remarks: list[str] = []
    drop_reason: str | None = None
    for header, field in HEADER_MAP.items():
        row[field] = raw.get(header)

    row["s_no"] = _parse_s_no(row.get("s_no"), warnings=warnings, row_remarks=row_remarks)
    row["order_number"] = _normalize_order_number(
        row.get("order_number"), warnings=warnings, row_remarks=row_remarks
    )
    row["invoice_number"] = _normalize_invoice_number(
        row.get("invoice_number"), warnings=warnings, row_remarks=row_remarks
    )
    row["invoice_date"] = _parse_datetime(
        row.get("invoice_date"), field="invoice_date", warnings=warnings, row_remarks=row_remarks
    )
    for field in NUMERIC_FIELDS:
        row[field] = _parse_numeric(row.get(field), warnings=warnings, field=field, row_remarks=row_remarks)
    row["mobile_number"] = _normalize_phone(
        row.get("mobile_number"),
        warnings=warnings,
        invalid_phone_numbers=invalid_phone_numbers,
        row_remarks=row_remarks,
    )
    raw_gstin = row.get("customer_gstin")
    row["customer_gstin"] = _normalize_gstin(raw_gstin, warnings=warnings, row_remarks=row_remarks)
    if raw_gstin in (None, ""):
        row_remarks.append("Customer GSTIN missing")

    if not row.get("order_number"):
        warning = "Skipping row with blank order_number"
        warnings.append(warning)
        drop_reason = warning
        return {}, row_remarks, drop_reason
    if row["invoice_date"] is None:
        warning = f"Skipping row with missing invoice_date for order {row['order_number']}"
        warnings.append(warning)
        drop_reason = warning
        return {}, row_remarks, drop_reason

    row["ingest_remarks"] = _build_ingest_remarks_payload(warnings=row_remarks)
    return row, row_remarks, drop_reason


def _read_workbook_rows(
    workbook_path: Path, *, warnings: list[str], logger: JsonLogger, store_code: str
) -> tuple[list[Dict[str, Any]], list[dict[str, Any]], list[dict[str, Any]], int]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True)
    sheet = wb.active
    header_cells = list(next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    headers = [cell for cell in header_cells if cell]
    missing = REQUIRED_HEADERS - set(headers)
    if missing:
        raise ValueError(f"UC GST workbook missing expected columns: {sorted(missing)}")

    rows: list[Dict[str, Any]] = []
    warning_rows: list[dict[str, Any]] = []
    dropped_rows: list[dict[str, Any]] = []
    data_rows = list(sheet.iter_rows(min_row=2, values_only=True))
    rows_downloaded = len(data_rows)

    invalid_phone_numbers: set[str] = set()

    for values in data_rows:
        if not any(value not in (None, "") for value in values):
            continue
        raw_row = {header: values[idx] if idx < len(values) else None for idx, header in enumerate(headers)}
        normalized, row_remarks, drop_reason = _coerce_row(
            raw_row, warnings=warnings, invalid_phone_numbers=invalid_phone_numbers
        )
        order_number = _stringify_value(raw_row.get("Booking ID"))
        if normalized:
            if normalized.get("ingest_remarks"):
                warning_rows.append(
                    {
                        "store_code": store_code,
                        "order_number": order_number,
                        "headers": headers,
                        "values": {header: _stringify_value(raw_row.get(header)) for header in headers},
                        "remarks": "; ".join(row_remarks) if row_remarks else None,
                        "ingest_remarks": normalized.get("ingest_remarks"),
                    }
                )
            rows.append(normalized)
        else:
            dropped_rows.append(
                {
                    "store_code": store_code,
                    "order_number": order_number,
                    "headers": headers,
                    "values": {header: _stringify_value(raw_row.get(header)) for header in headers},
                    "remarks": drop_reason or "; ".join(row_remarks) or "Row dropped due to missing required values",
                    "ingest_remarks": _build_ingest_remarks_payload(
                        warnings=row_remarks,
                        failures=[drop_reason] if drop_reason else None,
                    )
                    or drop_reason
                    or "; ".join(row_remarks)
                    or "Row dropped due to missing required values",
                }
            )
    return rows, warning_rows, dropped_rows, rows_downloaded


def _make_insert(table: sa.Table, values: Mapping[str, Any], *, use_sqlite: bool) -> sa.sql.dml.Insert:
    insert_fn = sqlite_insert if use_sqlite else pg_insert
    insert = insert_fn(table).values(**values)
    conflict_cols = [col.name for col in table.primary_key.columns] if table.primary_key else []
    for constraint in table.constraints:
        if isinstance(constraint, sa.UniqueConstraint):
            conflict_cols = [col.name for col in constraint.columns]
            break
    if not conflict_cols:
        return insert
    return insert.on_conflict_do_update(index_elements=conflict_cols, set_={key: insert.excluded[key] for key in values})


async def ingest_uc_orders_workbook(
    *,
    workbook_path: Path,
    store_code: str,
    cost_center: str,
    run_id: str,
    run_date: datetime,
    database_url: str,
    logger: JsonLogger,
) -> UcOrdersIngestResult:
    warnings: list[str] = []
    rows, warning_rows, dropped_rows, rows_downloaded = _read_workbook_rows(
        workbook_path, warnings=warnings, logger=logger, store_code=store_code
    )
    remark_entries = [
        {
            "store_code": store_code,
            "order_number": str(row.get("order_number")) if row.get("order_number") is not None else "",
            "ingest_remarks": str(row["ingest_remarks"]),
        }
        for row in rows
        if row.get("ingest_remarks")
    ]
    if not rows:
        log_event(
            logger=logger,
            phase="ingest",
            status="warn",
            message="No rows parsed from UC GST workbook",
            store_code=store_code,
            workbook=str(workbook_path),
        )
        return UcOrdersIngestResult(
            staging_rows=0,
            final_rows=0,
            staging_inserted=0,
            staging_updated=0,
            final_inserted=0,
            final_updated=0,
            warnings=warnings,
            ingest_remarks=remark_entries,
            rows_downloaded=rows_downloaded,
            dropped_rows=dropped_rows,
            warning_rows=warning_rows,
        )

    metadata = sa.MetaData()
    stg_table = _stg_uc_orders_table(metadata)
    final_table = _orders_table(metadata)
    use_sqlite = database_url.startswith("sqlite")

    async with session_scope(database_url) as session:
        bind = session.bind
        if isinstance(bind, sa.ext.asyncio.AsyncEngine):
            async with bind.begin() as conn:
                await conn.run_sync(metadata.create_all)
        elif isinstance(bind, sa.ext.asyncio.AsyncConnection):
            await bind.run_sync(metadata.create_all)
        else:
            raise TypeError(f"Unsupported SQLAlchemy bind for UC Orders ingest: {type(bind)!r}")

        stg_keys = {(store_code, row["order_number"], row["invoice_date"]) for row in rows}
        existing_stg: set[tuple[str, str, datetime | None]] = set()
        if stg_keys:
            key_tuple = sa.tuple_(stg_table.c.store_code, stg_table.c.order_number, stg_table.c.invoice_date)
            stmt = sa.select(stg_table.c.store_code, stg_table.c.order_number, stg_table.c.invoice_date).where(
                key_tuple.in_(list(stg_keys))
            )
            result = await session.execute(stmt)
            existing_stg = {tuple(row) for row in result.all()}
        staging_inserted = len(stg_keys - existing_stg)
        staging_updated = len(stg_keys & existing_stg)

        final_keys = {(cost_center, row["order_number"], row["invoice_date"]) for row in rows}
        existing_final: set[tuple[str, str, datetime | None]] = set()
        if final_keys:
            key_tuple = sa.tuple_(final_table.c.cost_center, final_table.c.order_number, final_table.c.order_date)
            stmt = sa.select(final_table.c.cost_center, final_table.c.order_number, final_table.c.order_date).where(
                key_tuple.in_(list(final_keys))
            )
            result = await session.execute(stmt)
            existing_final = {tuple(row) for row in result.all()}
        final_inserted = len(final_keys - existing_final)
        final_updated = len(final_keys & existing_final)

        staging_count = 0
        final_count = 0
        for row in rows:
            stg_values = {
                **{field: row.get(field) for field in STG_UC_ORDERS_COLUMNS},
                "run_id": run_id,
                "run_date": run_date,
                "cost_center": cost_center,
                "store_code": store_code,
            }
            stmt = _make_insert(stg_table, stg_values, use_sqlite=use_sqlite)
            await session.execute(stmt)
            staging_count += 1

            order_date = stg_values["invoice_date"]
            default_due_date = order_date + timedelta(days=3)
            due_days_delta = 0
            due_flag = "Normal Delivery"
            complete_processing_by = default_due_date - timedelta(days=1)

            final_values = {
                "cost_center": cost_center,
                "store_code": store_code,
                "source_system": "UClean",
                "order_number": stg_values.get("order_number"),
                "invoice_number": stg_values.get("invoice_number"),
                "order_date": order_date,
                "customer_code": None,
                "customer_name": stg_values.get("customer_name") or "",
                "mobile_number": stg_values.get("mobile_number") or "",
                "customer_gstin": stg_values.get("customer_gstin"),
                "customer_source": "Walk in Customer",
                "package_flag": False,
                "service_type": "UNKNOWN",
                "customer_address": None,
                "pieces": Decimal("0"),
                "weight": Decimal("0"),
                "due_date": default_due_date,
                "default_due_date": default_due_date,
                "due_days_delta": due_days_delta,
                "due_date_flag": due_flag,
                "complete_processing_by": complete_processing_by,
                "gross_amount": stg_values.get("gross_amount"),
                "discount_amount": Decimal("0"),
                "tax_amount": (stg_values.get("cgst") or Decimal("0")) + (stg_values.get("sgst") or Decimal("0")),
                "net_amount": stg_values.get("net_amount"),
                "payment_status": "Pending",
                "order_status": "Pending",
                "payment_mode": None,
                "payment_date": None,
                "payment_amount": None,
                "order_edited_flag": False,
                "system_order_status": "Active",
                "ingest_remarks": stg_values.get("ingest_remarks"),
                "google_maps_url": None,
                "latitude": None,
                "longitude": None,
                "created_by": 1,
                "created_at": run_date,
                "updated_by": None,
                "updated_at": None,
                "run_id": run_id,
                "run_date": run_date,
            }
            final_stmt = _make_insert(final_table, final_values, use_sqlite=use_sqlite)
            await session.execute(final_stmt)
            final_count += 1

        await session.commit()

    return UcOrdersIngestResult(
        staging_rows=staging_count,
        final_rows=final_count,
        staging_inserted=staging_inserted,
        staging_updated=staging_updated,
        final_inserted=final_inserted,
        final_updated=final_updated,
        warnings=warnings,
        ingest_remarks=remark_entries,
        rows_downloaded=rows_downloaded,
        dropped_rows=dropped_rows,
        warning_rows=warning_rows,
    )


def _expected_headers() -> Sequence[str]:
    return list(HEADER_MAP.keys())


__all__ = [
    "UcOrdersIngestResult",
    "ingest_uc_orders_workbook",
    "_stg_uc_orders_table",
    "_orders_table",
    "_expected_headers",
]
