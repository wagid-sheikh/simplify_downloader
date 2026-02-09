from __future__ import annotations

import hashlib
import re
from collections import Counter
from dataclasses import dataclass, field
from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any, Callable, Iterable, Mapping, Sequence

import openpyxl
import sqlalchemy as sa
from dateutil import parser
from sqlalchemy.dialects.postgresql import insert as pg_insert
from sqlalchemy.dialects.sqlite import insert as sqlite_insert

from app.common.db import session_scope
from app.common.date_utils import get_timezone
from app.dashboard_downloader.json_logger import JsonLogger, log_event

TABLE_ARCHIVE_BASE = "stg_uc_archive_orders_base"
TABLE_ARCHIVE_ORDER_DETAILS = "stg_uc_archive_order_details"
TABLE_ARCHIVE_PAYMENT_DETAILS = "stg_uc_archive_payment_details"

FILE_BASE = "base"
FILE_ORDER_DETAILS = "order_details"
FILE_PAYMENT_DETAILS = "payment_details"

REASON_MISSING_STORE_CODE = "missing_store_code"
REASON_MISSING_ORDER_CODE = "missing_order_code"
REASON_MISSING_LINE_HASH = "missing_line_hash"
REASON_MISSING_REQUIRED_FIELD = "missing_required_field"

EXPECTED_HEADERS: dict[str, tuple[str, ...]] = {
    FILE_BASE: (
        "store_code",
        "order_code",
        "pickup",
        "delivery",
        "customer_name",
        "customer_phone",
        "address",
        "payment_text",
        "instructions",
        "customer_source",
        "status",
        "status_date",
    ),
    FILE_ORDER_DETAILS: (
        "store_code",
        "order_code",
        "order_mode",
        "order_datetime",
        "pickup_datetime",
        "delivery_datetime",
        "service",
        "hsn_sac",
        "item_name",
        "rate",
        "quantity",
        "weight",
        "addons",
        "amount",
    ),
    FILE_PAYMENT_DETAILS: ("store_code", "order_code", "payment_mode", "amount", "payment_date", "transaction_id"),
}

PAYMENT_MODE_NORMALIZATION = {
    "UPI / WALLET": "UPI_WALLET",
    "UPI/WALLET": "UPI_WALLET",
    "UPI": "UPI",
    "WALLET": "WALLET",
    "CASH": "CASH",
    "CARD": "CARD",
}

STATUS_NORMALIZATION = {
    "PICKUP PENDING": "PICKUP_PENDING",
    "DELIVERY PENDING": "DELIVERY_PENDING",
    "DELIVERED": "DELIVERED",
    "CANCELLED": "CANCELLED",
}


@dataclass
class FileIngestResult:
    parsed: int = 0
    inserted: int = 0
    updated: int = 0
    rejected: int = 0
    warnings: int = 0
    reject_reasons: dict[str, int] = field(default_factory=dict)


@dataclass
class RejectRecord:
    source_file: str
    row_number: int
    reason_codes: list[str]


@dataclass
class ArchiveIngestResult:
    files: dict[str, FileIngestResult] = field(default_factory=dict)
    rejects: list[RejectRecord] = field(default_factory=list)


@dataclass
class ArchivePublishResult:
    orders_updated: int = 0
    sales_inserted: int = 0
    sales_updated: int = 0
    skipped: int = 0
    skip_reasons: dict[str, int] = field(default_factory=dict)


def _stg_uc_archive_orders_base_table(metadata: sa.MetaData) -> sa.Table:
    return sa.Table(
        TABLE_ARCHIVE_BASE,
        metadata,
        sa.Column("id", sa.Integer, primary_key=True),
        sa.Column("run_id", sa.Text),
        sa.Column("run_date", sa.DateTime(timezone=True)),
        sa.Column("cost_center", sa.String(length=8)),
        sa.Column("store_code", sa.String(length=8)),
        sa.Column("ingest_remarks", sa.Text),
        sa.Column("order_code", sa.String(length=24)),
        sa.Column("pickup_raw", sa.Text),
        sa.Column("delivery_raw", sa.Text),
        sa.Column("customer_name", sa.String(length=128)),
        sa.Column("customer_phone", sa.String(length=24)),
        sa.Column("address", sa.Text),
        sa.Column("payment_text", sa.Text),
        sa.Column("instructions", sa.Text),
        sa.Column("customer_source", sa.String(length=64)),
        sa.Column("status", sa.String(length=32)),
        sa.Column("status_date_raw", sa.Text),
        sa.Column("source_file", sa.Text),
    )


def _stg_uc_archive_order_details_table(metadata: sa.MetaData) -> sa.Table:
    return sa.Table(
        TABLE_ARCHIVE_ORDER_DETAILS,
        metadata,
        sa.Column("id", sa.Integer, primary_key=True),
        sa.Column("run_id", sa.Text),
        sa.Column("run_date", sa.DateTime(timezone=True)),
        sa.Column("cost_center", sa.String(length=8)),
        sa.Column("store_code", sa.String(length=8)),
        sa.Column("ingest_remarks", sa.Text),
        sa.Column("order_code", sa.String(length=24)),
        sa.Column("order_mode", sa.String(length=64)),
        sa.Column("order_datetime_raw", sa.Text),
        sa.Column("pickup_datetime_raw", sa.Text),
        sa.Column("delivery_datetime_raw", sa.Text),
        sa.Column("service", sa.Text),
        sa.Column("hsn_sac", sa.String(length=32)),
        sa.Column("item_name", sa.Text),
        sa.Column("rate", sa.Numeric(12, 2)),
        sa.Column("quantity", sa.Numeric(12, 2)),
        sa.Column("weight", sa.Numeric(12, 3)),
        sa.Column("addons", sa.Text),
        sa.Column("amount", sa.Numeric(12, 2)),
        sa.Column("line_hash", sa.String(length=64)),
        sa.Column("source_file", sa.Text),
    )


def _stg_uc_archive_payment_details_table(metadata: sa.MetaData) -> sa.Table:
    return sa.Table(
        TABLE_ARCHIVE_PAYMENT_DETAILS,
        metadata,
        sa.Column("id", sa.Integer, primary_key=True),
        sa.Column("run_id", sa.Text),
        sa.Column("run_date", sa.DateTime(timezone=True)),
        sa.Column("cost_center", sa.String(length=8)),
        sa.Column("store_code", sa.String(length=8)),
        sa.Column("ingest_remarks", sa.Text),
        sa.Column("order_code", sa.String(length=24)),
        sa.Column("payment_mode", sa.String(length=32)),
        sa.Column("amount", sa.Numeric(12, 2)),
        sa.Column("payment_date_raw", sa.Text),
        sa.Column("transaction_id", sa.String(length=128)),
        sa.Column("source_file", sa.Text),
    )


def validate_workbook_headers(headers: Sequence[Any], expected: Sequence[str]) -> dict[str, list[str]]:
    normalized = [str(h).strip() for h in headers if h is not None]
    mapped = [h for h in normalized if h in expected]
    follow_up = [h for h in normalized if h not in expected]
    ignored: list[str] = []
    return {"mapped": mapped, "ignored": ignored, "follow_up": follow_up}


def _non_blank_text(value: Any, *, upper: bool = False, collapse_spaces: bool = False) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if not text or text == "-":
        return None
    if collapse_spaces:
        text = re.sub(r"\s+", " ", text)
    if upper:
        text = text.upper()
    return text


def _parse_numeric(value: Any, remarks: list[str], reason_code: str) -> Decimal | None:
    if value is None:
        remarks.append(reason_code)
        return None
    text = str(value).strip()
    if not text:
        remarks.append(reason_code)
        return None
    cleaned = re.sub(r"[,₹]", "", text).replace("INR", "").strip()
    try:
        return Decimal(cleaned)
    except (InvalidOperation, ValueError):
        remarks.append(reason_code)
        return None


def _date_raw(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text if text and text != "-" else None


def _add_remarks(remarks: list[str]) -> str | None:
    if not remarks:
        return None
    deduped = sorted(set(remarks))
    return "; ".join(deduped)


def _inc_reason(counter: Counter[str], reasons: Iterable[str]) -> None:
    for reason in reasons:
        counter[reason] += 1


def _normalize_payment_mode(raw: Any, remarks: list[str]) -> str:
    token = _non_blank_text(raw, collapse_spaces=True)
    if not token:
        remarks.append("missing_payment_mode")
        return "UNKNOWN"
    upper = token.upper()
    normalized = PAYMENT_MODE_NORMALIZATION.get(upper)
    if normalized is None:
        normalized = re.sub(r"[^A-Z0-9]+", "_", upper).strip("_") or "UNKNOWN"
    if normalized != upper:
        remarks.append(f"payment_mode_normalized:{token}->{normalized}")
    return normalized


def _normalize_status(raw: Any, remarks: list[str]) -> str | None:
    token = _non_blank_text(raw, collapse_spaces=True)
    if not token:
        return None
    upper = token.upper()
    normalized = STATUS_NORMALIZATION.get(upper, re.sub(r"[^A-Z0-9]+", "_", upper).strip("_"))
    if normalized != upper:
        remarks.append(f"status_normalized:{token}->{normalized}")
    return normalized


def _generate_line_hash(row: Mapping[str, Any]) -> str:
    keys = ["store_code", "order_code", "service", "item_name", "rate", "quantity", "weight", "addons", "amount"]
    payload = "|".join("" if row.get(k) is None else str(row.get(k)) for k in keys)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _normalize_base_row(source: Mapping[str, Any], *, source_file: str, run_id: str, run_date: datetime, cost_center: str | None) -> tuple[dict[str, Any] | None, list[str], list[str]]:
    remarks: list[str] = []
    rejects: list[str] = []
    store_code = _non_blank_text(source.get("store_code"), upper=True)
    order_code = _non_blank_text(source.get("order_code"), upper=True, collapse_spaces=True)
    if not store_code:
        rejects.append(REASON_MISSING_STORE_CODE)
    if not order_code:
        rejects.append(REASON_MISSING_ORDER_CODE)
    if cost_center is None:
        remarks.append("missing_cost_center_mapping")
    phone = _non_blank_text(source.get("customer_phone"))
    if phone:
        cleaned = re.sub(r"\D", "", phone.replace("+91", ""))
        if cleaned:
            phone = cleaned
        if phone and len(phone) != 10:
            remarks.append("phone_format_warning")
    normalized = {
        "run_id": run_id,
        "run_date": run_date,
        "source_file": source_file,
        "store_code": store_code,
        "order_code": order_code,
        "pickup_raw": _date_raw(source.get("pickup")),
        "delivery_raw": _date_raw(source.get("delivery")),
        "customer_name": _non_blank_text(source.get("customer_name"), collapse_spaces=True),
        "customer_phone": phone,
        "address": _non_blank_text(source.get("address"), collapse_spaces=True),
        "payment_text": _non_blank_text(source.get("payment_text"), collapse_spaces=True),
        "instructions": _non_blank_text(source.get("instructions"), collapse_spaces=True),
        "customer_source": _non_blank_text(source.get("customer_source"), collapse_spaces=True),
        "status": _normalize_status(source.get("status"), remarks),
        "status_date_raw": _date_raw(source.get("status_date")),
        "cost_center": cost_center,
        "ingest_remarks": _add_remarks(remarks),
    }
    return (None if rejects else normalized), remarks, rejects


def _normalize_order_details_row(source: Mapping[str, Any], *, source_file: str, run_id: str, run_date: datetime, cost_center: str | None) -> tuple[dict[str, Any] | None, list[str], list[str]]:
    remarks: list[str] = []
    rejects: list[str] = []
    store_code = _non_blank_text(source.get("store_code"), upper=True)
    order_code = _non_blank_text(source.get("order_code"), upper=True, collapse_spaces=True)
    if not store_code:
        rejects.append(REASON_MISSING_STORE_CODE)
    if not order_code:
        rejects.append(REASON_MISSING_ORDER_CODE)
    service = _non_blank_text(source.get("service"), collapse_spaces=True)
    item_name = _non_blank_text(source.get("item_name"), collapse_spaces=True)
    if not service:
        remarks.append(f"{REASON_MISSING_REQUIRED_FIELD}:service")
    if not item_name:
        remarks.append(f"{REASON_MISSING_REQUIRED_FIELD}:item_name")
    if cost_center is None:
        remarks.append("missing_cost_center_mapping")
    normalized = {
        "run_id": run_id,
        "run_date": run_date,
        "source_file": source_file,
        "cost_center": cost_center,
        "store_code": store_code,
        "order_code": order_code,
        "order_mode": _non_blank_text(source.get("order_mode"), collapse_spaces=True),
        "order_datetime_raw": _date_raw(source.get("order_datetime")),
        "pickup_datetime_raw": _date_raw(source.get("pickup_datetime")),
        "delivery_datetime_raw": _date_raw(source.get("delivery_datetime")),
        "service": service,
        "hsn_sac": _non_blank_text(source.get("hsn_sac"), collapse_spaces=True),
        "item_name": item_name,
        "rate": _parse_numeric(source.get("rate"), remarks, "invalid_rate"),
        "quantity": _parse_numeric(source.get("quantity"), remarks, "invalid_quantity"),
        "weight": _parse_numeric(source.get("weight"), remarks, "invalid_weight"),
        "addons": _non_blank_text(source.get("addons"), collapse_spaces=True),
        "amount": _parse_numeric(source.get("amount"), remarks, "invalid_amount"),
    }
    if rejects:
        return None, remarks, rejects
    normalized["line_hash"] = _generate_line_hash(normalized)
    normalized["ingest_remarks"] = _add_remarks(remarks)
    if not normalized["line_hash"]:
        return None, remarks, [REASON_MISSING_LINE_HASH]
    return normalized, remarks, rejects


def _normalize_payment_row(source: Mapping[str, Any], *, source_file: str, run_id: str, run_date: datetime, cost_center: str | None) -> tuple[dict[str, Any] | None, list[str], list[str]]:
    remarks: list[str] = []
    rejects: list[str] = []
    store_code = _non_blank_text(source.get("store_code"), upper=True)
    order_code = _non_blank_text(source.get("order_code"), upper=True, collapse_spaces=True)
    if not store_code:
        rejects.append(REASON_MISSING_STORE_CODE)
    if not order_code:
        rejects.append(REASON_MISSING_ORDER_CODE)
    if cost_center is None:
        remarks.append("missing_cost_center_mapping")
    tx_id = _non_blank_text(source.get("transaction_id"), collapse_spaces=True)
    normalized = {
        "run_id": run_id,
        "run_date": run_date,
        "source_file": source_file,
        "cost_center": cost_center,
        "store_code": store_code,
        "order_code": order_code,
        "payment_mode": _normalize_payment_mode(source.get("payment_mode"), remarks),
        "amount": _parse_numeric(source.get("amount"), remarks, "invalid_payment_amount"),
        "payment_date_raw": _date_raw(source.get("payment_date")),
        "transaction_id": tx_id,
        "ingest_remarks": _add_remarks(remarks),
    }
    if normalized["payment_date_raw"] is None:
        remarks.append("missing_payment_date")
        normalized["ingest_remarks"] = _add_remarks(remarks)
    return (None if rejects else normalized), remarks, rejects


def _read_rows(path: Path, expected_headers: Sequence[str]) -> tuple[list[dict[str, Any]], dict[str, list[str]]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
    header_info = validate_workbook_headers(header_row, expected_headers)
    headers = [str(h).strip() if h is not None else "" for h in header_row]
    rows: list[dict[str, Any]] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        rows.append({headers[i]: row[i] if i < len(row) else None for i in range(len(headers)) if headers[i]})
    return rows, header_info


async def _resolve_cost_centers(database_url: str) -> dict[str, str]:
    async with session_scope(database_url) as session:
        try:
            rows = await session.execute(sa.text("SELECT store_code, cost_center FROM store_master"))
        except Exception:
            return {}
        return {str(r.store_code).strip().upper(): str(r.cost_center).strip() for r in rows if r.store_code}


def _build_upsert_statement(table: sa.Table, rows: list[dict[str, Any]], key_columns: Sequence[Any], use_sqlite: bool) -> Any:
    insert_fn = sqlite_insert if use_sqlite else pg_insert
    insert = insert_fn(table).values(rows)
    set_cols = {col.name: insert.excluded[col.name] for col in table.columns if col.name != "id"}
    return insert.on_conflict_do_update(index_elements=list(key_columns), set_=set_cols)


def _key_tuple(row: Mapping[str, Any], keys: Sequence[str]) -> tuple[Any, ...]:
    return tuple(("" if row.get(k) is None else row.get(k)) for k in keys)


async def _ingest_file(
    *,
    database_url: str,
    file_kind: str,
    source_path: Path,
    run_id: str,
    run_date: datetime,
    store_code: str | None,
    explicit_cost_center: str | None,
    cost_center_lookup: Mapping[str, str],
    logger: JsonLogger,
) -> tuple[FileIngestResult, list[RejectRecord]]:
    result = FileIngestResult()
    rejects: list[RejectRecord] = []
    rows, header_info = _read_rows(source_path, EXPECTED_HEADERS[file_kind])
    if header_info["follow_up"]:
        log_event(logger, "warning", "archive_ingest_header_follow_up", file=str(source_path), headers=header_info["follow_up"])
        result.warnings += len(header_info["follow_up"])
    use_store = (store_code or "").strip().upper() or None

    normalized_rows: list[dict[str, Any]] = []
    reject_counter: Counter[str] = Counter()
    for idx, row in enumerate(rows, start=2):
        result.parsed += 1
        row_store = _non_blank_text(row.get("store_code"), upper=True) or use_store
        if row_store:
            row["store_code"] = row_store
        resolved_cc = explicit_cost_center or (cost_center_lookup.get(row_store) if row_store else None)
        if file_kind == FILE_BASE:
            normalized, remarks, row_rejects = _normalize_base_row(row, source_file=source_path.name, run_id=run_id, run_date=run_date, cost_center=resolved_cc)
        elif file_kind == FILE_ORDER_DETAILS:
            normalized, remarks, row_rejects = _normalize_order_details_row(row, source_file=source_path.name, run_id=run_id, run_date=run_date, cost_center=resolved_cc)
        else:
            normalized, remarks, row_rejects = _normalize_payment_row(row, source_file=source_path.name, run_id=run_id, run_date=run_date, cost_center=resolved_cc)
        result.warnings += len(remarks)
        if row_rejects:
            result.rejected += 1
            _inc_reason(reject_counter, row_rejects)
            rejects.append(RejectRecord(source_file=source_path.name, row_number=idx, reason_codes=sorted(set(row_rejects))))
            continue
        assert normalized is not None
        normalized_rows.append(normalized)

    dedupe_keys_map = {
        FILE_BASE: ["store_code", "order_code"],
        FILE_ORDER_DETAILS: ["store_code", "order_code", "line_hash"],
        FILE_PAYMENT_DETAILS: ["store_code", "order_code", "payment_date_raw", "payment_mode", "amount", "transaction_id"],
    }
    unique_map: dict[tuple[Any, ...], dict[str, Any]] = {}
    for row in normalized_rows:
        key = _key_tuple(row, dedupe_keys_map[file_kind])
        unique_map[key] = row
    upsert_rows = list(unique_map.values())

    if not upsert_rows:
        result.reject_reasons = dict(reject_counter)
        return result, rejects

    async with session_scope(database_url) as session:
        bind = session.bind
        assert isinstance(bind, sa.ext.asyncio.AsyncEngine)
        metadata = sa.MetaData()
        if file_kind == FILE_BASE:
                table = _stg_uc_archive_orders_base_table(metadata)
                key_cols: Sequence[Any] = [table.c.store_code, table.c.order_code]
                key_names = ["store_code", "order_code"]
        elif file_kind == FILE_ORDER_DETAILS:
                table = _stg_uc_archive_order_details_table(metadata)
                key_cols = [table.c.store_code, table.c.order_code, table.c.line_hash]
                key_names = ["store_code", "order_code", "line_hash"]
        else:
                table = _stg_uc_archive_payment_details_table(metadata)
                key_cols = [
                    table.c.store_code,
                    table.c.order_code,
                    table.c.payment_date_raw,
                    table.c.payment_mode,
                    table.c.amount,
                    sa.text("coalesce(transaction_id, '')"),
                ]
                key_names = ["store_code", "order_code", "payment_date_raw", "payment_mode", "amount", "transaction_id"]

        where_clause = sa.or_(
            *[
                sa.and_(
                    *[
                        (sa.func.coalesce(getattr(table.c, c), "") == ("" if row.get(c) is None else row.get(c)))
                        for c in key_names
                        if c in table.c
                    ]
                )
                for row in upsert_rows
            ]
        )
        existing_rows = (await session.execute(sa.select(table).where(where_clause))).mappings().all()
        existing_keys = {_key_tuple(r, key_names) for r in existing_rows}
        use_sqlite = bind.url.get_backend_name().startswith("sqlite")
        stmt = _build_upsert_statement(table, upsert_rows, key_cols, use_sqlite=use_sqlite)
        await session.execute(stmt)
        await session.commit()

    for row in upsert_rows:
        if _key_tuple(row, key_names) in existing_keys:
            result.updated += 1
        else:
            result.inserted += 1
    result.reject_reasons = dict(reject_counter)
    return result, rejects


async def ingest_uc_archive_excels(
    *,
    database_url: str,
    run_id: str,
    run_date: datetime,
    store_code: str | None,
    cost_center: str | None,
    base_order_info_path: str | Path,
    order_details_path: str | Path,
    payment_details_path: str | Path,
    logger: JsonLogger,
    cost_center_resolver: Callable[[str], str | None] | None = None,
) -> ArchiveIngestResult:
    paths = {
        FILE_BASE: Path(base_order_info_path),
        FILE_ORDER_DETAILS: Path(order_details_path),
        FILE_PAYMENT_DETAILS: Path(payment_details_path),
    }
    cc_lookup = await _resolve_cost_centers(database_url)
    if cost_center_resolver is not None:
        for sc in list(cc_lookup.keys()):
            resolved = cost_center_resolver(sc)
            if resolved is not None:
                cc_lookup[sc] = resolved

    summary = ArchiveIngestResult(files={})
    for file_kind in [FILE_BASE, FILE_ORDER_DETAILS, FILE_PAYMENT_DETAILS]:
        file_result, file_rejects = await _ingest_file(
            database_url=database_url,
            file_kind=file_kind,
            source_path=paths[file_kind],
            run_id=run_id,
            run_date=run_date,
            store_code=store_code,
            explicit_cost_center=cost_center,
            cost_center_lookup=cc_lookup,
            logger=logger,
        )
        summary.files[file_kind] = file_result
        summary.rejects.extend(file_rejects)
    return summary


def _orders_table(metadata: sa.MetaData) -> sa.Table:
    return sa.Table(
        "orders",
        metadata,
        sa.Column("id", sa.BigInteger().with_variant(sa.Integer(), "sqlite"), primary_key=True, autoincrement=True),
        sa.Column("run_id", sa.Text()),
        sa.Column("run_date", sa.DateTime(timezone=True)),
        sa.Column("cost_center", sa.String(length=8)),
        sa.Column("store_code", sa.String(length=8)),
        sa.Column("order_number", sa.String(length=24)),
        sa.Column("order_date", sa.DateTime(timezone=True), nullable=False),
        sa.Column("customer_name", sa.String(length=128)),
        sa.Column("customer_address", sa.String(length=256)),
        sa.Column("mobile_number", sa.String(length=16)),
        sa.Column("service_type", sa.String(length=64)),
        sa.Column("pieces", sa.Numeric(12, 0)),
        sa.Column("weight", sa.Numeric(12, 2)),
    )


def _sales_table(metadata: sa.MetaData) -> sa.Table:
    return sa.Table(
        "sales",
        metadata,
        sa.Column("id", sa.BigInteger().with_variant(sa.Integer(), "sqlite"), primary_key=True, autoincrement=True),
        sa.Column("run_id", sa.Text()),
        sa.Column("run_date", sa.DateTime(timezone=True)),
        sa.Column("cost_center", sa.String(length=8)),
        sa.Column("store_code", sa.String(length=16)),
        sa.Column("order_date", sa.DateTime(timezone=True)),
        sa.Column("payment_date", sa.DateTime(timezone=True)),
        sa.Column("order_number", sa.String(length=16)),
        sa.Column("customer_code", sa.String(length=16)),
        sa.Column("customer_name", sa.String(length=128)),
        sa.Column("customer_address", sa.String(length=256)),
        sa.Column("mobile_number", sa.String(length=16)),
        sa.Column("payment_received", sa.Numeric(12, 2)),
        sa.Column("adjustments", sa.Numeric(12, 2)),
        sa.Column("balance", sa.Numeric(12, 2)),
        sa.Column("accepted_by", sa.String(length=64)),
        sa.Column("payment_mode", sa.String(length=32)),
        sa.Column("transaction_id", sa.String(length=64)),
        sa.Column("payment_made_at", sa.String(length=128)),
        sa.Column("order_type", sa.String(length=32)),
        sa.Column("is_duplicate", sa.Boolean()),
        sa.Column("is_edited_order", sa.Boolean()),
        sa.Column("ingest_remarks", sa.Text()),
    )


def _parse_payment_datetime(raw: Any) -> datetime | None:
    if raw in (None, ""):
        return None
    try:
        tz = get_timezone()
    except Exception:
        tz = timezone.utc
    if isinstance(raw, datetime):
        parsed = raw
    else:
        try:
            parsed = parser.parse(str(raw), dayfirst=True)
        except Exception:
            return None
    if parsed.tzinfo is None:
        return parsed.replace(tzinfo=tz)
    return parsed.astimezone(tz)


async def publish_uc_archive_to_orders_and_sales(
    *,
    database_url: str,
    logger: JsonLogger,
    run_id: str | None = None,
) -> ArchivePublishResult:
    result = ArchivePublishResult()
    async with session_scope(database_url) as session:
        bind = session.bind
        assert isinstance(bind, sa.ext.asyncio.AsyncEngine)
        metadata = sa.MetaData()
        orders_table = _orders_table(metadata)
        sales_table = _sales_table(metadata)
        details_table = _stg_uc_archive_order_details_table(metadata)
        payments_table = _stg_uc_archive_payment_details_table(metadata)

        details_stmt = sa.select(
            details_table.c.store_code,
            details_table.c.order_code,
            details_table.c.quantity,
            details_table.c.weight,
            details_table.c.service,
        )
        if run_id:
            details_stmt = details_stmt.where(details_table.c.run_id == run_id)
        details_rows = (await session.execute(details_stmt)).mappings().all()
        aggregates: dict[tuple[str, str], dict[str, Any]] = {}
        for row in details_rows:
            store_code = str(row.store_code or "").strip().upper()
            order_code = str(row.order_code or "").strip().upper()
            if not store_code or not order_code:
                continue
            key = (store_code, order_code)
            bucket = aggregates.setdefault(key, {"pieces": Decimal("0"), "weight": Decimal("0"), "services": set()})
            bucket["pieces"] += Decimal(str(row.quantity or 0))
            bucket["weight"] += Decimal(str(row.weight or 0))
            service = _non_blank_text(row.service, collapse_spaces=True)
            if service:
                bucket["services"].add(service)

        for (store_code, order_code), agg in aggregates.items():
            service_type = ", ".join(sorted(agg["services"])) if agg["services"] else None
            update_stmt = (
                sa.update(orders_table)
                .where(
                    orders_table.c.store_code == store_code,
                    orders_table.c.order_number == order_code,
                )
                .values(pieces=agg["pieces"], weight=agg["weight"], service_type=service_type)
            )
            update_result = await session.execute(update_stmt)
            result.orders_updated += int(update_result.rowcount or 0)

        context_rows = (
            await session.execute(
                sa.select(
                    orders_table.c.store_code,
                    orders_table.c.order_number,
                    orders_table.c.cost_center,
                    orders_table.c.order_date,
                    orders_table.c.customer_name,
                    orders_table.c.customer_address,
                    orders_table.c.mobile_number,
                )
            )
        ).mappings().all()
        order_context: dict[tuple[str, str], Mapping[str, Any]] = {}
        for ctx in context_rows:
            key = (str(ctx.store_code or "").strip().upper(), str(ctx.order_number or "").strip().upper())
            if not key[0] or not key[1]:
                continue
            current = order_context.get(key)
            if current is None or (ctx.order_date and (current.get("order_date") is None or ctx.order_date > current.get("order_date"))):
                order_context[key] = ctx

        payment_stmt = sa.select(payments_table)
        if run_id:
            payment_stmt = payment_stmt.where(payments_table.c.run_id == run_id)
        payment_rows = (await session.execute(payment_stmt)).mappings().all()

        deduped_sales: dict[tuple[str, str, datetime], dict[str, Any]] = {}
        for row in payment_rows:
            store_code = str(row.store_code or "").strip().upper()
            order_code = str(row.order_code or "").strip().upper()
            ctx = order_context.get((store_code, order_code))
            if ctx is None:
                result.skipped += 1
                result.skip_reasons["missing_order_context"] = result.skip_reasons.get("missing_order_context", 0) + 1
                log_event(logger=logger, phase="archive_publish_skip", status="warning", reason="missing_order_context", store_code=store_code, order_code=order_code)
                continue
            payment_date = _parse_payment_datetime(row.payment_date_raw)
            if payment_date is None:
                result.skipped += 1
                result.skip_reasons["payment_date_parse_failure"] = result.skip_reasons.get("payment_date_parse_failure", 0) + 1
                log_event(
                    logger=logger,
                    phase="archive_publish_skip",
                    status="warning",
                    reason="payment_date_parse_failure",
                    store_code=store_code,
                    order_code=order_code,
                    payment_date_raw=row.payment_date_raw,
                )
                continue
            if not ctx.get("cost_center") or not ctx.get("order_date"):
                result.skipped += 1
                result.skip_reasons["missing_required_context_fields"] = result.skip_reasons.get("missing_required_context_fields", 0) + 1
                log_event(logger=logger, phase="archive_publish_skip", status="warning", reason="missing_required_context_fields", store_code=store_code, order_code=order_code)
                continue
            key = (str(ctx["cost_center"]), order_code, payment_date)
            if key in deduped_sales:
                result.skipped += 1
                result.skip_reasons["key_conflict"] = result.skip_reasons.get("key_conflict", 0) + 1
                log_event(logger=logger, phase="archive_publish_skip", status="warning", reason="key_conflict", store_code=store_code, order_code=order_code, payment_date=payment_date.isoformat())
            deduped_sales[key] = {
                "run_id": row.run_id,
                "run_date": row.run_date,
                "cost_center": ctx["cost_center"],
                "store_code": store_code,
                "order_date": ctx["order_date"],
                "payment_date": payment_date,
                "order_number": order_code,
                "customer_name": ctx.get("customer_name"),
                "customer_address": ctx.get("customer_address"),
                "mobile_number": ctx.get("mobile_number"),
                "payment_received": row.amount,
                "payment_mode": row.payment_mode,
                "transaction_id": row.transaction_id,
                "payment_made_at": row.payment_date_raw,
            }

        upsert_rows = list(deduped_sales.values())
        if upsert_rows:
            key_names = ["cost_center", "order_number", "payment_date"]
            where_clause = sa.or_(
                *[
                    sa.and_(
                        sales_table.c.cost_center == row["cost_center"],
                        sales_table.c.order_number == row["order_number"],
                        sales_table.c.payment_date == row["payment_date"],
                    )
                    for row in upsert_rows
                ]
            )
            existing_rows = (await session.execute(sa.select(sales_table).where(where_clause))).mappings().all()
            existing_keys = {_key_tuple(r, key_names) for r in existing_rows}
            use_sqlite = bind.url.get_backend_name().startswith("sqlite")
            stmt = _build_upsert_statement(
                sales_table,
                upsert_rows,
                [sales_table.c.cost_center, sales_table.c.order_number, sales_table.c.payment_date],
                use_sqlite=use_sqlite,
            )
            await session.execute(stmt)
            for row in upsert_rows:
                if _key_tuple(row, key_names) in existing_keys:
                    result.sales_updated += 1
                else:
                    result.sales_inserted += 1

        await session.commit()

    return result
